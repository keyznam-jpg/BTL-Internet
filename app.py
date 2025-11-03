# -*- coding: utf-8 -*-
from datetime import datetime, date, timedelta
import os
import math
import json
import re
import smtplib
import hashlib
from functools import wraps
from urllib.parse import quote, urlencode, urljoin, urlparse
import uuid # Library to create unique tokens
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, login_user, login_required, logout_user, UserMixin, current_user, AnonymousUserMixin
)
from dotenv import load_dotenv
from sqlalchemy import func, extract, inspect, text, or_, case
from sqlalchemy.orm import joinedload
from collections import defaultdict
import calendar
import pandas as pd
import io
import unicodedata
import qrcode
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Import SocketIO and necessary functions
from flask_socketio import SocketIO, join_room
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
from werkzeug.security import generate_password_hash, check_password_hash

from flask_caching import Cache
from flask_migrate import Migrate
from flask_compress import Compress
from authlib.integrations.flask_client import OAuth

# Initialize Cache
# cache = Cache(app, config={'CACHE_TYPE': 'simple'})  # Use 'redis' if Redis available  # Moved below

# Load environment variables from the .env file
load_dotenv()

# Default payment session timeout (minutes)
DEFAULT_PAYMENT_TIMEOUT_MINUTES = 5

def build_mysql_uri():
    """Build the MySQL connection string from environment variables."""
    host = os.getenv("MYSQL_HOST", "localhost")
    port = os.getenv("MYSQL_PORT", "3306")
    user = os.getenv("MYSQL_USER", "root")
    pw   = os.getenv("MYSQL_PASSWORD", "")
    db   = os.getenv("MYSQL_DB", "khachsan")
    return f"mysql+pymysql://{user}:{pw}@{host}:{port}/{db}?charset=utf8mb4"

# Initialize Flask app
app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "khachsan-secret")
app.config["SQLALCHEMY_DATABASE_URI"] = build_mysql_uri()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_POOL_SIZE"] = 10
app.config["SQLALCHEMY_MAX_OVERFLOW"] = 20
app.config["SQLALCHEMY_POOL_RECYCLE"] = 3600
app.config['JSON_AS_ASCII'] = False  # Ensure UTF-8 encoding for JSON responses
app.config["AVATAR_UPLOAD_FOLDER"] = os.path.join(app.root_path, "static", "uploads", "avatars")
app.config["CHAT_UPLOAD_FOLDER"] = os.path.join(app.root_path, "static", "uploads", "chat")
app.config.setdefault("MAX_CONTENT_LENGTH", 2 * 1024 * 1024)
public_base_url = os.getenv("PUBLIC_BASE_URL")
if public_base_url and public_base_url.strip():
    app.config["PUBLIC_BASE_URL"] = public_base_url.strip().rstrip("/")
else:
    app.config["PUBLIC_BASE_URL"] = None
support_email = os.getenv("SUPPORT_EMAIL", "").strip()
app.config["SUPPORT_EMAIL"] = support_email or None

google_client_id = os.getenv("GOOGLE_CLIENT_ID", "").strip()
google_client_secret = os.getenv("GOOGLE_CLIENT_SECRET", "").strip()
app.config["GOOGLE_CLIENT_ID"] = google_client_id or None
app.config["GOOGLE_CLIENT_SECRET"] = google_client_secret or None
app.config["GOOGLE_OAUTH_ENABLED"] = bool(app.config["GOOGLE_CLIENT_ID"] and app.config["GOOGLE_CLIENT_SECRET"])

app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

# Initialize Cache
cache = Cache(app, config={'CACHE_TYPE': 'simple'})  # Use 'redis' if Redis available

# Disable caching in development
if app.debug:
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    cache.init_app(app, config={'CACHE_TYPE': 'null'})

# Initialize Compress
compress = Compress(app)

oauth = OAuth(app)
if app.config["GOOGLE_OAUTH_ENABLED"]:
    oauth.register(
        name="google",
        client_id=app.config["GOOGLE_CLIENT_ID"],
        client_secret=app.config["GOOGLE_CLIENT_SECRET"],
        server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
        client_kwargs={
            "scope": "openid email profile",
        },
    )
    app.logger.info("Google OAuth is enabled and configured.")
else:
    app.logger.info("Google OAuth is disabled; missing GOOGLE_CLIENT_ID/GOOGLE_CLIENT_SECRET.")

PAYMENT_METHODS = [
    ('cash', 'Tiền mặt'),
    ('qr', 'Chuyển khoản QR (VietQR)')
]
PAYMENT_METHOD_LABELS = {code: label for code, label in PAYMENT_METHODS}

PERMISSION_GROUPS = [
    (
        'dashboard',
        'Tổng quan',
        [
            ('dashboard.view', 'Xem bảng điều khiển & chỉ số nhanh'),
        ],
    ),
    (
        'bookings',
        'Đặt phòng & lịch phòng',
        [
            ('bookings.view_map', 'Xem sơ đồ phòng và trạng thái'),
            ('bookings.create_offline', 'Tạo đặt phòng tại quầy'),
            ('bookings.manage_online', 'Duyệt & xử lý đặt phòng online'),
            ('bookings.manage_waiting', 'Quản lý danh sách chờ & sắp xếp phòng'),
            ('bookings.checkin_checkout', 'Thực hiện nhận/trả phòng'),
            ('bookings.cancel', 'Hủy đặt phòng & giải phóng phòng'),
            ('bookings.extend', 'Gia hạn thời gian thuê phòng'),
        ],
    ),
    (
        'payments',
        'Thanh toán & hóa đơn',
        [
            ('payments.process', 'Xử lý thanh toán, đặt cọc, dịch vụ'),
            ('payments.invoices', 'Xem/in/gửi hóa đơn'),
            ('payments.export', 'Xuất báo cáo thanh toán'),
            ('payments.configure_timeout', 'Thiết lập thời hạn & phiên thanh toán'),
        ],
    ),
    (
        'customers',
        'Khách hàng & ưu đãi',
        [
            ('customers.view', 'Quản lý hồ sơ khách hàng'),
            ('customers.export', 'Xuất dữ liệu khách hàng'),
            ('customers.manage_accounts', 'Xóa hoặc vô hiệu hóa tài khoản khách hàng'),
            ('customers.vouchers', 'Cấu hình voucher & khuyến mãi'),
        ],
    ),
    (
        'services',
        'Dịch vụ & tài nguyên',
        [
            ('services.manage', 'Quản lý danh mục dịch vụ'),
            ('services.orders', 'Quản lý sử dụng dịch vụ & thanh toán'),
            ('room_types.manage', 'Quản lý loại phòng'),
            ('rooms.manage', 'Quản lý phòng (thêm/sửa/xóa)'),
        ],
    ),
    (
        'communications',
        'Liên lạc & email',
        [
            ('communications.chat', 'Xem & phản hồi tin nhắn khách hàng'),
            ('chat.delete', 'Xóa hội thoại khách hàng'),
            ('email.settings', 'Cấu hình email'),
            ('email.logs', 'Xem nhật ký email'),
        ],
    ),
    (
        'attendance',
        'Chấm công',
        [
            ('attendance.manage', 'Phê duyệt và báo cáo chấm công'),
        ],
    ),
    (
        'staff',
        'Nhân sự & lương thưởng',
        [
            ('staff.manage', 'Quản lý hồ sơ nhân viên'),
            ('payroll.configure', 'Cài đặt lương thưởng'),
        ],
    ),
    (
        'analytics',
        'Thống kê & báo cáo',
        [
            ('analytics.revenue', 'Xem và xuất báo cáo doanh thu'),
        ],
    ),
    (
        'operations',
        'Vận hành hệ thống',
        [
            ('roles.manage', 'Quản lý vai trò & phân quyền'),
            ('system.maintenance', 'Tác vụ bảo trì hệ thống'),
        ],
    ),

]
PERMISSION_META = {
    key: {
        'label': label,
        'group': group_key,
        'group_label': group_label,
    }
    for group_key, group_label, entries in PERMISSION_GROUPS
    for key, label in entries
}
ALL_PERMISSION_KEYS = set(PERMISSION_META.keys())
DEFAULT_ROLE_PERMISSIONS = {
    'admin': set(ALL_PERMISSION_KEYS),
    'nhanvien': set(),
}


def set_role_permissions(role, permission_keys):
    desired = set(permission_keys) & ALL_PERMISSION_KEYS
    current = {rp.permission for rp in role.permissions}
    changed = False
    for rp in list(role.permissions):
        if rp.permission not in desired:
            db.session.delete(rp)
            changed = True
    for perm in desired - current:
        role.permissions.append(RolePermission(permission=perm))
        changed = True
    return changed


def set_user_permissions(user, permission_keys):
    desired = set(permission_keys) & ALL_PERMISSION_KEYS
    current = {up.permission for up in user.personal_permissions}
    changed = False
    for up in list(user.personal_permissions):
        if up.permission not in desired:
            db.session.delete(up)
            changed = True
    for perm in desired - current:
        user.personal_permissions.append(UserPermission(permission=perm))
        changed = True
    return changed


def permission_required(*permissions):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Bạn cần đăng nhập để tiếp tục.', 'warning')
                return redirect(url_for('login'))
            if current_user.role and current_user.role.is_system:
                return fn(*args, **kwargs)
            if current_user.loai == 'admin':
                return fn(*args, **kwargs)
            if getattr(current_user, 'role_slug', None) == 'admin':
                return fn(*args, **kwargs)
            if not permissions:
                return fn(*args, **kwargs)
            if any(current_user.has_permission(perm) for perm in permissions):
                return fn(*args, **kwargs)
            support_email = app.config.get('SUPPORT_EMAIL')
            return render_template(
                'access_pending.html',
                contact_email=support_email,
                required_permissions=permissions
            ), 403
        return wrapper
    return deco


def customer_login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or not getattr(current_user, 'is_customer', False):
            flash('Vui lòng đăng nhập tài khoản khách hàng để tiếp tục.', 'warning')
            return redirect(url_for('khach_hang_dang_nhap', next=request.url))
        status = getattr(current_user, 'trang_thai_tai_khoan', 'hoat_dong')
        if status in {'khoa', 'da_xoa'}:
            if status == 'da_xoa':
                flash('Tai khoan cua ban da bi xoa. Vui long lien he ho tro hoac dang ky lai.', 'danger')
            else:
                flash('Tai khoan cua ban dang bi khoa. Vui long lien he ho tro.', 'danger')
            logout_user()
            return redirect(url_for('khach_hang_dang_nhap'))
        return fn(*args, **kwargs)
    return wrapper


def generate_role_slug(name):
    base = unicodedata.normalize('NFKD', (name or '').strip())
    base = base.encode('ascii', 'ignore').decode('ascii').lower()
    base = re.sub(r'[^a-z0-9]+', '-', base).strip('-')
    if not base:
        base = 'role'
    slug = base
    index = 1
    while Role.query.filter_by(slug=slug).first():
        slug = f"{base}-{index}"
        index += 1
    return slug


def ensure_default_roles():
    changed = False
    admin_role = Role.query.filter_by(slug='admin').first()
    if not admin_role:
        admin_role = Role(
            name='Quản trị viên',
            slug='admin',
            description='Toàn quyền quản trị hệ thống',
            is_system=True,
        )
        db.session.add(admin_role)
        changed = True
    elif not admin_role.is_system:
        admin_role.is_system = True
        changed = True
    staff_role = Role.query.filter_by(slug='nhanvien').first()
    if not staff_role:
        staff_role = Role(
            name='Nhân viên',
            slug='nhanvien',
            description='Vai trò mặc định cho nhân viên',
            is_system=False,
        )
        db.session.add(staff_role)
        changed = True
    db.session.flush()

    # Chỉ reset quyền admin, KHÔNG reset quyền các vai trò khác để giữ cấu hình tùy chỉnh
    if set_role_permissions(admin_role, DEFAULT_ROLE_PERMISSIONS['admin']):
        changed = True
    
    # KHÔNG gọi set_role_permissions cho staff_role để giữ nguyên quyền đã cấu hình
    # if set_role_permissions(staff_role, DEFAULT_ROLE_PERMISSIONS['nhanvien']):
    #     changed = True

    roles = {role.slug: role for role in Role.query.all()}
    default_role = roles.get('nhanvien') or staff_role
    updated_users = False
    for user in NguoiDung.query.all():
        slug = (user.loai or '').strip() or 'nhanvien'
        role = roles.get(slug, default_role)
        if user.role_id != role.id:
            user.role_id = role.id
            updated_users = True
        if user.loai != role.slug:
            user.loai = role.slug
            updated_users = True
    if changed or updated_users:
        db.session.commit()

VIETQR_BANK_ID = os.getenv('VIETQR_BANK_ID', '970423')
VIETQR_ACCOUNT_NO = os.getenv('VIETQR_ACCOUNT_NO', '99992162001')
VIETQR_BANK_NAME = os.getenv('VIETQR_BANK_NAME', 'TPBank')
VIETQR_ACCOUNT_NAME = os.getenv('VIETQR_ACCOUNT_NAME', 'Khách sạn PTIT')
DEPOSIT_PERCENT = float(os.getenv('DEPOSIT_PERCENT', '0.3'))
BOOKING_BLOCKING_STATUSES = ('dat', 'nhan', 'cho_xac_nhan', 'waiting')
BOOKING_STATUS_USER_MESSAGES = {
    'cho_xac_nhan': {
        'title': 'Đang chờ xác nhận',
        'message': 'Nhân viên đang kiểm tra khoản cọc của bạn. Chúng tôi sẽ thông báo ngay khi có kết quả.',
        'level': 'info'
    },
    'dat': {
        'title': 'Đã xác nhận',
        'message': 'Tiền cọc đã được xác nhận. Khách sạn sẽ giữ phòng cho bạn đến thời gian nhận phòng.',
        'level': 'success'
    },
    'nhan': {
        'title': 'Đã nhận phòng',
        'message': 'Bạn đã hoàn tất thủ tục nhận phòng. Chúc bạn có kỳ nghỉ tuyệt vời!',
        'level': 'success'
    },
    'da_thanh_toan': {
        'title': 'Đã thanh toán',
        'message': 'Đơn đặt phòng đã được thanh toán đầy đủ. Cảm ơn bạn đã tin tưởng khách sạn.',
        'level': 'success'
    },
    'huy': {
        'title': 'Đơn đặt phòng bị từ chối',
        'message': 'Rất tiếc chúng tôi không thể xác nhận đơn đặt phòng này. Vui lòng liên hệ lễ tân nếu cần hỗ trợ thêm.',
        'level': 'danger'
    },
    'waiting': {
        'title': 'Đang chờ',
        'message': 'Đơn đặt phòng của bạn đang trong danh sách chờ. Chúng tôi sẽ liên hệ khi phòng trống.',
        'level': 'warning'
    }
}
# Thông điệp dùng cho yêu cầu xác nhận cọc online
ONLINE_DEPOSIT_REQUEST_MESSAGE = 'Khách đã thanh toán tiền cọc online, vui lòng xác nhận.'
CUSTOMER_PENDING_CONFIRMATION_MESSAGE = (
    'Đang chờ xác nhận. Nhân viên đang kiểm tra khoản cọc của bạn. '
    'Chúng tôi sẽ thông báo ngay khi có kết quả.'
)
LOYALTY_POINT_RATE_VND = 100000  # 1 điểm cho mỗi 100.000đ đã thanh toán
LOYALTY_PERCENT_PER_POINT = 0.1  # 1 điểm = giảm 0.1%
LOYALTY_MAX_DISCOUNT_PERCENT = 40.0

# ==== CẤU HÌNH VOUCHER TOÀN CỤC ====
@cache.memoize(timeout=300)  # Cache voucher config for 5 minutes
def get_voucher_config():
    discount = HeThongCauHinh.query.filter_by(key='voucher_discount').first()
    expires = HeThongCauHinh.query.filter_by(key='voucher_expires').first()
    discount_percent = float(discount.value) if discount and discount.value else 10.0
    expires_days = int(expires.value) if expires and expires.value else 60
    return discount_percent, expires_days

# ==== ROUTE CÀI ĐẶT VOUCHER ====
@app.route('/cai-dat-voucher', methods=['POST'])
@login_required
@permission_required('customers.vouchers')
def cai_dat_voucher():
    try:
        discount_percent = float(request.form.get('discount_percent', 10))
        expires_days = int(request.form.get('expires_days', 60))
        if discount_percent < 0 or discount_percent > 100:
            raise ValueError("discount_percent_range")
        # Lưu vào bảng cấu hình
        for key, value in [('voucher_discount', discount_percent), ('voucher_expires', expires_days)]:
            cauhinh = HeThongCauHinh.query.filter_by(key=key).first()
            if cauhinh:
                cauhinh.value = str(value)
            else:
                cauhinh = HeThongCauHinh(key=key, value=str(value))
                db.session.add(cauhinh)
        db.session.commit()
        cache.delete_memoized(get_voucher_config)
        flash('Đã cập nhật cấu hình voucher. Các mã hiện có giữ nguyên thông tin; voucher mới sẽ dùng cấu hình mới.', 'success')
    except Exception:
        flash('Dữ liệu không hợp lệ!', 'danger')
    return redirect(url_for('quan_li_dich_vu'))
ALLOWED_AVATAR_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
os.makedirs(app.config["AVATAR_UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["CHAT_UPLOAD_FOLDER"], exist_ok=True)

ALLOWED_CHAT_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp", "pdf", "doc", "docx", "xls", "xlsx", "csv", "mp4", "mp3", "wav", "txt"}

BANK_ID = "970423"
BANK_ACCOUNT_NO = "99992162001"

# Initialize SocketIO
# Force threading mode because eventlet currently breaks on Python 3.13
socketio = SocketIO(app, async_mode="threading")

# Initialize SQLAlchemy and Flask-Login
db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = "login"


class AnonymousUser(AnonymousUserMixin):
    @property
    def is_customer(self):
        return False

    @property
    def is_staff(self):
        return False

    @property
    def avatar_path(self):
        return "img/ttcn.png"


login_manager.anonymous_user = AnonymousUser

# ========================= DATABASE MODELS =========================
class Role(db.Model):
    __tablename__ = "role"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    slug = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.String(255))
    is_system = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    users = db.relationship("NguoiDung", backref="role", lazy=True)
    permissions = db.relationship(
        "RolePermission",
        backref="role",
        lazy=True,
        cascade="all, delete-orphan"
    )

    def __repr__(self):
        return f"<Role {self.slug}>"


class RolePermission(db.Model):
    __tablename__ = "role_permission"
    id = db.Column(db.Integer, primary_key=True)
    role_id = db.Column(db.Integer, db.ForeignKey("role.id", ondelete="CASCADE"), nullable=False)
    permission = db.Column(db.String(100), nullable=False)

    __table_args__ = (
        db.UniqueConstraint("role_id", "permission", name="uq_role_permission"),
    )

    def __repr__(self):
        return f"<RolePermission role={self.role_id} perm={self.permission}>"


class UserPermission(db.Model):
    __tablename__ = "user_permission"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("nguoidung.id", ondelete="CASCADE"), nullable=False)
    permission = db.Column(db.String(100), nullable=False)

    __table_args__ = (
        db.UniqueConstraint("user_id", "permission", name="uq_user_permission"),
    )

    def __repr__(self):
        return f"<UserPermission user={self.user_id} perm={self.permission}>"


class NguoiDung(UserMixin, db.Model):
    __tablename__ = "nguoidung"
    id = db.Column(db.Integer, primary_key=True)
    ten_dang_nhap = db.Column(db.String(50), unique=True, nullable=False)
    mat_khau = db.Column(db.String(128), nullable=False)
    loai = db.Column(db.String(20), default="nhanvien")
    role_id = db.Column(db.Integer, db.ForeignKey("role.id"))
    ten  = db.Column(db.String(100), nullable=False)
    ngay_vao_lam = db.Column(db.Date)
    anh_dai_dien = db.Column(db.String(255))
    personal_permissions = db.relationship(
        "UserPermission",
        backref="user",
        lazy=True,
        cascade="all, delete-orphan"
    )
    def get_id(self): return str(self.id)

    @property
    def avatar_path(self):
        if self.anh_dai_dien:
            return self.anh_dai_dien.replace('\\', '/')
        return "img/ttcn.png"

    @property
    def role_slug(self):
        if self.role and self.role.slug:
            return self.role.slug
        return self.loai

    @property
    def role_name(self):
        if self.role and self.role.name:
            return self.role.name
        default_labels = {
            'admin': 'Quản trị viên',
            'nhanvien': 'Nhân viên',
        }
        if not self.loai:
            return ''
        return default_labels.get(self.loai, self.loai.replace('_', ' ').title())

    @property
    def is_customer(self):
        return False

    @property
    def is_staff(self):
        return True

    def has_permission(self, permission_key):
        if not permission_key:
            return True
        # System/admin roles always permitted
        if self.role and self.role.is_system:
            return True
        if self.loai == 'admin':
            return True
        if any(up.permission == permission_key for up in self.personal_permissions):
            return True
        if not self.role:
            return False
        return any(rp.permission == permission_key for rp in self.role.permissions)


class LoaiPhong(db.Model):
    __tablename__ = "loaiphong"
    id = db.Column(db.Integer, primary_key=True)
    ten = db.Column(db.String(100), nullable=False)
    so_nguoi_toi_da = db.Column(db.Integer, default=2)
    gia  = db.Column(db.BIGINT, default=0)
    mo_ta = db.Column(db.Text, nullable=True)
    co_voucher = db.Column(db.Boolean, default=False)
    phongs = db.relationship("Phong", backref="loai", lazy=True)

class Phong(db.Model):
    __tablename__ = "phong"
    id = db.Column(db.Integer, primary_key=True)
    ten = db.Column(db.String(50), unique=True, nullable=False)
    trang_thai = db.Column(db.String(20), default="trong")
    loai_id = db.Column(db.Integer, db.ForeignKey("loaiphong.id"), nullable=False)

class KhachHang(db.Model):
    __tablename__ = "khachhang"
    id = db.Column(db.Integer, primary_key=True)
    ho_ten = db.Column(db.String(100), nullable=False)
    cmnd   = db.Column(db.String(30), unique=True, nullable=False)
    sdt    = db.Column(db.String(30))
    email  = db.Column(db.String(120))
    dia_chi = db.Column(db.String(200))
    mat_khau_hash = db.Column(db.String(255))
    diem_tich_luy = db.Column(db.Integer, default=0)
    ngay_dang_ky = db.Column(db.DateTime)
    ngay_cap_nhat = db.Column(db.DateTime)
    lan_dang_nhap_cuoi = db.Column(db.DateTime)
    trang_thai_tai_khoan = db.Column(db.String(20), default='hoat_dong')  # hoat_dong, khoa, da_xoa
    deleted_at = db.Column(db.DateTime)
    deleted_reason = db.Column(db.String(255))
    deleted_by = db.Column(db.Integer, db.ForeignKey("nguoidung.id"))
    deleted_by_user = db.relationship('NguoiDung', foreign_keys=[deleted_by])

    def set_password(self, raw_password):
        if not raw_password:
            raise ValueError("Mật khẩu không được để trống.")
        self.mat_khau_hash = generate_password_hash(raw_password)
        if not self.ngay_dang_ky:
            self.ngay_dang_ky = datetime.now()
        self.ngay_cap_nhat = datetime.now()

    def check_password(self, raw_password):
        if not self.mat_khau_hash:
            return False
        return check_password_hash(self.mat_khau_hash, raw_password)

    @property
    def co_tai_khoan(self):
        return bool(self.mat_khau_hash or self.lan_dang_nhap_cuoi)

    def tang_diem(self, so_diem):
        if so_diem is None:
            return
        self.diem_tich_luy = (self.diem_tich_luy or 0) + int(so_diem)
        self.ngay_cap_nhat = datetime.now()

    def tru_diem(self, so_diem):
        if so_diem is None:
            return
        hien_tai = self.diem_tich_luy or 0
        moi = hien_tai - int(so_diem)
        self.diem_tich_luy = moi if moi > 0 else 0
        self.ngay_cap_nhat = datetime.now()


class CustomerUser(UserMixin):
    """Wrapper cho tài khoản khách hàng dùng chung với Flask-Login."""

    role = None
    personal_permissions = ()

    def __init__(self, khachhang: KhachHang):
        self.khachhang = khachhang

    def get_id(self):
        return f"customer:{self.khachhang.id}"

    @property
    def id(self):
        return self.khachhang.id

    @property
    def ten(self):
        return self.khachhang.ho_ten

    @property
    def ten_dang_nhap(self):
        return self.khachhang.email or self.khachhang.cmnd

    @property
    def loai(self):
        return 'khach'

    @property
    def role_slug(self):
        return 'khach'

    @property
    def role_name(self):
        return 'Khách hàng'

    @property
    def is_staff(self):
        return False

    @property
    def is_customer(self):
        return True

    @property
    def avatar_path(self):
        return "img/ttcn.png"

    def has_permission(self, permission_key):
        return False

    @property
    def is_active(self):
        return self.khachhang.trang_thai_tai_khoan not in {'khoa', 'da_xoa'}

    def __getattr__(self, item):
        return getattr(self.khachhang, item)

class DatPhong(db.Model):
    __tablename__ = "datphong"
    id = db.Column(db.Integer, primary_key=True)
    khachhang_id = db.Column(db.Integer, db.ForeignKey("khachhang.id"), nullable=False)
    phong_id = db.Column(db.Integer, db.ForeignKey("phong.id"), nullable=False)
    nhanvien_id = db.Column(db.Integer, db.ForeignKey("nguoidung.id"))
    hinh_thuc_thue = db.Column(db.String(10), default='ngay', nullable=False)
    ngay_nhan = db.Column(db.DateTime, nullable=False)
    ngay_tra  = db.Column(db.DateTime, nullable=False)
    thuc_te_nhan = db.Column(db.DateTime)
    thuc_te_tra  = db.Column(db.DateTime)
    so_dem = db.Column(db.Integer, default=1)
    trang_thai = db.Column(db.String(20), default="dat")
    chat_token = db.Column(db.String(36), unique=True)
    payment_token = db.Column(db.String(36), unique=True)
    tien_coc = db.Column(db.BIGINT, default=0)
    tien_phat = db.Column(db.BIGINT, default=0)
    tien_phong = db.Column(db.BIGINT, default=0)
    tien_dv = db.Column(db.BIGINT, default=0)
    tong_thanh_toan = db.Column(db.BIGINT, default=0)
    phuong_thuc_thanh_toan = db.Column(db.String(20))
    phuong_thuc_coc = db.Column(db.String(20))
    coc_da_thanh_toan = db.Column(db.Boolean, default=False)
    voucher_id = db.Column(db.Integer, db.ForeignKey("voucher.id"), nullable=True)
    auto_confirmed_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.now)
    diem_loyalty_da_cong = db.Column(db.Integer, default=0)
    voucher = db.relationship("Voucher")
    khachhang = db.relationship("KhachHang")
    phong = db.relationship("Phong")
    nhanvien = db.relationship("NguoiDung")

class DichVuLoai(db.Model):
    __tablename__ = "dichvuloai"
    id = db.Column(db.Integer, primary_key=True)
    ten = db.Column(db.String(100), nullable=False)
    dichvus = db.relationship("DichVu", backref="loai", lazy=True)

class DichVu(db.Model):
    __tablename__ = "dichvu"
    id = db.Column(db.Integer, primary_key=True)
    ten = db.Column(db.String(150), nullable=False)
    gia = db.Column(db.BIGINT, default=0)
    loai_id = db.Column(db.Integer, db.ForeignKey("dichvuloai.id"), nullable=False)

class SuDungDichVu(db.Model):
    __tablename__ = "sudungdv"
    id = db.Column(db.Integer, primary_key=True)
    datphong_id = db.Column(db.Integer, db.ForeignKey("datphong.id"), nullable=False)
    dichvu_id   = db.Column(db.Integer, db.ForeignKey("dichvu.id"), nullable=False)
    so_luong = db.Column(db.Integer, default=1)
    thoi_gian = db.Column(db.DateTime, default=datetime.now)
    trang_thai = db.Column(db.String(20), default='chua_thanh_toan')
    datphong = db.relationship("DatPhong")
    dichvu   = db.relationship("DichVu")

class TinNhan(db.Model):
    __tablename__ = "tinnhan"
    id = db.Column(db.Integer, primary_key=True)
    datphong_id = db.Column(db.Integer, db.ForeignKey("datphong.id"), nullable=False)
    nguoidung_id = db.Column(db.Integer, db.ForeignKey("nguoidung.id"))
    nguoi_gui = db.Column(db.String(10), nullable=False)
    noi_dung = db.Column(db.Text, nullable=False)
    thoi_gian = db.Column(db.DateTime, default=datetime.now)
    trang_thai = db.Column(db.String(20), default='chua_doc')
    datphong = db.relationship("DatPhong")
    nguoidung = db.relationship("NguoiDung")


class PaymentSession(db.Model):
    __tablename__ = "payment_session"
    id = db.Column(db.Integer, primary_key=True)
    token = db.Column(db.String(64), unique=True, nullable=False, index=True)
    kind = db.Column(db.String(20), nullable=False)
    payload = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)


def get_payment_timeout_minutes():
    setting = HeThongCauHinh.query.filter_by(key='payment_timeout_minutes').first()
    if setting and setting.value:
        try:
            minutes = int(setting.value)
            if 1 <= minutes <= 60:
                return minutes
            app.logger.warning(
                "payment_timeout_minutes out of range: %s",
                setting.value,
            )
        except (TypeError, ValueError):
            app.logger.warning(
                "Invalid payment_timeout_minutes value: %s",
                setting.value,
            )
    return DEFAULT_PAYMENT_TIMEOUT_MINUTES


def get_payment_session_ttl():
    return timedelta(minutes=get_payment_timeout_minutes())


def payment_session_expired(created_at, ttl=None):
    if not created_at:
        return True
    now = datetime.now()
    ttl = ttl or get_payment_session_ttl()
    diff = now - created_at
    expired = diff > ttl
    app.logger.info(
        "Payment session check: created_at=%s, now=%s, diff=%s, ttl=%s, expired=%s",
        created_at,
        now,
        diff,
        ttl,
        expired,
    )
    return expired


def payment_session_expires_at(created_at, ttl=None):
    base = created_at or datetime.now()
    ttl = ttl or get_payment_session_ttl()
    return base + ttl


def create_payment_session(token, kind, data_dict):
    session = PaymentSession(token=token, kind=kind, payload=json.dumps(data_dict))
    db.session.add(session)
    db.session.commit()
    return session


def get_payment_session(token):
    session = PaymentSession.query.filter_by(token=token).first()
    if not session:
        return None
    try:
        data = json.loads(session.payload) if session.payload else {}
    except Exception:
        data = {}
    return {'kind': session.kind, 'data': data, 'created_at': session.created_at}


def pop_payment_session(token):
    session = PaymentSession.query.filter_by(token=token).first()
    if session:
        db.session.delete(session)
        db.session.commit()
        return True
    return False


def invalidate_payment_sessions(kind, dat_id):
    removed = []
    for session in PaymentSession.query.filter_by(kind=kind).all():
        try:
            payload = json.loads(session.payload or "{}")
        except Exception:
            payload = {}
        if payload.get('dat_id') == dat_id:
            removed.append(session.token)
            db.session.delete(session)
    if removed:
        db.session.commit()
    return removed


def build_vietqr_url(amount, description):
    encoded_description = quote(description)
    return (
        f"https://img.vietqr.io/image/{VIETQR_BANK_ID}-{VIETQR_ACCOUNT_NO}-compact2.png"
        f"?amount={int(amount)}&addInfo={encoded_description}"
    )


def generate_qr_code(url):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


def _payment_session_common(token, expected_kind):
    session = get_payment_session(token)
    if not session or session['kind'] != expected_kind:
        return None, 'Phien thanh toan khong hop le hoac da het han.'
    if payment_session_expired(session['created_at']):
        pop_payment_session(token)
        return None, 'Phien thanh toan da het han (qua 5 phut).'
    expires_at = payment_session_expires_at(session['created_at'])
    remaining = max(0, int((expires_at - datetime.now()).total_seconds()))
    return {
        'data': session['data'],
        'created_at': session['created_at'],
        'expires_at': expires_at,
        'remaining_seconds': remaining,
    }, None


class LuongThuongCauHinh(db.Model):
    __tablename__ = "luongthuongcauhinh"
    id = db.Column(db.Integer, primary_key=True)
    moc_duoi = db.Column(db.BIGINT, nullable=False, default=0)
    moc_tren = db.Column(db.BIGINT)
    ty_le = db.Column(db.Float, nullable=False, default=0.0)
    ghi_chu = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.now)


class LuongNhanVien(db.Model):
    __tablename__ = "luongnhanvien"
    id = db.Column(db.Integer, primary_key=True)
    nguoidung_id = db.Column(db.Integer, db.ForeignKey("nguoidung.id"), unique=True, nullable=False)
    luong_co_ban = db.Column(db.BIGINT, default=0)
    phu_cap = db.Column(db.BIGINT, default=0)
    nguoidung = db.relationship("NguoiDung", backref=db.backref("thong_tin_luong", uselist=False))


class Voucher(db.Model):
    __tablename__ = "voucher"
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    khachhang_id = db.Column(db.Integer, db.ForeignKey("khachhang.id"), nullable=False)
    is_used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    expires_at = db.Column(db.DateTime)
    discount_percent = db.Column(db.Float, default=10.0)  # 10% mặc định, hỗ trợ số lẻ
    used_at = db.Column(db.DateTime)
    khachhang = db.relationship("KhachHang")

class HeThongCauHinh(db.Model):
    __tablename__ = "hethongcauhinh"
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.String(255))
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


class EmailTemplate(db.Model):
    __tablename__ = "emailtemplate"
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    subject = db.Column(db.String(255), nullable=False)
    body = db.Column(db.Text, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

# Bảng lưu lịch sử gửi email
class EmailLog(db.Model):
    __tablename__ = "email_log"
    id = db.Column(db.Integer, primary_key=True)
    recipient_email = db.Column(db.String(255), nullable=False)
    recipient_name = db.Column(db.String(255))
    template_key = db.Column(db.String(50))
    subject = db.Column(db.String(255), nullable=False)
    body = db.Column(db.Text)
    status = db.Column(db.String(20), default='success')  # success, failed, pending
    error_message = db.Column(db.Text)
    sent_at = db.Column(db.DateTime, default=datetime.now)
    sent_by = db.Column(db.Integer, db.ForeignKey("nguoidung.id"))
    datphong_id = db.Column(db.Integer, db.ForeignKey("datphong.id"))
    khachhang_id = db.Column(db.Integer, db.ForeignKey("khachhang.id"))
    
    # Relationships
    sender = db.relationship("NguoiDung", foreign_keys=[sent_by])
    booking = db.relationship("DatPhong", foreign_keys=[datphong_id])
    customer = db.relationship("KhachHang", foreign_keys=[khachhang_id])

# Bảng chấm công
class Attendance(db.Model):
    __tablename__ = "attendance"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("nguoidung.id"), nullable=False)
    checkin_time = db.Column(db.DateTime, nullable=False, default=datetime.now)
    status = db.Column(db.String(20), default="pending") # pending, approved, rejected
    note = db.Column(db.String(255))
    approved_by = db.Column(db.Integer, db.ForeignKey("nguoidung.id"))
    approved_time = db.Column(db.DateTime)
    user = db.relationship("NguoiDung", foreign_keys=[user_id])
    approver = db.relationship("NguoiDung", foreign_keys=[approved_by])


@login_manager.user_loader
def load_user(user_id):
    if not user_id:
        return None
    try:
        if user_id.startswith('customer:'):
            kh_id = int(user_id.split(':', 1)[1])
            kh = KhachHang.query.get(kh_id)
            return CustomerUser(kh) if kh else None
        if user_id.startswith('staff:'):
            user_id = user_id.split(':', 1)[1]
    except AttributeError:
        # user_id là dạng số nguyên cũ -> tiếp tục xử lý
        pass
    try:
        return NguoiDung.query.get(int(user_id))
    except (ValueError, TypeError):
        return None

# ========================= HELPER FUNCTIONS =========================
def vnd(n):
    return f"{n:,.0f} đ".replace(",", ".")

def vnd_short(n):
    """Format currency with abbreviations for large numbers"""
    if n >= 1000000000:  # Billion
        return f"{n/1000000000:.1f}B đ"
    elif n >= 1000000:  # Million
        return f"{n/1000000:.1f}M đ"
    elif n >= 1000:  # Thousand
        return f"{n/1000:.1f}K đ"
    else:
        return f"{n:,.0f} đ".replace(",", ".")

import random, string
def generate_voucher_code(length=8):
    while True:
        code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))
        if not Voucher.query.filter_by(code=code).first():
            return code

def tinh_diem_tich_luy_tu_thanh_toan(so_tien):
    if not so_tien or so_tien <= 0:
        return 0
    return max(0, int(so_tien // LOYALTY_POINT_RATE_VND))

def tinh_muc_giam_gia_tu_diem(so_diem):
    if not so_diem or so_diem <= 0:
        return 0.0
    giam = round(float(so_diem) * LOYALTY_PERCENT_PER_POINT, 1)
    return min(giam, LOYALTY_MAX_DISCOUNT_PERCENT)

def award_loyalty_points_for_booking(dp):
    if not dp or not dp.khachhang or not dp.khachhang.co_tai_khoan:
        return 0
    tong = dp.tong_thanh_toan or 0
    if tong <= 0:
        tong = (dp.tien_phong or 0) + (dp.tien_dv or 0) + (dp.tien_phat or 0)
    so_diem_du_kien = tinh_diem_tich_luy_tu_thanh_toan(tong)
    da_cong = dp.diem_loyalty_da_cong or 0
    if so_diem_du_kien <= da_cong:
        return 0
    moi = so_diem_du_kien - da_cong
    dp.khachhang.tang_diem(moi)
    dp.diem_loyalty_da_cong = da_cong + moi
    return moi

def get_current_customer():
    if not current_user.is_authenticated:
        return None
    if getattr(current_user, 'is_customer', False):
        if isinstance(current_user, CustomerUser):
            return current_user.khachhang
        return current_user
    return None

def resolve_next_url(default_endpoint='khach_hang_tai_khoan'):
    next_url = request.args.get('next') or request.form.get('next')
    if next_url and is_safe_local_url(next_url):
        return next_url
    return url_for(default_endpoint)

def is_safe_local_url(target):
    if not target:
        return False
    parsed = urlparse(target)
    return not parsed.netloc and not parsed.scheme

def normalize_email(email_value):
    if not email_value:
        return ''
    return email_value.strip().lower()

def format_percent(value):
    try:
        value_float = float(value)
    except (TypeError, ValueError):
        return '0'
    if value_float.is_integer():
        return str(int(value_float))
    text = f"{value_float:.1f}"
    return text.rstrip('0').rstrip('.')


def issue_voucher_for_khachhang(khachhang_id, discount_percent=None, expires_days=None, auto_commit=True):
    # Luôn lấy cấu hình mới nhất từ database
    discount_percent_db, expires_days_db = get_voucher_config()
    if discount_percent is None:
        discount_percent = discount_percent_db
    if expires_days is None:
        expires_days = expires_days_db
    code = generate_voucher_code()
    expires_at = datetime.now() + timedelta(days=expires_days)
    voucher = Voucher(code=code, khachhang_id=khachhang_id, discount_percent=discount_percent, expires_at=expires_at)
    db.session.add(voucher)
    if auto_commit:
        db.session.commit()
    return voucher

def fmt_dt(dt):
    return dt.strftime("%H:%M:%S %d/%m/%Y") if dt else ''

app.jinja_env.filters["vnd"] = vnd
app.jinja_env.filters["fmt_dt"] = fmt_dt
app.jinja_env.filters["vnd_short"] = vnd_short
app.jinja_env.filters["percent"] = format_percent


def allowed_avatar(filename):
    return bool(filename and "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_AVATAR_EXTENSIONS)


def allowed_chat_file(filename):
    return bool(filename and "." in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_CHAT_EXTENSIONS)


def build_message_payload(raw_text):
    """Parse stored noi_dung into a structured payload."""
    if not raw_text:
        return {"type": "text", "text": ""}
    try:
        data = json.loads(raw_text)
        if isinstance(data, dict) and data.get('type') in {'file', 'order', 'system'}:
            return data
    except (json.JSONDecodeError, TypeError):
        pass
    return {"type": "text", "text": raw_text}


def create_file_message(file_storage, uploader_role):
    """Save uploaded file and return structured payload dict."""
    filename = secure_filename(file_storage.filename or '')
    if not filename or not allowed_chat_file(filename):
        raise ValueError('Định dạng tệp không được hỗ trợ.')

    ext = filename.rsplit('.', 1)[1].lower()
    unique_name = f"{uploader_role}_{uuid.uuid4().hex}.{ext}"
    save_path = os.path.join(app.config['CHAT_UPLOAD_FOLDER'], unique_name)
    file_storage.save(save_path)
    rel_path = os.path.join('uploads', 'chat', unique_name).replace('\\', '/')
    return {
        'type': 'file',
        'name': file_storage.filename,
        'mime': file_storage.mimetype,
        'path': rel_path,
        'size': os.path.getsize(save_path)
    }


def persist_message(datphong_id, sender, content_dict_or_text, user_id=None):
    if isinstance(content_dict_or_text, dict):
        raw = json.dumps(content_dict_or_text, ensure_ascii=False)
    else:
        raw = content_dict_or_text
    msg = TinNhan(
        datphong_id=datphong_id,
        nguoidung_id=user_id,
        nguoi_gui=sender,
        noi_dung=raw,
        thoi_gian=datetime.now()
    )
    db.session.add(msg)
    db.session.commit()
    return msg


def serialize_message(msg):
    payload = build_message_payload(msg.noi_dung)
    staff_name = msg.nguoidung.ten if msg.nguoidung else ''
    staff_avatar = ''
    if msg.nguoidung and hasattr(msg.nguoidung, 'anh_dai_dien') and msg.nguoidung.anh_dai_dien:
        staff_avatar = url_for('static', filename=msg.nguoidung.anh_dai_dien, _external=False)
    
    guest_name = ''
    if msg.datphong and getattr(msg.datphong, 'khachhang', None):
        guest_name = msg.datphong.khachhang.ho_ten or ''
    
    sender_role = (msg.nguoi_gui or '').lower()
    if sender_role in ('nhanvien', 'staff'):
        sender_name = staff_name or 'Nhân viên'
    elif sender_role in ('khach', 'guest'):
        sender_name = guest_name or 'Khách'
    elif sender_role in ('he_thong', 'system'):
        sender_name = 'Hệ thống'
    else:
        sender_name = staff_name or guest_name or ''
    
    base = {
        'nguoi_gui': msg.nguoi_gui,
        'thoi_gian': msg.thoi_gian.strftime('%H:%M %d/%m'),
        'ten_nhan_vien': staff_name,
        'ten_khach': guest_name,
        'ten_khach_hang': guest_name,
        'ten_nguoi_gui': sender_name,
        'avatar_nhan_vien': staff_avatar
    }
    if payload.get('type') == 'file':
        payload['url'] = url_for('static', filename=payload['path'], _external=False)
    return {**base, **payload}


def get_active_booking_by_token(token):
    if not token:
        return None
    return DatPhong.query.filter_by(chat_token=token, trang_thai='nhan').first()

# ========================= PHÂN QUYỀN =========================
# ========================= CHẤM CÔNG =========================
@app.route('/attendance', methods=['GET', 'POST'])
@login_required
def attendance_checkin():
    if request.method == 'POST':
        today = datetime.now().date()
        existing = Attendance.query.filter(
            Attendance.user_id == current_user.id,
            func.date(Attendance.checkin_time) == today
        ).first()
        if existing:
            flash('Bạn đã gửi yêu cầu chấm công hôm nay!', 'warning')
        else:
            note = request.form.get('note')
            att = Attendance(user_id=current_user.id, checkin_time=datetime.now(), note=note)
            db.session.add(att)
            db.session.commit()
            flash('Đã gửi yêu cầu chấm công!', 'success')
        return redirect(url_for('attendance_checkin'))
    attendances = Attendance.query.filter_by(user_id=current_user.id).order_by(Attendance.checkin_time.desc()).all()
    return render_template('attendance.html', attendances=attendances)

# Quản trị viên phê duyệt chấm công
@app.route('/attendance/admin', methods=['GET'])
@login_required
@permission_required('attendance.manage')
def attendance_admin():
    attendances = Attendance.query.order_by(Attendance.checkin_time.desc()).all()
    return render_template('attendance_admin.html', attendances=attendances)

@app.route('/attendance/approve/<int:att_id>', methods=['POST'])
@login_required
@permission_required('attendance.manage')
def attendance_approve(att_id):
    att = Attendance.query.get_or_404(att_id)
    action = request.form.get('action')
    if att.status != 'pending':
        flash('Yêu cầu đã được xử lý!', 'warning')
        return redirect(url_for('attendance_admin'))
    if action == 'approve':
        att.status = 'approved'
        att.approved_by = current_user.id
        att.approved_time = datetime.now()
        flash('Đã duyệt chấm công!', 'success')
    elif action == 'reject':
        att.status = 'rejected'
        att.approved_by = current_user.id
        att.approved_time = datetime.now()
        flash('Đã từ chối chấm công!', 'danger')
    db.session.commit()
    return redirect(url_for('attendance_admin'))

@app.context_processor
def inject_globals():
    unread_count = 0
    pending_online_count = 0
    if current_user.is_authenticated:
        if current_user.has_permission('communications.chat'):
            unread_count = TinNhan.query.join(DatPhong).filter(
                TinNhan.trang_thai == 'chua_doc',
                TinNhan.nguoi_gui == 'khach',
                DatPhong.trang_thai == 'nhan'
            ).count()
        if current_user.has_permission('bookings.manage_online'):
            pending_online_count = DatPhong.query.join(
                TinNhan,
                db.and_(
                    TinNhan.datphong_id == DatPhong.id,
                    TinNhan.nguoi_gui == 'khach',
                    TinNhan.noi_dung == ONLINE_DEPOSIT_REQUEST_MESSAGE
                )
            ).filter(
                DatPhong.trang_thai == 'cho_xac_nhan'
            ).distinct().count()
    return dict(
        now=datetime.now,
        unread_messages=unread_count,
        pending_online_count=pending_online_count,
        vnd=vnd,
        hotel_info=get_hotel_profile()
    )


def ensure_tables_exist():
    try:
        db.create_all()
        try:
            inspector = inspect(db.engine)
            user_columns = {col['name'] for col in inspector.get_columns('nguoidung')}
            if 'role_id' not in user_columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE nguoidung ADD COLUMN role_id INT NULL'))
            columns = {col['name'] for col in inspector.get_columns('khachhang')}
            if 'email' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN email VARCHAR(120)'))
            if 'mat_khau_hash' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN mat_khau_hash VARCHAR(255)'))
            if 'diem_tich_luy' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN diem_tich_luy INT DEFAULT 0'))
            if 'ngay_dang_ky' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN ngay_dang_ky DATETIME NULL'))
            if 'ngay_cap_nhat' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN ngay_cap_nhat DATETIME NULL'))
            if 'lan_dang_nhap_cuoi' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN lan_dang_nhap_cuoi DATETIME NULL'))
            if 'trang_thai_tai_khoan' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text("ALTER TABLE khachhang ADD COLUMN trang_thai_tai_khoan VARCHAR(20) DEFAULT 'hoat_dong'"))
            if 'deleted_at' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN deleted_at DATETIME NULL'))
            if 'deleted_reason' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN deleted_reason VARCHAR(255) NULL'))
            if 'deleted_by' not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE khachhang ADD COLUMN deleted_by INT NULL'))
            columns_datphong = {col['name'] for col in inspector.get_columns('datphong')}
            if 'created_at' not in columns_datphong:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE datphong ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP'))
            if 'diem_loyalty_da_cong' not in columns_datphong:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE datphong ADD COLUMN diem_loyalty_da_cong INT DEFAULT 0'))
            voucher_columns = inspector.get_columns('voucher')
            discount_column = next((col for col in voucher_columns if col['name'] == 'discount_percent'), None)
            if discount_column and discount_column['type'].__class__.__name__.lower() in {'integer', 'smallinteger', 'bigint'}:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE voucher MODIFY COLUMN discount_percent FLOAT DEFAULT 10'))
        except Exception as exc:
            app.logger.warning("Không thể đảm bảo cột cho bảng: %s", exc)
        ensure_default_roles()
    except Exception as exc:
        app.logger.warning("Không thể tạo bảng tự động: %s", exc)


def ensure_customer_email_templates():
    templates = {
        'customer_welcome': {
            'subject': 'Chào mừng {{ ho_ten }} trở thành thành viên {{ ten_khach_san }}',
            'body': """<!DOCTYPE html>
<html lang="vi">
  <body style="margin:0;background-color:#f8fafc;font-family:'Helvetica Neue',Arial,sans-serif;color:#0f172a;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="padding:32px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="max-width:560px;background-color:#ffffff;border-radius:18px;overflow:hidden;box-shadow:0 20px 44px -26px rgba(15,23,42,0.32);">
            <tr>
              <td style="background:linear-gradient(135deg,#2563eb,#0ea5e9);padding:28px 36px;color:#ffffff;">
                <h1 style="margin:0;font-size:24px;">Chào mừng thành viên mới</h1>
                <p style="margin:10px 0 0;font-size:14px;opacity:0.9;">{{ ten_khach_san }}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:32px 36px;">
                <p style="margin:0 0 16px;font-size:15px;">Xin chào {{ ho_ten }},</p>
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">
                  Cảm ơn bạn đã gia nhập cộng đồng thành viên của <strong>{{ ten_khach_san }}</strong>. Từ bây giờ, bạn có thể:
                </p>
                <ul style="margin:0 0 20px 20px;padding:0;font-size:15px;line-height:1.6;">
                  <li>Theo dõi lịch sử đặt phòng và trạng thái đơn hàng.</li>
                  <li>Tích lũy điểm thưởng và đổi lấy voucher ưu đãi.</li>
                  <li>Cập nhật thông tin cá nhân để nhận các chương trình dành riêng cho bạn.</li>
                </ul>
                <p style="margin:0 0 12px;font-size:15px;font-weight:600;">Thông tin đăng nhập:</p>
                <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;margin:0 0 24px;border-collapse:separate;border-spacing:0 10px;">
                  <tr>
                    <td style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:12px 16px;">
                      <span style="display:block;font-size:13px;color:#64748b;">CMND/CCCD</span>
                      <span style="font-size:16px;font-weight:600;color:#2563eb;">{{ cmnd }}</span>
                    </td>
                  </tr>
                  <tr>
                    <td style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:12px 16px;">
                      <span style="display:block;font-size:13px;color:#64748b;">Email</span>
                      <span style="font-size:16px;font-weight:600;color:#2563eb;">{{ email or 'Chưa cập nhật' }}</span>
                    </td>
                  </tr>
                </table>
                {% if tai_khoan_url %}
                <div style="margin:24px 0;text-align:center;">
                  <a href="{{ tai_khoan_url }}" style="display:inline-block;padding:12px 28px;border-radius:999px;background:#2563eb;color:#ffffff;text-decoration:none;font-weight:600;">Khám phá tài khoản</a>
                </div>
                {% endif %}
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">
                  Đừng quên đăng nhập và cập nhật thông tin cá nhân để chúng tôi phục vụ bạn tốt hơn. Nếu cần hỗ trợ, đội ngũ {{ ten_khach_san }} luôn sẵn sàng!
                </p>
              </td>
            </tr>
            <tr>
              <td style="padding:22px 36px;background:#f3f4f6;border-top:1px solid #e2e8f0;font-size:13px;color:#64748b;">
                <p style="margin:0 0 6px;">Trân trọng,<br>{{ ten_khach_san }}</p>
                <p style="margin:0;">Email được gửi tự động, vui lòng không trả lời thư này.</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""
        },
        'customer_account_locked': {
            'subject': 'Tài khoản của bạn tại {{ ten_khach_san }} đã bị khóa',
            'body': """<!DOCTYPE html>
<html lang="vi">
  <body style="margin:0;background-color:#fff7ed;font-family:'Helvetica Neue',Arial,sans-serif;color:#1f2937;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="padding:32px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="max-width:560px;background-color:#ffffff;border-radius:18px;overflow:hidden;box-shadow:0 20px 44px -28px rgba(180,83,9,0.4);">
            <tr>
              <td style="background:linear-gradient(135deg,#f97316,#fb923c);padding:26px 34px;color:#ffffff;">
                <h1 style="margin:0;font-size:22px;">Thông báo khóa tài khoản</h1>
                <p style="margin:8px 0 0;font-size:14px;opacity:0.9;">{{ ten_khach_san }}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:30px 34px;">
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">Xin chào {{ ho_ten }},</p>
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">
                  Tài khoản khách hàng của bạn tại <strong>{{ ten_khach_san }}</strong> đã tạm thời bị khóa để đảm bảo an toàn.
                </p>
                <div style="margin:0 0 20px;padding:16px;border-radius:16px;background:rgba(251, 191, 36, 0.12);border:1px solid rgba(251, 146, 60, 0.35);">
                  <strong style="display:block;margin-bottom:8px;font-size:15px;">Lý do:</strong>
                  <span style="font-size:14px;line-height:1.6;">{{ ly_do }}</span>
                </div>
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">
                  Nếu bạn cần hỗ trợ thêm hoặc muốn mở lại tài khoản, vui lòng liên hệ với chúng tôi qua các kênh hỗ trợ của khách sạn.
                </p>
              </td>
            </tr>
            <tr>
              <td style="padding:20px 34px;background:#fff7ed;border-top:1px solid rgba(251, 146, 60, 0.2);font-size:13px;color:#b45309;">
                <p style="margin:0;">Trân trọng,<br>{{ ten_khach_san }}</p>
                <p style="margin:8px 0 0;font-size:12px;opacity:0.85;">Email được gửi tự động, vui lòng không trả lời thư này.</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""
        },
        'customer_account_unlocked': {
            'subject': 'Tài khoản của bạn tại {{ ten_khach_san }} đã được mở khóa',
            'body': """<!DOCTYPE html>
<html lang="vi">
  <body style="margin:0;background-color:#ecfeff;font-family:'Helvetica Neue',Arial,sans-serif;color:#0f172a;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="padding:32px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="max-width:560px;background-color:#ffffff;border-radius:18px;overflow:hidden;box-shadow:0 20px 44px -26px rgba(14,165,233,0.35);">
            <tr>
              <td style="background:linear-gradient(135deg,#0ea5e9,#2563eb);padding:26px 34px;color:#ffffff;">
                <h1 style="margin:0;font-size:22px;">Tài khoản đã hoạt động trở lại</h1>
                <p style="margin:8px 0 0;font-size:14px;opacity:0.9;">{{ ten_khach_san }}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:30px 34px;">
                <p style="margin:0 0 16px;font-size:15px;">Xin chào {{ ho_ten }},</p>
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">
                  Tài khoản khách hàng của bạn tại <strong>{{ ten_khach_san }}</strong> đã được mở khóa. Bạn có thể đăng nhập và sử dụng các tiện ích như trước đây.
                </p>
                {% if dang_nhap_url %}
                <div style="text-align:center;margin:24px 0;">
                  <a href="{{ dang_nhap_url }}" style="display:inline-block;padding:12px 26px;border-radius:999px;background:#2563eb;color:#ffffff;text-decoration:none;font-weight:600;">Đăng nhập ngay</a>
                </div>
                {% endif %}
                <p style="margin:0;font-size:15px;line-height:1.6;">
                  Nếu bạn cần thêm hỗ trợ, đừng ngần ngại liên hệ với chúng tôi. Chúc bạn có những trải nghiệm tuyệt vời tại {{ ten_khach_san }}!
                </p>
              </td>
            </tr>
            <tr>
              <td style="padding:20px 34px;background:#ecfeff;border-top:1px solid rgba(14,165,233,0.25);font-size:13px;color:#1e3a8a;">
                <p style="margin:0;">Trân trọng,<br>{{ ten_khach_san }}</p>
                <p style="margin:8px 0 0;font-size:12px;opacity:0.85;">Email được gửi tự động, vui lòng không trả lời thư này.</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""
        },
        'customer_account_deleted': {
            'subject': 'Tai khoan cua ban tai {{ ten_khach_san }} da bi xoa',
            'body': """<!DOCTYPE html>
<html lang="vi">
  <body style="margin:0;background-color:#fff5f5;font-family:'Helvetica Neue',Arial,sans-serif;color:#1f2937;">
    <div style="max-width:640px;margin:0 auto;padding:28px;">
      <div style="background:#ffffff;border-radius:16px;padding:28px;box-shadow:0 18px 38px -24px rgba(190,24,60,0.25);">
        <h1 style="margin:0 0 12px;color:#b91c1c;font-size:22px;">Tai khoan da bi xoa</h1>
        <p style="margin:0 0 12px;font-size:15px;line-height:1.6;">Xin chao {{ ho_ten }},</p>
        <p style="margin:0 0 12px;font-size:15px;line-height:1.6;">Tai khoan khach hang cua ban tai <strong>{{ ten_khach_san }}</strong> da bi xoa boi quan tri vien.</p>
        {% if ly_do %}
        <div style="margin:16px 0;padding:16px;border-radius:12px;background:rgba(248,113,113,0.12);border:1px solid rgba(248,113,113,0.35);">
          <strong style="display:block;margin-bottom:6px;">Ly do:</strong>
          <span style="font-size:14px;line-height:1.6;">{{ ly_do }}</span>
        </div>
        {% endif %}
        <p style="margin:0;font-size:15px;line-height:1.6;">Neu ban can ho tro them, vui long lien he lai voi chung toi.</p>
      </div>
      <p style="margin:18px 0 0;font-size:12px;color:#b91c1c;text-align:center;">Email duoc gui tu dong, vui long khong tra loi.</p>
    </div>
  </body>
</html>"""
        },
        'customer_account_self_deleted': {
            'subject': 'Ban da xoa tai khoan tai {{ ten_khach_san }}',
            'body': """<!DOCTYPE html>
<html lang="vi">
  <body style="margin:0;background-color:#ecfdf5;font-family:'Helvetica Neue',Arial,sans-serif;color:#065f46;">
    <div style="max-width:640px;margin:0 auto;padding:28px;">
      <div style="background:#ffffff;border-radius:16px;padding:28px;box-shadow:0 18px 38px -24px rgba(5,150,105,0.25);">
        <h1 style="margin:0 0 12px;color:#047857;font-size:22px;">Xac nhan xoa tai khoan</h1>
        <p style="margin:0 0 12px;font-size:15px;line-height:1.6;">Xin chao {{ ho_ten }},</p>
        <p style="margin:0 0 12px;font-size:15px;line-height:1.6;">Chung toi da ghi nhan yeu cau xoa tai khoan cua ban tai <strong>{{ ten_khach_san }}</strong>.</p>
        {% if ly_do %}
        <div style="margin:16px 0;padding:16px;border-radius:12px;background:rgba(16,185,129,0.12);border:1px solid rgba(16,185,129,0.35);color:#047857;">
          <strong style="display:block;margin-bottom:6px;">Ly do ban cung cap:</strong>
          <span style="font-size:14px;line-height:1.6;">{{ ly_do }}</span>
        </div>
        {% endif %}
        <p style="margin:0;font-size:15px;line-height:1.6;">Ban co the dang ky lai bat ky luc nao bang CMND/CCCD va email nay neu muon. Neu can ho tro, hay lien he voi chung toi.</p>
      </div>
      <p style="margin:18px 0 0;font-size:12px;color:#047857;text-align:center;">Email duoc gui tu dong, vui long khong tra loi.</p>
    </div>
  </body>
</html>"""
        },
        'customer_password_reset': {
            'subject': 'Yêu cầu đặt lại mật khẩu tại {{ ten_khach_san }}',
            'body': """<!DOCTYPE html>
<html lang="vi">
  <body style="margin:0;background-color:#f3f4f6;font-family:'Helvetica Neue',Arial,sans-serif;color:#0f172a;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="padding:32px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="max-width:520px;background-color:#ffffff;border-radius:16px;overflow:hidden;box-shadow:0 18px 40px -22px rgba(15,23,42,0.35);">
            <tr>
              <td style="background:linear-gradient(135deg,#2563eb,#1d4ed8);padding:24px 32px;color:#ffffff;">
                <h1 style="margin:0;font-size:22px;">Yêu cầu đặt lại mật khẩu</h1>
                <p style="margin:8px 0 0;font-size:14px;opacity:0.85;">{{ ten_khach_san }}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:28px 32px;">
                <p style="margin:0 0 16px;font-size:15px;">Xin chào {{ ho_ten }},</p>
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">
                  Chúng tôi vừa nhận được yêu cầu đặt lại mật khẩu cho tài khoản khách hàng của bạn tại {{ ten_khach_san }}.
                </p>
                <p style="margin:0 0 12px;font-size:15px;line-height:1.6;">Mật khẩu tạm thời của bạn là:</p>
                <div style="margin:0 0 24px;padding:16px;border-radius:12px;background:#f8fafc;border:1px solid #e2e8f0;text-align:center;">
                  <span style="display:inline-block;font-size:20px;font-weight:700;letter-spacing:1px;color:#2563eb;">{{ mat_khau_moi }}</span>
                </div>
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">
                  Vui lòng đăng nhập bằng mật khẩu trên và đổi sang mật khẩu mới ngay sau khi truy cập để đảm bảo an toàn.
                </p>
                {% if dang_nhap_url %}
                <div style="margin:24px 0;text-align:center;">
                  <a href="{{ dang_nhap_url }}" style="display:inline-block;padding:12px 24px;border-radius:999px;background:#2563eb;color:#ffffff;text-decoration:none;font-weight:600;">Đăng nhập ngay</a>
                </div>
                {% endif %}
                <p style="margin:0 0 16px;font-size:15px;line-height:1.6;">
                  Nếu bạn không gửi yêu cầu này, hãy liên hệ với chúng tôi để được hỗ trợ và bảo vệ tài khoản.
                </p>
              </td>
            </tr>
            <tr>
              <td style="padding:20px 32px;background:#f3f4f6;border-top:1px solid #e2e8f0;font-size:13px;color:#64748b;">
                <p style="margin:0 0 8px;">Trân trọng,<br>{{ ten_khach_san }}</p>
                <p style="margin:0;">Email được gửi tự động, vui lòng không trả lời thư này.</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""
        }
    }
    changed = False
    for key, tpl in templates.items():
        existing = EmailTemplate.query.filter_by(key=key).first()
        if existing:
            if existing.subject != tpl['subject'] or existing.body != tpl['body']:
                existing.subject = tpl['subject']
                existing.body = tpl['body']
                changed = True
        else:
            db.session.add(EmailTemplate(key=key, subject=tpl['subject'], body=tpl['body']))
            changed = True
    if changed:
        db.session.commit()


with app.app_context():
    ensure_tables_exist()
    ensure_customer_email_templates()


def tinh_thuong_doanh_thu(doanh_thu, tiers):
    applicable = None
    sorted_tiers = sorted(tiers, key=lambda t: (t.moc_duoi or 0))
    for tier in sorted_tiers:
        lower = tier.moc_duoi or 0
        upper = tier.moc_tren if tier.moc_tren is not None else float('inf')
        if doanh_thu >= lower and doanh_thu <= upper:
            applicable = tier
    if applicable is None and sorted_tiers:
        last = sorted_tiers[-1]
        if last.moc_tren is None and doanh_thu >= (last.moc_duoi or 0):
            applicable = last
    rate = applicable.ty_le if applicable else 0
    thuong = int(round(float(doanh_thu) * rate / 100)) if rate else 0
    return thuong, rate

TOP_REVENUE_BONUS_DEFAULT = 500_000


def get_config_value(key, default=''):
    setting = HeThongCauHinh.query.filter_by(key=key).first()
    if setting and setting.value is not None:
        return setting.value
    return default


def resolve_public_base_url():
    configured = app.config.get("PUBLIC_BASE_URL")
    if configured:
        return configured

    try:
        stored = get_config_value('PUBLIC_BASE_URL', '').strip()
    except Exception:
        stored = ''
    if stored:
        stored = stored.rstrip("/")
        app.config["PUBLIC_BASE_URL"] = stored
        return stored

    forwarded_host = request.headers.get("X-Forwarded-Host")
    if forwarded_host:
        forwarded_host = forwarded_host.split(",")[0].strip()
        if forwarded_host:
            proto = request.headers.get("X-Forwarded-Proto", request.scheme)
            if proto:
                proto = proto.split(",")[0].strip()
            else:
                proto = request.scheme
            forwarded_port = request.headers.get("X-Forwarded-Port")
            if forwarded_port:
                forwarded_port = forwarded_port.split(",")[0].strip()
                host_with_port = f"{forwarded_host}:{forwarded_port}"
            else:
                host_with_port = forwarded_host
            return f"{proto}://{host_with_port}".rstrip("/")

    return request.host_url.rstrip("/")


def set_config_values(pairs):
    dirty = False
    for key, value in pairs.items():
        setting = HeThongCauHinh.query.filter_by(key=key).first()
        if not setting:
            setting = HeThongCauHinh(key=key)
            db.session.add(setting)
        new_value = '' if value is None else str(value)
        if setting.value != new_value:
            setting.value = new_value
            dirty = True
    if dirty:
        db.session.commit()


def get_config_int(key, default):
    setting = HeThongCauHinh.query.filter_by(key=key).first()
    if setting:
        try:
            return max(0, int(setting.value))
        except (ValueError, TypeError):
            pass
    return default

def set_config_int(key, value):
    value = max(0, int(value))
    setting = HeThongCauHinh.query.filter_by(key=key).first()
    if not setting:
        setting = HeThongCauHinh(key=key)
        db.session.add(setting)
    setting.value = str(value)
    db.session.commit()

def get_top_bonus():
    return get_config_int('TOP_REVENUE_BONUS', TOP_REVENUE_BONUS_DEFAULT)

# Số ngày công tối thiểu để nhận phụ cấp
MIN_WORK_DAYS_KEY = 'MIN_WORK_DAYS'
MIN_WORK_DAYS_DEFAULT = 22

def get_min_work_days():
    return get_config_int(MIN_WORK_DAYS_KEY, MIN_WORK_DAYS_DEFAULT)

def set_min_work_days(val):
    set_config_int(MIN_WORK_DAYS_KEY, val)

SALARY_MODE_KEY = 'SALARY_CALC_MODE'
SALARY_MODE_MONTHLY = 'monthly'
SALARY_MODE_DAILY = 'daily'
SALARY_DAILY_DIVISOR = 28

def _round_divide(numerator, denominator):
    if denominator == 0:
        return 0
    quotient, remainder = divmod(numerator, denominator)
    if remainder * 2 >= denominator:
        quotient += 1
    return quotient

def get_salary_mode():
    value = (get_config_value(SALARY_MODE_KEY, SALARY_MODE_MONTHLY) or SALARY_MODE_MONTHLY).strip().lower()
    return SALARY_MODE_DAILY if value == SALARY_MODE_DAILY else SALARY_MODE_MONTHLY

def set_salary_mode(mode):
    normalized = (mode or '').strip().lower()
    if normalized not in {SALARY_MODE_MONTHLY, SALARY_MODE_DAILY}:
        raise ValueError('invalid salary mode')
    set_config_values({SALARY_MODE_KEY: normalized})

def compute_daily_rate(base_salary):
    base_salary = max(0, int(base_salary or 0))
    if base_salary <= 0:
        return 0
    return _round_divide(base_salary, SALARY_DAILY_DIVISOR)

def compute_effective_base_salary(base_salary, work_days, mode=None):
    mode = (mode or get_salary_mode()).strip().lower()
    base_salary = max(0, int(base_salary or 0))
    work_days = max(0, int(work_days or 0))
    if mode == SALARY_MODE_DAILY:
        # Chế độ lương ngày: nếu chưa chấm công (work_days = 0) thì lương = 0
        if work_days == 0:
            return 0
        if base_salary > 0 and work_days > 0:
            return _round_divide(base_salary * work_days, SALARY_DAILY_DIVISOR)
    return base_salary


DEFAULT_SMTP_CONFIG = {
    'SMTP_HOST': os.getenv('DEFAULT_SMTP_HOST', 'smtp.office365.com'),
    'SMTP_PORT': os.getenv('DEFAULT_SMTP_PORT', '587'),
    'SMTP_USERNAME': os.getenv('DEFAULT_SMTP_USERNAME', 'hotelservice@88c0c7.onmicrosoft.com'),
    'SMTP_SENDER_EMAIL': os.getenv('DEFAULT_SMTP_SENDER_EMAIL', 'hotelservice@88c0c7.onmicrosoft.com'),
    'SMTP_USE_TLS': os.getenv('DEFAULT_SMTP_USE_TLS', '1'),
    'SMTP_USE_SSL': os.getenv('DEFAULT_SMTP_USE_SSL', '0')
}


def _default_smtp_value(key, fallback=''):
    env_value = os.getenv(key)
    if env_value is not None and str(env_value).strip() != '':
        return env_value
    return fallback


def _interpret_bool(value, fallback='0'):
    candidate = value if value is not None and str(value).strip() != '' else fallback
    return str(candidate).strip().lower() in {'1', 'true', 'yes', 'on'}


EMAIL_TEMPLATE_DEFAULTS = {
    'booking_confirmation': {
        'name': 'Email xác nhận booking',
        'subject': 'Xác nhận đặt phòng #{{ ma_dat_phong }} - {{ ten_khach_san }}',
        'body': '''
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Xác nhận đặt phòng</title>
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f4;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f4f4f4; padding: 20px 0;">
        <tr>
            <td align="center">
                <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); overflow: hidden;">
                    <!-- Header -->
                    <tr>
                        <td style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 40px 30px; text-align: center;">
                            <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 600;">✅ Xác Nhận Đặt Phòng</h1>
                            <p style="color: #ffffff; margin: 10px 0 0 0; font-size: 16px; opacity: 0.95;">{{ ten_khach_san }}</p>
                        </td>
                    </tr>
                    
                    <!-- Body -->
                    <tr>
                        <td style="padding: 40px 30px;">
                            <p style="font-size: 16px; color: #333333; margin: 0 0 20px 0; line-height: 1.6;">
                                Kính thưa <strong style="color: #667eea;">{{ ten_khach_hang }}</strong>,
                            </p>
                            <p style="font-size: 15px; color: #555555; margin: 0 0 30px 0; line-height: 1.6;">
                                Cảm ơn quý khách đã tin tưởng và đặt phòng tại <strong>{{ ten_khach_san }}</strong>. Chúng tôi rất vui được phục vụ quý khách!
                            </p>
                            
                            <!-- Booking Info Card -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f8f9fa; border-radius: 8px; border-left: 4px solid #667eea; margin-bottom: 25px;">
                                <tr>
                                    <td style="padding: 20px;">
                                        <h2 style="color: #667eea; margin: 0 0 15px 0; font-size: 18px; font-weight: 600;">📋 Thông Tin Đặt Phòng</h2>
                                        <table width="100%" cellpadding="8" cellspacing="0">
                                            <tr>
                                                <td style="color: #666666; font-size: 14px; padding: 5px 0;">Mã đặt phòng:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right; padding: 5px 0;">#{{ ma_dat_phong }}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #666666; font-size: 14px; padding: 5px 0;">Phòng:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right; padding: 5px 0;">{{ ten_phong }}{% if loai_phong %} <span style="color: #667eea;">({{ loai_phong }})</span>{% endif %}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #666666; font-size: 14px; padding: 5px 0;">Check-in:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right; padding: 5px 0;">{{ thoi_gian_nhan }}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #666666; font-size: 14px; padding: 5px 0;">Check-out:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right; padding: 5px 0;">{{ thoi_gian_tra }}</td>
                                            </tr>
                                            {% set so_luu_tru_value = so_luong_luu_tru or so_dem %}
                                            {% set don_vi_luu_tru_display = don_vi_luu_tru or ("đêm" if so_luu_tru_value else "") %}
                                            {% set nhan_luu_tru_display = nhan_luu_tru or "Số đêm lưu trú" %}
                                            {% if so_luu_tru_value %}
                                            <tr>
                                                <td style="color: #666666; font-size: 14px; padding: 5px 0;">{{ nhan_luu_tru_display }}:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right; padding: 5px 0;">{{ so_luu_tru_value }}{% if don_vi_luu_tru_display %} {{ don_vi_luu_tru_display }}{% endif %}</td>
                                            </tr>
                                            {% endif %}
                                            {% if tien_coc %}
                                            <tr>
                                                <td style="color: #666666; font-size: 14px; padding: 5px 0;">Tiền cọc:</td>
                                                <td style="color: #28a745; font-size: 14px; font-weight: 600; text-align: right; padding: 5px 0;">{{ tien_coc }}</td>
                                            </tr>
                                            {% endif %}
                                            <tr style="border-top: 2px solid #dee2e6;">
                                                <td style="color: #333333; font-size: 15px; font-weight: 600; padding: 10px 0 5px 0;">Tổng tiền dự kiến:</td>
                                                <td style="color: #667eea; font-size: 18px; font-weight: 700; text-align: right; padding: 10px 0 5px 0;">{{ tong_tien }}</td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>
                            
                            <!-- Note -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #fff3cd; border-radius: 8px; margin-bottom: 25px;">
                                <tr>
                                    <td style="padding: 15px;">
                                        <p style="margin: 0; color: #856404; font-size: 14px; line-height: 1.5;">
                                            ⏰ <strong>Lưu ý:</strong> Vui lòng đến nhận phòng đúng giờ. Nếu có thắc mắc, xin liên hệ với chúng tôi.
                                        </p>
                                    </td>
                                </tr>
                            </table>
                            
                            <!-- Contact Info -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #e7f3ff; border-radius: 8px;">
                                <tr>
                                    <td style="padding: 20px;">
                                        <h3 style="color: #0066cc; margin: 0 0 12px 0; font-size: 16px; font-weight: 600;">📞 Thông Tin Liên Hệ</h3>
                                        <p style="margin: 5px 0; color: #333333; font-size: 14px;">
                                            <strong>Hotline:</strong> {{ so_dien_thoai_khach_san }}
                                        </p>
                                        <p style="margin: 5px 0; color: #333333; font-size: 14px;">
                                            <strong>Địa chỉ:</strong> {{ dia_chi_khach_san }}
                                        </p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Footer -->
                    <tr>
                        <td style="background-color: #f8f9fa; padding: 25px 30px; text-align: center; border-top: 1px solid #dee2e6;">
                            <p style="margin: 0 0 5px 0; color: #666666; font-size: 14px;">Trân trọng,</p>
                            <p style="margin: 0; color: #333333; font-size: 16px; font-weight: 600;">{{ ten_khach_san }}</p>
                            <p style="margin: 10px 0 0 0; color: #999999; font-size: 12px;">
                                Email này được gửi tự động, vui lòng không trả lời.
                            </p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>
        '''
    },
    'checkin_notification': {
        'name': 'Email check-in',
        'subject': 'Chào mừng đến với {{ ten_khach_san }} 🏨',
        'body': '''
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Chào mừng check-in</title>
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f4;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f4f4f4; padding: 20px 0;">
        <tr>
            <td align="center">
                <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); overflow: hidden;">
                    <!-- Header -->
                    <tr>
                        <td style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); padding: 40px 30px; text-align: center;">
                            <h1 style="color: #ffffff; margin: 0; font-size: 32px; font-weight: 600;">🎉 Chào Mừng!</h1>
                            <p style="color: #ffffff; margin: 10px 0 0 0; font-size: 18px; opacity: 0.95;">Quý khách đã nhận phòng thành công</p>
                        </td>
                    </tr>
                    
                    <!-- Body -->
                    <tr>
                        <td style="padding: 40px 30px;">
                            <p style="font-size: 16px; color: #333333; margin: 0 0 15px 0; line-height: 1.6;">
                                Kính thưa <strong style="color: #f5576c;">{{ ten_khach_hang }}</strong>,
                            </p>
                            <p style="font-size: 15px; color: #555555; margin: 0 0 30px 0; line-height: 1.6;">
                                Chúng tôi rất vui được chào đón quý khách tại <strong>{{ ten_khach_san }}</strong>! Hy vọng quý khách có một kỳ nghỉ tuyệt vời và thoải mái nhất.
                            </p>
                            
                            <!-- Check-in Success Card -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background: linear-gradient(135deg, #d4fc79 0%, #96e6a1 100%); border-radius: 12px; margin-bottom: 25px;">
                                <tr>
                                    <td style="padding: 25px; text-align: center;">
                                        <div style="font-size: 48px; margin-bottom: 10px;">✅</div>
                                        <h2 style="color: #2d5016; margin: 0 0 8px 0; font-size: 20px; font-weight: 700;">Đã Nhận Phòng Thành Công</h2>
                                        <p style="color: #3d6622; margin: 0; font-size: 14px;">Mã đặt phòng: <strong>#{{ ma_dat_phong }}</strong></p>
                                    </td>
                                </tr>
                            </table>
                            
                            <!-- Room Info -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f8f9fa; border-radius: 8px; margin-bottom: 25px;">
                                <tr>
                                    <td style="padding: 20px;">
                                        <h3 style="color: #333333; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">🏨 Thông Tin Phòng</h3>
                                        <table width="100%" cellpadding="6" cellspacing="0">
                                            <tr>
                                                <td style="color: #666666; font-size: 14px;">Phòng:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right;">{{ ten_phong }}{% if loai_phong %} <span style="color: #f5576c;">({{ loai_phong }})</span>{% endif %}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #666666; font-size: 14px;">Nhận phòng:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right;">{{ thoi_gian_nhan_thuc_te }}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #666666; font-size: 14px;">Dự kiến trả:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right;">{{ thoi_gian_tra }}</td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>
                            
                            <!-- Services Info -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #e3f2fd; border-radius: 8px; border-left: 4px solid #2196f3; margin-bottom: 25px;">
                                <tr>
                                    <td style="padding: 20px;">
                                        <h3 style="color: #1976d2; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">📱 Hỗ Trợ & Dịch Vụ</h3>
                                        <div style="margin-bottom: 12px;">
                                            <div style="display: inline-block; background-color: #ffffff; padding: 8px 12px; border-radius: 6px; margin: 4px 0;">
                                                <span style="color: #1976d2; font-weight: 600;">📱</span>
                                                <span style="color: #333333; font-size: 13px; margin-left: 5px;">Quét QR code trong phòng để chat với lễ tân</span>
                                            </div>
                                        </div>
                                        <div style="margin-bottom: 12px;">
                                            <div style="display: inline-block; background-color: #ffffff; padding: 8px 12px; border-radius: 6px; margin: 4px 0;">
                                                <span style="color: #1976d2; font-weight: 600;">☎️</span>
                                                <span style="color: #333333; font-size: 13px; margin-left: 5px;">Gọi lễ tân: Nhấn số <strong>0</strong> trên điện thoại phòng</span>
                                            </div>
                                        </div>
                                        <div style="margin-bottom: 12px;">
                                            <div style="display: inline-block; background-color: #ffffff; padding: 8px 12px; border-radius: 6px; margin: 4px 0;">
                                                <span style="color: #1976d2; font-weight: 600;">📞</span>
                                                <span style="color: #333333; font-size: 13px; margin-left: 5px;">Hotline: <strong>{{ so_dien_thoai_khach_san }}</strong></span>
                                            </div>
                                        </div>
                                        <div>
                                            <div style="display: inline-block; background-color: #ffffff; padding: 8px 12px; border-radius: 6px; margin: 4px 0;">
                                                <span style="color: #1976d2; font-weight: 600;">🍽️</span>
                                                <span style="color: #333333; font-size: 13px; margin-left: 5px;">Đặt dịch vụ: Qua chat hoặc gọi lễ tân</span>
                                            </div>
                                        </div>
                                    </td>
                                </tr>
                            </table>
                            
                            {% if voucher_moi %}
                            <table width="100%" cellpadding="0" cellspacing="0" style="background: linear-gradient(135deg, #fff5d7 0%, #f9d787 100%); border-radius: 8px; border-left: 4px solid #f39c12; margin-bottom: 25px;">
                                <tr>
                                    <td style="padding: 20px;">
                                        <h3 style="color: #b36b00; margin: 0 0 12px 0; font-size: 16px; font-weight: 600;">🎁 Ưu đãi dành riêng cho bạn</h3>
                                        <p style="margin: 5px 0; color: #5d3b00; font-size: 14px;">Mã voucher: <strong>{{ voucher_moi_code }}</strong></p>
                                        {% if voucher_moi_discount %}
                                        <p style="margin: 5px 0; color: #5d3b00; font-size: 14px;">Giảm giá: <strong>{{ voucher_moi_discount }}%</strong></p>
                                        {% endif %}
                                        {% if voucher_moi_han %}
                                        <p style="margin: 5px 0; color: #5d3b00; font-size: 14px;">Hạn sử dụng: <strong>{{ voucher_moi_han }}</strong></p>
                                        {% endif %}
                                        <p style="margin: 10px 0 0 0; color: #5d3b00; font-size: 13px;">Giữ lại mã này và nhập khi đặt phòng lần tiếp theo để nhận ưu đãi hấp dẫn.</p>
                                    </td>
                                </tr>
                            </table>
                            {% endif %}
                            
                            <!-- Wish -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background: linear-gradient(135deg, #ffecd2 0%, #fcb69f 100%); border-radius: 8px;">
                                <tr>
                                    <td style="padding: 20px; text-align: center;">
                                        <p style="margin: 0; color: #8b4513; font-size: 16px; font-weight: 600; line-height: 1.6;">
                                            🎊 Chúc quý khách có kỳ nghỉ tuyệt vời!
                                        </p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Footer -->
                    <tr>
                        <td style="background-color: #f8f9fa; padding: 25px 30px; text-align: center; border-top: 1px solid #dee2e6;">
                            <p style="margin: 0 0 5px 0; color: #666666; font-size: 14px;">Trân trọng,</p>
                            <p style="margin: 0; color: #333333; font-size: 16px; font-weight: 600;">{{ ten_khach_san }}</p>
                            <p style="margin: 10px 0 0 0; color: #999999; font-size: 12px;">
                                Chúc quý khách có những trải nghiệm đáng nhớ!
                            </p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>
        '''
    },
    'invoice_notice': {
        'name': 'Email hóa đơn',
        'subject': 'Hóa đơn thanh toán #{{ ma_dat_phong }} - {{ ten_khach_san }}',
        'body': '''
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Hóa đơn thanh toán</title>
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f4;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f4f4f4; padding: 20px 0;">
        <tr>
            <td align="center">
                <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); overflow: hidden;">
                    <!-- Header -->
                    <tr>
                        <td style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 40px 30px; text-align: center;">
                            <div style="font-size: 48px; margin-bottom: 10px;">💰</div>
                            <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 600;">HÓA ĐƠN THANH TOÁN</h1>
                            <p style="color: #ffffff; margin: 10px 0 0 0; font-size: 18px; opacity: 0.95;">#{{ ma_dat_phong }}</p>
                        </td>
                    </tr>
                    
                    <!-- Body -->
                    <tr>
                        <td style="padding: 40px 30px;">
                            <p style="font-size: 16px; color: #333333; margin: 0 0 15px 0; line-height: 1.6;">
                                Kính thưa <strong style="color: #667eea;">{{ ten_khach_hang }}</strong>,
                            </p>
                            <p style="font-size: 15px; color: #555555; margin: 0 0 30px 0; line-height: 1.6;">
                                Cảm ơn quý khách đã tin tưởng và sử dụng dịch vụ của <strong>{{ ten_khach_san }}</strong>! Dưới đây là thông tin hóa đơn thanh toán của quý khách.
                            </p>
                            
                            <!-- Booking Info -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f8f9fa; border-radius: 8px; margin-bottom: 20px;">
                                <tr>
                                    <td style="padding: 20px;">
                                        <h3 style="color: #333333; margin: 0 0 15px 0; font-size: 16px; font-weight: 600;">📋 Thông Tin Đặt Phòng</h3>
                                        <table width="100%" cellpadding="6" cellspacing="0">
                                            <tr>
                                                <td style="color: #666666; font-size: 14px;">Phòng:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right;">{{ ten_phong }}{% if loai_phong %} <span style="color: #667eea;">({{ loai_phong }})</span>{% endif %}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #666666; font-size: 14px;">Check-in:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right;">{{ thoi_gian_nhan }}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #666666; font-size: 14px;">Check-out:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right;">{{ thoi_gian_tra }}</td>
                                            </tr>
                                            {% set so_luu_tru_value = so_luong_luu_tru or so_dem %}
                                            {% set don_vi_luu_tru_display = don_vi_luu_tru or ("đêm" if so_luu_tru_value else "") %}
                                            {% set nhan_luu_tru_display = nhan_luu_tru or "Số đêm lưu trú" %}
                                            {% if so_luu_tru_value %}
                                            <tr>
                                                <td style="color: #666666; font-size: 14px;">{{ nhan_luu_tru_display }}:</td>
                                                <td style="color: #333333; font-size: 14px; font-weight: 600; text-align: right;">{{ so_luu_tru_value }}{% if don_vi_luu_tru_display %} {{ don_vi_luu_tru_display }}{% endif %}</td>
                                            </tr>
                                            {% endif %}
                                        </table>
                                    </td>
                                </tr>
                            </table>
                            
                            <!-- Payment Details -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%); border-radius: 8px; margin-bottom: 20px;">
                                <tr>
                                    <td style="padding: 25px;">
                                        <h3 style="color: #2c3e50; margin: 0 0 20px 0; font-size: 18px; font-weight: 700; text-align: center;">💵 Chi Tiết Thanh Toán</h3>
                                        <table width="100%" cellpadding="10" cellspacing="0">
                                            {% if tien_phong %}
                                            <tr>
                                                <td style="color: #34495e; font-size: 15px; padding: 8px 0; border-bottom: 1px solid #bdc3c7;">Tiền phòng</td>
                                                <td style="color: #2c3e50; font-size: 15px; font-weight: 600; text-align: right; padding: 8px 0; border-bottom: 1px solid #bdc3c7;">{{ tien_phong }}</td>
                                            </tr>
                                            {% endif %}
                                            {% if tien_dich_vu and tien_dich_vu != "0 ₫" %}
                                            <tr>
                                                <td style="color: #34495e; font-size: 15px; padding: 8px 0; border-bottom: 1px solid #bdc3c7;">Tiền dịch vụ</td>
                                                <td style="color: #2c3e50; font-size: 15px; font-weight: 600; text-align: right; padding: 8px 0; border-bottom: 1px solid #bdc3c7;">{{ tien_dich_vu }}</td>
                                            </tr>
                                            {% endif %}
                                            {% if tien_phat and tien_phat != "0 ₫" %}
                                            <tr>
                                                <td style="color: #34495e; font-size: 15px; padding: 8px 0; border-bottom: 1px solid #bdc3c7;">Phí phát sinh</td>
                                                <td style="color: #e74c3c; font-size: 15px; font-weight: 600; text-align: right; padding: 8px 0; border-bottom: 1px solid #bdc3c7;">{{ tien_phat }}</td>
                                            </tr>
                                            {% endif %}
                                            <tr style="background-color: rgba(255,255,255,0.3);">
                                                <td style="color: #2c3e50; font-size: 17px; font-weight: 700; padding: 12px 0;">TỔNG CỘNG</td>
                                                <td style="color: #27ae60; font-size: 20px; font-weight: 700; text-align: right; padding: 12px 0;">{{ tong_tien }}</td>
                                            </tr>
                                            {% if tien_coc_display %}
                                            <tr>
                                                <td style="color: #34495e; font-size: 15px; padding: 8px 0; border-top: 2px dashed #bdc3c7;">Tiền cọc đã thanh toán</td>
                                                <td style="color: #27ae60; font-size: 15px; font-weight: 600; text-align: right; padding: 8px 0; border-top: 2px dashed #bdc3c7;">{{ tien_coc_display }}{% if hinh_thuc_coc %} ({{ hinh_thuc_coc }}){% endif %}</td>
                                            </tr>
                                            {% endif %}
                                            {% if so_tien_da_thanh_toan and so_tien_da_thanh_toan != "0 ₫" %}
                                            <tr>
                                                <td style="color: #34495e; font-size: 15px; padding: 8px 0;">Đã thanh toán</td>
                                                <td style="color: #2c3e50; font-size: 15px; font-weight: 600; text-align: right; padding: 8px 0;">{{ so_tien_da_thanh_toan }}</td>
                                            </tr>
                                            {% endif %}
                                            {% if hinh_thuc_thanh_toan %}
                                            <tr>
                                                <td style="color: #34495e; font-size: 15px; padding: 8px 0;">Hình thức thanh toán</td>
                                                <td style="color: #2c3e50; font-size: 15px; font-weight: 600; text-align: right; padding: 8px 0;">{{ hinh_thuc_thanh_toan }}</td>
                                            </tr>
                                            {% endif %}
                                            {% if tien_dich_vu_da_thanh_toan %}
                                            <tr>
                                                <td style="color: #34495e; font-size: 15px; padding: 8px 0;">Đã thanh toán dịch vụ</td>
                                                <td style="color: #27ae60; font-size: 15px; font-weight: 600; text-align: right; padding: 8px 0;">-{{ tien_dich_vu_da_thanh_toan }}</td>
                                            </tr>
                                            {% endif %}
                                            {% if con_lai and con_lai != "0 ₫" %}
                                            <tr style="background: linear-gradient(135deg, #ffd89b 0%, #19547b 100%);">
                                                <td style="color: #ffffff; font-size: 18px; font-weight: 700; padding: 15px 10px; border-radius: 6px;">CÒN LẠI PHẢI TRẢ</td>
                                                <td style="color: #ffffff; font-size: 22px; font-weight: 700; text-align: right; padding: 15px 10px; border-radius: 6px;">{{ con_lai }}</td>
                                            </tr>
                                            {% endif %}
                                        </table>
                                    </td>
                                </tr>
                            </table>
                            
                            <!-- Service Details -->
                            {% if chi_tiet_dich_vu and chi_tiet_dich_vu != "Không sử dụng dịch vụ" %}
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #e8f5e9; border-radius: 8px; border-left: 4px solid #4caf50; margin-bottom: 25px;">
                                <tr>
                                    <td style="padding: 20px;">
                                        <h3 style="color: #2e7d32; margin: 0 0 12px 0; font-size: 16px; font-weight: 600;">📋 Chi Tiết Dịch Vụ Đã Sử Dụng</h3>
                                        <div style="color: #1b5e20; font-size: 13px; line-height: 1.8; white-space: pre-line;">{{ chi_tiet_dich_vu }}</div>
                                    </td>
                                </tr>
                            </table>
                            {% endif %}
                            
                            <!-- Thank You Message -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background: linear-gradient(135deg, #ffeaa7 0%, #fdcb6e 100%); border-radius: 8px; margin-bottom: 20px;">
                                <tr>
                                    <td style="padding: 20px; text-align: center;">
                                        <p style="margin: 0; color: #8b4513; font-size: 16px; font-weight: 600; line-height: 1.8;">
                                            ⭐ Cảm ơn quý khách đã tin tưởng!<br>
                                            ⭐ Rất mong được phục vụ quý khách trong tương lai!
                                        </p>
                                    </td>
                                </tr>
                            </table>
                            
                            <!-- Contact Info -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #e3f2fd; border-radius: 8px;">
                                <tr>
                                    <td style="padding: 20px;">
                                        <h3 style="color: #1976d2; margin: 0 0 12px 0; font-size: 16px; font-weight: 600;">📞 Thông Tin Liên Hệ</h3>
                                        <p style="margin: 5px 0; color: #333333; font-size: 14px;">
                                            <strong>Hotline:</strong> {{ so_dien_thoai_khach_san }}
                                        </p>
                                        <p style="margin: 5px 0; color: #333333; font-size: 14px;">
                                            <strong>Địa chỉ:</strong> {{ dia_chi_khach_san }}
                                        </p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Footer -->
                    <tr>
                        <td style="background-color: #f8f9fa; padding: 25px 30px; text-align: center; border-top: 1px solid #dee2e6;">
                            <p style="margin: 0 0 5px 0; color: #666666; font-size: 14px;">Trân trọng,</p>
                            <p style="margin: 0 0 3px 0; color: #333333; font-size: 16px; font-weight: 600;">{{ ten_khach_san }}</p>
                            <p style="margin: 0; color: #999999; font-size: 13px;">📞 {{ so_dien_thoai_khach_san }} | 📍 {{ dia_chi_khach_san }}</p>
                            <p style="margin: 15px 0 0 0; color: #999999; font-size: 12px;">
                                Email này được gửi tự động từ hệ thống quản lý khách sạn
                            </p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>

        '''
    }
}


def ensure_email_templates():
    """
    Ensure email templates exist in database and update them with latest defaults.
    This allows code changes to EMAIL_TEMPLATE_DEFAULTS to propagate to database.
    """
    updated = False
    for key, meta in EMAIL_TEMPLATE_DEFAULTS.items():
        existing = EmailTemplate.query.filter_by(key=key).first()
        if not existing:
            # Create new template
            db.session.add(EmailTemplate(key=key, subject=meta['subject'], body=meta['body']))
            updated = True
        else:
            # Update existing template if content changed
            if existing.subject != meta['subject'] or existing.body != meta['body']:
                existing.subject = meta['subject']
                existing.body = meta['body']
                updated = True
    if updated:
        db.session.commit()
        app.logger.info('Email templates updated from defaults.')


def get_email_settings():
    default_host = _default_smtp_value('SMTP_HOST', DEFAULT_SMTP_CONFIG['SMTP_HOST'])
    default_port = _default_smtp_value('SMTP_PORT', DEFAULT_SMTP_CONFIG['SMTP_PORT'])
    default_username = _default_smtp_value('SMTP_USERNAME', DEFAULT_SMTP_CONFIG['SMTP_USERNAME'])
    default_sender = _default_smtp_value('SMTP_SENDER_EMAIL', DEFAULT_SMTP_CONFIG['SMTP_SENDER_EMAIL'])
    default_tls = _default_smtp_value('SMTP_USE_TLS', DEFAULT_SMTP_CONFIG['SMTP_USE_TLS'])
    default_ssl = _default_smtp_value('SMTP_USE_SSL', DEFAULT_SMTP_CONFIG['SMTP_USE_SSL'])
    return {
        'smtp_host': get_config_value('SMTP_HOST', default_host),
        'smtp_port': get_config_value('SMTP_PORT', default_port),
        'smtp_username': get_config_value('SMTP_USERNAME', default_username),
        'smtp_password': get_config_value('SMTP_PASSWORD', os.getenv('SMTP_PASSWORD', '')),
        'sender_email': get_config_value('SMTP_SENDER_EMAIL', default_sender),
        'smtp_use_tls': _interpret_bool(get_config_value('SMTP_USE_TLS', default_tls), default_tls),
        'smtp_use_ssl': _interpret_bool(get_config_value('SMTP_USE_SSL', default_ssl), default_ssl)
    }


def get_hotel_profile():
    return {
        'name': get_config_value('HOTEL_NAME', 'Khách sạn PTIT'),
        'phone': get_config_value('HOTEL_PHONE', '028 8888 9999'),
        'address': get_config_value('HOTEL_ADDRESS', 'Khách sạn Ptit, Km10 Hà Đông'),
        'email': get_config_value('HOTEL_EMAIL', '')
    }


def build_booking_email_context(dp):
    """
    Build comprehensive email context with all necessary variables.
    Enhanced with more details for better customer communication.
    """
    hotel = get_hotel_profile()
    
    # Calculate number of nights
    so_dem = 0
    so_gio = 0
    if dp.ngay_nhan and dp.ngay_tra:
        delta_hours = (dp.ngay_tra - dp.ngay_nhan).total_seconds() / 3600
        if dp.hinh_thuc_thue == 'gio':
            so_gio = max(1, math.ceil(delta_hours))
        else:
            so_dem = max(1, math.ceil(delta_hours / 24))
    
    # Calculate remaining amount
    con_lai = (dp.tong_thanh_toan or 0) - (dp.tien_coc or 0)
    
    # Get room type name
    loai_phong = ''
    if dp.phong and dp.phong.loai:
        loai_phong = dp.phong.loai.ten
    
    # Build service details
    chi_tiet_dich_vu = ''
    if hasattr(dp, 'dich_vu_su_dung') and dp.dich_vu_su_dung:
        dv_lines = []
        for sudung in dp.dich_vu_su_dung:
            if sudung.dich_vu:
                line = f"  - {sudung.dich_vu.ten}: {sudung.so_luong} x {vnd(sudung.dich_vu.gia)} = {vnd(sudung.so_luong * sudung.dich_vu.gia)}"
                dv_lines.append(line)
        chi_tiet_dich_vu = '\n'.join(dv_lines) if dv_lines else 'Không sử dụng dịch vụ nào.'
    
    context = {
        # Customer info
        'ten_khach_hang': dp.khachhang.ho_ten if dp.khachhang else 'Quý khách',
        
        # Hotel info
        'ten_khach_san': hotel['name'],
        'so_dien_thoai_khach_san': hotel['phone'],
        'dia_chi_khach_san': hotel['address'],
        
        # Booking info
        'ma_dat_phong': dp.id,
        'ten_phong': dp.phong.ten if dp.phong else '',
        'loai_phong': loai_phong,
        
        # Time info
        'thoi_gian_nhan': fmt_dt(dp.ngay_nhan),
        'thoi_gian_tra': fmt_dt(dp.ngay_tra),
        'thoi_gian_nhan_thuc_te': fmt_dt(dp.thuc_te_nhan) if dp.thuc_te_nhan else '',
        'so_dem': so_dem,
        'nhan_luu_tru': 'Số giờ lưu trú' if dp.hinh_thuc_thue == 'gio' else 'Số đêm lưu trú',
        'don_vi_luu_tru': 'giờ' if dp.hinh_thuc_thue == 'gio' else 'đêm',
        'so_luong_luu_tru': (so_gio if dp.hinh_thuc_thue == 'gio' else so_dem) or '',
        
        # Money info
        'tong_tien': vnd(dp.tong_thanh_toan or 0),
        'tien_coc': vnd(dp.tien_coc or 0) if dp.tien_coc else '',
        'tien_phong': vnd(dp.tien_phong or 0) if hasattr(dp, 'tien_phong') and dp.tien_phong else '',
        'tien_dich_vu': vnd(dp.tien_dv or 0) if hasattr(dp, 'tien_dv') and dp.tien_dv else '',
        'tien_phat': vnd(dp.tien_phat or 0) if hasattr(dp, 'tien_phat') and dp.tien_phat else '',
        'con_lai': vnd(con_lai) if con_lai > 0 else '',
        'tien_dich_vu_da_thanh_toan': '',
        'voucher_moi': False,
        'voucher_moi_code': '',
        'voucher_moi_discount': '',
        'voucher_moi_han': '',
        
        # Service details
        'chi_tiet_dich_vu': chi_tiet_dich_vu,
    }
    
    # Add voucher info if exists
    if dp.voucher:
        context['voucher_discount_percent'] = dp.voucher.discount_percent
        context['voucher_code'] = dp.voucher.code
    
    return context


def _pdf_safe_text(value):
    if value is None:
        return ''
    if not isinstance(value, str):
        value = str(value)
    try:
        value.encode('latin-1')
        return value
    except UnicodeEncodeError:
        normalized = unicodedata.normalize('NFKD', value)
        return ''.join(ch for ch in normalized if not unicodedata.combining(ch))


def generate_invoice_pdf(dp, dich_vu_su_dung):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError('Chưa cài đặt thư viện reportlab để tạo file PDF hóa đơn.') from exc

    so_dem, tien_phong, tong_tien_dv, tien_phat, tong, checkin, checkout, don_vi_tinh, so_luong_tinh = snapshot_and_bill(dp)
    tong_thanh_toan = dp.tong_thanh_toan or tong
    tien_coc = dp.tien_coc or 0
    tien_dv_da_thanh_toan = sum((dv.dichvu.gia * dv.so_luong) for dv in dich_vu_su_dung)
    tien_con_lai = max(0, tong_thanh_toan - tien_coc - tien_dv_da_thanh_toan)

    hotel = get_hotel_profile()

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    text = pdf.beginText(40, height - 50)
    text.setFont('Helvetica-Bold', 16)
    text.textLine(_pdf_safe_text(hotel.get('name') or 'Hoa don khach san'))
    text.setFont('Helvetica', 10)
    text.textLine(_pdf_safe_text(f"Dia chi: {hotel.get('address', '')}"))
    if hotel.get('phone'):
        text.textLine(_pdf_safe_text(f"Hotline: {hotel['phone']}"))
    if hotel.get('email'):
        text.textLine(_pdf_safe_text(f"Email: {hotel['email']}"))

    text.textLine('')
    text.setFont('Helvetica-Bold', 13)
    text.textLine(_pdf_safe_text(f"Hoa don dat phong #{dp.id}"))
    text.setFont('Helvetica', 10)
    text.textLine(_pdf_safe_text(f"Ngay lap: {fmt_dt(datetime.now())}"))

    text.textLine('')
    text.setFont('Helvetica-Bold', 11)
    text.textLine('Thong tin khach hang')
    text.setFont('Helvetica', 10)
    text.textLine(_pdf_safe_text(f"Ho ten: {dp.khachhang.ho_ten}"))
    text.textLine(_pdf_safe_text(f"CMND/CCCD: {dp.khachhang.cmnd}"))
    if dp.khachhang.sdt:
        text.textLine(_pdf_safe_text(f"So dien thoai: {dp.khachhang.sdt}"))
    if dp.khachhang.email:
        text.textLine(_pdf_safe_text(f"Email: {dp.khachhang.email}"))
    if dp.khachhang.dia_chi:
        text.textLine(_pdf_safe_text(f"Dia chi: {dp.khachhang.dia_chi}"))

    text.textLine('')
    text.setFont('Helvetica-Bold', 11)
    text.textLine('Thong tin luu tru')
    text.setFont('Helvetica', 10)
    text.textLine(_pdf_safe_text(f"Phong: {dp.phong.ten} - {dp.phong.loai.ten}"))
    text.textLine(_pdf_safe_text(f"Hinh thuc: {'Theo gio' if dp.hinh_thuc_thue == 'gio' else 'Theo ngay'}"))
    text.textLine(_pdf_safe_text(f"Thoi gian nhan: {fmt_dt(dp.ngay_nhan)}"))
    text.textLine(_pdf_safe_text(f"Thoi gian tra: {fmt_dt(dp.ngay_tra)}"))
    if dp.thuc_te_nhan:
        text.textLine(_pdf_safe_text(f"Nhan thuc te: {fmt_dt(dp.thuc_te_nhan)}"))
    if dp.thuc_te_tra:
        text.textLine(_pdf_safe_text(f"Tra thuc te: {fmt_dt(dp.thuc_te_tra)}"))

    text.textLine('')
    text.setFont('Helvetica-Bold', 11)
    text.textLine('Chi tiet dich vu')
    text.setFont('Helvetica', 10)
    if dich_vu_su_dung:
        for dv in dich_vu_su_dung:
            dong = f"- {dv.dichvu.ten}: {dv.so_luong} x {vnd(dv.dichvu.gia)} = {vnd(dv.dichvu.gia * dv.so_luong)}"
            text.textLine(_pdf_safe_text(dong))
    else:
        text.textLine(_pdf_safe_text('Khong su dung dich vu nao.'))

    text.textLine('')
    text.setFont('Helvetica-Bold', 11)
    text.textLine('Tong ket chi phi')
    text.setFont('Helvetica', 10)
    text.textLine(_pdf_safe_text(f"Tien phong ({so_luong_tinh} {don_vi_tinh}): {vnd(tien_phong)}"))
    text.textLine(_pdf_safe_text(f"Tien dich vu: {vnd(tong_tien_dv)}"))
    text.textLine(_pdf_safe_text(f"Tien phat: {vnd(tien_phat)}"))
    if dp.voucher:
        text.textLine(_pdf_safe_text(f"Voucher ap dung: {dp.voucher.code} ({dp.voucher.discount_percent}%)"))
    text.textLine(_pdf_safe_text(f"Tong thanh toan: {vnd(tong_thanh_toan)}"))
    text.textLine(_pdf_safe_text(f"Tien coc da thu: {vnd(tien_coc)}"))
    text.textLine(_pdf_safe_text(f"Tien dich vu da thanh toan: {vnd(tien_dv_da_thanh_toan)}"))
    text.textLine(_pdf_safe_text(f"So tien con lai: {vnd(tien_con_lai)}"))

    text.textLine('')
    text.setFont('Helvetica', 9)
    text.textLine(_pdf_safe_text('Xin cam on quy khach da lua chon khach san cua chung toi!'))

    pdf.drawText(text)
    pdf.showPage()
    pdf.save()

    buffer.seek(0)
    return buffer.read()


def render_email_content(template_key, context):
    tpl = EmailTemplate.query.filter_by(key=template_key).first()
    if not tpl:
        raise ValueError(f'Không tìm thấy mẫu email với khóa "{template_key}".')
    subject_tpl = app.jinja_env.from_string(tpl.subject or '')
    body_tpl = app.jinja_env.from_string(tpl.body or '')
    subject = subject_tpl.render(**context)
    body = body_tpl.render(**context)
    return subject.strip(), body


def send_email_with_template(template_key, recipient_email, context, attachments=None, datphong_id=None, khachhang_id=None):
    """
    Gửi email với template và lưu lịch sử vào database
    
    Args:
        template_key: Key của template email
        recipient_email: Email người nhận
        context: Dictionary chứa dữ liệu để render template
        attachments: List các file đính kèm (optional)
        datphong_id: ID của đặt phòng liên quan (optional)
        khachhang_id: ID của khách hàng (optional)
    
    Returns:
        True nếu gửi thành công
    
    Raises:
        ValueError: Nếu email không hợp lệ
        RuntimeError: Nếu chưa cấu hình SMTP
    """
    if not recipient_email:
        raise ValueError('Địa chỉ email người nhận không hợp lệ.')
    settings = get_email_settings()
    if not settings['smtp_host'] or not settings['sender_email']:
        raise RuntimeError('Chưa cấu hình SMTP hoặc địa chỉ gửi đi.')

    subject, body = render_email_content(template_key, context)
    
    # Lấy tên người nhận từ context nếu có
    recipient_name = context.get('ten_khach') or context.get('ho_ten') or None
    
    # Tạo email log với status pending
    email_log = EmailLog(
        recipient_email=recipient_email,
        recipient_name=recipient_name,
        template_key=template_key,
        subject=subject,
        body=body,
        status='pending',
        sent_by=current_user.id if current_user.is_authenticated else None,
        datphong_id=datphong_id,
        khachhang_id=khachhang_id
    )
    db.session.add(email_log)
    db.session.commit()
    
    # Check if body is HTML - improved detection
    is_html = (
        '<!DOCTYPE html>' in body or 
        '<html' in body.lower() or 
        '<div' in body.lower() or 
        '<table' in body.lower() or
        '<p>' in body.lower() or
        '<br' in body.lower()
    )
    
    # Create proper MIME structure
    # If we have attachments, we need a 'mixed' container
    # Otherwise, for HTML we use 'alternative'
    if attachments:
        # Root message is 'mixed' to hold both content and attachments
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = settings['sender_email']
        msg['To'] = recipient_email
        
        if is_html:
            # Create 'alternative' part for text/html content
            msg_alternative = MIMEMultipart('alternative')
            
            # Add plain text version first
            plain_text = "Vui lòng xem email ở định dạng HTML trong trình đọc email hỗ trợ HTML."
            text_part = MIMEText(plain_text, 'plain', 'utf-8')
            msg_alternative.attach(text_part)
            
            # Add HTML version last (preferred)
            html_part = MIMEText(body, 'html', 'utf-8')
            msg_alternative.attach(html_part)
            
            # Attach the alternative part to root message
            msg.attach(msg_alternative)
        else:
            # Just plain text
            text_part = MIMEText(body, 'plain', 'utf-8')
            msg.attach(text_part)
            
    else:
        # No attachments
        if is_html:
            # Use 'alternative' for HTML emails without attachments
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = settings['sender_email']
            msg['To'] = recipient_email
            
            # Add plain text version first (fallback)
            plain_text = "Vui lòng xem email ở định dạng HTML trong trình đọc email hỗ trợ HTML."
            text_part = MIMEText(plain_text, 'plain', 'utf-8')
            msg.attach(text_part)
            
            # Add HTML version last (preferred)
            html_part = MIMEText(body, 'html', 'utf-8')
            msg.attach(html_part)
        else:
            # Plain text email without attachments
            msg = MIMEMultipart()
            msg['Subject'] = subject
            msg['From'] = settings['sender_email']
            msg['To'] = recipient_email
            text_part = MIMEText(body, 'plain', 'utf-8')
            msg.attach(text_part)
    
    # Add attachments if any
    if attachments:
        for attachment in attachments:
            filename = attachment.get('filename')
            content = attachment.get('content')
            mime_type = attachment.get('mime_type', 'application/octet-stream')
            if not filename or content is None:
                continue
            
            # Create attachment part
            maintype, subtype = mime_type.split('/', 1) if '/' in mime_type else ('application', 'octet-stream')
            part = MIMEBase(maintype, subtype)
            part.set_payload(content)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)

    smtp_port = settings['smtp_port'] or '587'
    try:
        port = int(smtp_port)
    except (TypeError, ValueError):
        port = 587

    try:
        server = None
        if settings['smtp_use_ssl']:
            server = smtplib.SMTP_SSL(settings['smtp_host'], port, timeout=30)
        else:
            server = smtplib.SMTP(settings['smtp_host'], port, timeout=30)
            if settings['smtp_use_tls']:
                server.starttls()

        username = settings['smtp_username'] or settings['sender_email']
        password = settings['smtp_password']
        if username and password:
            server.login(username, password)

        # Send email as string for MIME messages  
        server.sendmail(settings['sender_email'], recipient_email, msg.as_string())
        
        # Cập nhật log: gửi thành công
        email_log.status = 'success'
        email_log.error_message = None
        db.session.commit()
        
    except Exception as e:
        # Cập nhật log: gửi thất bại
        email_log.status = 'failed'
        email_log.error_message = str(e)
        db.session.commit()
        raise
    finally:
        if 'server' in locals() and server:
            try:
                server.quit()
            except Exception:
                pass

    return True

def snapshot_and_bill(dp, now=None):
    if now is None: now = datetime.now()
    checkin = dp.thuc_te_nhan or dp.ngay_nhan
    actual_checkout = dp.thuc_te_tra or now
    scheduled_duration_hours = (dp.ngay_tra - dp.ngay_nhan).total_seconds() / 3600
    don_vi_tinh, so_luong_tinh, tien_phong = '', 0, 0
    if dp.hinh_thuc_thue == 'gio':
        don_vi_tinh = 'giờ'
        so_luong_tinh = max(1, math.ceil(scheduled_duration_hours))
        gia_phong_theo_gio = int(dp.phong.loai.gia * 0.2)
        tien_phong = so_luong_tinh * gia_phong_theo_gio
    else:
        don_vi_tinh = 'đêm'
        so_luong_tinh = max(1, math.ceil(scheduled_duration_hours / 24))
        tien_phong = dp.phong.loai.gia * so_luong_tinh
    tien_phat = 0
    # Phí phạt chỉ được tính nếu khách trả phòng MUỘN HƠN lịch đặt (dp.ngay_tra)
    if actual_checkout > dp.ngay_tra:
        gio_qua_han = math.ceil((actual_checkout - dp.ngay_tra).total_seconds() / 3600)
        if gio_qua_han > 0:
            tien_phat = gio_qua_han * 300000

    actual_duration_hours = (actual_checkout - checkin).total_seconds() / 3600
    so_dem = max(1, math.ceil(actual_duration_hours / 24)) # Vẫn giữ để tham khảo
    
    # Tính TỔNG tiền dịch vụ CHỈ từ các dịch vụ ĐÃ THANH TOÁN
    # Không tính các dịch vụ: chua_thanh_toan, cho_xac_nhan (chưa xác nhận)
    dich_vu_da_thanh_toan = SuDungDichVu.query.filter_by(
        datphong_id=dp.id, 
        trang_thai='da_thanh_toan'
    ).all()
    tong_tien_dv = sum(r.dichvu.gia * r.so_luong for r in dich_vu_da_thanh_toan)

    tong = tien_phong + tong_tien_dv + tien_phat

    return so_dem, tien_phong, tong_tien_dv, tien_phat, tong, checkin, actual_checkout, don_vi_tinh, so_luong_tinh


def build_invoice_context(dp):
    dich_vu_su_dung = SuDungDichVu.query.filter_by(datphong_id=dp.id, trang_thai='da_thanh_toan').options(joinedload(SuDungDichVu.dichvu)).all()
    so_dv_chua_tinh = SuDungDichVu.query.filter(
        SuDungDichVu.datphong_id == dp.id,
        SuDungDichVu.trang_thai != 'da_thanh_toan'
    ).count()
    tien_dv_da_thanh_toan = sum(dv.dichvu.gia * dv.so_luong for dv in dich_vu_su_dung)

    voucher_obj = dp.voucher
    voucher_discount_amount = 0
    voucher_applied_percent = voucher_obj.discount_percent if voucher_obj else 0

    so_dem, tien_phong_goc, tong_tien_dv_calc, tien_phat_calc, tong_calc_raw, checkin, checkout, don_vi_tinh, so_luong_tinh = snapshot_and_bill(dp)

    tien_phong_calc = tien_phong_goc
    if voucher_obj:
        voucher_discount_amount = int(tien_phong_goc * voucher_obj.discount_percent / 100)
        tien_phong_calc = max(0, tien_phong_calc - voucher_discount_amount)
    tong_calc = tien_phong_calc + tong_tien_dv_calc + tien_phat_calc

    if dp.trang_thai == 'da_thanh_toan':
        display_tien_phong = dp.tien_phong or tien_phong_calc
        display_tien_dv = dp.tien_dv or tong_tien_dv_calc
        display_tien_phat = dp.tien_phat or tien_phat_calc
        display_tong = dp.tong_thanh_toan or tong_calc
    else:
        display_tien_phong = tien_phong_calc
        display_tien_dv = tong_tien_dv_calc
        display_tien_phat = tien_phat_calc
        display_tong = tong_calc

    remaining_amount = max(0, display_tong - (dp.tien_coc or 0) - tien_dv_da_thanh_toan)
    qr_amount = 0 if dp.trang_thai in ('da_thanh_toan', 'huy') else remaining_amount

    qr_code_url = None
    qr_description = quote(f"TT HD{dp.id} {dp.khachhang.ho_ten}")
    if qr_amount > 0:
        qr_code_url = (
            f"https://img.vietqr.io/image/{VIETQR_BANK_ID}-{VIETQR_ACCOUNT_NO}-compact2.png"
            f"?amount={int(qr_amount)}&addInfo={qr_description}"
        )

    voucher_policy_percent, _ = get_voucher_config()

    template_ctx = {
        'dich_vu_su_dung': dich_vu_su_dung,
        'so_dv_chua_tinh': so_dv_chua_tinh,
        'tien_da_tra_truoc': tien_dv_da_thanh_toan,
        'voucher_obj': voucher_obj,
        'voucher_discount_amount': voucher_discount_amount,
        'voucher_applied_percent': voucher_applied_percent,
        'voucher_policy_percent': voucher_policy_percent,
        'so_dem': so_dem,
        'don_vi_tinh': don_vi_tinh,
        'so_luong_tinh': so_luong_tinh,
        'checkin': checkin,
        'checkout': checkout,
        'tien_phong': display_tien_phong,
        'tien_phong_goc': tien_phong_goc,
        'tien_dv': display_tien_dv,
        'tien_phat': display_tien_phat,
        'tong': display_tong,
        'tien_con_lai': remaining_amount,  # Always show the correct remaining amount
        'qr_code_url': qr_code_url,
        'qr_amount': qr_amount,
        'qr_description': qr_description,
        'qr_bank_id': VIETQR_BANK_ID,
        'qr_bank_name': VIETQR_BANK_NAME,
        'qr_account_no': VIETQR_ACCOUNT_NO,
        'qr_account_name': VIETQR_ACCOUNT_NAME,
        'payment_method': dp.phuong_thuc_thanh_toan,
        'payment_method_label': PAYMENT_METHOD_LABELS.get(dp.phuong_thuc_thanh_toan),
        'payment_confirmed': dp.trang_thai in ('da_thanh_toan', 'huy'),
        'tien_coc': dp.tien_phat if dp.trang_thai == 'huy' else (dp.tien_coc or 0),
        'coc_da_thanh_toan': bool(dp.coc_da_thanh_toan),
        'phuong_thuc_coc_label': PAYMENT_METHOD_LABELS.get(dp.phuong_thuc_coc),
        'ghi_chu_mat_coc': 'Mất cọc do không đến nhận phòng đúng giờ.' if dp.trang_thai == 'huy' else None
    }
    template_ctx['tien_da_thanh_toan'] = max(0, template_ctx['tong'] - template_ctx['tien_con_lai'])

    calc_values = {
        'tien_phong': tien_phong_calc,
        'tien_dv': tong_tien_dv_calc,
        'tien_phat': tien_phat_calc,
        'tong': tong_calc
    }

    return template_ctx, calc_values

def cancel_booking_for_no_show(dp, auto_cancel_minutes=None):
    """Mark a booking as cancelled due to no-show if it is past the configured deadline."""
    if not dp or dp.trang_thai != 'dat' or dp.thuc_te_nhan is not None:
        return False

    # Nếu phòng vẫn đang có khách ở (quá hạn trả), không tự động hủy
    if dp.phong and dp.phong.trang_thai == 'dang_o':
        active_stay = DatPhong.query.filter(
            DatPhong.phong_id == dp.phong_id,
            DatPhong.trang_thai == 'nhan',
            DatPhong.id != dp.id
        ).first()
        if active_stay:
            return False

    minutes = auto_cancel_minutes if auto_cancel_minutes is not None else get_config_int('auto_cancel_minutes', 5)
    if minutes <= 0:
        return False

    reference_time = dp.auto_confirmed_at or dp.created_at or dp.ngay_nhan
    if not reference_time:
        return False

    deadline = reference_time + timedelta(minutes=minutes)
    if datetime.now() < deadline:
        return False

    now = datetime.now()
    dp.trang_thai = 'huy'
    dp.tong_thanh_toan = dp.tien_coc
    dp.tien_phat = dp.tien_coc
    dp.tien_phong = 0
    dp.tien_coc = 0
    dp.thuc_te_tra = now
    dp.phuong_thuc_thanh_toan = 'qr'
    dp.coc_da_thanh_toan = True

    other_booking = DatPhong.query.filter(
        DatPhong.phong_id == dp.phong_id,
        DatPhong.id != dp.id,
        DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES),
        ~db.or_(DatPhong.ngay_tra <= dp.ngay_nhan, DatPhong.ngay_nhan >= dp.ngay_tra)
    ).first()

    dp.phong.trang_thai = 'trong' if not other_booking else 'da_dat'
    return True


def huy_dat_phong_khong_den():
    with app.app_context():
        minutes = get_config_int('auto_cancel_minutes', 5)
        bookings = DatPhong.query.filter(
            DatPhong.trang_thai == 'dat',
            DatPhong.thuc_te_nhan.is_(None)
        ).all()
        if not bookings:
            return
        count = 0
        for dp in bookings:
            if cancel_booking_for_no_show(dp, minutes):
                db.session.commit()
                count += 1
        if count > 0:
            app.logger.info(
                f'Đã tự động hủy {count} đặt phòng không đến trong {minutes} phút. '
                'Doanh thu từ tiền cọc đã được ghi nhận.'
            )
            # flash(f'Đã tự động hủy {count} đặt phòng không đến trong 5 phút. Doanh thu từ tiền cọc đã được ghi nhận.', 'success')


@app.route('/api/bookings/<int:dat_id>/auto-cancel', methods=['POST'])
@login_required
@permission_required('bookings.cancel')
def api_auto_cancel_booking(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    minutes = get_config_int('auto_cancel_minutes', 5)
    if cancel_booking_for_no_show(dp, minutes):
        db.session.commit()
        socketio.emit('booking_auto_cancelled', {'booking_id': dp.id})
        return jsonify({'status': 'cancelled'})
    return jsonify({'status': 'pending'})


@app.route('/api/bookings/<int:dat_id>/mark-waiting', methods=['POST'])
@login_required
@permission_required('bookings.manage_waiting')
def api_mark_booking_waiting(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if dp.trang_thai not in {'dat', 'waiting'}:
        return jsonify({'status': 'invalid'}), 400

    dp.trang_thai = 'waiting'
    dp.auto_confirmed_at = None
    reason_note = 'Chuyển sang booking chờ do phòng chưa sẵn sàng.'
    dp.ghi_chu = ((dp.ghi_chu or '') + ' ' + reason_note).strip()

    if dp.khachhang:
        tn = TinNhan(
            datphong_id=dp.id,
            nguoi_gui='he_thong',
            noi_dung='Booking được chuyển sang danh sách chờ vì phòng chưa sẵn sàng. Bộ phận lễ tân sẽ liên hệ lại.',
            thoi_gian=datetime.now(),
            trang_thai='chua_doc'
        )
        db.session.add(tn)

    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route('/api/bookings/<int:dat_id>/refund-overstay', methods=['POST'])
@login_required
@permission_required('bookings.cancel')
def api_refund_overstay_booking(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if dp.trang_thai not in {'dat', 'waiting'}:
        return jsonify({'status': 'invalid'}), 400

    refund_amount = dp.tien_coc or 0
    dp.trang_thai = 'huy'
    dp.tien_phong = 0
    dp.tien_dv = 0
    dp.tien_phat = 0
    dp.tong_thanh_toan = 0
    dp.phuong_thuc_thanh_toan = None
    dp.phuong_thuc_coc = None
    dp.coc_da_thanh_toan = False
    dp.tien_coc = 0
    dp.thuc_te_nhan = None
    dp.thuc_te_tra = datetime.now()
    refund_note = f'Hoàn tiền do phòng bị chiếm dụng, hoàn {vnd(refund_amount)}.'
    dp.ghi_chu = ((dp.ghi_chu or '') + ' ' + refund_note).strip()

    if dp.khachhang:
        tn = TinNhan(
            datphong_id=dp.id,
            nguoi_gui='he_thong',
            noi_dung='Đơn đặt phòng được hoàn tiền vì phòng chưa sẵn sàng. Chúng tôi xin lỗi về sự bất tiện.',
            thoi_gian=datetime.now(),
            trang_thai='chua_doc'
        )
        db.session.add(tn)

    db.session.commit()
    return jsonify({'status': 'ok'})


def clean_expired_payment_sessions():
    with app.app_context():
        ttl = get_payment_session_ttl()
        cutoff = datetime.now() - ttl
        expired = PaymentSession.query.filter(PaymentSession.created_at < cutoff).all()
        if not expired:
            return

        deposit_tokens = [ps.token for ps in expired if ps.kind == 'deposit']
        for ps in expired:
            db.session.delete(ps)

        if deposit_tokens:
            DatPhong.query.filter(DatPhong.payment_token.in_(deposit_tokens)).update(
                {'payment_token': None},
                synchronize_session=False
            )

        db.session.commit()
        print(f"Cleaned {len(expired)} expired payment sessions")



def cleanup_expired_data():
    """Cleanup expired payment sessions and vouchers."""
    now = datetime.now()
    ttl = get_payment_session_ttl()
    expired_sessions = PaymentSession.query.filter(
        PaymentSession.created_at < now - ttl
    ).all()
    for session in expired_sessions:
        db.session.delete(session)
    
    expired_vouchers = Voucher.query.filter(Voucher.expires_at < now, Voucher.is_used == False).all()
    for voucher in expired_vouchers:
        db.session.delete(voucher)
    
    db.session.commit()
    app.logger.info(f"Cleaned up {len(expired_sessions)} expired sessions and {len(expired_vouchers)} expired vouchers")

def huy_dat_phong_timeout():
    """Tự động hủy các booking có payment session đã hết thời gian."""
    with app.app_context():
        # Lấy thời gian timeout từ cấu hình
        timeout_setting = HeThongCauHinh.query.filter_by(key='payment_timeout_minutes').first()
        timeout_minutes = int(timeout_setting.value) if timeout_setting and timeout_setting.value else 5
        timeout_minutes = max(1, timeout_minutes)
        timeout_timedelta = timedelta(minutes=timeout_minutes)

        now = datetime.now()
        cutoff = now - timeout_timedelta
        
        # Tìm các booking đang chờ xác nhận
        pending_bookings = DatPhong.query.filter_by(trang_thai='cho_xac_nhan').all()
        
        cancelled_count = 0
        for dp in pending_bookings:
            should_cancel = False
            
            if dp.payment_token:
                # Có payment token, kiểm tra session có expired không
                session = PaymentSession.query.filter_by(token=dp.payment_token).first()
                if not session or session.created_at < cutoff:
                    should_cancel = True
            else:
                # Không có payment token, kiểm tra thời gian tạo booking
                if dp.created_at < cutoff:
                    should_cancel = True
            
            if should_cancel:
                app.logger.info(f"Cancelling expired booking ID {dp.id}, created at {dp.created_at}")
                
                # Hủy booking với trạng thái đặc biệt để không hiển thị trong quản lý hóa đơn
                dp.trang_thai = 'huy_timeout'
                dp.thuc_te_tra = now
                dp.tong_thanh_toan = 0  # Không ghi nhận doanh thu vì chưa thanh toán
                dp.tien_phat = 0
                dp.tien_phong = 0
                dp.tien_coc = 0
                dp.phuong_thuc_thanh_toan = None
                dp.coc_da_thanh_toan = False
                
                # Cập nhật trạng thái phòng
                other_booking = DatPhong.query.filter(
                    DatPhong.phong_id == dp.phong_id,
                    DatPhong.id != dp.id,
                    DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES),
                    ~db.or_(DatPhong.ngay_tra <= dp.ngay_nhan, DatPhong.ngay_nhan >= dp.ngay_tra)
                ).first()
                if not other_booking:
                    dp.phong.trang_thai = 'trong'
                else:
                    dp.phong.trang_thai = 'da_dat'
                
                # Xóa payment session nếu có
                if dp.payment_token:
                    session = PaymentSession.query.filter_by(token=dp.payment_token).first()
                    if session:
                        db.session.delete(session)
                
                cancelled_count += 1
        
        if cancelled_count > 0:
            db.session.commit()
            app.logger.info(f'Đã tự động hủy {cancelled_count} đặt phòng do hết thời gian thanh toán.')

# Initialize Background Scheduler after function definition
scheduler = BackgroundScheduler()
scheduler.add_job(func=huy_dat_phong_khong_den, trigger="interval", minutes=1)
scheduler.add_job(func=huy_dat_phong_timeout, trigger="interval", minutes=1)
scheduler.add_job(func=cleanup_expired_data, trigger="interval", hours=1)  # Run every hour
scheduler.start()

# Ensure scheduler shuts down properly on exit
atexit.register(lambda: scheduler.shutdown() if scheduler.running else None)

# ========================= MAIN ROUTES =========================
@app.route('/')
def index():
    return redirect(url_for('dashboard')) if current_user.is_authenticated else redirect(url_for('login'))

@app.route('/test-timeout-cleanup')
@login_required
@permission_required('system.maintenance')
def test_timeout_cleanup():
    """Route để test việc dọn dẹp timeout thủ công"""
    huy_dat_phong_timeout()
    flash('Đã chạy cleanup timeout thủ công', 'info')
    return redirect(url_for('thanh_toan_chua_hoan_tat'))

@app.route('/khach-hang/dang-ky', methods=['GET', 'POST'])
def khach_hang_dang_ky():
    if current_user.is_authenticated:
        if getattr(current_user, 'is_customer', False):
            return redirect(url_for('khach_hang_tai_khoan'))
        flash('Bạn đang đăng nhập bằng tài khoản nhân viên.', 'info')
        return redirect(url_for('dashboard'))

    form = {
        'ho_ten': '',
        'cmnd': '',
        'email': '',
        'sdt': '',
        'dia_chi': ''
    }
    errors = {}

    prefill = None
    if request.method != 'POST':
        prefill = session.pop('google_registration', None)
        if prefill:
            form['ho_ten'] = prefill.get('name') or form['ho_ten']
            form['email'] = prefill.get('email') or form['email']

    if request.method == 'POST':
        ho_ten = (request.form.get('ho_ten') or '').strip()
        cmnd = (request.form.get('cmnd') or '').strip()
        email_raw = (request.form.get('email') or '').strip()
        email_normalized = normalize_email(email_raw)
        sdt = (request.form.get('sdt') or '').strip()
        dia_chi = (request.form.get('dia_chi') or '').strip()
        mat_khau = request.form.get('mat_khau') or ''
        mat_khau_xac_nhan = request.form.get('mat_khau_xac_nhan') or ''
        form.update({'ho_ten': ho_ten, 'cmnd': cmnd, 'email': email_raw, 'sdt': sdt, 'dia_chi': dia_chi})

        # Kiểm tra xem có đang đăng ký từ Google không
        google_reg = session.get('google_registration')
        is_google_registration = google_reg and google_reg.get('email') == email_raw

        if not ho_ten:
            errors['ho_ten'] = 'Vui lòng nhập họ tên đầy đủ.'
        if not cmnd:
            errors['cmnd'] = 'Vui lòng nhập số CMND/CCCD.'
        elif not re.fullmatch(r'\d{9,12}', cmnd):
            errors['cmnd'] = 'CMND/CCCD phải gồm 9-12 chữ số.'
        if not email_raw:
            errors['email'] = 'Vui lòng nhập email để nhận thông tin.'
        elif not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', email_raw):
            errors['email'] = 'Địa chỉ email không hợp lệ.'
        
        # Nếu không phải đăng ký từ Google, bắt buộc phải có mật khẩu
        if not is_google_registration:
            if len(mat_khau) < 8:
                errors['mat_khau'] = 'Mật khẩu phải có tối thiểu 8 ký tự.'
            if mat_khau != mat_khau_xac_nhan:
                errors['mat_khau_xac_nhan'] = 'Mật khẩu xác nhận chưa khớp.'

        existing = KhachHang.query.filter_by(cmnd=cmnd).first() if cmnd else None
        if existing and existing.co_tai_khoan:
            errors['cmnd'] = 'CMND/CCCD này đã được đăng ký tài khoản khách hàng.'

        if email_normalized:
            email_conflict = KhachHang.query.filter(
                func.lower(KhachHang.email) == email_normalized,
                KhachHang.cmnd != cmnd,
                or_(
                    KhachHang.mat_khau_hash.isnot(None),
                    KhachHang.lan_dang_nhap_cuoi.isnot(None)
                ),
                KhachHang.trang_thai_tai_khoan != 'da_xoa'
            ).first()
            if email_conflict:
                errors['email'] = 'Email này đã được sử dụng cho tài khoản khác.'

        if not errors:
            try:
                if existing:
                    kh = existing
                else:
                    kh = KhachHang(cmnd=cmnd)
                    db.session.add(kh)
                kh.ho_ten = ho_ten
                kh.email = email_normalized or None
                kh.sdt = sdt or None
                kh.dia_chi = dia_chi or None
                
                # Nếu đăng ký từ Google và không có mật khẩu, không set password
                if is_google_registration and not mat_khau:
                    kh.mat_khau_hash = None  # Đăng nhập chỉ qua Google
                else:
                    kh.set_password(mat_khau)
                
                kh.trang_thai_tai_khoan = 'hoat_dong'
                kh.diem_tich_luy = kh.diem_tich_luy or 0
                now = datetime.now()
                kh.ngay_dang_ky = kh.ngay_dang_ky or now
                kh.ngay_cap_nhat = now
                kh.lan_dang_nhap_cuoi = now
                db.session.flush()
                points_added = 0
                if kh.id:
                    thanh_toan_truoc = DatPhong.query.filter_by(khachhang_id=kh.id, trang_thai='da_thanh_toan').all()
                    for dp in thanh_toan_truoc:
                        points_added += award_loyalty_points_for_booking(dp)
                db.session.commit()
                
                # Xóa session google_registration sau khi hoàn tất
                session.pop('google_registration', None)
                
                if kh.email:
                    try:
                        hotel = get_hotel_profile()
                        send_email_with_template(
                            'customer_welcome',
                            kh.email,
                            {
                                'ho_ten': kh.ho_ten,
                                'ten_khach_san': hotel['name'],
                                'cmnd': kh.cmnd,
                                'email': kh.email,
                                'tai_khoan_url': url_for('khach_hang_tai_khoan', _external=True)
                            },
                            khachhang_id=kh.id
                        )
                    except Exception as exc:
                        app.logger.warning('Không thể gửi email chào mừng khách hàng mới: %s', exc)
                login_user(CustomerUser(kh))
                flash('Đăng ký tài khoản thành công! Bạn có thể theo dõi đặt phòng và tích điểm ngay bây giờ.', 'success')
                if points_added:
                    flash(f'Đã cộng thêm {points_added} điểm từ những lần lưu trú trước đây.', 'info')
                return redirect(url_for('khach_hang_tai_khoan'))
            except Exception as exc:
                app.logger.error('Lỗi đăng ký tài khoản khách hàng: %s', exc)
                db.session.rollback()
                flash('Không thể đăng ký tài khoản. Vui lòng thử lại sau ít phút.', 'danger')

    # Kiểm tra xem có session Google không để hiển thị form phù hợp
    google_reg = session.get('google_registration')
    is_from_google = bool(google_reg)
    
    return render_template('khach_hang_dang_ky.html', form=form, errors=errors, is_from_google=is_from_google, google_data=google_reg)


@app.route('/khach-hang/dang-nhap/google')
def khach_hang_dang_nhap_google():
    if not app.config.get('GOOGLE_OAUTH_ENABLED'):
        flash('Chức năng đăng nhập bằng Google chưa được cấu hình.', 'warning')
        return redirect(url_for('khach_hang_dang_nhap'))
    if current_user.is_authenticated:
        if getattr(current_user, 'is_customer', False):
            return redirect(url_for('khach_hang_tai_khoan'))
        flash('Bạn đang đăng nhập bằng tài khoản nhân viên.', 'info')
        return redirect(url_for('dashboard'))

    next_url = request.args.get('next')
    if is_safe_local_url(next_url):
        session['google_login_next'] = next_url
    redirect_uri = url_for('khach_hang_dang_nhap_google_callback', _external=True)
    app.logger.info(f'Redirect URI for Google OAuth: {redirect_uri}')
    try:
        return oauth.google.authorize_redirect(redirect_uri)
    except Exception as exc:
        app.logger.error('Khởi tạo đăng nhập Google thất bại: %s', exc, exc_info=True)
        flash('Không thể kết nối tới Google. Vui lòng thử lại sau.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))


@app.route('/khach-hang/dang-nhap/google/callback')
def khach_hang_dang_nhap_google_callback():
    if not app.config.get('GOOGLE_OAUTH_ENABLED'):
        flash('Chức năng đăng nhập bằng Google chưa được cấu hình.', 'warning')
        return redirect(url_for('khach_hang_dang_nhap'))

    try:
        app.logger.info('Đang yêu cầu token từ Google...')
        token = oauth.google.authorize_access_token()
        app.logger.info('Nhận token thành công từ Google')
    except Exception as exc:
        app.logger.error('Google OAuth token exchange failed: %s', exc, exc_info=True)
        flash('Không thể xác thực với Google. Vui lòng thử lại.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))

    # Lấy thông tin user từ token
    user_info = None
    try:
        app.logger.info('Đang lấy userinfo từ token...')
        # Authlib tự động lấy userinfo khi có token
        user_info = token.get('userinfo')
        if user_info:
            app.logger.info(f'Lấy userinfo từ token thành công: {user_info.get("email")}')
    except Exception as exc:
        app.logger.warning('Không thể lấy userinfo từ token: %s', exc)
        user_info = None

    # Nếu không có trong token, gọi API userinfo
    if not user_info:
        try:
            app.logger.info('Đang gọi Google userinfo API...')
            resp = oauth.google.get('https://www.googleapis.com/oauth2/v3/userinfo')
            if resp.status_code == 200:
                user_info = resp.json()
                app.logger.info(f'Lấy userinfo từ API thành công: {user_info.get("email")}')
            else:
                app.logger.error(f'Google userinfo API trả về lỗi: {resp.status_code}')
        except Exception as exc:
            app.logger.error('Google userinfo request failed: %s', exc)

    if not user_info:
        flash('Không đọc được thông tin người dùng từ Google.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))

    email = user_info.get('email')
    if not email:
        flash('Google không cung cấp email. Không thể đăng nhập.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))

    email_normalized = normalize_email(email)
    if not email_normalized:
        flash('Không thể chuẩn hóa email Google.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))

    if user_info.get('email_verified') is False:
        flash('Email Google của bạn chưa được xác minh.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))

    kh = KhachHang.query.filter(func.lower(KhachHang.email) == email_normalized).first()
    
    # Nếu chưa có tài khoản, chuyển đến trang đăng ký để điền CMND/CCCD
    if not kh:
        # Lưu thông tin Google vào session để tự động điền form
        session['google_registration'] = {
            'email': email,
            'name': user_info.get('name') or '',
            'picture': user_info.get('picture'),
            'email_verified': True
        }
        flash('Email này chưa có tài khoản. Vui lòng hoàn tất đăng ký bằng cách điền CMND/CCCD.', 'warning')
        app.logger.info(f'Email Google chưa có tài khoản, chuyển đến đăng ký: {email_normalized}')
        return redirect(url_for('khach_hang_dang_ky'))

    # Tài khoản đã có - kiểm tra trạng thái
    if kh.trang_thai_tai_khoan == 'khoa':
        flash('Tài khoản của bạn đang bị khóa. Vui lòng liên hệ hỗ trợ.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))
    if kh.trang_thai_tai_khoan == 'da_xoa':
        flash('Tài khoản của bạn đã bị xóa. Bạn có thể đăng ký lại nếu muốn tiếp tục sử dụng.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))

    # Cập nhật thông tin đăng nhập
    now = datetime.now()
    kh.lan_dang_nhap_cuoi = now
    kh.ngay_cap_nhat = now
    if not kh.ho_ten and user_info.get('name'):
        kh.ho_ten = user_info.get('name')
    if not kh.email:
        kh.email = email_normalized
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        app.logger.error('Cập nhật hồ sơ khách hàng khi đăng nhập Google thất bại: %s', exc)
        flash('Không thể cập nhật thông tin tài khoản. Vui lòng thử lại.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))

    login_user(CustomerUser(kh))
    flash('Đăng nhập Google thành công.', 'success')
    next_url = session.pop('google_login_next', None)
    if not is_safe_local_url(next_url):
        next_url = url_for('khach_hang_tai_khoan')
    return redirect(next_url)


@app.route('/khach-hang/dang-nhap', methods=['GET', 'POST'])
def khach_hang_dang_nhap():
    if current_user.is_authenticated:
        if getattr(current_user, 'is_customer', False):
            return redirect(url_for('khach_hang_tai_khoan'))
        flash('Bạn đang đăng nhập bằng tài khoản nhân viên.', 'info')
        return redirect(url_for('dashboard'))

    error = None
    form_username = ''
    if request.method == 'POST':
        identifier = (request.form.get('ten_dang_nhap') or '').strip()
        mat_khau = request.form.get('mat_khau') or ''
        remember = bool(request.form.get('ghi_nho'))
        form_username = identifier
        if not identifier or not mat_khau:
            error = 'Vui lòng nhập đủ thông tin đăng nhập.'
        else:
            identifier_lower = normalize_email(identifier)
            kh = KhachHang.query.filter(
                db.or_(
                    KhachHang.cmnd == identifier,
                    func.lower(KhachHang.email) == identifier_lower
                )
            ).first()
            if not kh:
                error = 'Không tìm thấy tài khoản phù hợp.'
            elif not kh.co_tai_khoan:
                error = 'CMND/CCCD này chưa được đăng ký tài khoản khách hàng.'
            elif kh.trang_thai_tai_khoan == 'khoa':
                error = 'Tài khoản của bạn đang bị khóa. Vui lòng liên hệ hỗ trợ.'
            elif kh.trang_thai_tai_khoan == 'da_xoa':
                error = 'Tài khoản của bạn đã bị xóa. Bạn có thể đăng ký lại hoặc liên hệ hỗ trợ nếu cần.'
            elif not kh.check_password(mat_khau):
                error = 'Mật khẩu không chính xác.'
            else:
                kh.lan_dang_nhap_cuoi = datetime.now()
                kh.ngay_cap_nhat = datetime.now()
                db.session.commit()
                login_user(CustomerUser(kh), remember=remember)
                flash('Đăng nhập thành công.', 'success')
                return redirect(resolve_next_url())

    return render_template('khach_hang_dang_nhap.html', error=error, form_username=form_username)


@app.route('/khach-hang/dang-xuat')
@customer_login_required
def khach_hang_dang_xuat():
    logout_user()
    flash('Bạn đã đăng xuất khỏi tài khoản khách hàng.', 'info')
    return redirect(url_for('khach_hang_dang_nhap'))


@app.route('/khach-hang/quen-mat-khau', methods=['GET', 'POST'])
def khach_hang_quen_mat_khau():
    if current_user.is_authenticated and getattr(current_user, 'is_customer', False):
        return redirect(url_for('khach_hang_tai_khoan'))

    form = {'cmnd': '', 'email': ''}
    error = None

    if request.method == 'POST':
        cmnd = (request.form.get('cmnd') or '').strip()
        email_raw = (request.form.get('email') or '').strip()
        email_normalized = normalize_email(email_raw)
        form.update({'cmnd': cmnd, 'email': email_raw})

        if not cmnd or not email_raw:
            error = 'Vui lòng nhập đầy đủ CMND/CCCD và email đã đăng ký.'
        else:
            kh = KhachHang.query.filter_by(cmnd=cmnd).first()
            if not kh or not kh.co_tai_khoan:
                error = 'Thông tin chưa khớp với tài khoản nào.'
            elif not kh.email:
                error = 'Tài khoản này chưa cập nhật email. Vui lòng liên hệ lễ tân để thay đổi mật khẩu.'
            elif kh.email.lower() != email_normalized:
                error = 'Email không trùng khớp với thông tin đã đăng ký.'
            else:
                new_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                try:
                    kh.set_password(new_password)
                    kh.ngay_cap_nhat = datetime.now()
                    db.session.commit()
                except Exception as exc:
                    db.session.rollback()
                    app.logger.error('Không thể đặt lại mật khẩu cho khách hàng %s: %s', kh.id, exc)
                    error = 'Không thể đặt lại mật khẩu lúc này. Vui lòng thử lại sau.'
                else:
                    email_sent = False
                    if kh.email:
                        try:
                            hotel = get_hotel_profile()
                            send_email_with_template(
                                'customer_password_reset',
                                kh.email,
                                {
                                    'ho_ten': kh.ho_ten,
                                    'ten_khach_san': hotel['name'],
                                    'mat_khau_moi': new_password,
                                    'dang_nhap_url': url_for('khach_hang_dang_nhap', _external=True)
                                },
                                khachhang_id=kh.id
                            )
                            email_sent = True
                        except Exception as exc:
                            app.logger.warning('Không thể gửi email đặt lại mật khẩu: %s', exc)
                            flash('Đã tạo mật khẩu mới nhưng không thể gửi email thông báo. Vui lòng liên hệ lễ tân để nhận hỗ trợ.', 'warning')
                    if email_sent:
                        flash('Mật khẩu mới đã được gửi tới email của bạn. Vui lòng kiểm tra hộp thư và đổi lại mật khẩu sau khi đăng nhập.', 'success')
                    else:
                        flash('Mật khẩu mới đã được tạo. Vui lòng liên hệ lễ tân để nhận hỗ trợ đăng nhập.', 'success')
                    flash('Vui lòng đăng nhập và đổi mật khẩu để đảm bảo an toàn.', 'info')
                    return redirect(url_for('khach_hang_dang_nhap'))

    return render_template('khach_hang_quen_mat_khau.html', form=form, error=error)


@app.route('/khach-hang/tai-khoan', methods=['GET', 'POST'])
@customer_login_required
def khach_hang_tai_khoan():
    kh = get_current_customer()
    if not kh:
        flash('Không tìm thấy thông tin tài khoản.', 'danger')
        return redirect(url_for('khach_hang_dang_nhap'))

    errors = {}
    active_tab = request.form.get('hanh_dong') or request.args.get('tab') or 'thong_tin'

    if request.method == 'POST':
        action = request.form.get('hanh_dong')
        if action == 'cap_nhat_thong_tin':
            ho_ten = (request.form.get('ho_ten') or '').strip()
            sdt = (request.form.get('sdt') or '').strip()
            dia_chi = (request.form.get('dia_chi') or '').strip()
            email_raw = (request.form.get('email') or '').strip()
            email_normalized = normalize_email(email_raw)
            if not ho_ten:
                errors['ho_ten'] = 'Họ tên không được để trống.'
            if email_raw and not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', email_raw):
                errors['email'] = 'Định dạng email không hợp lệ.'
            if email_normalized:
                email_conflict = KhachHang.query.filter(
                    func.lower(KhachHang.email) == email_normalized,
                    KhachHang.id != kh.id,
                    or_(
                        KhachHang.mat_khau_hash.isnot(None),
                        KhachHang.lan_dang_nhap_cuoi.isnot(None)
                    ),
                    KhachHang.trang_thai_tai_khoan != 'da_xoa'
                ).first()
                if email_conflict:
                    errors['email'] = 'Email này đã được dùng cho tài khoản khác.'
            if not errors:
                kh.ho_ten = ho_ten
                kh.sdt = sdt or None
                kh.dia_chi = dia_chi or None
                kh.email = email_normalized or None
                kh.ngay_cap_nhat = datetime.now()
                db.session.commit()
                flash('Đã cập nhật thông tin cá nhân.', 'success')
                return redirect(url_for('khach_hang_tai_khoan', tab='thong_tin'))
            active_tab = 'thong_tin'

        elif action == 'doi_mat_khau':
            mat_khau_cu = request.form.get('mat_khau_cu') or ''
            mat_khau_moi = request.form.get('mat_khau_moi') or ''
            mat_khau_moi_xac_nhan = request.form.get('mat_khau_moi_xac_nhan') or ''
            if not kh.check_password(mat_khau_cu):
                errors['mat_khau_cu'] = 'Mật khẩu hiện tại không đúng.'
            if len(mat_khau_moi) < 8:
                errors['mat_khau_moi'] = 'Mật khẩu mới phải có ít nhất 8 ký tự.'
            if mat_khau_moi != mat_khau_moi_xac_nhan:
                errors['mat_khau_moi_xac_nhan'] = 'Mật khẩu xác nhận chưa khớp.'
            if not errors:
                kh.set_password(mat_khau_moi)
                kh.ngay_cap_nhat = datetime.now()
                db.session.commit()
                flash('Đã đổi mật khẩu thành công.', 'success')
                return redirect(url_for('khach_hang_tai_khoan', tab='doi_mat_khau'))
            active_tab = 'doi_mat_khau'

        elif action == 'xoa_tai_khoan':
            mat_khau_xoa = request.form.get('mat_khau_xoa') or ''
            ly_do_xoa = (request.form.get('ly_do_xoa') or '').strip()
            if not kh.check_password(mat_khau_xoa):
                errors['mat_khau_xoa'] = 'Mat khau khong chinh xac.'
            if not errors:
                ly_do_short = ly_do_xoa[:255] if ly_do_xoa else None
                kh.deleted_at = datetime.now()
                kh.deleted_reason = ly_do_short
                kh.deleted_by = None
                kh.trang_thai_tai_khoan = 'da_xoa'
                kh.mat_khau_hash = None
                kh.ngay_cap_nhat = datetime.now()
                try:
                    db.session.commit()
                except Exception as exc:
                    db.session.rollback()
                    app.logger.error('Không thể xóa tài khoản khách hàng %s: %s', kh.id, exc)
                    flash('Không thể xóa tài khoản ngay lúc này. Vui lòng thử lại.', 'danger')
                    return redirect(url_for('khach_hang_tai_khoan', tab='xoa_tai_khoan'))

                if kh.email:
                    try:
                        hotel = get_hotel_profile()
                        send_email_with_template(
                            'customer_account_self_deleted',
                            kh.email,
                            {
                                'ho_ten': kh.ho_ten,
                                'ten_khach_san': hotel.get('name', 'Khach san'),
                                'ly_do': ly_do_short
                            },
                            khachhang_id=kh.id
                        )
                    except Exception as exc:
                        app.logger.warning('Không thể gửi email từ xóa tài khoản cho %s: %s', kh.email, exc)
                logout_user()
                flash('Tài khoản của bạn đã được xóa thành công.', 'success')
                return redirect(url_for('khach_hang_dang_nhap'))
            active_tab = 'xoa_tai_khoan'

        elif action == 'doi_diem':
            try:
                so_diem = int(request.form.get('so_diem', '0'))
            except ValueError:
                so_diem = 0
            if so_diem <= 0:
                errors['so_diem'] = 'Vui lòng nhập số điểm muốn đổi lớn hơn 0.'
            elif so_diem > (kh.diem_tich_luy or 0):
                errors['so_diem'] = 'Bạn không đủ điểm để đổi.'
            discount_percent = tinh_muc_giam_gia_tu_diem(so_diem)
            if discount_percent <= 0:
                errors['so_diem'] = 'Số điểm chưa đủ để tạo voucher.'
            if not errors:
                try:
                    voucher = issue_voucher_for_khachhang(kh.id, discount_percent=discount_percent, auto_commit=False)
                    kh.tru_diem(so_diem)
                    db.session.commit()
                    flash(
                        f'Đã đổi {so_diem} điểm lấy voucher {voucher.code} giảm {discount_percent}%.',
                        'success'
                    )
                    return redirect(url_for('khach_hang_tai_khoan', tab='uu_dai'))
                except Exception as exc:
                    db.session.rollback()
                    app.logger.error('Không thể đổi điểm sang voucher: %s', exc)
                    errors['so_diem'] = 'Không thể tạo voucher lúc này, vui lòng thử lại.'
            active_tab = 'uu_dai'

    lich_su_dat_phong = DatPhong.query.filter_by(khachhang_id=kh.id).order_by(DatPhong.ngay_nhan.desc()).all()
    vouchers = Voucher.query.filter_by(khachhang_id=kh.id).order_by(Voucher.created_at.desc()).all()
    now_dt = datetime.now()
    voucher_con_han = []
    vouchers_khong_con_hieu_luc = []
    for voucher in vouchers:
        if not voucher.is_used and (voucher.expires_at is None or voucher.expires_at >= now_dt):
            voucher_con_han.append(voucher)
        else:
            status_text = 'Đã sử dụng' if voucher.is_used else 'Đã hết hạn'
            vouchers_khong_con_hieu_luc.append((voucher, status_text))

    return render_template(
        'khach_hang_tai_khoan.html',
        kh=kh,
        errors=errors,
        active_tab=active_tab,
        lich_su_dat_phong=lich_su_dat_phong,
        vouchers_con_han=voucher_con_han,
        vouchers_khong_con_hieu_luc=vouchers_khong_con_hieu_luc,
        loyalty_percent_per_point=LOYALTY_PERCENT_PER_POINT,
        loyalty_max_discount=LOYALTY_MAX_DISCOUNT_PERCENT,
        tinh_muc_giam_gia_tu_diem=tinh_muc_giam_gia_tu_diem
    )


@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        u = NguoiDung.query.filter_by(ten_dang_nhap=request.form['ten_dang_nhap'], mat_khau=request.form['mat_khau']).first()
        if u:
            login_user(u)
            return redirect(url_for('dashboard'))
        flash('Sai tên đăng nhập hoặc mật khẩu','danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
# @cache.cached(timeout=300)  # Cache for 5 minutes
@permission_required('dashboard.view')
def dashboard():
    today = date.today()

    checkins_today = DatPhong.query.filter(
        DatPhong.trang_thai == 'dat',
        func.date(DatPhong.ngay_nhan) == today
    ).count()

    checkouts_today = DatPhong.query.filter(
        DatPhong.trang_thai == 'nhan',
        func.date(DatPhong.ngay_tra) == today
    ).count()

    occupied_rooms = Phong.query.filter_by(trang_thai='dang_o').count()
    total_rooms = Phong.query.count()

    unread_query = TinNhan.query.join(DatPhong).filter(
        TinNhan.trang_thai == 'chua_doc',
        TinNhan.nguoi_gui == 'khach',
        DatPhong.trang_thai == 'nhan'
    ).options(joinedload(TinNhan.datphong).joinedload(DatPhong.phong))
    total_unread_messages = unread_query.count()
    recent_unread_messages = unread_query.order_by(TinNhan.thoi_gian.desc()).limit(5).all()

    return render_template('dashboard.html',
                           checkins_today=checkins_today,
                           checkouts_today=checkouts_today,
                           occupied_rooms=occupied_rooms,
                           total_rooms=total_rooms,
                           total_unread_messages=total_unread_messages,
                           recent_unread_messages=recent_unread_messages)

@app.route('/thanh-toan-chua-hoan-tat')
@login_required
@permission_required('payments.process')
def thanh_toan_chua_hoan_tat():
    # Chạy cleanup timeout trước khi hiển thị
    huy_dat_phong_timeout()

    # Lấy thời gian timeout từ cấu hình
    timeout_setting = HeThongCauHinh.query.filter_by(key='payment_timeout_minutes').first()
    timeout_minutes = int(timeout_setting.value) if timeout_setting and timeout_setting.value else 5
    timeout_minutes = max(1, timeout_minutes)
    payment_session_ttl = timedelta(minutes=timeout_minutes)

    now = datetime.now()
    pending_sessions = []
    
    # 1. Lấy các booking chưa thanh toán (cho_xac_nhan)
    pending_bookings = DatPhong.query.filter_by(trang_thai='cho_xac_nhan').all()
    for dp in pending_bookings:
        # Kiểm tra xem booking có payment session active không
        has_active_session = False
        if dp.payment_token:
            session = PaymentSession.query.filter_by(token=dp.payment_token).first()
            if session and not payment_session_expired(session.created_at, ttl=payment_session_ttl):
                has_active_session = True
        
        # Nếu không có payment session hoặc session đã expired, hiển thị booking này
        if not has_active_session:
            expires_at = dp.created_at + payment_session_ttl
            remaining_seconds = max(0, int((expires_at - now).total_seconds()))
            
            info = {
                'type': 'booking',
                'id': dp.id,
                'khach_hang': dp.khachhang.ho_ten,
                'phong': dp.phong.ten,
                'so_tien': dp.tien_coc or 0,
                'created_at': dp.created_at,
                'expires_at': expires_at,
                'remaining_seconds': remaining_seconds,
                'ngay_nhan': dp.ngay_nhan,
                'ngay_tra': dp.ngay_tra
            }
            pending_sessions.append(info)
    
    # 2. Lấy các payment sessions đang pending
    sessions = PaymentSession.query.filter(
        PaymentSession.created_at > now - payment_session_ttl
    ).all()
    
    for sess in sessions:
        try:
            data = json.loads(sess.payload or '{}')
            if not data.get('completed'):
                # Lấy thông tin khách hàng và số tiền
                info = {
                    'type': 'payment_session',
                    'token': sess.token, 
                    'kind': sess.kind, 
                    'created_at': sess.created_at
                }
                if sess.kind == 'deposit':
                    dp = DatPhong.query.get(data.get('dat_id'))
                    if dp:
                        info['khach_hang'] = dp.khachhang.ho_ten
                        info['so_tien'] = data.get('amount', dp.tien_coc or 0)
                        info['phong'] = dp.phong.ten
                        info['dat_id'] = dp.id
                elif sess.kind == 'service':
                    dp = DatPhong.query.get(data.get('dat_id'))
                    if dp:
                        info['khach_hang'] = dp.khachhang.ho_ten
                        info['so_tien'] = data.get('tong', 0)
                        info['phong'] = dp.phong.ten
                        info['dat_id'] = dp.id
                elif sess.kind == 'room':
                    dp = DatPhong.query.get(data.get('dat_id'))
                    if dp:
                        info['khach_hang'] = dp.khachhang.ho_ten
                        info['so_tien'] = data.get('amount_due', 0)
                        info['phong'] = dp.phong.ten
                        info['dat_id'] = dp.id
                expires_at = sess.created_at + timedelta(minutes=timeout_minutes)
                info['expires_at'] = expires_at
                info['remaining_seconds'] = max(0, int((expires_at - now).total_seconds()))
                pending_sessions.append(info)
        except Exception as e:
            app.logger.warning(f'Error parsing session {sess.token}: {e}')
    
    # Sắp xếp theo thời gian tạo (mới nhất trước)
    pending_sessions.sort(key=lambda x: x['created_at'], reverse=True)
    
    return render_template('thanh_toan_chua_hoan_tat.html', sessions=pending_sessions, timeout_minutes=timeout_minutes)

@app.route('/cap-nhat-timeout-thanh-toan', methods=['POST'])
@login_required
@permission_required('payments.configure_timeout')
def cap_nhat_timeout_thanh_toan():
    """Cập nhật thời gian timeout cho thanh toán"""
    try:
        timeout_minutes = int(request.form.get('timeout_minutes', 5))

        # Validate timeout (1-60 phút)
        if timeout_minutes < 1 or timeout_minutes > 60:
            flash('Thời gian timeout phải từ 1 đến 60 phút', 'danger')
            return redirect(url_for('thanh_toan_chua_hoan_tat'))

        # Lưu vào database
        setting = HeThongCauHinh.query.filter_by(key='payment_timeout_minutes').first()
        if setting:
            setting.value = str(timeout_minutes)
        else:
            setting = HeThongCauHinh(key='payment_timeout_minutes', value=str(timeout_minutes))
            db.session.add(setting)

        db.session.commit()
        flash(f'Đã cập nhật thời gian timeout thanh toán thành {timeout_minutes} phút', 'success')

    except ValueError:
        flash('Thời gian timeout phải là số nguyên', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi khi cập nhật: {str(e)}', 'danger')

    return redirect(url_for('thanh_toan_chua_hoan_tat'))

@app.route('/huy-dat-phong/<int:dat_id>')
@login_required
@permission_required('bookings.cancel')
def huy_dat_phong(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    
    # Chỉ cho phép hủy nếu chưa thanh toán
    if dp.trang_thai != 'cho_xac_nhan':
        flash('Không thể hủy đặt phòng này vì đã được xác nhận.', 'danger')
        return redirect(url_for('thanh_toan_chua_hoan_tat'))
    
    # Lưu thông tin trước khi xóa
    khach_hang = dp.khachhang.ho_ten
    ten_phong = dp.phong.ten
    
    # Xóa payment session nếu có
    if dp.payment_token:
        session = PaymentSession.query.filter_by(token=dp.payment_token).first()
        if session:
            db.session.delete(session)
    
    # Xóa booking
    db.session.delete(dp)
    db.session.commit()
    
    flash(f'Đã hủy đặt phòng của khách {khach_hang} cho phòng {ten_phong}.', 'success')
    return redirect(url_for('thanh_toan_chua_hoan_tat'))

@app.route('/huy-thanh-toan/<token>')
@login_required
@permission_required('payments.process')
def huy_thanh_toan(token):
    sess = PaymentSession.query.filter_by(token=token).first()
    if sess:
        db.session.delete(sess)
        db.session.commit()
        flash('Đã hủy phiên thanh toán thành công.', 'success')
    else:
        flash('Không tìm thấy phiên thanh toán.', 'error')
    return redirect(url_for('thanh_toan_chua_hoan_tat'))

@app.route('/dat-phong', methods=['GET', 'POST'])
@login_required
@permission_required('bookings.create_offline')
def dat_phong():
    # ...existing code...
    discount_percent, _ = get_voucher_config()
    if request.method == 'POST':
        # Lấy thông tin khách hàng
        email_input = (request.form.get('email') or '').strip()
        email_normalized = email_input.lower() if email_input else None
        sdt_input = (request.form.get('sdt') or '').strip()
        dia_chi_input = (request.form.get('dia_chi') or '').strip()
        kh = KhachHang.query.filter_by(cmnd=request.form['cmnd']).first()
        if not kh:
            kh = KhachHang(
                ho_ten=request.form['ho_ten'],
                cmnd=request.form['cmnd'],
                sdt=sdt_input or None,
                email=email_normalized,
                dia_chi=dia_chi_input or None
            )
            db.session.add(kh)
        else:
            kh.ho_ten = request.form['ho_ten']
            kh.sdt = sdt_input or kh.sdt
            kh.dia_chi = dia_chi_input or kh.dia_chi
            if email_normalized:
                kh.email = email_normalized
        db.session.flush()

        phong_id = int(request.form['phong_id'])
        phong = Phong.query.get(phong_id)
        ngay_nhan = datetime.fromisoformat(request.form['ngay_gio_nhan'])
        ngay_tra = datetime.fromisoformat(request.form['ngay_gio_tra'])
        
        # Kiểm tra overlap - nếu có thì đặt waiting
        overlap = DatPhong.query.filter(
            DatPhong.phong_id == phong_id,
            DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES),
            ~db.or_(DatPhong.ngay_tra <= ngay_nhan, DatPhong.ngay_nhan >= ngay_tra)
        ).first()
        
        is_waiting = overlap is not None
        if not is_waiting:
            current_stay = DatPhong.query.filter(
                DatPhong.phong_id == phong_id,
                DatPhong.trang_thai == 'nhan'
            ).order_by(DatPhong.ngay_nhan.desc()).first()
            if current_stay:
                is_waiting = True
        
        hinh_thuc = request.form.get('hinh_thuc_thue', 'ngay')
        duration_seconds = (ngay_tra - ngay_nhan).total_seconds()
        tien_phong_du_kien = 0
        if hinh_thuc == 'ngay':
            so_dem = max(1, math.ceil(duration_seconds / 3600 / 24))
            tien_phong_du_kien = phong.loai.gia * so_dem
        elif hinh_thuc == 'gio':
            so_gio = max(1, math.ceil(duration_seconds / 3600))
            tien_phong_du_kien = so_gio * int(phong.loai.gia * 0.2)

        voucher_code = request.form.get('voucher_code', '').strip().upper()
        voucher_obj = None
        voucher_id = None
        discount_applied = 0
        if voucher_code:
            voucher_obj = Voucher.query.filter_by(code=voucher_code, is_used=False).first()
            if not voucher_obj:
                flash('Mã voucher không hợp lệ hoặc đã sử dụng.', 'danger')
            elif voucher_obj.expires_at < datetime.now():
                flash('Mã voucher đã hết hạn.', 'danger')
            else:
                discount_applied = int(tien_phong_du_kien * voucher_obj.discount_percent / 100)
                tien_phong_du_kien -= discount_applied
                voucher_obj.is_used = True
                voucher_obj.used_at = datetime.now()
                db.session.commit()
                voucher_id = voucher_obj.id
                flash(f'Áp dụng voucher thành công! Giảm {voucher_obj.discount_percent}% ({vnd(discount_applied)}).', 'success')

        tien_coc_tinh_toan = int(tien_phong_du_kien * DEPOSIT_PERCENT)
        so_dem_for_db = max(1, math.ceil(duration_seconds / 3600 / 24))
        dp = DatPhong(
            khachhang_id=kh.id,
            phong_id=phong_id,
            ngay_nhan=ngay_nhan,
            ngay_tra=ngay_tra,
            so_dem=so_dem_for_db,
            hinh_thuc_thue=hinh_thuc,
            tien_coc=tien_coc_tinh_toan,
            voucher_id=voucher_id,
            trang_thai='waiting' if is_waiting else 'cho_xac_nhan'
        )
        db.session.add(dp)
        # Không set phong.trang_thai = 'da_dat' ở đây - chỉ set khi thanh toán thành công
        db.session.flush()

        booking_email_context = None
        if kh.email:
            booking_email_context = build_booking_email_context(dp)
            booking_email_context['tong_tien'] = vnd(dp.tien_coc or 0)

        db.session.commit()

        if kh.email and booking_email_context:
            try:
                send_email_with_template('booking_confirmation', kh.email, booking_email_context, 
                                        datphong_id=dp.id, khachhang_id=kh.id)
            except Exception as exc:
                app.logger.warning('Không thể gửi email xác nhận booking: %s', exc)

        socketio.emit('new_booking_notification', {
            'phong': dp.phong.ten,
            'khach': dp.khachhang.ho_ten,
            'message': f'Phòng {dp.phong.ten} vừa được đặt bởi khách {dp.khachhang.ho_ten}.'
        })
        
        if is_waiting:
            flash('Đặt phòng của bạn đã được chuyển sang trạng thái "Đang chờ" và sẽ được xử lý khi phòng trống.', 'warning')
        else:
            flash('Yêu cầu đặt phòng đã được gửi. Vui lòng thanh toán tiền cọc để hoàn tất đặt phòng.', 'info')
        
        return redirect(url_for('thanh_toan_coc', dat_id=dp.id))
    ds_dat_hom_nay = DatPhong.query.filter(db.func.date(DatPhong.ngay_nhan) == datetime.today().date()).all()
    return render_template('dat_phong.html', loais=LoaiPhong.query.all(), ds_dat=ds_dat_hom_nay, voucher_discount=discount_percent)

@app.route('/dat-phong-online', methods=['GET', 'POST'])
def dat_phong_online():
    loais = LoaiPhong.query.order_by(LoaiPhong.ten.asc()).all()
    if request.method == 'POST':
        try:
            ho_ten = (request.form.get('ho_ten') or '').strip()
            cmnd = (request.form.get('cmnd') or '').strip()
            email = (request.form.get('email') or '').strip()
            sdt = (request.form.get('sdt') or '').strip()
            dia_chi = (request.form.get('dia_chi') or '').strip()
            loai_id = int(request.form.get('loai_id'))
            phong_id = int(request.form.get('phong_id'))
            ngay_nhan_str = request.form.get('ngay_nhan')
            ngay_tra_str = request.form.get('ngay_tra')
            hinh_thuc = (request.form.get('hinh_thuc') or 'ngay').strip()
        except (TypeError, ValueError):
            flash('Thông tin đặt phòng chưa đầy đủ.', 'danger')
            return redirect(url_for('dat_phong_online'))

        if not ho_ten or not cmnd or not ngay_nhan_str or not ngay_tra_str:
            flash('Vui lòng nhập đầy đủ Họ tên, CMND/CCCD và thời gian nhận/trả phòng.', 'danger')
            return redirect(url_for('dat_phong_online'))

        try:
            ngay_nhan = datetime.fromisoformat(ngay_nhan_str)
            ngay_tra = datetime.fromisoformat(ngay_tra_str)
        except ValueError:
            flash('Định dạng ngày giờ không hợp lệ.', 'danger')
            return redirect(url_for('dat_phong_online'))

        if ngay_tra <= ngay_nhan:
            flash('Ngày trả phòng phải sau ngày nhận phòng.', 'danger')
            return redirect(url_for('dat_phong_online'))

        phong = Phong.query.get(phong_id)
        if not phong or phong.loai_id != loai_id:
            flash('Phòng đã chọn không tồn tại.', 'danger')
            return redirect(url_for('dat_phong_online'))

        availability = compute_available_rooms(loai_id, ngay_nhan, ngay_tra)
        selected_room = next((room for room in availability if room['id'] == phong_id), None)
        if not selected_room or not selected_room['available']:
            flash('Phòng đã được đặt trong khoảng thời gian này. Vui lòng chọn phòng khác.', 'warning')
            return redirect(url_for('dat_phong_online'))

        overlap = DatPhong.query.filter(
            DatPhong.phong_id == phong_id,
            DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES),
            ~db.or_(DatPhong.ngay_tra <= ngay_nhan, DatPhong.ngay_nhan >= ngay_tra)
        ).first()
        if overlap:
            flash(
                f"Phòng đã có khách từ {fmt_dt(overlap.ngay_nhan)} - {fmt_dt(overlap.ngay_tra)}. Vui lòng chọn phòng khác.",
                'warning'
            )
            return redirect(url_for('dat_phong_online'))

        current_stay = DatPhong.query.filter(
            DatPhong.phong_id == phong_id,
            DatPhong.trang_thai == 'nhan'
        ).first()
        requires_waiting_after_confirm = current_stay is not None

        duration_seconds = (ngay_tra - ngay_nhan).total_seconds()
        nights = max(1, math.ceil(duration_seconds / (24 * 60 * 60)))
        hours = max(1, math.ceil(duration_seconds / 3600))
        hinh_thuc = 'gio' if hinh_thuc == 'gio' else 'ngay'

        estimated_total = 0
        if hinh_thuc == 'gio':
            price_per_hour = int(round(phong.loai.gia * 0.2))
            estimated_total = price_per_hour * hours
            so_dem_for_db = 1
        else:
            estimated_total = phong.loai.gia * nights
            so_dem_for_db = nights

        voucher_code = request.form.get('voucher_code', '').strip().upper()
        voucher_obj = None
        voucher_id = None
        discount_applied = 0
        if voucher_code:
            voucher_obj = Voucher.query.filter_by(code=voucher_code, is_used=False).first()
            if not voucher_obj:
                flash('Mã voucher không hợp lệ hoặc đã sử dụng.', 'danger')
            elif voucher_obj.expires_at and voucher_obj.expires_at < datetime.now():
                flash('Mã voucher đã hết hạn.', 'danger')
            else:
                discount_applied = int(estimated_total * voucher_obj.discount_percent / 100)
                estimated_total -= discount_applied
                voucher_obj.is_used = True
                voucher_obj.used_at = datetime.now()
                db.session.commit()
                voucher_id = voucher_obj.id
                flash(f'Áp dụng voucher thành công! Giảm {voucher_obj.discount_percent}% ({vnd(discount_applied)}).', 'success')

        tien_coc = max(50000, int(estimated_total * DEPOSIT_PERCENT))

        kh = KhachHang.query.filter_by(cmnd=cmnd).first()
        if not kh:
            kh = KhachHang(cmnd=cmnd)
            db.session.add(kh)
        kh.ho_ten = ho_ten
        kh.sdt = sdt or kh.sdt
        kh.dia_chi = dia_chi or kh.dia_chi
        if email:
            kh.email = email.lower()
        db.session.flush()

        dp = DatPhong(
            khachhang_id=kh.id,
            phong_id=phong_id,
            ngay_nhan=ngay_nhan,
            ngay_tra=ngay_tra,
            so_dem=so_dem_for_db,
            hinh_thuc_thue=hinh_thuc,
            tien_coc=tien_coc,
            tong_thanh_toan=estimated_total,
            trang_thai='cho_xac_nhan',
            chat_token=str(uuid.uuid4()),
            phuong_thuc_coc='qr',
            coc_da_thanh_toan=False,
            voucher_id=voucher_id
        )
        db.session.add(dp)
        db.session.commit()

        if requires_waiting_after_confirm:
            flash('Đặt phòng thành công! Khi nhân viên xác nhận cọc, yêu cầu sẽ chuyển sang Booking chờ cho đến khi phòng trống.', 'info')
        else:
            flash('Đặt phòng thành công! Vui lòng chuyển khoản tiền cọc để giữ phòng.', 'success')
        return redirect(url_for('dat_phong_online_dat_coc', token=dp.chat_token))

    return render_template(
        'dat_phong_online.html',
        loais=loais,
        deposit_percent=int(DEPOSIT_PERCENT * 100),
        active_tab='booking'
    )
@app.route('/dat-phong-online/<token>/dat-coc')
def dat_phong_online_dat_coc(token):
    dp = DatPhong.query.filter_by(chat_token=token).first_or_404()
    last_customer_request = TinNhan.query.filter_by(
        datphong_id=dp.id,
        nguoi_gui='khach',
        noi_dung=ONLINE_DEPOSIT_REQUEST_MESSAGE
    ).order_by(TinNhan.thoi_gian.desc()).first()
    pending_confirmation_requested = bool(last_customer_request) and dp.trang_thai == 'cho_xac_nhan'
    qr_amount = int(dp.tien_coc or max(50000, int((dp.tong_thanh_toan or 0) * DEPOSIT_PERCENT)))
    description = quote(f"Coc HD{dp.id} {dp.khachhang.ho_ten}")
    qr_code_url = (f"https://img.vietqr.io/image/{VIETQR_BANK_ID}-{VIETQR_ACCOUNT_NO}-compact2.png"
                   f"?amount={qr_amount}&addInfo={description}")
    diff_seconds = max(0, int((dp.ngay_tra - dp.ngay_nhan).total_seconds())) if dp.ngay_nhan and dp.ngay_tra else 0
    duration_text = None
    if diff_seconds:
        if dp.hinh_thuc_thue == 'gio':
            hours = max(1, math.ceil(diff_seconds / 3600))
            duration_text = f"{hours} giờ"
        else:
            nights = max(1, math.ceil(diff_seconds / (24 * 3600)))
            duration_text = f"{nights} đêm"
    can_request = dp.trang_thai == 'cho_xac_nhan' and not dp.coc_da_thanh_toan
    status_source = BOOKING_STATUS_USER_MESSAGES.get(dp.trang_thai)
    if status_source:
        status_info = dict(status_source)
    else:
        readable = (dp.trang_thai or 'không xác định').replace('_', ' ').title()
        status_info = {
            'title': readable,
            'message': f'Trạng thái đơn đặt phòng hiện tại: {readable}.',
            'level': 'info'
        }
    loais = LoaiPhong.query.order_by(LoaiPhong.ten.asc()).all()
    return render_template(
        'dat_phong_online.html',
        loais=loais,
        deposit_percent=int(DEPOSIT_PERCENT * 100),
        active_tab='payment',
        dp=dp,
        qr_code_url=qr_code_url,
        qr_amount=qr_amount,
        can_request=can_request,
        duration_text=duration_text,
        status_info=status_info,
        has_pending_deposit_request=pending_confirmation_requested,
        pending_deposit_message=CUSTOMER_PENDING_CONFIRMATION_MESSAGE if pending_confirmation_requested else '',
        pending_deposit_message_full=CUSTOMER_PENDING_CONFIRMATION_MESSAGE
    )


@app.route('/dat-phong-online/<token>/yeu-cau-xac-nhan', methods=['POST'])
def dat_phong_online_request_confirmation(token):
    dp = DatPhong.query.filter_by(chat_token=token).first_or_404()
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    payload = {}
    if request.is_json:
        payload = request.get_json(silent=True) or {}
    confirm_raw = payload.get('confirm') if request.is_json else request.form.get('confirm')
    if confirm_raw is None:
        confirm_flag = True
    else:
        confirm_flag = str(confirm_raw).lower() in ('true', '1', 'yes', 'on')
    if not confirm_flag:
        message = 'Vui lòng nhấn nút "Tôi đã thanh toán cọc" sau khi chuyển khoản.'
        status = 'warning'
        if wants_json:
            return jsonify({'message': message, 'status': status}), 400
        flash(message, status)
        return redirect(url_for('dat_phong_online_dat_coc', token=token))
    if dp.trang_thai != 'cho_xac_nhan':
        message = 'Đơn đặt phòng đã được xử lý.'
        status = 'info'
    else:
        recent_request = TinNhan.query.filter_by(datphong_id=dp.id, nguoi_gui='khach')\
            .order_by(TinNhan.thoi_gian.desc()).first()
        if recent_request and (datetime.now() - recent_request.thoi_gian).total_seconds() < 120:
            message = 'Bạn đã gửi yêu cầu gần đây. Vui lòng chờ nhân viên kiểm tra.'
            status = 'warning'
        else:
            msg = ONLINE_DEPOSIT_REQUEST_MESSAGE
            tn = TinNhan(datphong_id=dp.id, nguoi_gui='khach', noi_dung=msg,
                         thoi_gian=datetime.now(), trang_thai='chua_doc')
            db.session.add(tn)
            db.session.commit()
            socketio.emit('online_booking_deposit_request', {
                'booking_id': dp.id,
                'phong': dp.phong.ten,
                'khach': dp.khachhang.ho_ten
            })
            message = CUSTOMER_PENDING_CONFIRMATION_MESSAGE
            status = 'success'
    if wants_json:
        return jsonify({'message': message, 'status': status})
    flash(message, status if status in ('success', 'info', 'warning', 'danger') else 'info')
    return redirect(url_for('dat_phong_online_dat_coc', token=token))
@app.route('/quan-ly-dat-phong-online')
@login_required
@permission_required('bookings.manage_online')
def quan_ly_dat_phong_online():
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Items per page
    pending_query = DatPhong.query.options(
        joinedload(DatPhong.khachhang),
        joinedload(DatPhong.phong).joinedload(Phong.loai)
    ).join(
        TinNhan,
        db.and_(
            TinNhan.datphong_id == DatPhong.id,
            TinNhan.nguoi_gui == 'khach',
            TinNhan.noi_dung == ONLINE_DEPOSIT_REQUEST_MESSAGE
        )
    ).filter(
        DatPhong.trang_thai == 'cho_xac_nhan'
    ).distinct().order_by(DatPhong.ngay_nhan.asc())
    pagination = pending_query.paginate(page=page, per_page=per_page, error_out=False)
    pending = pagination.items
    return render_template('quan_ly_dat_phong_online.html', pending=pending,
                           pagination=pagination, deposit_percent=int(DEPOSIT_PERCENT * 100))


@app.route('/quan-ly-dat-phong-online/<int:dat_id>/xac-nhan', methods=['POST'])
@login_required
@permission_required('bookings.manage_online')
def quan_ly_dat_phong_online_xac_nhan(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if dp.trang_thai != 'cho_xac_nhan':
        flash('Đơn đặt phòng đã được xử lý.', 'info')
        return redirect(url_for('quan_ly_dat_phong_online'))
    dp.coc_da_thanh_toan = True
    dp.phuong_thuc_coc = 'qr'
    active_stay = DatPhong.query.filter(
        DatPhong.phong_id == dp.phong_id,
        DatPhong.id != dp.id,
        DatPhong.trang_thai == 'nhan'
    ).first()
    if active_stay:
        dp.trang_thai = 'waiting'
        dp.auto_confirmed_at = None
    else:
        dp.trang_thai = 'dat'
        dp.auto_confirmed_at = datetime.now()
        if dp.phong.trang_thai == 'trong':
            dp.phong.trang_thai = 'da_dat'
    tn = TinNhan(datphong_id=dp.id, nguoi_gui='he_thong',
                 noi_dung='Đã xác nhận tiền cọc đặt phòng online.',
                 thoi_gian=datetime.now(), trang_thai='chua_doc')
    db.session.add(tn)
    db.session.commit()

    if dp.khachhang.email:
        try:
            ctx = build_booking_email_context(dp)
            ctx['tong_tien'] = vnd(dp.tong_thanh_toan or 0)
            send_email_with_template('booking_confirmation', dp.khachhang.email, ctx,
                                     datphong_id=dp.id, khachhang_id=dp.khachhang_id)
        except Exception as exc:
            app.logger.warning('Không thể gửi email xác nhận đặt phòng online: %s', exc)

    socketio.emit('online_booking_confirmed', {
        'booking_id': dp.id,
        'phong': dp.phong.ten,
        'khach': dp.khachhang.ho_ten
    })
    flash('Đã xác nhận tiền cọc và giữ phòng cho khách.', 'success')
    return redirect(url_for('quan_ly_dat_phong_online'))


@app.route('/quan-ly-dat-phong-online/<int:dat_id>/tu-choi', methods=['POST'])
@login_required
@permission_required('bookings.manage_online')
def quan_ly_dat_phong_online_tu_choi(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if dp.trang_thai != 'cho_xac_nhan':
        flash('Đơn đặt phòng đã được xử lý.', 'info')
        return redirect(url_for('quan_ly_dat_phong_online'))
    dp.trang_thai = 'huy'
    dp.coc_da_thanh_toan = False
    dp.phuong_thuc_coc = None
    tn = TinNhan(datphong_id=dp.id, nguoi_gui='he_thong',
                 noi_dung='Đã từ chối yêu cầu đặt phòng online.',
                 thoi_gian=datetime.now(), trang_thai='chua_doc')
    db.session.add(tn)
    db.session.commit()
    socketio.emit('online_booking_rejected', {
        'booking_id': dp.id,
        'phong': dp.phong.ten,
        'khach': dp.khachhang.ho_ten
    })
    flash('Đã từ chối yêu cầu đặt phòng.', 'info')
    return redirect(url_for('quan_ly_dat_phong_online'))


@app.route('/quan-ly-booking-cho')
@login_required
@permission_required('bookings.manage_waiting')
def quan_ly_booking_cho():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    waiting_query = DatPhong.query.options(
        joinedload(DatPhong.khachhang),
        joinedload(DatPhong.phong).joinedload(Phong.loai),
        joinedload(DatPhong.nhanvien)
    ).filter(
        DatPhong.trang_thai == 'waiting'
    ).order_by(DatPhong.created_at.desc())
    pagination = waiting_query.paginate(page=page, per_page=per_page, error_out=False)
    waiting = pagination.items
    return render_template('quan_ly_booking_cho.html', waiting=waiting,
                           pagination=pagination)


@app.route('/quan-ly-booking-cho/<int:dat_id>/tu-choi', methods=['POST'])
@login_required
@permission_required('bookings.manage_waiting')
def quan_ly_booking_cho_tu_choi(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if dp.trang_thai != 'waiting':
        flash('Booking đã được xử lý.', 'info')
        return redirect(url_for('quan_ly_booking_cho'))
    dp.trang_thai = 'huy'
    tn = TinNhan(datphong_id=dp.id, nguoi_gui='he_thong',
                 noi_dung='Booking chờ của bạn đã bị từ chối.',
                 thoi_gian=datetime.now(), trang_thai='chua_doc')
    db.session.add(tn)
    db.session.commit()
    flash('Đã từ chối booking chờ.', 'info')
    return redirect(url_for('quan_ly_booking_cho'))


@app.route('/nhan-phong', methods=['GET','POST'])
@login_required
@permission_required('bookings.checkin_checkout')
def nhan_phong():
    huy_dat_phong_khong_den()
    # Lấy tab từ query parameter (mặc định là checkin)
    active_tab = request.args.get('tab', 'checkin')
    
    if request.method == 'POST':
        dp = DatPhong.query.get(int(request.form['dat_id']))
        if dp:
            now = datetime.now()
            # Nếu là booking waiting đã auto_confirmed, chuyển thành dat trước
            if dp.trang_thai == 'waiting':
                # Use raw SQL to check auto_confirmed_at since SQLAlchemy metadata isn't refreshed
                check_query = db.text('SELECT auto_confirmed_at FROM datphong WHERE id = :id')
                result = db.session.execute(check_query, {'id': dp.id}).fetchone()
                if result and result[0]:  # auto_confirmed_at is not NULL
                    dp.trang_thai = 'dat'
                else:
                    # Waiting booking not auto-confirmed, cannot check-in
                    flash('Đặt phòng đang chờ xác nhận, không thể nhận phòng.', 'danger')
                    return redirect(url_for('nhan_phong'))

            # Proceed with check-in for confirmed bookings
            if dp.trang_thai == 'dat':
                dp.trang_thai = 'nhan'
                dp.thuc_te_nhan = now
                dp.nhanvien_id = current_user.id
                dp.phong.trang_thai = 'dang_o'
                if not dp.chat_token:
                    dp.chat_token = str(uuid.uuid4())
                email_recipient = dp.khachhang.email
                email_context = None
                voucher_new = None
                if email_recipient:
                    email_context = build_booking_email_context(dp)
                    email_context['tong_tien'] = vnd((dp.tong_thanh_toan or dp.tien_coc or 0))
                # Phát voucher nếu loại phòng được cấu hình tặng voucher
                loai_phong = getattr(dp.phong, 'loai', None)
                if loai_phong and getattr(loai_phong, 'co_voucher', False):
                    voucher_new = issue_voucher_for_khachhang(dp.khachhang_id)
                    # Gửi tin nhắn voucher cho khách
                    msg = f"Chúc mừng! Bạn nhận được voucher giảm giá {voucher_new.discount_percent}% cho lần đặt tiếp theo. Mã: {voucher_new.code}. HSD: {voucher_new.expires_at.strftime('%d/%m/%Y')}"
                    tn = TinNhan(datphong_id=dp.id, nguoi_gui='he_thong', noi_dung=msg, thoi_gian=datetime.now(), trang_thai='chua_doc')
                    db.session.add(tn)
                db.session.commit()

                if voucher_new and email_context:
                    email_context['voucher_moi'] = True
                    email_context['voucher_moi_code'] = voucher_new.code
                    email_context['voucher_moi_discount'] = voucher_new.discount_percent
                    email_context['voucher_moi_han'] = voucher_new.expires_at.strftime('%d/%m/%Y')

                if email_recipient and email_context:
                    try:
                        send_email_with_template('checkin_notification', email_recipient, email_context,
                                                datphong_id=dp.id, khachhang_id=dp.khachhang_id)
                    except Exception as exc:
                        app.logger.warning('Không thể gửi email check-in: %s', exc)
                flash(f'Đã nhận phòng lúc {fmt_dt(now)}','success')
            else:
                flash('Trạng thái đặt phòng không hợp lệ để nhận phòng.', 'danger')
        else:
            flash('Không tìm thấy thông tin đặt phòng.', 'danger')
        return redirect(url_for('nhan_phong'))
    discount_percent, _ = get_voucher_config()
    voucher_room_types = (
        LoaiPhong.query
        .filter_by(co_voucher=True)
        .order_by(LoaiPhong.ten.asc())
        .all()
    )
    voucher_room_names = [
        loai.ten for loai in voucher_room_types
        if getattr(loai, 'ten', None)
    ]
    if voucher_room_names:
        if len(voucher_room_names) == 1:
            voucher_room_label = voucher_room_names[0]
        elif len(voucher_room_names) == 2:
            voucher_room_label = f"{voucher_room_names[0]} hoặc {voucher_room_names[1]}"
        else:
            voucher_room_label = ", ".join(voucher_room_names[:-1]) + f" hoặc {voucher_room_names[-1]}"
    else:
        voucher_room_label = ''
    voucher_rooms_enabled = bool(voucher_room_names)
    cancel_setting = HeThongCauHinh.query.filter_by(key='auto_cancel_minutes').first()
    try:
        auto_cancel_minutes = int(cancel_setting.value) if cancel_setting and cancel_setting.value else 5
    except ValueError:
        auto_cancel_minutes = 5
    # Use raw SQL to avoid SQLAlchemy model attribute issues after schema changes
    ds_nhan_query = db.text("""
        SELECT * FROM datphong 
        WHERE (trang_thai = 'dat' OR (trang_thai = 'waiting' AND auto_confirmed_at IS NOT NULL))
        AND DATE(ngay_nhan) <= CURDATE()
        ORDER BY ngay_nhan
    """)
    ds_nhan = db.session.execute(ds_nhan_query).fetchall()
    # Convert to objects for template compatibility
    ds_nhan_ids = [row.id for row in ds_nhan]
    if ds_nhan_ids:
        ds_nhan = (
            DatPhong.query
            .filter(DatPhong.id.in_(ds_nhan_ids))
            .order_by(DatPhong.ngay_nhan.asc())
            .all()
        )
    else:
        ds_nhan = []
    active_stays = DatPhong.query.filter(DatPhong.trang_thai == 'nhan').all()
    active_by_room = {}
    for stay in active_stays:
        active_by_room.setdefault(stay.phong_id, []).append(stay)

    cancel_data = {}
    for booking in ds_nhan:
        blockers = [
            stay for stay in active_by_room.get(booking.phong_id, [])
            if stay.id != booking.id
        ]
        blocked = bool(blockers)
        booking.blocked_by_overstay = blocked
        booking.overstay_blockers = blockers

        if blocked:
            blocker = sorted(
                blockers,
                key=lambda x: x.ngay_nhan or datetime.min,
                reverse=True
            )[0]
            blocker_checkout = blocker.thuc_te_tra or blocker.ngay_tra
            cancel_data[booking.id] = {
                'blocked': True,
                'blocker_name': blocker.khachhang.ho_ten if blocker.khachhang else 'Khách đang ở',
                'blocker_checkout': fmt_dt(blocker_checkout) if blocker_checkout else '---',
                'blocker_booking_id': blocker.id
            }
            continue

        confirmed_at = booking.auto_confirmed_at or booking.created_at or booking.ngay_nhan
        if not confirmed_at:
            confirmed_at = datetime.now()
        deadline = confirmed_at + timedelta(minutes=auto_cancel_minutes)
        cancel_data[booking.id] = {
            'blocked': False,
            'confirmed_at': confirmed_at.strftime('%Y-%m-%dT%H:%M:%S'),
            'deadline': deadline.strftime('%Y-%m-%dT%H:%M:%S'),
            'confirmed_display': confirmed_at.strftime('%H:%M %d/%m/%Y'),
            'deadline_display': deadline.strftime('%H:%M %d/%m/%Y'),
            'minutes': auto_cancel_minutes,
        }
    ds_thue = DatPhong.query.filter_by(trang_thai='nhan').order_by(DatPhong.ngay_nhan.desc()).all()
    return render_template(
        'nhan_phong.html',
        ds_nhan=ds_nhan,
        ds_thue=ds_thue,
        voucher_discount=discount_percent,
        voucher_room_label=voucher_room_label,
        voucher_rooms_enabled=voucher_rooms_enabled,
        auto_cancel_minutes=auto_cancel_minutes,
        cancel_data=cancel_data,
        active_tab=active_tab
    )


def build_service_booking_payload(dat_phong):
    """Return booking/service details for async updates."""
    if not dat_phong:
        return None

    rows = (
        SuDungDichVu.query
        .filter_by(datphong_id=dat_phong.id, trang_thai='chua_thanh_toan')
        .order_by(SuDungDichVu.thoi_gian.asc())
        .all()
    )
    services = []
    total = 0
    for row in rows:
        unit_price = row.dichvu.gia if row.dichvu else 0
        quantity = row.so_luong or 0
        amount = unit_price * quantity
        total += amount
        services.append({
            'id': row.id,
            'ten': row.dichvu.ten if row.dichvu else '',
            'gia': unit_price,
            'so_luong': quantity,
            'thanh_tien': amount
        })

    return {
        'id': dat_phong.id,
        'phong_id': dat_phong.phong_id,
        'phong_ten': dat_phong.phong.ten if dat_phong.phong else '',
        'khach_ten': dat_phong.khachhang.ho_ten if dat_phong.khachhang else '',
        'khach_sdt': getattr(dat_phong.khachhang, 'dien_thoai', '') if dat_phong.khachhang else '',
        'services': services,
        'tong': total
    }


def build_salary_settings_context(selected_id=None):
    """Return context data used by the salary settings page."""
    all_staffs = NguoiDung.query.order_by(NguoiDung.ten.asc()).all()
    if selected_id:
        staffs = [s for s in all_staffs if s.id == selected_id]
        if not staffs:
            selected_id = None
            staffs = all_staffs
    else:
        staffs = all_staffs

    bonus_amount = get_top_bonus()
    salary_mode = get_salary_mode()
    salary_mode = get_salary_mode()

    now = datetime.now()
    start_month = datetime(now.year, now.month, 1)
    next_month = (start_month + timedelta(days=32)).replace(day=1)
    month_rows = db.session.query(
        DatPhong.nhanvien_id.label('nv_id'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('doanh_thu')
    ).filter(
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).group_by(DatPhong.nhanvien_id).all()
    top_staff_ids = set()
    if month_rows:
        top_max = max(int(row.doanh_thu or 0) for row in month_rows)
        if top_max > 0:
            top_staff_ids = {row.nv_id for row in month_rows if int(row.doanh_thu or 0) == top_max}

    salary_records = {item.nguoidung_id: item for item in LuongNhanVien.query.all()}
    tiers = LuongThuongCauHinh.query.order_by(LuongThuongCauHinh.moc_duoi.asc()).all()
    salary_mode = get_salary_mode()

    work_days_map = {}
    if all_staffs:
        if salary_mode == SALARY_MODE_DAILY:
            for staff in all_staffs:
                work_days = db.session.query(func.count()).filter(
                    Attendance.user_id == staff.id,
                    Attendance.status == 'approved',
                    Attendance.checkin_time >= start_month,
                    Attendance.checkin_time < next_month
                ).scalar() or 0
                work_days_map[staff.id] = int(work_days or 0)
        else:
            work_days_map = {staff.id: 0 for staff in all_staffs}

    total_base = 0
    for record in salary_records.values():
        work_days = work_days_map.get(record.nguoidung_id, 0)
        if salary_mode == SALARY_MODE_DAILY:
            total_base += compute_effective_base_salary(record.luong_co_ban, work_days, salary_mode)
        else:
            total_base += record.luong_co_ban or 0
    total_allowance = sum((record.phu_cap or 0) for record in salary_records.values())
    configured_count = len(salary_records)

    min_days = get_min_work_days()
    auto_cancel_minutes = get_config_int('auto_cancel_minutes', 5)

    return {
        'staffs': staffs,
        'all_staffs': all_staffs,
        'selected_id': selected_id,
        'bonus_amount': bonus_amount,
        'top_staff_ids': top_staff_ids,
        'top_staffs': [s for s in all_staffs if s.id in top_staff_ids],
        'salary_records': salary_records,
        'tiers': tiers,
        'total_base': total_base,
        'total_allowance': total_allowance,
        'configured_count': configured_count,
        'min_days': min_days,
        'auto_cancel_minutes': auto_cancel_minutes,
        'salary_mode': salary_mode,
        'salary_daily_divisor': SALARY_DAILY_DIVISOR,
    }

@app.route('/dich-vu-thanh-toan')
@login_required
@permission_required('payments.process', 'services.orders')
def dich_vu_thanh_toan():
    dat_id = request.args.get("dat_id")
    chon_dat = DatPhong.query.get(int(dat_id)) if dat_id else None
    hoa_don_dv = SuDungDichVu.query.filter_by(datphong_id=dat_id, trang_thai='chua_thanh_toan').all() if dat_id else []
    initial_booking = build_service_booking_payload(chon_dat) if chon_dat else None
    
    phongs = Phong.query.all()
    now = datetime.now()
    overdue_bookings = DatPhong.query.filter(
        DatPhong.trang_thai == 'nhan',
        DatPhong.ngay_tra < now
    ).all()
    overdue_phong_ids = {booking.phong_id for booking in overdue_bookings}

    for p in phongs:
        base_status = p.trang_thai or 'trong'
        if p.id in overdue_phong_ids:
            p.calculated_status = 'qua_gio'
        else:
            p.calculated_status = base_status
        p.effective_status = 'dang_o' if p.calculated_status == 'qua_gio' else p.calculated_status
            
    return render_template(
        'dichvu_thanhtoan.html',
        loais=DichVuLoai.query.all(),
        phongs=phongs,
        chon_dat=chon_dat,
        hoa_don_dv=hoa_don_dv,
        initial_booking=initial_booking
    )

@app.route('/thanh-toan-dv/<int:dat_id>', methods=['GET', 'POST'])
@login_required
@permission_required('payments.process', 'services.orders')
def thanh_toan_dv(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    rows = (
        SuDungDichVu.query
        .filter_by(datphong_id=dat_id, trang_thai='chua_thanh_toan')
        .order_by(SuDungDichVu.thoi_gian.asc())
        .all()
    )
    if not rows:
        flash('Khong co dich vu moi de thanh toan.', 'warning')
        return redirect(url_for('dich_vu_thanh_toan', dat_id=dat_id))

    items = []
    tong = 0
    for row in rows:
        thanh_tien = row.dichvu.gia * row.so_luong
        items.append({
            'id': row.id,
            'ten': row.dichvu.ten,
            'gia': row.dichvu.gia,
            'so_luong': row.so_luong,
            'thanh_tien': thanh_tien
        })
        tong += thanh_tien

    if request.method == 'POST':
        payment_method = request.form.get('payment_method')
        if payment_method not in ['cash', 'qr']:
            return redirect(url_for('thanh_toan_dv', dat_id=dat_id))

        if payment_method == 'cash':
            # Clear any pending QR attempts so they no longer appear as unfinished payments
            invalidate_payment_sessions('service', dat_id)
            payment_token = uuid.uuid4().hex
            create_payment_session(payment_token, 'service', {
                'dat_id': dat_id,
                'tong': tong,
                'usage_ids': [item['id'] for item in items],
                'items': [
                    {
                        'ten': item['ten'],
                        'so_luong': item['so_luong'],
                        'thanh_tien': item['thanh_tien']
                    }
                    for item in items
                ],
                'completed': True,
                'completed_at': datetime.now().isoformat(),
                'payment_method': 'cash'
            })

            for row in rows:
                row.trang_thai = 'da_thanh_toan'
            dp.tien_dv = (dp.tien_dv or 0) + tong
            db.session.commit()
            flash('Da ghi nhan thanh toan dich vu bang tien mat.', 'success')
            return redirect(url_for('in_hoa_don_dv', token=payment_token))

        invalidate_payment_sessions('service', dat_id)
        payment_token = uuid.uuid4().hex
        create_payment_session(payment_token, 'service', {
            'dat_id': dat_id,
            'tong': tong,
            'usage_ids': [item['id'] for item in items],
            'items': [
                {
                    'ten': item['ten'],
                    'so_luong': item['so_luong'],
                    'thanh_tien': item['thanh_tien']
                }
                for item in items
            ]
        })
        return redirect(url_for('show_qr_service', token=payment_token))

    return render_template('thanh_toan_dv.html', dp=dp, items=items, tong=tong)

@app.route('/thanh-toan-coc/<int:dat_id>', methods=['GET', 'POST'])
@login_required
@permission_required('payments.process')
def thanh_toan_coc(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if dp.coc_da_thanh_toan:
        flash('Dat phong nay da thanh toan tien coc.', 'info')
        return redirect(url_for('dat_phong'))

    if request.method == 'POST':
        payment_method = request.form.get('payment_method')
        if payment_method not in ['cash', 'qr']:
            return redirect(url_for('thanh_toan_coc', dat_id=dat_id))

        if payment_method == 'cash':
            # Remove any QR payment sessions that might still be linked to this booking
            invalidate_payment_sessions('deposit', dat_id)
            if dp.payment_token:
                dp.payment_token = None
            dp.phuong_thuc_coc = 'cash'
            dp.coc_da_thanh_toan = True
            active_stay = DatPhong.query.filter(
                DatPhong.phong_id == dp.phong_id,
                DatPhong.id != dp.id,
                DatPhong.trang_thai == 'nhan'
            ).first()
            if active_stay:
                dp.trang_thai = 'waiting'
                dp.auto_confirmed_at = None
            else:
                dp.trang_thai = 'dat'
                dp.auto_confirmed_at = datetime.now()
                if dp.phong.trang_thai == 'trong':
                    dp.phong.trang_thai = 'da_dat'
            db.session.commit()
            flash('Da ghi nhan thanh toan tien coc bang tien mat.', 'success')
            return redirect(url_for('in_hoa_don_coc', dat_id=dat_id))

        invalidate_payment_sessions('deposit', dat_id)
        if dp.payment_token:
            pop_payment_session(dp.payment_token)
        payment_token = uuid.uuid4().hex
        dp.payment_token = payment_token
        dp.phuong_thuc_coc = 'qr'
        dp.coc_da_thanh_toan = False
        create_payment_session(payment_token, 'deposit', {
            'dat_id': dat_id,
            'amount': int(dp.tien_coc or 0),
            'customer_name': dp.khachhang.ho_ten,
            'booking_code': dp.id
        })
        db.session.commit()
        return redirect(url_for('show_qr_deposit', token=payment_token))

    discount_percent, _ = get_voucher_config()
    return render_template('thanh_toan_coc.html', dp=dp, voucher_discount=discount_percent)

def process_waiting_bookings(phong_id, previous_checkout_time=None):
    """Finalize waiting bookings when the room becomes available."""
    waiting_bookings = DatPhong.query.filter(
        DatPhong.phong_id == phong_id,
        DatPhong.trang_thai == 'waiting'
    ).order_by(DatPhong.created_at.asc()).all()
    updated = False

    for wb in waiting_bookings:
        overlap = DatPhong.query.filter(
            DatPhong.phong_id == phong_id,
            DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES),
            DatPhong.id != wb.id,
            ~db.or_(DatPhong.ngay_tra <= wb.ngay_nhan, DatPhong.ngay_nhan >= wb.ngay_tra)
        ).first()

        if not overlap:
            update_query = db.text('UPDATE datphong SET auto_confirmed_at = :timestamp, nhanvien_id = :nhanvien_id WHERE id = :id')
            db.session.execute(update_query, {
                'timestamp': datetime.now(),
                'nhanvien_id': current_user.id if current_user and hasattr(current_user, 'id') else None,
                'id': wb.id
            })
            updated = True

            if previous_checkout_time and previous_checkout_time > wb.ngay_nhan:
                delay = previous_checkout_time - wb.ngay_nhan
                wb.ngay_tra = wb.ngay_tra + delay
                wb.ngay_nhan = previous_checkout_time
                updated = True

            if previous_checkout_time and previous_checkout_time > wb.ngay_nhan:
                msg = (
                    "Booking cho cua ban da duoc tu dong xac nhan. "
                    f"Do phong trong muon, thoi gian nhan phong duoc dieu chinh thanh "
                    f"{wb.ngay_nhan.strftime('%d/%m/%Y %H:%M')} va tra phong {wb.ngay_tra.strftime('%d/%m/%Y %H:%M')}."
                )
            else:
                msg = 'Booking cho cua ban da duoc tu dong xac nhan va se xuat hien trong check-in khi den ngay nhan phong.'

            tn = TinNhan(datphong_id=wb.id, nguoi_gui='he_thong',
                         noi_dung=msg,
                         thoi_gian=datetime.now(), trang_thai='chua_doc')
            db.session.add(tn)
            updated = True

            socketio.emit('booking_confirmed', {
                'booking_id': wb.id,
                'phong': wb.phong.ten,
                'khach': wb.khachhang.ho_ten
            })
            break

    if updated:
        db.session.commit()


@app.route('/thanh-toan/<int:dat_id>', methods=['GET', 'POST'])
@login_required
@permission_required('payments.process')
def thanh_toan(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    email_prefill = request.args.get('email', '')
    if not dp.thuc_te_tra:
        dp.thuc_te_tra = datetime.now()
        db.session.commit()

    invoice_ctx, calc_values = build_invoice_context(dp)
    selected_method = request.form.get('payment_method') or dp.phuong_thuc_thanh_toan or 'qr'

    if request.method == 'POST':
        if dp.trang_thai == 'da_thanh_toan':
            flash('Dat phong nay da duoc thanh toan truoc do.', 'info')
            return redirect(url_for('thanh_toan', dat_id=dat_id))

        payment_method = request.form.get('payment_method')
        if payment_method not in PAYMENT_METHOD_LABELS:
            return redirect(url_for('thanh_toan', dat_id=dat_id))

        if payment_method == 'cash':
            # Remove any existing QR payment sessions that are no longer needed
            invalidate_payment_sessions('room', dat_id)
            dp.thuc_te_tra = dp.thuc_te_tra or datetime.now()
            dp.tien_phong = calc_values['tien_phong']
            dp.tien_dv = calc_values['tien_dv']
            dp.tien_phat = calc_values['tien_phat']
            dp.tong_thanh_toan = calc_values['tong']
            dp.phuong_thuc_thanh_toan = 'cash'
            dp.phong.trang_thai = 'trong'
            dp.trang_thai = 'da_thanh_toan'
            dp.nhanvien_id = current_user.id
            diem_moi = award_loyalty_points_for_booking(dp)
            db.session.commit()
            
            # Kiểm tra và chuyển waiting bookings
            process_waiting_bookings(dp.phong_id, dp.thuc_te_tra)
            
            flash('Da ghi nhan thanh toan phong bang tien mat.', 'success')
            if diem_moi:
                flash(f'Khách hàng được cộng thêm {diem_moi} điểm thưởng.', 'info')
            return redirect(url_for('in_hoa_don', dat_id=dat_id, return_tab='checkout'))

        amount_due = max(0, invoice_ctx.get('tien_con_lai', 0))
        if amount_due <= 0:
            flash('Khong con so tien phai thanh toan.', 'info')
            return redirect(url_for('thanh_toan', dat_id=dat_id))

        invalidate_payment_sessions('room', dat_id)
        payment_token = uuid.uuid4().hex
        create_payment_session(payment_token, 'room', {
            'dat_id': dat_id,
            'calc_values': calc_values,
            'amount_due': int(amount_due),
            'tien_da_tra_truoc': invoice_ctx.get('tien_da_tra_truoc', 0),
            'summary': {
                'tien_phong': invoice_ctx.get('tien_phong', 0),
                'tien_dv': invoice_ctx.get('tien_dv', 0),
                'tien_phat': invoice_ctx.get('tien_phat', 0),
                'tien_coc': dp.tien_coc or 0,
                'tong': invoice_ctx.get('tong', 0)
            }
        })
        return redirect(url_for('show_qr_room', token=payment_token))

    invoice_ctx, _ = build_invoice_context(dp)
    invoice_ctx['selected_payment_method'] = selected_method
    invoice_ctx['payment_methods'] = PAYMENT_METHODS

    return render_template(
        'thanh_toan_phong.html',
        dp=dp,
        email_prefill=email_prefill,
        now=datetime.now(),
        **invoice_ctx
    )

@app.route('/qr/confirm/<token>')
def qr_confirm(token):
    session = get_payment_session(token)
    if not session:
        return render_template('payment_confirm.html', error='Phiên thanh toán không hợp lệ hoặc đã hết hạn.')

    ttl = get_payment_session_ttl()
    timeout_minutes = max(1, int(ttl.total_seconds() // 60))

    if payment_session_expired(session['created_at'], ttl=ttl):
        pop_payment_session(token)
        return render_template(
            'payment_confirm.html',
            error=f'Phiên thanh toán đã hết hạn (quá {timeout_minutes} phút).',
            timeout_minutes=timeout_minutes,
        )

    data = session['data']
    if data.get('completed'):
        return render_template('payment_confirm.html', error='Phiên thanh toán đã được xác nhận. Vui lòng tạo mã QR mới nếu cần.')

    kind = session['kind']
    if kind == 'deposit':
        dp = DatPhong.query.get_or_404(data['dat_id'])
        amount = data.get('amount', int(dp.tien_coc or 0))
        description = f"Tien coc dat phong #{dp.id}"
    elif kind == 'service':
        dp = DatPhong.query.get_or_404(data['dat_id'])
        amount = data.get('tong', 0)
        description = f"Dich vu cho don #{dp.id}"
    elif kind == 'room':
        dp = DatPhong.query.get_or_404(data['dat_id'])
        amount = data.get('amount_due', 0)
        description = f"Thanh toan phong #{dp.id}"
    else:
        return render_template('payment_confirm.html', error='Loại thanh toán không được hỗ trợ.')

    expires_at = payment_session_expires_at(session['created_at'], ttl=ttl)
    remaining = max(0, int((expires_at - datetime.now()).total_seconds()))

    return render_template(
        'payment_confirm.html',
        token=token,
        payment_kind=kind,
        amount=amount,
        description=description,
        expires_at=expires_at,
        expires_at_iso=expires_at.isoformat(),
        countdown_seconds=remaining,
        timeout_minutes=timeout_minutes,
    )


@app.route('/qr-image/<token>')
def qr_image(token):
    session = get_payment_session(token)
    if not session or payment_session_expired(session['created_at']):
        return send_file(io.BytesIO(b'Invalid or expired session'), mimetype='image/png')
    
    confirm_url = url_for('qr_confirm', token=token, _external=True)
    qr_buf = generate_qr_code(confirm_url)
    return send_file(qr_buf, mimetype='image/png')


@app.route('/cam-on/<token>')
def cam_on(token):
    session = get_payment_session(token)
    if not session or not session['data'].get('completed'):
        return render_template('payment_confirm.html', error='Phiên thanh toán không hợp lệ hoặc chưa hoàn tất.')
    
    data = session['data']
    kind = session['kind']
    if kind == 'deposit':
        message = 'Cảm ơn bạn đã thanh toán tiền cọc. Đặt phòng của bạn đã được xác nhận.'
    elif kind == 'service':
        message = 'Cảm ơn bạn đã thanh toán dịch vụ. Dịch vụ của bạn đã được xác nhận.'
    elif kind == 'room':
        message = 'Cảm ơn bạn đã hoàn tất thanh toán. Thủ tục trả phòng đã hoàn tất.'
    else:
        message = 'Cảm ơn bạn đã hoàn tất thanh toán.'
    
    return render_template('cam_on.html', message=message)


@app.route('/qr/deposit/<token>')
def show_qr_deposit(token):
    session = get_payment_session(token)
    if not session:
        flash('Phiên thanh toán không hợp lệ hoặc đã hết hạn.', 'danger')
        return redirect(url_for('dat_phong'))

    ttl = get_payment_session_ttl()
    timeout_minutes = max(1, int(ttl.total_seconds() // 60))

    if payment_session_expired(session['created_at'], ttl=ttl):
        pop_payment_session(token)
        flash(
            f'Phiên thanh toán đã hết hạn (quá {timeout_minutes} phút).',
            'danger',
        )
        return redirect(url_for('dat_phong'))

    data = session['data']
    dp = DatPhong.query.get_or_404(data['dat_id'])
    amount = data.get('amount', int(dp.tien_coc or 0))
    qr_url = url_for('qr_image', token=token)
    confirm_url = url_for('qr_confirm', token=token, _external=True)
    status_url = url_for('api_payment_status', token=token)
    invoice_url = url_for('in_hoa_don_coc', dat_id=dp.id)
    if data.get('completed'):
        flash('Phiên thanh toán đã được xác nhận. Vui lòng tạo mã QR mới nếu cần.', 'info')
        return redirect(invoice_url)
    expires_at = payment_session_expires_at(session['created_at'], ttl=ttl)
    remaining = max(0, int((expires_at - datetime.now()).total_seconds()))

    return render_template(
        'qr_deposit.html',
        dp=dp,
        amount=amount,
        qr_url=qr_url,
        confirm_url=confirm_url,
        status_url=status_url,
        invoice_url=invoice_url,
        countdown_seconds=remaining,
        timeout_minutes=timeout_minutes,
    )


@app.route('/qr/service/<token>')
def show_qr_service(token):
    session = get_payment_session(token)
    if not session:
        flash('Phiên thanh toán không hợp lệ hoặc đã hết hạn.', 'danger')
        return redirect(url_for('dich_vu_thanh_toan'))

    ttl = get_payment_session_ttl()
    timeout_minutes = max(1, int(ttl.total_seconds() // 60))

    if payment_session_expired(session['created_at'], ttl=ttl):
        pop_payment_session(token)
        flash(
            f'Phiên thanh toán đã hết hạn (quá {timeout_minutes} phút).',
            'danger',
        )
        return redirect(url_for('dich_vu_thanh_toan'))

    data = session['data']
    dp = DatPhong.query.get_or_404(data['dat_id'])
    amount = data.get('tong', 0)
    qr_url = url_for('qr_image', token=token)
    confirm_url = url_for('qr_confirm', token=token, _external=True)
    status_url = url_for('api_payment_status', token=token)
    invoice_url = url_for('in_hoa_don_dv', token=token)
    if data.get('completed'):
        flash('Phiên thanh toán đã được xác nhận. Vui lòng tạo mã QR mới nếu cần.', 'info')
        return redirect(invoice_url)
    expires_at = payment_session_expires_at(session['created_at'], ttl=ttl)
    remaining = max(0, int((expires_at - datetime.now()).total_seconds()))

    return render_template(
        'qr_service.html',
        dp=dp,
        amount=amount,
        qr_url=qr_url,
        confirm_url=confirm_url,
        status_url=status_url,
        invoice_url=invoice_url,
        countdown_seconds=remaining,
        items=data.get('items', []),
        timeout_minutes=timeout_minutes,
    )


@app.route('/qr/room/<token>')
def show_qr_room(token):
    session = get_payment_session(token)
    if not session:
        flash('Phiên thanh toán không hợp lệ hoặc đã hết hạn.', 'danger')
        return redirect(url_for('thanh_toan', dat_id=session['data'].get('dat_id', 0)))

    ttl = get_payment_session_ttl()
    timeout_minutes = max(1, int(ttl.total_seconds() // 60))

    if payment_session_expired(session['created_at'], ttl=ttl):
        pop_payment_session(token)
        flash(
            f'Phiên thanh toán đã hết hạn (quá {timeout_minutes} phút).',
            'danger',
        )
        return redirect(url_for('thanh_toan', dat_id=session['data'].get('dat_id', 0)))

    data = session['data']
    dp = DatPhong.query.get_or_404(data['dat_id'])
    amount = data.get('amount_due', 0)
    tien_da_tra_truoc = data.get('tien_da_tra_truoc', 0)
    qr_url = url_for('qr_image', token=token)
    confirm_url = url_for('qr_confirm', token=token, _external=True)
    status_url = url_for('api_payment_status', token=token)
    invoice_url = url_for('in_hoa_don', dat_id=dp.id)
    if data.get('completed'):
        flash('Phiên thanh toán đã được xác nhận. Vui lòng tạo mã QR mới nếu cần.', 'info')
        return redirect(invoice_url)
    expires_at = payment_session_expires_at(session['created_at'], ttl=ttl)
    remaining = max(0, int((expires_at - datetime.now()).total_seconds()))
    calc_values = data.get('calc_values', {})

    return render_template(
        'qr_room.html',
        dp=dp,
        amount_due=amount,
        tien_da_tra_truoc=tien_da_tra_truoc,
        qr_url=qr_url,
        confirm_url=confirm_url,
        status_url=status_url,
        invoice_url=invoice_url,
        countdown_seconds=remaining,
        calc_values=calc_values,
        timeout_minutes=timeout_minutes,
    )


@app.route('/api/payment/confirm/<token>', methods=['POST'])
def api_confirm_payment(token):
    session_model = PaymentSession.query.filter_by(token=token).first()
    if not session_model:
        return jsonify({'success': False, 'message': 'Phiên thanh toán không hợp lệ.'}), 404

    if payment_session_expired(session_model.created_at):
        pop_payment_session(token)
        return jsonify({'success': False, 'message': 'Phiên thanh toán đã hết hạn.'})

    try:
        data = json.loads(session_model.payload or '{}')
    except Exception:
        data = {}

    if data.get('completed'):
        return jsonify({'success': True, 'redirect_url': data.get('redirect_url')})

    kind = session_model.kind

    try:
        if kind == 'deposit':
            dat_id = data['dat_id']
            dp = DatPhong.query.get_or_404(dat_id)
            data['redirect_url'] = url_for('cam_on', token=token)
            data['message'] = 'Cảm ơn bạn đã thanh toán tiền cọc. Đặt phòng của bạn đã được xác nhận.'
            data['completed'] = True
            dp.coc_da_thanh_toan = True
            dp.phuong_thuc_coc = 'qr'
            dp.payment_token = None
            active_stay = DatPhong.query.filter(
                DatPhong.phong_id == dp.phong_id,
                DatPhong.id != dp.id,
                DatPhong.trang_thai == 'nhan'
            ).first()
            if active_stay:
                dp.trang_thai = 'waiting'
                dp.auto_confirmed_at = None
            else:
                dp.trang_thai = 'dat'
                dp.auto_confirmed_at = datetime.now()
                if dp.phong.trang_thai == 'trong':
                    dp.phong.trang_thai = 'da_dat'
            session_model.payload = json.dumps(data)
            db.session.commit()
            socketio.emit('deposit_payment_confirmed', {'dat_id': dat_id})
            return jsonify({'success': True, 'redirect_url': data['redirect_url']})

        if kind == 'service':
            dat_id = data['dat_id']
            usage_ids = data.get('usage_ids', [])
            dp = DatPhong.query.get_or_404(dat_id)
            query = SuDungDichVu.query.filter(
                SuDungDichVu.id.in_(usage_ids),
                SuDungDichVu.trang_thai == 'chua_thanh_toan'
            )
            updated = 0
            amount_total = 0
            for usage in query:
                usage.trang_thai = 'da_thanh_toan'
                amount_total += usage.dichvu.gia * usage.so_luong
                updated += 1
            if updated == 0:
                db.session.rollback()
                return jsonify({'success': False, 'message': 'Không tìm thấy dịch vụ cần xác nhận.'})
            dp.tien_dv = (dp.tien_dv or 0) + amount_total
            data['redirect_url'] = url_for('cam_on', token=token)
            data['message'] = 'Cảm ơn bạn đã thanh toán dịch vụ. Dịch vụ của bạn đã được xác nhận.'
            data['completed'] = True
            session_model.payload = json.dumps(data)
            db.session.commit()
            socketio.emit('service_payment_confirmed', {'dat_id': dat_id})
            return jsonify({'success': True, 'redirect_url': data['redirect_url']})

        if kind == 'room':
            dat_id = data['dat_id']
            calc_values = data.get('calc_values', {})
            dp = DatPhong.query.get_or_404(dat_id)
            if dp.trang_thai == 'da_thanh_toan':
                data['redirect_url'] = url_for('cam_on', token=token)
                data['message'] = 'Cảm ơn bạn đã hoàn tất thanh toán. Thủ tục trả phòng đã được hoàn tất.'
                data['completed'] = True
                session_model.payload = json.dumps(data)
                db.session.commit()
                return jsonify({'success': True, 'redirect_url': data['redirect_url']})
            dp.thuc_te_tra = dp.thuc_te_tra or datetime.now()
            dp.tien_phong = calc_values.get('tien_phong', dp.tien_phong or 0)
            dp.tien_dv = calc_values.get('tien_dv', dp.tien_dv or 0)
            dp.tien_phat = calc_values.get('tien_phat', dp.tien_phat or 0)
            dp.tong_thanh_toan = calc_values.get('tong', dp.tong_thanh_toan or 0)
            dp.phuong_thuc_thanh_toan = 'qr'
            dp.phong.trang_thai = 'trong'
            dp.trang_thai = 'da_thanh_toan'
            if current_user.is_authenticated:
                dp.nhanvien_id = current_user.id
            diem_moi = award_loyalty_points_for_booking(dp)
            data['redirect_url'] = url_for('cam_on', token=token)
            if diem_moi:
                data['message'] = f'Cảm ơn bạn đã hoàn tất thanh toán. Bạn vừa được cộng {diem_moi} điểm tích lũy.'
            else:
                data['message'] = 'Cảm ơn bạn đã hoàn tất thanh toán. Thủ tục trả phòng đã được hoàn tất.'
            data['completed'] = True
            session_model.payload = json.dumps(data)
            db.session.commit()
            socketio.emit('room_payment_confirmed', {'dat_id': dat_id})
            return jsonify({'success': True, 'redirect_url': data['redirect_url']})

        return jsonify({'success': False, 'message': 'Loại thanh toán không được hỗ trợ.'}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 500


@app.route('/api/payment/status/<token>')
def api_payment_status(token):
    session = get_payment_session(token)
    if session:
        data = session['data']
        if data.get('completed'):
            redirect_url = data.get('redirect_url')
            # Không pop session để /cam-on có thể truy cập
            return jsonify({'status': 'completed', 'redirect_url': redirect_url})
        if payment_session_expired(session['created_at']):
            pop_payment_session(token)
            return jsonify({'status': 'expired'})
        kind = session['kind']
        if kind == 'deposit':
            dp = DatPhong.query.get(data['dat_id'])
            if dp and dp.coc_da_thanh_toan:
                pop_payment_session(token)
                return jsonify({'status': 'completed', 'redirect_url': url_for('in_hoa_don_coc', dat_id=dp.id)})
        elif kind == 'service':
            remaining = SuDungDichVu.query.filter_by(datphong_id=data['dat_id'], trang_thai='chua_thanh_toan').count()
            if remaining == 0:
                pop_payment_session(token)
                return jsonify({'status': 'completed', 'redirect_url': url_for('in_hoa_don_dv', token=token)})
        elif kind == 'room':
            dp = DatPhong.query.get(data['dat_id'])
            if dp and dp.trang_thai == 'da_thanh_toan':
                pop_payment_session(token)
                return jsonify({'status': 'completed', 'redirect_url': url_for('in_hoa_don', dat_id=dp.id)})
        expires_at = payment_session_expires_at(session['created_at'])
        return jsonify({'status': 'pending', 'expires_at': expires_at.isoformat()})

    dp = DatPhong.query.filter_by(payment_token=token).first()
    if dp and dp.coc_da_thanh_toan:
        return jsonify({'status': 'completed', 'redirect_url': url_for('in_hoa_don_coc', dat_id=dp.id)})
    return jsonify({'status': 'invalid'})


@app.route('/gui-hoa-don-email/<int:dat_id>', methods=['POST'])
@login_required
@permission_required('payments.invoices')
def gui_hoa_don_email(dat_id):
    email_to = (request.form.get('email_to') or '').strip()
    if not email_to:
        flash('Vui lòng nhập địa chỉ email khách hàng.', 'warning')
        return redirect(url_for('thanh_toan', dat_id=dat_id))

    dp = DatPhong.query.get_or_404(dat_id)
    dich_vu_su_dung = SuDungDichVu.query.filter_by(datphong_id=dp.id, trang_thai='da_thanh_toan').all()
    tien_dich_vu_da_thanh_toan = sum(
        (dv.so_luong * dv.dichvu.gia) for dv in dich_vu_su_dung if dv.dichvu
    )

    # Calculate actual amounts using snapshot_and_bill
    so_dem, tien_phong, tong_tien_dv, tien_phat, tong, checkin, checkout, don_vi_tinh, so_luong_tinh = snapshot_and_bill(dp)
    
    # Build service details
    chi_tiet_dv_lines = []
    for dv in dich_vu_su_dung:
        line = f"  - {dv.dichvu.ten}: {dv.so_luong} x {vnd(dv.dichvu.gia)} = {vnd(dv.so_luong * dv.dichvu.gia)}"
        chi_tiet_dv_lines.append(line)
    
    # Calculate remaining amount
    tien_con_lai = max(0, tong - (dp.tien_coc or 0) - tien_dich_vu_da_thanh_toan)
    
    # Build context with calculated values
    context = build_booking_email_context(dp)
    context['chi_tiet_dich_vu'] = '\n'.join(chi_tiet_dv_lines) if chi_tiet_dv_lines else 'Không sử dụng dịch vụ'
    context['tong_tien'] = vnd(tong)  # Use calculated total
    context['tien_phong'] = vnd(tien_phong)
    context['tien_dich_vu'] = vnd(tong_tien_dv)
    context['tien_phat'] = vnd(tien_phat) if tien_phat > 0 else ''
    context['tien_dich_vu_da_thanh_toan'] = vnd(tien_dich_vu_da_thanh_toan) if tien_dich_vu_da_thanh_toan > 0 else ''
    context['con_lai'] = vnd(tien_con_lai) if tien_con_lai > 0 else ''
    context['so_tien_da_thanh_toan'] = vnd(max(0, tong - tien_con_lai)) if tong - tien_con_lai > 0 else vnd(0)
    context['tien_coc_display'] = vnd(dp.tien_coc or 0) if dp.tien_coc else ''
    context['hinh_thuc_coc'] = context.get('phuong_thuc_coc_label') or ''
    context['hinh_thuc_thanh_toan'] = context.get('payment_method_label') or ''
    context['so_dem'] = so_dem
    if dp.hinh_thuc_thue == 'gio':
        context['nhan_luu_tru'] = 'Số giờ lưu trú'
        context['don_vi_luu_tru'] = 'giờ'
        context['so_luong_luu_tru'] = so_luong_tinh
    else:
        context['nhan_luu_tru'] = 'Số đêm lưu trú'
        context['don_vi_luu_tru'] = 'đêm'
        context['so_luong_luu_tru'] = so_dem

    # Send email without PDF attachment - use formatted email body instead
    try:
        send_email_with_template('invoice_notice', email_to, context, attachments=None,
                                datphong_id=dp.id, khachhang_id=dp.khachhang_id)
        flash(f'Đã gửi hóa đơn tới {email_to}.', 'success')
    except (RuntimeError, ValueError) as exc:
        flash(str(exc), 'danger')
    except (smtplib.SMTPException, OSError) as exc:
        app.logger.exception('Không thể gửi email hóa đơn: %s', exc)
        flash('Không thể gửi email hóa đơn. Vui lòng kiểm tra cấu hình SMTP và thử lại.', 'danger')
    except Exception as exc:
        app.logger.exception('Lỗi không xác định khi gửi hóa đơn: %s', exc)
        flash('Hệ thống gặp lỗi khi gửi email. Vui lòng thử lại sau.', 'danger')

    # Redirect based on booking status
    if dp.trang_thai == 'da_thanh_toan':
        return redirect(url_for('in_hoa_don', dat_id=dat_id))
    else:
        return redirect(url_for('thanh_toan', dat_id=dat_id, email=email_to))

@app.route('/in-hoa-don-coc/<int:dat_id>')
@login_required
@permission_required('payments.invoices')
def in_hoa_don_coc(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if not dp.coc_da_thanh_toan:
        flash('Đặt phòng này chưa thanh toán tiền cọc.', 'warning')
        return redirect(url_for('dat_phong'))
    
    discount_percent, _ = get_voucher_config()
    # Since deposit is already paid, don't generate QR code
    # Pass payment method and payment time
    payment_method = dp.phuong_thuc_coc or 'Chưa xác định'
    
    # Get actual payment time from payment session if available
    payment_time = dp.created_at  # Default to booking creation time
    if dp.payment_token:
        session = PaymentSession.query.filter_by(token=dp.payment_token).first()
        if session:
            payment_time = session.created_at
    
    return render_template('hoa_don_coc.html', dp=dp, now=datetime.now(), voucher_discount=discount_percent, 
                          payment_method=payment_method, payment_time=payment_time)

@app.route('/in-hoa-don-dv/<token>')
@login_required
@permission_required('payments.invoices')
def in_hoa_don_dv(token):
    # Try to get from payment session first
    session = get_payment_session(token)
    if session and session['kind'] == 'service':
        data = session['data']
        dat_id = data['dat_id']
        dp = DatPhong.query.get_or_404(dat_id)
        
        # Get services from payment session
        usage_ids = data.get('usage_ids', [])
        if usage_ids:
            dich_vu_su_dung = SuDungDichVu.query.filter(
                SuDungDichVu.id.in_(usage_ids),
                SuDungDichVu.trang_thai == 'da_thanh_toan'
            ).all()
        else:
            # Fallback: get all paid services for this booking
            dich_vu_su_dung = SuDungDichVu.query.filter_by(
                datphong_id=dat_id, 
                trang_thai='da_thanh_toan'
            ).all()
    else:
        # Fallback for old links: get dat_id from token (assuming it's dat_id)
        try:
            dat_id = int(token)
            dp = DatPhong.query.get_or_404(dat_id)
            dich_vu_su_dung = SuDungDichVu.query.filter_by(
                datphong_id=dat_id, 
                trang_thai='da_thanh_toan'
            ).all()
        except (ValueError, TypeError):
            flash('Token không hợp lệ.', 'danger')
            return redirect(url_for('dich_vu_thanh_toan'))
    
    if not dich_vu_su_dung:
        flash('Không có dịch vụ đã thanh toán để in hóa đơn.', 'warning')
        return redirect(url_for('dich_vu_thanh_toan', dat_id=dat_id))
    
    # Tạo danh sách items cho template
    items = []
    tong_tien = 0
    for dv in dich_vu_su_dung:
        thanh_tien = dv.so_luong * dv.dichvu.gia
        items.append({
            'ten': dv.dichvu.ten,
            'gia': dv.dichvu.gia,
            'so_luong': dv.so_luong,
            'thanh_tien': thanh_tien
        })
        tong_tien += thanh_tien

    qr_code_url = build_vietqr_url(tong_tien, f"Dich vu cho don #{dp.id} {dp.khachhang.ho_ten}")
    
    # Check if there are unpaid services - if so, show QR for additional payments
    unpaid_services_count = SuDungDichVu.query.filter_by(datphong_id=dat_id, trang_thai='chua_thanh_toan').count()
    show_payment_qr = unpaid_services_count > 0
    
    # Get actual payment time from payment session
    paid_at = datetime.now()  # Default to current time
    if session:
        paid_at = session['created_at']
    
    return render_template('hoa_don_dv.html', dp=dp, items=items, tong=tong_tien, qr_code_url=qr_code_url, paid_at=paid_at, show_payment_qr=show_payment_qr)

@app.route('/in-hoa-don/<int:dat_id>')
@login_required
@permission_required('payments.invoices')
def in_hoa_don(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if dp.trang_thai != 'da_thanh_toan':
        flash('Đặt phòng này chưa hoàn tất thanh toán.', 'warning')
        return redirect(url_for('thanh_toan', dat_id=dat_id))
    
    invoice_ctx, calc_values = build_invoice_context(dp)
    # Add payment time for completed invoices
    payment_time = dp.thuc_te_tra or dp.created_at  # Use check-out time or booking creation time
    
    # Try to get actual payment time from payment session
    if dp.payment_token:
        session = PaymentSession.query.filter_by(token=dp.payment_token).first()
        if session:
            payment_time = session.created_at
    
    return render_template('hoa_don.html', dp=dp, now=datetime.now(), payment_time=payment_time, **invoice_ctx)

@app.route('/xoa-sudung-dichvu/<int:sudungdv_id>', methods=['POST'])
@login_required
@permission_required('payments.process', 'services.orders')
def xoa_sudung_dichvu(sudungdv_id):
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    sd_item = SuDungDichVu.query.get_or_404(sudungdv_id)
    datphong_id = sd_item.datphong_id
    db.session.delete(sd_item)
    db.session.commit()
    booking_payload = build_service_booking_payload(DatPhong.query.get(datphong_id))
    message = 'Đã xóa dịch vụ khỏi hóa đơn.'
    if wants_json:
        return jsonify({'message': message, 'booking': booking_payload})
    flash(message, 'success')
    return redirect(url_for('dich_vu_thanh_toan', dat_id=datphong_id))

@app.route('/tra-phong/<int:dat_id>', methods=['POST'])
@login_required
@permission_required('bookings.checkin_checkout')
def tra_phong(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if not dp.thuc_te_tra:
        dp.thuc_te_tra = datetime.now()
        db.session.commit()
    return redirect(url_for('thanh_toan', dat_id=dat_id))


# ============================================
# GIA HẠN PHÒNG - EXTEND BOOKING
# ============================================
@app.route('/gia-han-phong/<int:dat_id>', methods=['GET', 'POST'])
@login_required
@permission_required('bookings.extend')
def gia_han_phong(dat_id):
    """
    Gia hạn thời gian thuê phòng cho khách hàng đang ở.
    
    Điều kiện:
    - Booking phải ở trạng thái 'nhan' (đang ở)
    - Không được trùng với booking khác đã đặt phòng này
    - Gia hạn không bị tính phí phạt
    """
    dp = DatPhong.query.options(
        db.joinedload(DatPhong.khachhang),
        db.joinedload(DatPhong.phong).joinedload(Phong.loai)
    ).get_or_404(dat_id)
    
    # Kiểm tra điều kiện gia hạn
    if dp.trang_thai != 'nhan':
        flash('Chỉ có thể gia hạn cho khách đang ở phòng.', 'danger')
        return redirect(url_for('nhan_phong', tab='checkout'))
    
    if request.method == 'POST':
        try:
            # Lấy thông tin gia hạn
            ngay_tra_moi_str = request.form.get('ngay_tra_moi')
            if not ngay_tra_moi_str:
                flash('Vui lòng chọn ngày trả mới.', 'danger')
                return redirect(url_for('gia_han_phong', dat_id=dat_id))
            
            # Parse ngày trả mới
            ngay_tra_moi = datetime.strptime(ngay_tra_moi_str, '%Y-%m-%dT%H:%M')
            ngay_tra_hien_tai = dp.ngay_tra
            
            # Validate ngày trả mới
            if ngay_tra_moi <= ngay_tra_hien_tai:
                flash('Ngày trả mới phải sau ngày trả hiện tại.', 'danger')
                return redirect(url_for('gia_han_phong', dat_id=dat_id))
            
            if ngay_tra_moi <= datetime.now():
                flash('Ngày trả mới phải sau thời điểm hiện tại.', 'danger')
                return redirect(url_for('gia_han_phong', dat_id=dat_id))
            
            # Kiểm tra xung đột với booking khác
            conflicting_bookings = DatPhong.query.filter(
                DatPhong.phong_id == dp.phong_id,
                DatPhong.id != dat_id,
                DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES),
                # Kiểm tra overlap: booking khác bắt đầu trước khi kết thúc gia hạn
                DatPhong.ngay_nhan < ngay_tra_moi,
                # Và kết thúc sau ngày trả hiện tại
                DatPhong.ngay_tra > ngay_tra_hien_tai
            ).all()
            
            if conflicting_bookings:
                # Tìm booking gần nhất
                nearest = min(conflicting_bookings, key=lambda b: b.ngay_nhan)
                flash(
                    f'Không thể gia hạn đến {ngay_tra_moi.strftime("%d/%m/%Y %H:%M")}. '
                    f'Phòng đã có khách đặt từ {nearest.ngay_nhan.strftime("%d/%m/%Y %H:%M")} '
                    f'(Khách: {nearest.khachhang.ho_ten}). '
                    f'Bạn chỉ có thể gia hạn đến trước thời điểm này.',
                    'warning'
                )
                return redirect(url_for('gia_han_phong', dat_id=dat_id))
            
            # Tính số đêm thêm
            delta = ngay_tra_moi - ngay_tra_hien_tai
            so_dem_them = max(1, int(delta.total_seconds() / 86400))
            
            # Tính tiền phòng thêm (KHÔNG CÓ PHÍ PHẠT)
            gia_phong = dp.phong.loai.gia
            tien_phong_them = gia_phong * so_dem_them
            
            # Cập nhật booking
            dp.ngay_tra = ngay_tra_moi
            dp.so_dem += so_dem_them
            dp.tien_phong = (dp.tien_phong or 0) + tien_phong_them
            
            # Reset tổng thanh toán để tính lại khi thanh toán
            # (sẽ bao gồm tiền phòng mới + dịch vụ + cọc đã trả)
            dp.tong_thanh_toan = 0
            
            db.session.commit()
            
            # Gửi thông báo qua chat (nếu có)
            if dp.chat_token:
                try:
                    msg = f"Phòng của bạn đã được gia hạn đến {ngay_tra_moi.strftime('%d/%m/%Y %H:%M')}. Thêm {so_dem_them} đêm, tiền phòng thêm: {vnd(tien_phong_them)}."
                    tn = TinNhan(
                        datphong_id=dat_id,
                        nguoi_gui='he_thong',
                        noi_dung=msg,
                        thoi_gian=datetime.now(),
                        trang_thai='chua_doc'
                    )
                    db.session.add(tn)
                    db.session.commit()
                    
                    # Emit SocketIO event
                    socketio.emit('new_message', {
                        'booking_id': dat_id,
                        'message': msg,
                        'sender': 'system'
                    })
                except Exception as e:
                    app.logger.warning(f'Không thể gửi thông báo gia hạn: {e}')
            
            flash(
                f'Đã gia hạn phòng thành công! '
                f'Ngày trả mới: {ngay_tra_moi.strftime("%d/%m/%Y %H:%M")}. '
                f'Thêm {so_dem_them} đêm, tiền phòng thêm: {vnd(tien_phong_them)}.',
                'success'
            )
            return redirect(url_for('nhan_phong', tab='checkout'))
            
        except ValueError as e:
            flash(f'Định dạng ngày giờ không hợp lệ: {e}', 'danger')
            return redirect(url_for('gia_han_phong', dat_id=dat_id))
        except Exception as e:
            app.logger.error(f'Lỗi gia hạn phòng: {e}')
            db.session.rollback()
            flash('Có lỗi xảy ra khi gia hạn phòng. Vui lòng thử lại.', 'danger')
            return redirect(url_for('gia_han_phong', dat_id=dat_id))
    
    # GET request - hiển thị form gia hạn
    # Tìm booking tiếp theo sau booking hiện tại (nếu có)
    next_booking = DatPhong.query.filter(
        DatPhong.phong_id == dp.phong_id,
        DatPhong.id != dat_id,
        DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES),
        DatPhong.ngay_nhan >= dp.ngay_tra
    ).order_by(DatPhong.ngay_nhan.asc()).first()
    
    # Tính thời gian gia hạn tối đa
    max_extend_date = None
    if next_booking:
        # Chỉ gia hạn đến trước khi khách tiếp theo nhận phòng
        max_extend_date = next_booking.ngay_nhan - timedelta(hours=2)  # Buffer 2 giờ
    
    # Tính min date (1 giờ sau ngày trả hiện tại)
    min_extend_date = dp.ngay_tra + timedelta(hours=1)
    
    return render_template(
        'gia_han_phong.html',
        dp=dp,
        next_booking=next_booking,
        max_extend_date=max_extend_date,
        min_extend_date=min_extend_date,
        gia_phong=dp.phong.loai.gia
    )


@app.route('/khach-hang')
@login_required
@permission_required('customers.view')
def khach_hang():
    ds_khach = DatPhong.query.order_by(DatPhong.ngay_nhan.asc()).all()
    return render_template('khach_hang.html', ds_khach=ds_khach)


@app.route('/quan-ly-tai-khoan-khach-hang')
@login_required
@permission_required('customers.view')
def quan_ly_tai_khoan_khach_hang():
    keyword = (request.args.get('q') or '').strip()
    trang_thai = request.args.get('trang_thai') or ''
    
    # Pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = 3  # 3 customers per page
    
    query = KhachHang.query.filter(
        or_(
            KhachHang.mat_khau_hash.isnot(None),
            KhachHang.lan_dang_nhap_cuoi.isnot(None)
        ),
        KhachHang.trang_thai_tai_khoan != 'da_xoa'
    )
    if keyword:
        like_pattern = f"%{keyword}%"
        query = query.filter(
            or_(
                KhachHang.ho_ten.ilike(like_pattern),
                KhachHang.cmnd.ilike(like_pattern),
                KhachHang.email.ilike(like_pattern)
            )
        )
    if trang_thai in {'hoat_dong', 'khoa'}:
        query = query.filter(KhachHang.trang_thai_tai_khoan == trang_thai)
    
    # Get all accounts for stats
    all_accounts = query.all()
    
    # Apply pagination
    query = query.order_by(
        KhachHang.trang_thai_tai_khoan.asc(),
        case((KhachHang.ngay_cap_nhat.is_(None), 1), else_=0),
        KhachHang.ngay_cap_nhat.desc(),
        KhachHang.ho_ten.asc()
    )
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    accounts = pagination.items
    
    return render_template(
        'quan_ly_tai_khoan_khach_hang.html',
        accounts=accounts,
        all_accounts=all_accounts,
        pagination=pagination,
        keyword=keyword,
        selected_status=trang_thai
    )


@app.route('/quan-ly-tai-khoan-khach-hang/<int:kh_id>/trang-thai', methods=['POST'])
@login_required
@permission_required('customers.view')
def cap_nhat_trang_thai_khach_hang(kh_id):
    kh = KhachHang.query.get_or_404(kh_id)
    trang_thai = request.form.get('trang_thai')
    ly_do = (request.form.get('ly_do') or '').strip()
    redirect_to = request.form.get('redirect_to')

    if trang_thai not in {'hoat_dong', 'khoa'}:
        flash('Trạng thái tài khoản không hợp lệ.', 'danger')
        return _redirect_after_status_change(redirect_to)

    try:
        hotel = get_hotel_profile()
    except Exception:
        hotel = {'name': 'Khách sạn'}

    if trang_thai == 'khoa':
        if not ly_do:
            flash('Vui lòng nhập lý do khóa tài khoản.', 'warning')
            return _redirect_after_status_change(redirect_to)
        kh.trang_thai_tai_khoan = 'khoa'
        kh.ngay_cap_nhat = datetime.now()
        db.session.commit()
        if kh.email:
            try:
                send_email_with_template(
                    'customer_account_locked',
                    kh.email,
                    {
                        'ho_ten': kh.ho_ten,
                        'ten_khach_san': hotel.get('name', 'Khách sạn'),
                        'ly_do': ly_do
                    },
                    khachhang_id=kh.id
                )
            except Exception as exc:
                app.logger.warning('Không thể gửi email khóa tài khoản: %s', exc)
        flash('Đã khóa tài khoản khách hàng và gửi thông báo.', 'success')
    else:
        kh.trang_thai_tai_khoan = 'hoat_dong'
        kh.ngay_cap_nhat = datetime.now()
        db.session.commit()
        if kh.email:
            try:
                send_email_with_template(
                    'customer_account_unlocked',
                    kh.email,
                    {
                        'ho_ten': kh.ho_ten,
                        'ten_khach_san': hotel.get('name', 'Khách sạn'),
                        'dang_nhap_url': url_for('khach_hang_dang_nhap', _external=True)
                    },
                    khachhang_id=kh.id
                )
            except Exception as exc:
                app.logger.warning('Không thể gửi email mở khóa tài khoản: %s', exc)
        flash('Đã mở khóa tài khoản khách hàng và gửi thông báo.', 'success')

    return _redirect_after_status_change(redirect_to)


def _redirect_after_status_change(redirect_to):
    if redirect_to:
        parsed = urlparse(redirect_to)
        if not parsed.netloc and not parsed.scheme:
            return redirect(redirect_to)
    return redirect(url_for('quan_ly_tai_khoan_khach_hang'))


@app.route('/quan-ly-tai-khoan-khach-hang/<int:kh_id>/xoa', methods=['POST'])
@login_required
@permission_required('customers.manage_accounts')
def xoa_tai_khoan_khach_hang(kh_id):
    kh = KhachHang.query.get_or_404(kh_id)
    redirect_to = request.form.get('redirect_to')
    ly_do = (request.form.get('ly_do_xoa') or '').strip()

    if not ly_do:
        flash('Vui lòng nhập lý do xóa tài khoản.', 'warning')
        return _redirect_after_status_change(redirect_to)

    if not kh.co_tai_khoan:
        flash('Tài khoản này đã được xóa hoặc không có thông tin đăng nhập.', 'info')
        return _redirect_after_status_change(redirect_to)

    kh.deleted_at = datetime.now()
    kh.deleted_reason = ly_do
    kh.deleted_by = current_user.id if getattr(current_user, 'id', None) else None
    kh.trang_thai_tai_khoan = 'da_xoa'
    kh.mat_khau_hash = None
    kh.ngay_cap_nhat = datetime.now()

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        app.logger.error('Không thể xóa tài khoản khách hàng %s: %s', kh.id, exc)
        flash('Không thể xóa tài khoản ngay lúc này. Vui lòng thử lại.', 'danger')
        return _redirect_after_status_change(redirect_to)

    if kh.email:
        try:
            hotel = get_hotel_profile()
            send_email_with_template(
                'customer_account_deleted',
                kh.email,
                {
                    'ho_ten': kh.ho_ten,
                    'ten_khach_san': hotel.get('name', 'Khách sạn'),
                    'ly_do': ly_do
                },
                khachhang_id=kh.id
            )
        except Exception as exc:
            app.logger.warning('Không thể gửi email xóa tài khoản cho %s: %s', kh.email, exc)

    flash('Đã xóa tài khoản khách hàng và thông báo tới khách hàng.', 'success')
    return _redirect_after_status_change(redirect_to)

@app.route('/thong-tin-ca-nhan', methods=['GET', 'POST'])
@login_required
def thong_tin_ca_nhan():
    if request.method == 'POST':
        form_name = request.form.get('form_name')
        if form_name == 'change_password':
            if current_user.mat_khau != request.form.get('mat_khau_cu'):
                flash('Mật khẩu cũ không chính xác.', 'danger')
            elif request.form.get('mat_khau_moi') != request.form.get('xac_nhan_mk'):
                flash('Mật khẩu mới và xác nhận không khớp.', 'danger')
            else:
                current_user.mat_khau = request.form['mat_khau_moi']
                db.session.commit()
                flash('Đổi mật khẩu thành công.', 'success')
        elif form_name == 'change_avatar':
            file = request.files.get('avatar')
            if not file or not file.filename:
                flash('Vui lòng chọn ảnh để tải lên.', 'warning')
            elif not allowed_avatar(file.filename):
                flash('Định dạng ảnh không được hỗ trợ.', 'danger')
            else:
                filename = secure_filename(file.filename)
                ext = filename.rsplit('.', 1)[1].lower()
                unique_name = f"user_{current_user.id}_{int(datetime.now().timestamp())}.{ext}"
                upload_dir = app.config['AVATAR_UPLOAD_FOLDER']
                os.makedirs(upload_dir, exist_ok=True)
                file_path = os.path.join(upload_dir, unique_name)
                file.save(file_path)

                old_relative = current_user.anh_dai_dien
                if old_relative and old_relative.startswith('uploads/avatars/'):
                    old_path = os.path.join(app.root_path, 'static', old_relative.replace('/', os.sep))
                    if os.path.exists(old_path):
                        try:
                            os.remove(old_path)
                        except OSError:
                            pass

                current_user.anh_dai_dien = f"uploads/avatars/{unique_name}"
                db.session.commit()
                flash('Cập nhật ảnh đại diện thành công.', 'success')
        else:
            flash('Yêu cầu không hợp lệ.', 'warning')
        return redirect(url_for('thong_tin_ca_nhan'))
    return render_template('thong_tin_ca_nhan.html')



@app.route('/luong-thuong')
@login_required
def luong_thuong():
    luong_record = LuongNhanVien.query.filter_by(nguoidung_id=current_user.id).first()
    luong_co_ban = luong_record.luong_co_ban if luong_record else 0
    phu_cap = luong_record.phu_cap if luong_record else 0
    salary_mode = get_salary_mode()

    # Kiểm tra số ngày công trong tháng
    now = datetime.now()
    first_day = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
    work_days = db.session.query(func.count()).filter(
        Attendance.user_id == current_user.id,
        Attendance.status == 'approved',
        Attendance.checkin_time >= first_day,
        Attendance.checkin_time <= last_day
    ).scalar() or 0
    work_days = int(work_days or 0)
    min_days = get_min_work_days()
    if work_days < min_days:
        phu_cap = 0
    base_effective = compute_effective_base_salary(luong_co_ban, work_days, salary_mode)
    daily_rate = compute_daily_rate(luong_co_ban) if salary_mode == SALARY_MODE_DAILY else 0

    now = datetime.now()
    start_month = datetime(now.year, now.month, 1)
    next_month = (start_month + timedelta(days=32)).replace(day=1)

    doanh_thu = db.session.query(func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0)).filter(
        DatPhong.nhanvien_id == current_user.id,
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).scalar()
    doanh_thu = int(doanh_thu or 0)

    tiers = LuongThuongCauHinh.query.order_by(LuongThuongCauHinh.moc_duoi.asc()).all()
    thuong, ty_le = tinh_thuong_doanh_thu(doanh_thu, tiers)
    bonus_amount = get_top_bonus()

    bonus_amount = get_top_bonus()

    bonus_amount = get_top_bonus()

    top_rows = db.session.query(
        DatPhong.nhanvien_id.label('nv_id'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('doanh_thu')
    ).filter(
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).group_by(DatPhong.nhanvien_id).all()
    top_bonus = 0
    if top_rows:
        top_max = max(int(row.doanh_thu or 0) for row in top_rows)
        if top_max > 0 and any(int(row.nv_id) == current_user.id and int(row.doanh_thu or 0) == top_max for row in top_rows):
            top_bonus = bonus_amount

    salary_info = {
        'luong_co_ban': base_effective,
        'base_monthly': luong_co_ban,
        'salary_mode': salary_mode,
        'daily_rate': daily_rate,
        'daily_divisor': SALARY_DAILY_DIVISOR,
        'work_days': work_days,
        'min_days': min_days,
        'phu_cap': phu_cap,
        'thuong_thang': thuong,
        'ty_le': ty_le,
        'top_bonus': top_bonus,
        'is_top': bool(top_bonus),
        'tong': base_effective + phu_cap + thuong + top_bonus,
        'doanh_thu': doanh_thu,
        'configured_top_bonus': bonus_amount
    }

    return render_template('luong_thuong.html', salary_info=salary_info, tiers=tiers, start_month=start_month, bonus_amount=bonus_amount)

@app.route('/tai-xuong-luong-excel')
@login_required
def tai_xuong_luong_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    # Debug: kiểm tra vai trò hiện tại
    print(f"DEBUG: current_user.role_slug = '{current_user.role_slug}'")
    print(f"DEBUG: current_user.ten = '{current_user.ten}'")
    print(f"DEBUG: current_user.id = {current_user.id}")

    # Tính toán thông tin lương giống như route luong_thuong
    luong_record = LuongNhanVien.query.filter_by(nguoidung_id=current_user.id).first()
    luong_co_ban = luong_record.luong_co_ban if luong_record else 0
    phu_cap = luong_record.phu_cap if luong_record else 0
    salary_mode = get_salary_mode()

    # Kiểm tra số ngày công trong tháng
    now = datetime.now()
    first_day = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
    work_days = db.session.query(func.count()).filter(
        Attendance.user_id == current_user.id,
        Attendance.status == 'approved',
        Attendance.checkin_time >= first_day,
        Attendance.checkin_time <= last_day
    ).scalar() or 0
    work_days = int(work_days or 0)
    min_days = get_min_work_days()
    if work_days < min_days:
        phu_cap = 0
    base_effective = compute_effective_base_salary(luong_co_ban, work_days, salary_mode)
    daily_rate = compute_daily_rate(luong_co_ban) if salary_mode == SALARY_MODE_DAILY else 0

    start_month = datetime(now.year, now.month, 1)
    next_month = (start_month + timedelta(days=32)).replace(day=1)

    doanh_thu = db.session.query(func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0)).filter(
        DatPhong.nhanvien_id == current_user.id,
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).scalar()
    doanh_thu = int(doanh_thu or 0)

    tiers = LuongThuongCauHinh.query.order_by(LuongThuongCauHinh.moc_duoi.asc()).all()
    thuong, ty_le = tinh_thuong_doanh_thu(doanh_thu, tiers)
    bonus_amount = get_top_bonus()

    top_rows = db.session.query(
        DatPhong.nhanvien_id.label('nv_id'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('doanh_thu')
    ).filter(
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).group_by(DatPhong.nhanvien_id).all()
    top_bonus = 0
    if top_rows:
        top_max = max(int(row.doanh_thu or 0) for row in top_rows)
        if top_max > 0 and any(int(row.nv_id) == current_user.id and int(row.doanh_thu or 0) == top_max for row in top_rows):
            top_bonus = bonus_amount

    tong_luong = base_effective + phu_cap + thuong + top_bonus

    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng lương cá nhân"

    # Định nghĩa styles
    title_font = Font(name='Arial', size=16, bold=True, color='2F7D5A')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F7D5A', end_color='2F7D5A', fill_type='solid')
    data_font = Font(name='Arial', size=11)
    total_font = Font(name='Arial', size=12, bold=True, color='2F7D5A')
    total_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Tiêu đề
    ws['A1'] = f'BẢNG LƯƠNG CÁ NHÂN - THÁNG {start_month.strftime("%m/%Y")}'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')

    # Thông tin nhân viên
    ws['A3'] = 'Họ và tên:'
    ws['B3'] = current_user.ten
    ws['A4'] = 'Chức vụ:'
    ws['B4'] = current_user.role_name or ''
    ws['A5'] = 'Tháng:'
    ws['B5'] = start_month.strftime('%m/%Y')

    # Header cho bảng lương
    headers = ['STT', 'Khoản mục', 'Số tiền', 'Ghi chú']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Dữ liệu lương
    base_note = f'Lương tháng {now.month}/{now.year}'
    if salary_mode == SALARY_MODE_DAILY and work_days > 0:
        base_note += f' | {work_days} ngày x {daily_rate:,.0f} VNĐ (chia {SALARY_DAILY_DIVISOR})'
    data = [
        (1, 'Lương cơ bản', f"{base_effective:,.0f} VNĐ", base_note),
        (2, 'Phụ cấp', f"{phu_cap:,.0f} VNĐ", f'Ngày công: {work_days}/{min_days}'),
        (3, 'Doanh thu tháng', f"{doanh_thu:,.0f} VNĐ", ''),
        (4, 'Thưởng doanh thu', f"{thuong:,.0f} VNĐ", f'Tỷ lệ: {ty_le:.1f}%' if ty_le else ''),
        (5, 'Thưởng top doanh thu', f"{top_bonus:,.0f} VNĐ", 'Top 1 doanh thu' if top_bonus else ''),
        (6, 'TỔNG CỘNG', f"{tong_luong:,.0f} VNĐ", '')
    ]

    for row_num, (stt, item, amount, note) in enumerate(data, 8):
        # STT
        cell = ws.cell(row=row_num, column=1)
        cell.value = stt
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # Khoản mục
        cell = ws.cell(row=row_num, column=2)
        cell.value = item
        cell.font = total_font if 'TỔNG' in item else data_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

        # Số tiền
        cell = ws.cell(row=row_num, column=3)
        cell.value = amount
        cell.font = total_font if 'TỔNG' in item else data_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = thin_border
        if 'TỔNG' in item:
            cell.fill = total_fill

        # Ghi chú
        cell = ws.cell(row=row_num, column=4)
        cell.value = note
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

    # Thang thưởng (nếu có)
    if tiers:
        ws['A15'] = 'THANG THƯỞNG DOANH THU'
        ws['A15'].font = Font(name='Arial', size=14, bold=True, color='2F7D5A')
        ws.merge_cells('A15:D15')

        tier_headers = ['Mốc doanh thu', 'Tỷ lệ thưởng', 'Ghi chú']
        for col_num, header in enumerate(tier_headers, 1):
            cell = ws.cell(row=17, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_num, tier in enumerate(tiers, 18):
            # Mốc doanh thu
            moc_text = f"Từ {tier.moc_duoi:,.0f}"
            if tier.moc_tren:
                moc_text += f" đến {tier.moc_tren:,.0f}"
            else:
                moc_text += "+"
            ws.cell(row=row_num, column=1).value = moc_text
            ws.cell(row=row_num, column=1).font = data_font
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_num, column=1).border = thin_border

            # Tỷ lệ thưởng
            ws.cell(row=row_num, column=2).value = f"{tier.ty_le}%"
            ws.cell(row=row_num, column=2).font = data_font
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_num, column=2).border = thin_border

            # Ghi chú
            ws.cell(row=row_num, column=3).value = tier.ghi_chu or ''
            ws.cell(row=row_num, column=3).font = data_font
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_num, column=3).border = thin_border

    # Tự động điều chỉnh độ rộng cột
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 30)  # Giới hạn tối đa 30
        ws.column_dimensions[column_letter].width = adjusted_width

    # Tạo response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Bảng lương {current_user.ten.replace(" ", "_")}_{start_month.strftime("%m_%Y")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/so-do-phong')
@login_required
@permission_required('bookings.view_map')
def so_do_phong():
    phongs = Phong.query.order_by(Phong.ten).all()
    loai_phongs = LoaiPhong.query.all()
    now = datetime.now()

    overdue_bookings = DatPhong.query.filter(
        DatPhong.trang_thai == 'nhan',
        DatPhong.ngay_tra < now
    ).all()
    overdue_phong_ids = {booking.phong_id for booking in overdue_bookings}

    phong_ids = [p.id for p in phongs]

    active_by_room = {}
    upcoming_by_room = {}

    if phong_ids:
        active_bookings = (
            DatPhong.query.options(joinedload(DatPhong.khachhang))
            .filter(DatPhong.phong_id.in_(phong_ids), DatPhong.trang_thai == 'nhan')
            .order_by(DatPhong.ngay_nhan.desc())
            .all()
        )
        for booking in active_bookings:
            active_by_room.setdefault(booking.phong_id, booking)

        upcoming_bookings = (
            DatPhong.query.options(joinedload(DatPhong.khachhang))
            .filter(
                DatPhong.phong_id.in_(phong_ids),
                DatPhong.trang_thai.in_(['dat', 'cho_xac_nhan']),
                DatPhong.ngay_tra >= now
            )
            .order_by(DatPhong.ngay_nhan.asc())
            .all()
        )
        
        # Filter out expired cho_xac_nhan bookings
        valid_upcoming_bookings = []
        for booking in upcoming_bookings:
            if booking.trang_thai == 'dat':
                valid_upcoming_bookings.append(booking)
            elif booking.trang_thai == 'cho_xac_nhan':
                # Check if booking has active payment session or is within 5 minutes of creation
                is_valid = False
                if booking.payment_token:
                    session = PaymentSession.query.filter_by(token=booking.payment_token).first()
                    if session and not payment_session_expired(session.created_at):
                        is_valid = True
                else:
                    # No payment session yet, check if booking is within 5 minutes
                    if not payment_session_expired(booking.created_at):
                        is_valid = True
                
                if is_valid:
                    valid_upcoming_bookings.append(booking)
                else:
                    # Mark expired cho_xac_nhan as cancelled
                    booking.trang_thai = 'huy'
                    db.session.commit()
        
        upcoming_bookings = valid_upcoming_bookings
        for booking in upcoming_bookings:
            upcoming_by_room.setdefault(booking.phong_id, booking)

    for p in phongs:
        p.current_booking = active_by_room.get(p.id)
        p.upcoming_booking = upcoming_by_room.get(p.id)

        if p.id in overdue_phong_ids:
            p.calculated_status = 'qua_gio'
        elif p.current_booking:
            p.calculated_status = 'dang_o'
        elif p.upcoming_booking and p.upcoming_booking.trang_thai == 'dat':
            p.calculated_status = 'da_dat'
        elif p.upcoming_booking and p.upcoming_booking.trang_thai == 'cho_xac_nhan':
            p.calculated_status = 'cho_thanh_toan'
        else:
            p.calculated_status = 'trong'

    # Calculate statistics
    stats = {
        'trong': sum(1 for p in phongs if p.calculated_status == 'trong'),
        'dang_o': sum(1 for p in phongs if p.calculated_status == 'dang_o'),
        'da_dat': sum(1 for p in phongs if p.calculated_status == 'da_dat'),
        'cho_thanh_toan': sum(1 for p in phongs if p.calculated_status == 'cho_thanh_toan'),
        'qua_gio': sum(1 for p in phongs if p.calculated_status == 'qua_gio')
    }

    return render_template('so_do_phong.html', phongs=phongs, loai_phongs=loai_phongs, stats=stats)

# ========================= STATIC PAGES (FIX) =========================
@app.route('/quy-dinh')
@login_required
def quy_dinh():
    return render_template('simple.html', title='Quy Định Khách Sạn')

@app.route('/tro-giup')
@login_required
def tro_giup():
    return render_template('simple.html', title='Trợ Giúp')

@app.route('/gioi-thieu')
@login_required
def gioi_thieu():
    return render_template('simple.html', title='Giới Thiệu')

# ========================= ADMIN ROUTES =========================
# File: app.py

# REMOVED: Chức năng quản lý phòng đã bị xóa bỏ
# @app.route('/quan-li-phong') - đã xóa

@app.route('/nhan-vien', methods=['GET','POST'])
@login_required
@permission_required('staff.manage')
def nhan_vien():
    roles = Role.query.order_by(Role.name.asc()).all()
    default_role = next((r for r in roles if r.slug == 'nhanvien'), roles[0] if roles else None)

    if request.method == 'POST':
        ten_dn = request.form['ten_dn']
        if NguoiDung.query.filter_by(ten_dang_nhap=ten_dn).first():
            flash(f"Tên đăng nhập '{ten_dn}' đã tồn tại.", 'danger')
        else:
            role = None
            try:
                role_id = int(request.form.get('role_id') or 0)
            except (TypeError, ValueError):
                role_id = 0
            if role_id:
                role = Role.query.get(role_id)
            if not role:
                role = default_role
            if not role:
                flash('Chưa cấu hình vai trò. Vui lòng tạo vai trò trước.', 'danger')
                return redirect(url_for('quan_ly_vai_tro'))
            if role and role.slug == 'admin':
                first_admin = NguoiDung.query.filter(
                    or_(
                        NguoiDung.loai == 'admin',
                        NguoiDung.role.has(Role.slug == 'admin')
                    )
                ).order_by(NguoiDung.id.asc()).first()
                if first_admin and current_user.id != first_admin.id:
                    flash('Chỉ quản trị viên đầu tiên mới được tạo tài khoản quản trị.', 'warning')
                    return redirect(url_for('nhan_vien'))
            nv = NguoiDung(
                ten_dang_nhap=ten_dn,
                mat_khau=request.form['mat_khau'],
                ten=request.form['ten'],
                loai=role.slug if role else 'nhanvien',
                role_id=role.id if role else None,
                ngay_vao_lam=datetime.now().date()
            )
            db.session.add(nv)
            db.session.commit()
            flash('Đã thêm nhân viên mới!', 'success')
        return redirect(url_for('nhan_vien'))

    staffs = NguoiDung.query.order_by(NguoiDung.ten.asc()).all()
    bonus_amount = get_top_bonus()
    now = datetime.now()
    start_month = datetime(now.year, now.month, 1)
    next_month = (start_month + timedelta(days=32)).replace(day=1)

    salary_mode = get_salary_mode()

    base_stats = {
        'tong_hoa_don': 0,
        'tong_doanh_thu': 0,
        'tong_tien_phong': 0,
        'tong_tien_dv': 0,
        'tong_tien_phat': 0,
        'khach_phuc_vu': 0
    }
    stats_map = {s.id: base_stats.copy() for s in staffs}

    totals_rows = db.session.query(
        DatPhong.nhanvien_id.label('nv_id'),
        func.count(DatPhong.id).label('tong_hoa_don'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('tong_doanh_thu'),
        func.coalesce(func.sum(DatPhong.tien_phong), 0).label('tong_tien_phong'),
        func.coalesce(func.sum(DatPhong.tien_dv), 0).label('tong_tien_dv'),
        func.coalesce(func.sum(DatPhong.tien_phat), 0).label('tong_tien_phat'),
        func.count(func.distinct(DatPhong.khachhang_id)).label('khach_phuc_vu')
    ).filter(
        DatPhong.nhanvien_id.isnot(None),
        DatPhong.trang_thai == 'da_thanh_toan'
    ).group_by(DatPhong.nhanvien_id).all()

    for row in totals_rows:
        stats_map[row.nv_id] = {
            'tong_hoa_don': row.tong_hoa_don,
            'tong_doanh_thu': int(row.tong_doanh_thu or 0),
            'tong_tien_phong': int(row.tong_tien_phong or 0),
            'tong_tien_dv': int(row.tong_tien_dv or 0),
            'tong_tien_phat': int(row.tong_tien_phat or 0),
            'khach_phuc_vu': row.khach_phuc_vu
        }

    month_stats = {s.id: {'so_hd': 0, 'doanh_thu': 0} for s in staffs}
    month_rows = db.session.query(
        DatPhong.nhanvien_id.label('nv_id'),
        func.count(DatPhong.id).label('so_hd'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('doanh_thu')
    ).filter(
        DatPhong.nhanvien_id.isnot(None),
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).group_by(DatPhong.nhanvien_id).all()

    top_revenue = 0
    top_staff_ids = set()
    if month_rows:
        top_revenue = max(int(row.doanh_thu or 0) for row in month_rows)
        if top_revenue > 0:
            top_staff_ids = {row.nv_id for row in month_rows if int(row.doanh_thu or 0) == top_revenue}

    for row in month_rows:
        month_stats[row.nv_id] = {
            'so_hd': row.so_hd,
            'doanh_thu': int(row.doanh_thu or 0)
        }

    tiers = LuongThuongCauHinh.query.order_by(LuongThuongCauHinh.moc_duoi.asc()).all()
    top_staffs = [s for s in staffs if s.id in top_staff_ids]

    salary_records = {item.nguoidung_id: item for item in LuongNhanVien.query.all()}
    salary_preview = {}
    # Khởi tạo work_days_map trước khi sử dụng
    first_day = start_month.replace(hour=0, minute=0, second=0, microsecond=0)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
    work_days_map = {}
    for staff in staffs:
        work_days = db.session.query(func.count()).filter(
            Attendance.user_id == staff.id,
            Attendance.status == 'approved',
            Attendance.checkin_time >= first_day,
            Attendance.checkin_time <= last_day
        ).scalar() or 0
        work_days_map[staff.id] = int(work_days or 0)

    min_days = get_min_work_days()
    for staff in staffs:
        record = salary_records.get(staff.id)
        base_monthly = record.luong_co_ban if record else 0
        work_days = int(work_days_map.get(staff.id, 0) or 0)
        # Nếu chưa đủ ngày công thì phụ cấp = 0
        allowance = (record.phu_cap if record and work_days >= min_days else 0)
        doanh_thu = month_stats.get(staff.id, {}).get('doanh_thu', 0)
        thuong, rate = tinh_thuong_doanh_thu(doanh_thu, tiers)
        top_bonus = bonus_amount if (top_staff_ids and staff.id in top_staff_ids) else 0
        base_effective = compute_effective_base_salary(base_monthly, work_days, salary_mode)
        daily_rate = compute_daily_rate(base_monthly) if salary_mode == SALARY_MODE_DAILY else 0
        salary_preview[staff.id] = {
            'luong_co_ban': base_effective,
            'base_monthly': base_monthly,
            'salary_mode': salary_mode,
            'daily_rate': daily_rate,
            'daily_divisor': SALARY_DAILY_DIVISOR,
            'phu_cap': allowance,
            'thuong': thuong,
            'ty_le': rate,
            'top_bonus': top_bonus,
            'is_top': staff.id in top_staff_ids,
            'tong': base_effective + allowance + thuong + top_bonus,
            'doanh_thu': doanh_thu,
            'work_days': work_days,
            'min_days': min_days
        }

    top_staffs = [s for s in staffs if s.id in top_staff_ids]

    # Tính số ngày công cho từng nhân viên trong tháng hiện tại
    first_day = start_month.replace(hour=0, minute=0, second=0, microsecond=0)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
    work_days_map = {}
    for staff in staffs:
        work_days = db.session.query(func.count()).filter(
            Attendance.user_id == staff.id,
            Attendance.status == 'approved',
            Attendance.checkin_time >= first_day,
            Attendance.checkin_time <= last_day
        ).scalar() or 0
        work_days_map[staff.id] = int(work_days or 0)

    min_days = get_min_work_days()
    summary = {
        'total': len(staffs),
        'admins': sum(1 for s in staffs if s.role_slug == 'admin'),
        'employees': sum(1 for s in staffs if s.role_slug != 'admin'),
        'new_this_month': sum(1 for s in staffs if s.ngay_vao_lam and s.ngay_vao_lam >= start_month.date()),
        'handled_this_month': sum(month_stats.get(s.id, {}).get('so_hd', 0) for s in staffs)
    }

    return render_template(
        'nhan_vien.html',
        staffs=staffs,
        stats_map=stats_map,
        month_stats=month_stats,
        summary=summary,
        start_month=start_month,
        bonus_amount=bonus_amount,
        top_staffs=top_staffs,
        tiers=tiers,
        salary_preview=salary_preview,
        salary_records=salary_records,
        work_days_map=work_days_map,
        min_days=min_days,
        attendance_month_default=start_month.strftime('%Y-%m'),
        salary_mode=salary_mode,
        roles=roles
    )


def build_role_management_context(selected_id=None):
    roles = Role.query.order_by(Role.name.asc()).all()
    selected_role = None

    if roles:
        if selected_id:
            selected_role = next((role for role in roles if role.id == selected_id), None)
        if not selected_role:
            selected_role = roles[0]
            selected_id = selected_role.id

    selected_permissions = {rp.permission for rp in selected_role.permissions} if selected_role else set()
    permission_groups_ui = []
    for group_key, group_label, entries in PERMISSION_GROUPS:
        permissions = [key for key, _ in entries]
        total = len(permissions)
        selected_count = sum(1 for key in permissions if key in selected_permissions)
        permission_groups_ui.append({
            'key': group_key,
            'label': group_label,
            'entries': entries,
            'permissions': permissions,
            'total': total,
            'selected_count': selected_count,
            'all_selected': bool(total and selected_count == total),
            'partially_selected': bool(0 < selected_count < total),
        })
    assigned_counts = {role.id: NguoiDung.query.filter_by(role_id=role.id).count() for role in roles}
    system_role_ids = {role.id for role in roles if role.is_system}
    protected_role_ids = {role.id for role in roles if role.is_system or role.slug == 'admin'}

    return {
        'roles': roles,
        'selected_role': selected_role,
        'selected_id': selected_role.id if selected_role else None,
        'selected_permissions': selected_permissions,
        'assigned_counts': assigned_counts,
        'system_role_ids': system_role_ids,
        'protected_role_ids': protected_role_ids,
        'permission_groups': PERMISSION_GROUPS,
        'permission_groups_ui': permission_groups_ui,
        'permission_meta': PERMISSION_META,
    }


def process_role_action(action, form):
    try:
        if action == 'create':
            name = (form.get('name') or '').strip()
            description = (form.get('description') or '').strip()
            if not name:
                return {'success': False, 'message': 'Tên vai trò không được để trống.', 'category': 'danger'}
            existing = Role.query.filter(func.lower(Role.name) == name.lower()).first()
            if existing:
                return {
                    'success': False,
                    'message': 'Tên vai trò đã tồn tại.',
                    'category': 'danger',
                    'selected_id': existing.id
                }
            slug = generate_role_slug(name)
            role = Role(name=name, slug=slug, description=description, is_system=False)
            db.session.add(role)
            db.session.flush()
            permissions = set(form.getlist('permissions')) & ALL_PERMISSION_KEYS
            if permissions:
                set_role_permissions(role, permissions)
            db.session.commit()
            return {
                'success': True,
                'message': 'Đã tạo vai trò mới.',
                'category': 'success',
                'selected_id': role.id
            }

        elif action == 'update':
            try:
                role_id = int(form.get('role_id') or 0)
            except (TypeError, ValueError):
                role_id = 0
            role = Role.query.get(role_id)
            if not role:
                return {'success': False, 'message': 'Không tìm thấy vai trò cần cập nhật.', 'category': 'danger'}
            name = (form.get('name') or '').strip()
            description = (form.get('description') or '').strip()
            permissions = set(form.getlist('permissions')) & ALL_PERMISSION_KEYS
            if name and name.lower() != (role.name or '').lower():
                duplicate = Role.query.filter(func.lower(Role.name) == name.lower(), Role.id != role.id).first()
                if duplicate:
                    return {
                        'success': False,
                        'message': 'Tên vai trò đã được sử dụng.',
                        'category': 'danger',
                        'selected_id': role.id
                    }
                role.name = name
            role.description = description
            if role.is_system and role.slug == 'admin':
                set_role_permissions(role, ALL_PERMISSION_KEYS)
            else:
                set_role_permissions(role, permissions)
            db.session.commit()
            return {
                'success': True,
                'message': 'Đã cập nhật vai trò.',
                'category': 'success',
                'selected_id': role.id
            }

        elif action == 'delete':
            try:
                role_id = int(form.get('role_id') or 0)
            except (TypeError, ValueError):
                role_id = 0
            role = Role.query.get(role_id)
            if not role:
                return {'success': False, 'message': 'Không tìm thấy vai trò để xoá.', 'category': 'danger'}
            if role.slug == 'admin':
                return {
                    'success': False,
                    'message': 'Khong the xoa vai tro quan tri vien.',
                    'category': 'warning',
                    'selected_id': role.id
                }
            if role.is_system:
                return {
                    'success': False,
                    'message': 'Khong the xoa vai tro he thong.',
                    'category': 'warning',
                    'selected_id': role.id
                }
            if NguoiDung.query.filter_by(role_id=role.id).count() > 0:
                return {
                    'success': False,
                    'message': 'Không thể xoá vai trò đang được gán cho người dùng.',
                    'category': 'warning',
                    'selected_id': role.id
                }
            RolePermission.query.filter_by(role_id=role.id).delete()
            db.session.delete(role)
            db.session.commit()
            return {
                'success': True,
                'message': 'Đã xoá vai trò.',
                'category': 'success',
                'selected_id': None
            }

        else:
            return {'success': False, 'message': 'Yêu cầu không hợp lệ.', 'category': 'danger'}

    except Exception as exc:
        db.session.rollback()
        app.logger.exception('Không thể xử lý thao tác vai trò: %s', exc)
        return {
            'success': False,
            'message': 'Hệ thống gặp lỗi trong khi xử lý yêu cầu.',
            'category': 'danger'
        }


def wants_role_json_response():
    requested_with = request.headers.get('X-Requested-With', '')
    if requested_with.lower() == 'xmlhttprequest':
        return True
    accept_header = request.headers.get('Accept', '')
    return 'application/json' in accept_header.lower()


@app.route('/quan-ly-vai-tro', methods=['GET', 'POST'])
@login_required
@permission_required('roles.manage')
def quan_ly_vai_tro():
    selected_id = request.args.get('role_id', type=int)
    expects_json = wants_role_json_response()

    if request.method == 'POST':
        action = (request.form.get('action') or '').strip()
        result = process_role_action(action, request.form)
        context = build_role_management_context(result.get('selected_id') if result.get('selected_id') is not None else selected_id)

        if expects_json:
            status = 'success' if result.get('success') else 'error'
            status_code = 200 if result.get('success') else 400
            fragments = {
                'roleList': render_template('partials/role_list_items.html', **context),
                'roleDetail': render_template('partials/role_detail.html', **context),
            }
            return jsonify({
                'status': status,
                'message': result.get('message'),
                'fragments': fragments,
                'selectedRoleId': context.get('selected_id'),
            }), status_code

        flash(result.get('message'), result.get('category', 'info'))
        redirect_id = context.get('selected_id')
        if redirect_id:
            return redirect(url_for('quan_ly_vai_tro', role_id=redirect_id))
        return redirect(url_for('quan_ly_vai_tro'))

    context = build_role_management_context(selected_id)
    if expects_json:
        fragments = {
            'roleList': render_template('partials/role_list_items.html', **context),
            'roleDetail': render_template('partials/role_detail.html', **context),
        }
        return jsonify({
            'status': 'success',
            'fragments': fragments,
            'selectedRoleId': context.get('selected_id'),
        })

    return render_template('quan_ly_vai_tro.html', **context)


@app.route('/nhan-vien/export-cham-cong')
@login_required
@permission_required('staff.manage')
def export_attendance_overview():
    month_str = request.args.get('month', '').strip()
    try:
        if month_str:
            target_month = datetime.strptime(month_str, '%Y-%m')
        else:
            now = datetime.now()
            target_month = datetime(now.year, now.month, 1)
    except ValueError:
        now = datetime.now()
        target_month = datetime(now.year, now.month, 1)

    start_date = target_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_date = (start_date + timedelta(days=32)).replace(day=1)

    staffs = NguoiDung.query.order_by(NguoiDung.ten.asc()).all()
    if not staffs:
        flash('Chưa có dữ liệu nhân viên để xuất chấm công.', 'warning')
        return redirect(url_for('nhan_vien'))

    days = []
    cursor = start_date
    while cursor < end_date:
        days.append(cursor.date())
        cursor += timedelta(days=1)

    attendance_rows = Attendance.query.filter(
        Attendance.checkin_time >= start_date,
        Attendance.checkin_time < end_date
    ).all()

    status_priority = {'approved': 3, 'pending': 2, 'rejected': 1}
    status_symbols = {'approved': '✓', 'pending': '•', 'rejected': '✘'}
    status_fills = {
        'approved': ('16A34A', 'FFFFFF'),
        'pending': ('F59E0B', '1F2937'),
        'rejected': ('EF4444', 'FFFFFF')
    }

    attendance_map = {staff.id: {} for staff in staffs}
    for record in attendance_rows:
        day = record.checkin_time.date()
        status = record.status or 'pending'
        if status not in status_priority:
            continue
        user_map = attendance_map.setdefault(record.user_id, {})
        current_status = user_map.get(day)
        if not current_status or status_priority[status] > status_priority.get(current_status, 0):
            user_map[day] = status

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cham cong'

    total_columns = 3 + len(days) + 1

    title = f"Tổng hợp chấm công - Tháng {start_date.strftime('%m/%Y')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = GradientFill(stop=("16A34A", "2F7D5A"))

    subtitle = f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}  •  Người xuất: {current_user.ten}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(size=11, color='2F4F4F')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    subtitle_cell.fill = PatternFill('solid', fgColor='E8F5E9')

    headers = ["STT", "Họ tên", "Chức vụ"] + [day.strftime('%d') for day in days] + ["Ngày công"]
    header_row = 3
    header_fill = PatternFill('solid', fgColor='DCFCE7')
    header_font = Font(bold=True, color='14532D')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    for idx, staff in enumerate(staffs, start=1):
        row_index = header_row + idx
        ws.cell(row=row_index, column=1, value=idx).alignment = Alignment(horizontal='center')
        ws.cell(row=row_index, column=1).border = thin_border

        name_cell = ws.cell(row=row_index, column=2, value=staff.ten)
        name_cell.font = Font(bold=True, color='1F2937')
        name_cell.border = thin_border
        name_cell.alignment = Alignment(vertical='center')

        role_display = staff.role_name or ''
        role_cell = ws.cell(row=row_index, column=3, value=role_display)
        role_cell.alignment = Alignment(horizontal='center')
        role_cell.border = thin_border

        approvals = 0
        for day_idx, day in enumerate(days, start=1):
            status = attendance_map.get(staff.id, {}).get(day)
            col = 3 + day_idx
            cell = ws.cell(row=row_index, column=col)
            if status:
                symbol = status_symbols.get(status, '')
                cell.value = symbol
                fill_color, font_color = status_fills.get(status, ('E2E8F0', '1F2937'))
                cell.fill = PatternFill('solid', fgColor=fill_color)
                cell.font = Font(color=font_color, bold=True)
                if status == 'approved':
                    approvals += 1
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        total_cell = ws.cell(row=row_index, column=total_columns, value=approvals)
        total_cell.font = Font(bold=True, color='14532D')
        total_cell.alignment = Alignment(horizontal='center')
        total_cell.fill = PatternFill('solid', fgColor='F0FDF4')
        total_cell.border = thin_border

    legend_row = header_row + len(staffs) + 2
    legend_text = "Ký hiệu: ✓ Đã duyệt | • Chờ duyệt | ✘ Từ chối"
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=total_columns)
    legend_cell = ws.cell(row=legend_row, column=1, value=legend_text)
    legend_cell.font = Font(italic=True, color='374151')
    legend_cell.alignment = Alignment(horizontal='left', vertical='center')

    column_widths = {
        1: 6,
        2: 26,
        3: 16,
        total_columns: 12
    }
    for idx_day in range(len(days)):
        column_widths[4 + idx_day] = 4.5
    for col_idx in range(1, total_columns + 1):
        width = column_widths.get(col_idx, 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = ws.cell(row=4, column=4)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Chấm công_{start_date.strftime('%Y-%m')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/nhan-vien/<int:nhanvien_id>')
@login_required
@permission_required('staff.manage')
def nhan_vien_chi_tiet(nhanvien_id):
    nv = NguoiDung.query.get_or_404(nhanvien_id)

    totals = db.session.query(
        func.count(DatPhong.id).label('tong_hoa_don'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('tong_doanh_thu'),
        func.coalesce(func.sum(DatPhong.tien_phong), 0).label('tong_tien_phong'),
        func.coalesce(func.sum(DatPhong.tien_dv), 0).label('tong_tien_dv'),
        func.coalesce(func.sum(DatPhong.tien_phat), 0).label('tong_tien_phat'),
        func.count(func.distinct(DatPhong.khachhang_id)).label('khach_phuc_vu')
    ).filter(
        DatPhong.nhanvien_id == nv.id,
        DatPhong.trang_thai == 'da_thanh_toan'
    ).one()

    overall_stats = {
        'tong_hoa_don': totals.tong_hoa_don or 0,
        'tong_doanh_thu': int(totals.tong_doanh_thu or 0),
        'tong_tien_phong': int(totals.tong_tien_phong or 0),
        'tong_tien_dv': int(totals.tong_tien_dv or 0),
        'tong_tien_phat': int(totals.tong_tien_phat or 0),
        'khach_phuc_vu': totals.khach_phuc_vu or 0
    }

    now = datetime.now()
    start_month = datetime(now.year, now.month, 1)
    next_month = (start_month + timedelta(days=32)).replace(day=1)

    month_totals = db.session.query(
        func.count(DatPhong.id).label('tong_hoa_don'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('tong_doanh_thu'),
        func.coalesce(func.sum(DatPhong.tien_phong), 0).label('tong_tien_phong'),
        func.coalesce(func.sum(DatPhong.tien_dv), 0).label('tong_tien_dv'),
        func.coalesce(func.sum(DatPhong.tien_phat), 0).label('tong_tien_phat')
    ).filter(
        DatPhong.nhanvien_id == nv.id,
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).one()

    month_stats = {
        'tong_hoa_don': month_totals.tong_hoa_don or 0,
        'tong_doanh_thu': int(month_totals.tong_doanh_thu or 0),
        'tong_tien_phong': int(month_totals.tong_tien_phong or 0),
        'tong_tien_dv': int(month_totals.tong_tien_dv or 0),
        'tong_tien_phat': int(month_totals.tong_tien_phat or 0)
    }

    bonus_amount = get_top_bonus()

    top_rows = db.session.query(
        DatPhong.nhanvien_id.label('nv_id'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('doanh_thu')
    ).filter(
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).group_by(DatPhong.nhanvien_id).all()
    top_bonus = 0
    if top_rows:
        top_max = max(int(row.doanh_thu or 0) for row in top_rows)
        if top_max > 0 and any(int(row.nv_id) == nv.id and int(row.doanh_thu or 0) == top_max for row in top_rows):
            top_bonus = bonus_amount

    salary_mode = get_salary_mode()
    luong_record = LuongNhanVien.query.filter_by(nguoidung_id=nv.id).first()
    luong_co_ban = luong_record.luong_co_ban if luong_record else 0
    phu_cap = luong_record.phu_cap if luong_record else 0
    tiers = LuongThuongCauHinh.query.order_by(LuongThuongCauHinh.moc_duoi.asc()).all()
    thuong_thang, ty_le_thuong = tinh_thuong_doanh_thu(month_stats['tong_doanh_thu'], tiers)
    # Tính số ngày công cho nhân viên này trong tháng hiện tại
    first_day = start_month.replace(hour=0, minute=0, second=0, microsecond=0)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
    work_days = db.session.query(func.count()).filter(
        Attendance.user_id == nv.id,
        Attendance.status == 'approved',
        Attendance.checkin_time >= first_day,
        Attendance.checkin_time <= last_day
    ).scalar() or 0
    work_days = int(work_days or 0)
    min_days = get_min_work_days()
    # Nếu chưa đủ ngày công thì phụ cấp = 0
    phu_cap_display = phu_cap if work_days >= min_days else 0
    base_effective = compute_effective_base_salary(luong_co_ban, work_days, salary_mode)
    daily_rate = compute_daily_rate(luong_co_ban) if salary_mode == SALARY_MODE_DAILY else 0
    salary_info = {
        'luong_co_ban': base_effective,
        'base_monthly': luong_co_ban,
        'salary_mode': salary_mode,
        'daily_rate': daily_rate,
        'daily_divisor': SALARY_DAILY_DIVISOR,
        'phu_cap': phu_cap_display,
        'thuong_thang': thuong_thang,
        'ty_le': ty_le_thuong,
        'top_bonus': top_bonus,
        'is_top': bool(top_bonus),
        'tong': base_effective + phu_cap_display + thuong_thang + top_bonus,
        'doanh_thu': month_stats['tong_doanh_thu'],
        'work_days': work_days,
        'min_days': min_days
    }

    average_ticket = int(overall_stats['tong_doanh_thu'] / overall_stats['tong_hoa_don']) if overall_stats['tong_hoa_don'] else 0
    service_ratio = round((overall_stats['tong_tien_dv'] / overall_stats['tong_doanh_thu']) * 100, 1) if overall_stats['tong_doanh_thu'] else 0
    penalty_ratio = round((overall_stats['tong_tien_phat'] / overall_stats['tong_doanh_thu']) * 100, 1) if overall_stats['tong_doanh_thu'] else 0

    recent_bookings = (
        DatPhong.query.options(joinedload(DatPhong.khachhang), joinedload(DatPhong.phong))
        .filter(
            DatPhong.nhanvien_id == nv.id,
            DatPhong.trang_thai == 'da_thanh_toan'
        )
        .order_by(DatPhong.thuc_te_tra.desc())
        .limit(6)
        .all()
    )

    top_customers = db.session.query(
        KhachHang.ho_ten.label('ten'),
        func.count(DatPhong.id).label('so_lan'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('tong_chi')
    ).join(DatPhong, KhachHang.id == DatPhong.khachhang_id)
    top_customers = top_customers.filter(
        DatPhong.nhanvien_id == nv.id,
        DatPhong.trang_thai == 'da_thanh_toan'
    ).group_by(KhachHang.id).order_by(func.count(DatPhong.id).desc()).limit(5).all()

    def subtract_months(dt, months):
        year = dt.year
        month = dt.month - months
        while month <= 0:
            month += 12
            year -= 1
        return datetime(year, month, 1)

    trend_start = subtract_months(start_month, 5)
    trend_rows = db.session.query(
        func.extract('year', DatPhong.thuc_te_tra).label('year'),
        func.extract('month', DatPhong.thuc_te_tra).label('month'),
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('doanh_thu')
    ).filter(
        DatPhong.nhanvien_id == nv.id,
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= trend_start,
        DatPhong.thuc_te_tra < next_month
    ).group_by('year', 'month').order_by('year', 'month').all()

    trend_lookup = {(int(row.year), int(row.month)): int(row.doanh_thu or 0) for row in trend_rows}
    trend_months = []
    cursor = trend_start
    while cursor < next_month:
        trend_months.append((cursor.year, cursor.month))
        cursor = (cursor + timedelta(days=32)).replace(day=1)
    revenue_trend = [
        {
            'label': f"{month:02d}/{year}",
            'value': trend_lookup.get((year, month), 0)
        }
        for year, month in trend_months
    ]

    role_permissions = {rp.permission for rp in nv.role.permissions} if nv.role else set()
    personal_permissions = {up.permission for up in nv.personal_permissions}

    return render_template(
        'nhan_vien_chi_tiet.html',
        nv=nv,
        overall_stats=overall_stats,
        month_stats=month_stats,
        average_ticket=average_ticket,
        service_ratio=service_ratio,
        penalty_ratio=penalty_ratio,
        recent_bookings=recent_bookings,
        top_customers=top_customers,
        revenue_trend=revenue_trend,
        start_month=start_month,
        salary_info=salary_info,
        tiers=tiers,
        roles=Role.query.order_by(Role.name.asc()).all(),
        role_permissions=role_permissions,
        personal_permissions=personal_permissions,
        permission_groups=PERMISSION_GROUPS,
        permission_meta=PERMISSION_META
    )


@app.route('/nhan-vien/<int:nhanvien_id>/role', methods=['POST'])
@login_required
@permission_required('staff.manage')
def cap_nhat_vai_tro_nhan_vien(nhanvien_id):
    nv = NguoiDung.query.get_or_404(nhanvien_id)
    payload = request.get_json(silent=True)
    if payload is None:
        payload = request.form
    wants_json = (
        request.is_json
        or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or request.accept_mimetypes['application/json'] >= request.accept_mimetypes['text/html']
    )

    def respond_error(message, status=400):
        if wants_json:
            return jsonify({'status': 'error', 'message': message}), status
        flash(message, 'danger')
        return redirect(url_for('nhan_vien_chi_tiet', nhanvien_id=nv.id))

    role = None
    role_id = payload.get('role_id') or payload.get('role')
    role_slug = payload.get('role_slug') or payload.get('slug')
    if role_id:
        try:
            role = Role.query.get(int(role_id))
        except (TypeError, ValueError):
            role = None
    if not role and role_slug:
        slug_value = str(role_slug).strip().lower()
        if slug_value:
            role = Role.query.filter(func.lower(Role.slug) == slug_value).first()
    if not role:
        return respond_error('Khong tim thay vai tro yeu cau.', 404)

    current_is_admin = (nv.loai == 'admin') or (nv.role and nv.role.slug == 'admin')
    target_is_admin = role.slug == 'admin'
    admin_filter = or_(
        NguoiDung.loai == 'admin',
        NguoiDung.role.has(Role.slug == 'admin')
    )
    first_admin = NguoiDung.query.filter(admin_filter).order_by(NguoiDung.id.asc()).first()

    if target_is_admin and first_admin and current_user.id != first_admin.id:
        return respond_error('Chi quan tri vien dau tien moi co the gan vai tro quan tri.', 403)

    if current_is_admin and not target_is_admin:
        if first_admin and nv.id == first_admin.id:
            return respond_error('Khong the thay doi vai tro cua quan tri vien ban dau.', 403)
        if first_admin and current_user.id not in (first_admin.id, nv.id):
            return respond_error('Chi quan tri vien dau tien hoac chinh tai khoan do moi duoc thay doi vai tro nay.', 403)
        remaining_admins = NguoiDung.query.filter(
            NguoiDung.id != nv.id,
            admin_filter
        ).count()
        if remaining_admins == 0:
            return respond_error('Khong the chuyen vai tro cua quan tri vien cuoi cung.', 403)

    changed = False
    if nv.role_id != role.id:
        nv.role_id = role.id
        changed = True
    if nv.loai != role.slug:
        nv.loai = role.slug
        changed = True

    if not changed:
        response_data = {
            'status': 'success',
            'message': 'Vai tro khong thay doi.',
            'role': {
                'id': role.id,
                'name': role.name,
                'slug': role.slug
            }
        }
        if wants_json:
            return jsonify(response_data)
        flash(response_data['message'], 'info')
        return redirect(url_for('nhan_vien_chi_tiet', nhanvien_id=nv.id))

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return respond_error('Khong the cap nhat vai tro. Vui long thu lai sau.', 500)

    response_payload = {
        'status': 'success',
        'message': 'Da cap nhat vai tro.',
        'role': {
            'id': role.id,
            'name': role.name,
            'slug': role.slug
        }
    }
    if wants_json:
        return jsonify(response_payload)
    flash(response_payload['message'], 'success')
    return redirect(url_for('nhan_vien_chi_tiet', nhanvien_id=nv.id))


@app.route('/nhan-vien/<int:nhanvien_id>/permissions', methods=['POST'])
@login_required
@permission_required('staff.manage')
def cap_nhat_quyen_ca_nhan(nhanvien_id):
    nv = NguoiDung.query.get_or_404(nhanvien_id)
    payload = request.get_json(silent=True)
    if payload is None:
        payload = request.form
    permissions = payload.get('permissions', [])
    if isinstance(permissions, str):
        permissions = [permissions]
    if not isinstance(permissions, (list, tuple, set)):
        return jsonify({'status': 'error', 'message': 'Dữ liệu không hợp lệ.'}), 400
    desired = [perm for perm in permissions if isinstance(perm, str)]
    set_user_permissions(nv, desired)
    db.session.commit()
    personal_permissions = [up.permission for up in nv.personal_permissions]
    role_permissions = [rp.permission for rp in nv.role.permissions] if nv.role else []
    return jsonify({
        'status': 'success',
        'message': 'Đã cập nhật quyền cá nhân.',
        'personalPermissions': personal_permissions,
        'rolePermissions': role_permissions
    })

@app.route('/nhan-vien/xoa/<int:nhanvien_id>', methods=['POST'])
@login_required
@permission_required('staff.manage')
def xoa_nhan_vien(nhanvien_id):
    if current_user.id == nhanvien_id:
        flash('Bạn không thể tự xoá tài khoản của mình.', 'warning')
        return redirect(url_for('nhan_vien'))

    nv = NguoiDung.query.get_or_404(nhanvien_id)

    if nv.role_slug == 'admin':
        first_admin = NguoiDung.query.filter(
            or_(
                NguoiDung.loai == 'admin',
                NguoiDung.role.has(Role.slug == 'admin')
            )
        ).order_by(NguoiDung.id.asc()).first()

        if first_admin:
            if nv.id == first_admin.id:
                flash('Không thể xoá quản trị viên ban đầu.', 'warning')
                return redirect(url_for('nhan_vien'))

            if current_user.id != first_admin.id:
                flash('Chỉ quản trị viên đầu tiên mới có quyền xoá các quản trị viên khác.', 'warning')
                return redirect(url_for('nhan_vien'))

        remaining_admins = NguoiDung.query.filter(
            NguoiDung.id != nv.id,
            or_(
                NguoiDung.loai == 'admin',
                NguoiDung.role.has(Role.slug == 'admin')
            )
        ).count()
        if remaining_admins == 0:
            flash('Không thể xoá quản trị viên cuối cùng.', 'warning')
            return redirect(url_for('nhan_vien'))

    DatPhong.query.filter_by(nhanvien_id=nv.id).update({'nhanvien_id': None})
    TinNhan.query.filter_by(nguoidung_id=nv.id).update({'nguoidung_id': None})
    Attendance.query.filter_by(user_id=nv.id).delete()
    Attendance.query.filter_by(approved_by=nv.id).update({'approved_by': None})
    EmailLog.query.filter_by(sent_by=nv.id).update({'sent_by': None})
    LuongNhanVien.query.filter_by(nguoidung_id=nv.id).delete()

    try:
        db.session.delete(nv)
        db.session.commit()
        flash('Đã xoá nhân viên.', 'success')
    except Exception:
        db.session.rollback()
        flash('Không thể xoá nhân viên. Vui lòng thử lại.', 'danger')

    return redirect(url_for('nhan_vien'))

@app.route('/cai-dat-luong-thuong', methods=['GET', 'POST'])
@login_required
@permission_required('payroll.configure')
def cai_dat_luong_thuong():
    selected_id = request.args.get('staff_id', type=int)
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == 'POST':
        form_name = request.form.get('form_name')
        status = 'success'
        message = ''
        http_status = 200

        if form_name == 'save_salary':
            try:
                nguoidung_id = int(request.form['nguoidung_id'])
                luong_co_ban = int(request.form.get('luong_co_ban', 0) or 0)
                phu_cap = int(request.form.get('phu_cap', 0) or 0)
                record = LuongNhanVien.query.filter_by(nguoidung_id=nguoidung_id).first()
                if not record:
                    record = LuongNhanVien(nguoidung_id=nguoidung_id)
                    db.session.add(record)
                record.luong_co_ban = max(0, luong_co_ban)
                record.phu_cap = max(0, phu_cap)
                db.session.commit()
                message = 'Đã cập nhật lương cho nhân viên.'
            except Exception:
                db.session.rollback()
                status = 'danger'
                http_status = 400
                message = 'Không thể cập nhật lương cho nhân viên.'
        elif form_name == 'set_min_work_days':
            try:
                min_days = int(request.form.get('min_work_days', 0) or 0)
                set_min_work_days(min_days)
                message = 'Đã cập nhật số ngày công tối thiểu nhận phụ cấp.'
            except Exception:
                status = 'danger'
                http_status = 400
                message = 'Cập nhật số ngày công thất bại.'
        elif form_name == 'save_tier':
            try:
                tier_id = request.form.get('tier_id')
                moc_duoi = int(request.form.get('moc_duoi', 0) or 0)
                moc_tren_raw = request.form.get('moc_tren')
                moc_tren = int(moc_tren_raw) if moc_tren_raw else None
                ty_le = float(request.form.get('ty_le', 0) or 0)
                ghi_chu = request.form.get('ghi_chu', '')
            except (TypeError, ValueError):
                status = 'danger'
                http_status = 400
                message = 'Dữ liệu cấu hình không hợp lệ.'
            else:
                try:
                    if tier_id:
                        tier = LuongThuongCauHinh.query.get(int(tier_id))
                        if tier:
                            tier.moc_duoi = moc_duoi
                            tier.moc_tren = moc_tren
                            tier.ty_le = ty_le
                            tier.ghi_chu = ghi_chu
                    else:
                        tier = LuongThuongCauHinh(moc_duoi=moc_duoi, moc_tren=moc_tren, ty_le=ty_le, ghi_chu=ghi_chu)
                        db.session.add(tier)
                    db.session.commit()
                    message = 'Đã lưu mức thưởng.'
                except Exception:
                    db.session.rollback()
                    status = 'danger'
                    http_status = 400
                    message = 'Không thể lưu mức thưởng.'
        elif form_name == 'delete_tier':
            try:
                tier_id = int(request.form['tier_id'])
                tier = LuongThuongCauHinh.query.get_or_404(tier_id)
                db.session.delete(tier)
                db.session.commit()
                message = 'Đã xóa mức thưởng.'
            except Exception:
                db.session.rollback()
                status = 'danger'
                http_status = 400
                message = 'Không thể xóa mức thưởng.'
        elif form_name == 'save_top_bonus':
            try:
                bonus_value = max(0, int(request.form.get('top_bonus', 0) or 0))
                set_config_int('TOP_REVENUE_BONUS', bonus_value)
                cache.delete_memoized(get_top_bonus)
                message = 'Đã cập nhật thưởng top doanh thu.'
            except ValueError:
                status = 'danger'
                http_status = 400
                message = 'Giá trị không hợp lệ.'
        elif form_name == 'set_salary_mode':
            try:
                set_salary_mode(request.form.get('salary_mode'))
                message = 'Đã cập nhật chế độ tính lương.'
            except ValueError:
                status = 'danger'
                http_status = 400
                message = 'Chế độ tính lương không hợp lệ.'
        elif form_name == 'save_auto_cancel':
            try:
                minutes = int(request.form.get('auto_cancel_minutes', 5) or 5)
                if minutes < 1:
                    minutes = 5
                set_config_int('auto_cancel_minutes', minutes)
                message = 'Đã cập nhật thời gian tự động hủy đặt phòng.'
            except ValueError:
                status = 'danger'
                http_status = 400
                message = 'Giá trị không hợp lệ.'
        else:
            status = 'warning'
            http_status = 400
            message = 'Yêu cầu không hợp lệ.'

        if wants_json:
            context = build_salary_settings_context(selected_id)
            payload = {
                'message': message,
                'status': status,
                'stats_html': render_template('partials/salary_stats_grid.html', **context),
                'employees_html': render_template('partials/salary_employee_list.html', **context),
                'tiers_html': render_template('partials/salary_tier_list.html', **context),
            }
            return jsonify(payload), http_status

        flash(message, status)
        redirect_kwargs = {}
        if selected_id:
            redirect_kwargs['staff_id'] = selected_id
        return redirect(url_for('cai_dat_luong_thuong', **redirect_kwargs))

    context = build_salary_settings_context(selected_id)
    return render_template('cai_dat_luong_thuong.html', **context)


@app.route('/cai-dat-email', methods=['GET', 'POST'])
@login_required
@permission_required('email.settings')
def cai_dat_email():
    ensure_email_templates()
    if request.method == 'POST':
        form_name = request.form.get('form_name') or request.form.get('action')
        if form_name == 'save_smtp':
            smtp_host = request.form.get('smtp_host', '').strip()
            smtp_port = request.form.get('smtp_port', '').strip() or '587'
            smtp_username = request.form.get('smtp_username', '').strip()
            smtp_password = request.form.get('smtp_password', '')
            sender_email = request.form.get('sender_email', '').strip()
            use_tls = '1' if request.form.get('smtp_use_tls') else '0'
            use_ssl = '1' if request.form.get('smtp_use_ssl') else '0'

            set_config_values({
                'SMTP_HOST': smtp_host,
                'SMTP_PORT': smtp_port,
                'SMTP_USERNAME': smtp_username,
                'SMTP_PASSWORD': smtp_password,
                'SMTP_SENDER_EMAIL': sender_email,
                'SMTP_USE_TLS': use_tls,
                'SMTP_USE_SSL': use_ssl
            })
            flash('Đã cập nhật cấu hình SMTP.', 'success')
        elif form_name == 'save_template':
            template_key = request.form.get('template_key')
            subject = request.form.get('template_subject', '').strip()
            body = request.form.get('template_body', '').strip()
            tpl = EmailTemplate.query.filter_by(key=template_key).first()
            if not tpl:
                flash('Không tìm thấy mẫu email yêu cầu.', 'danger')
            elif not subject or not body:
                flash('Vui lòng nhập đầy đủ tiêu đề và nội dung mẫu.', 'warning')
            else:
                tpl.subject = subject
                tpl.body = body
                db.session.commit()
                flash('Đã lưu mẫu email.', 'success')
        elif form_name == 'reset_template':
            template_key = request.form.get('template_key')
            defaults = EMAIL_TEMPLATE_DEFAULTS.get(template_key)
            tpl = EmailTemplate.query.filter_by(key=template_key).first()
            if not defaults or not tpl:
                flash('Không thể khôi phục mẫu email.', 'danger')
            else:
                tpl.subject = defaults['subject']
                tpl.body = defaults['body']
                db.session.commit()
                flash('Đã khôi phục mẫu email về mặc định.', 'success')
        else:
            flash('Yêu cầu không hợp lệ.', 'warning')
        return redirect(url_for('cai_dat_email'))

    email_settings = get_email_settings()
    template_rows = {tpl.key: tpl for tpl in EmailTemplate.query.all()}
    templates = []
    for key, meta in EMAIL_TEMPLATE_DEFAULTS.items():
        tpl = template_rows.get(key)
        templates.append({
            'key': key,
            'name': meta['name'],
            'subject': tpl.subject if tpl else meta['subject'],
            'body': tpl.body if tpl else meta['body']
        })

    return render_template(
        'cai_dat_email.html',
        email_settings=email_settings,
        templates=templates
    )


@app.route('/lich-su-email')
@login_required
@permission_required('email.logs')
def lich_su_email():
    """Trang xem lịch sử gửi email"""
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Lọc theo trạng thái
    status_filter = request.args.get('status', '')
    query = EmailLog.query
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    # Sắp xếp theo thời gian mới nhất
    query = query.order_by(EmailLog.sent_at.desc())
    
    # Phân trang
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    logs = pagination.items
    
    # Thống kê
    total_sent = EmailLog.query.count()
    total_success = EmailLog.query.filter_by(status='success').count()
    total_failed = EmailLog.query.filter_by(status='failed').count()
    total_pending = EmailLog.query.filter_by(status='pending').count()
    
    return render_template(
        'lich_su_email.html',
        logs=logs,
        pagination=pagination,
        status_filter=status_filter,
        total_sent=total_sent,
        total_success=total_success,
        total_failed=total_failed,
        total_pending=total_pending
    )


@app.route('/chi-tiet-email/<int:log_id>')
@login_required
@permission_required('email.logs')
def chi_tiet_email(log_id):
    """Xem chi tiết một email log"""
    log = EmailLog.query.get_or_404(log_id)
    return render_template('chi_tiet_email.html', log=log)


@app.route('/thong-ke-doanh-thu')
@login_required
@permission_required('analytics.revenue')
def thong_ke_doanh_thu():
    now = datetime.now()
    view_type = request.args.get('view', 'month')  # month, quarter, year, custom
    month = int(request.args.get('thang', now.month))
    year = int(request.args.get('nam', now.year))
    quarter = int(request.args.get('quy', 1))
    
    # === LẤY DỮ LIỆU THEO LOẠI VIEW ===
    if view_type == 'year':
        # Thống kê cả năm
        ds_doanh_thu = DatPhong.query.filter(
            db.or_(DatPhong.trang_thai == 'da_thanh_toan', DatPhong.trang_thai == 'huy'),
            extract('year', DatPhong.thuc_te_tra) == year
        ).order_by(DatPhong.thuc_te_tra.asc()).all()
        
        # Doanh thu theo tháng
        revenue_by_month = defaultdict(float)
        room_revenue_by_month = defaultdict(float)
        service_revenue_by_month = defaultdict(float)
        for d in ds_doanh_thu:
            if d.thuc_te_tra:
                m = d.thuc_te_tra.month
                revenue_by_month[m] += d.tong_thanh_toan
                room_revenue_by_month[m] += d.tien_phong
                service_revenue_by_month[m] += d.tien_dv
        
        chart_labels = [f"T{m}" for m in range(1, 13)]
        chart_data = [revenue_by_month.get(m, 0) for m in range(1, 13)]
        chart_room_data = [room_revenue_by_month.get(m, 0) for m in range(1, 13)]
        chart_service_data = [service_revenue_by_month.get(m, 0) for m in range(1, 13)]
        
    elif view_type == 'quarter':
        # Thống kê theo quý
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        ds_doanh_thu = DatPhong.query.filter(
            db.or_(DatPhong.trang_thai == 'da_thanh_toan', DatPhong.trang_thai == 'huy'),
            extract('year', DatPhong.thuc_te_tra) == year,
            extract('month', DatPhong.thuc_te_tra).between(start_month, end_month)
        ).order_by(DatPhong.thuc_te_tra.asc()).all()
        
        revenue_by_month = defaultdict(float)
        room_revenue_by_month = defaultdict(float)
        service_revenue_by_month = defaultdict(float)
        for d in ds_doanh_thu:
            if d.thuc_te_tra:
                m = d.thuc_te_tra.month
                revenue_by_month[m] += d.tong_thanh_toan
                room_revenue_by_month[m] += d.tien_phong
                service_revenue_by_month[m] += d.tien_dv
        
        chart_labels = [f"Tháng {m}" for m in range(start_month, end_month + 1)]
        chart_data = [revenue_by_month.get(m, 0) for m in range(start_month, end_month + 1)]
        chart_room_data = [room_revenue_by_month.get(m, 0) for m in range(start_month, end_month + 1)]
        chart_service_data = [service_revenue_by_month.get(m, 0) for m in range(start_month, end_month + 1)]
        
    else:  # month (default)
        ds_doanh_thu = DatPhong.query.filter(
            db.or_(DatPhong.trang_thai == 'da_thanh_toan', DatPhong.trang_thai == 'huy'),
            extract('year', DatPhong.thuc_te_tra) == year,
            extract('month', DatPhong.thuc_te_tra) == month
        ).order_by(DatPhong.thuc_te_tra.asc()).all()
        
        days_in_month = calendar.monthrange(year, month)[1]
        chart_labels = [f"{d:02d}" for d in range(1, days_in_month + 1)]
        revenue_by_day = defaultdict(float)
        room_revenue_by_day = defaultdict(float)
        service_revenue_by_day = defaultdict(float)
        for d in ds_doanh_thu:
            if d.thuc_te_tra:
                day = d.thuc_te_tra.day
                revenue_by_day[day] += d.tong_thanh_toan
                room_revenue_by_day[day] += d.tien_phong
                service_revenue_by_day[day] += d.tien_dv
        
        chart_data = [revenue_by_day.get(d, 0) for d in range(1, days_in_month + 1)]
        chart_room_data = [room_revenue_by_day.get(d, 0) for d in range(1, days_in_month + 1)]
        chart_service_data = [service_revenue_by_day.get(d, 0) for d in range(1, days_in_month + 1)]
    
    # === TÍNH TỔNG ===
    tong_cong = {
        'phong': sum(d.tien_phong for d in ds_doanh_thu),
        'dv': sum(d.tien_dv for d in ds_doanh_thu),
        'phat': sum(d.tien_phat for d in ds_doanh_thu),
        'coc': sum(d.tien_coc for d in ds_doanh_thu),
        'tong': sum(d.tong_thanh_toan for d in ds_doanh_thu),
        'so_booking': len(ds_doanh_thu)
    }
    
    # === SO SÁNH VỚI KỲ TRƯỚC ===
    if view_type == 'month':
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        prev_data = DatPhong.query.filter(
            db.or_(DatPhong.trang_thai == 'da_thanh_toan', DatPhong.trang_thai == 'huy'),
            extract('year', DatPhong.thuc_te_tra) == prev_year,
            extract('month', DatPhong.thuc_te_tra) == prev_month
        ).all()
    else:
        prev_year = year - 1
        prev_data = DatPhong.query.filter(
            db.or_(DatPhong.trang_thai == 'da_thanh_toan', DatPhong.trang_thai == 'huy'),
            extract('year', DatPhong.thuc_te_tra) == prev_year
        ).all()
    
    prev_total = sum(d.tong_thanh_toan for d in prev_data)
    growth_rate = ((tong_cong['tong'] - prev_total) / prev_total * 100) if prev_total > 0 else 0
    
    # === TOP KHÁCH HÀNG ===
    customer_revenue = defaultdict(lambda: {'total': 0, 'count': 0, 'name': '', 'id': 0})
    for d in ds_doanh_thu:
        customer_revenue[d.khachhang_id]['total'] += d.tong_thanh_toan
        customer_revenue[d.khachhang_id]['count'] += 1
        customer_revenue[d.khachhang_id]['name'] = d.khachhang.ho_ten
        customer_revenue[d.khachhang_id]['id'] = d.khachhang_id
    
    top_customers = sorted(customer_revenue.values(), key=lambda x: x['total'], reverse=True)[:5]
    
    # === TOP PHÒNG ===
    room_revenue = defaultdict(lambda: {'total': 0, 'count': 0, 'name': '', 'type': ''})
    for d in ds_doanh_thu:
        room_revenue[d.phong_id]['total'] += d.tong_thanh_toan
        room_revenue[d.phong_id]['count'] += 1
        room_revenue[d.phong_id]['name'] = d.phong.ten
        room_revenue[d.phong_id]['type'] = d.phong.loai.ten
    
    top_rooms = sorted(room_revenue.values(), key=lambda x: x['total'], reverse=True)[:5]
    
    # === DOANH THU THEO LOẠI PHÒNG ===
    room_type_revenue = defaultdict(float)
    for d in ds_doanh_thu:
        room_type_revenue[d.phong.loai.ten] += d.tong_thanh_toan
    
    pie_labels = list(room_type_revenue.keys())
    pie_data = list(room_type_revenue.values())
    
    # === TRUNG BÌNH ===
    avg_revenue_per_booking = tong_cong['tong'] / tong_cong['so_booking'] if tong_cong['so_booking'] > 0 else 0
    
    return render_template('thong_ke_doanh_thu.html',
        ds_doanh_thu=ds_doanh_thu,
        tong_cong=tong_cong,
        current_month=month,
        current_year=year,
        current_quarter=quarter,
        view_type=view_type,
        chart_labels=chart_labels,
        chart_data=chart_data,
        chart_room_data=chart_room_data,
        chart_service_data=chart_service_data,
        pie_labels=pie_labels,
        pie_data=pie_data,
        growth_rate=growth_rate,
        prev_total=prev_total,
        top_customers=top_customers,
        top_rooms=top_rooms,
        avg_revenue_per_booking=avg_revenue_per_booking
    )

@app.route('/quan-li-hoa-don')
@login_required
@permission_required('payments.invoices')
def quan_li_hoa_don():
    query = DatPhong.query.filter(
        db.or_(
            DatPhong.trang_thai == 'da_thanh_toan',
            db.and_(DatPhong.trang_thai == 'huy', DatPhong.tong_thanh_toan > 0)
        )
    )
    search_kh, search_phong = request.args.get('khach_hang', ''), request.args.get('phong', '')
    tu_ngay_str, den_ngay_str = request.args.get('tu_ngay', ''), request.args.get('den_ngay', '')
    if search_kh:
        query = query.join(KhachHang).filter(db.or_(KhachHang.ho_ten.ilike(f'%{search_kh}%'), KhachHang.cmnd.ilike(f'%{search_kh}%')))
    if search_phong:
        query = query.join(Phong).filter(Phong.ten.ilike(f'%{search_phong}%'))
    if tu_ngay_str:
        query = query.filter(DatPhong.thuc_te_tra >= datetime.strptime(tu_ngay_str, '%Y-%m-%d'))
    if den_ngay_str:
        query = query.filter(DatPhong.thuc_te_tra <= datetime.strptime(den_ngay_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    ds_hoa_don = query.order_by(DatPhong.thuc_te_tra.desc()).all()
    return render_template('quan_li_hoa_don.html', ds_hoa_don=ds_hoa_don, search_values=request.args)

@app.route('/xuat-excel-hoa-don')
@login_required
@permission_required('payments.export')
def xuat_excel_hoa_don():
    from io import BytesIO
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference, Series
    
    # Lấy dữ liệu giống như quan_li_hoa_don
    query = DatPhong.query.filter(
        db.or_(
            DatPhong.trang_thai == 'da_thanh_toan',
            db.and_(DatPhong.trang_thai == 'huy', DatPhong.tong_thanh_toan > 0)
        )
    )
    search_kh, search_phong = request.args.get('khach_hang', ''), request.args.get('phong', '')
    tu_ngay_str, den_ngay_str = request.args.get('tu_ngay', ''), request.args.get('den_ngay', '')
    if search_kh:
        query = query.join(KhachHang).filter(db.or_(KhachHang.ho_ten.ilike(f'%{search_kh}%'), KhachHang.cmnd.ilike(f'%{search_kh}%')))
    if search_phong:
        query = query.join(Phong).filter(Phong.ten.ilike(f'%{search_phong}%'))
    if tu_ngay_str:
        query = query.filter(DatPhong.thuc_te_tra >= datetime.strptime(tu_ngay_str, '%Y-%m-%d'))
    if den_ngay_str:
        query = query.filter(DatPhong.thuc_te_tra <= datetime.strptime(den_ngay_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    ds_hoa_don = query.order_by(DatPhong.thuc_te_tra.desc()).all()
    
    # Tạo dữ liệu cho Excel
    data = []
    for d in ds_hoa_don:
        trang_thai = 'Đã thanh toán' if d.trang_thai == 'da_thanh_toan' else 'Hủy (mất cọc)'
        data.append({
            'Mã HĐ': d.id,
            'Khách hàng': f"{d.khachhang.ho_ten} ({d.khachhang.cmnd})",
            'Phòng': d.phong.ten,
            'Ngày hoàn tất': d.thuc_te_tra.strftime('%d/%m/%Y %H:%M') if d.thuc_te_tra else '',
            'Trạng thái': trang_thai,
            'Tiền phòng (VNĐ)': d.tien_phong,
            'Tiền dịch vụ (VNĐ)': d.tien_dv,
            'Phí phạt (VNĐ)': d.tien_phat,
            'Tổng cộng (VNĐ)': d.tong_thanh_toan
        })
    
    # Tạo DataFrame
    df = pd.DataFrame(data)
    
    # Tính toán thống kê cho biểu đồ
    tong_da_thanh_toan = len([d for d in ds_hoa_don if d.trang_thai == 'da_thanh_toan'])
    tong_huy_mat_coc = len([d for d in ds_hoa_don if d.trang_thai == 'huy'])
    
    tong_tien_phong = sum(d.tien_phong for d in ds_hoa_don)
    tong_tien_dv = sum(d.tien_dv for d in ds_hoa_don)
    tong_tien_phat = sum(d.tien_phat for d in ds_hoa_don)
    tong_doanh_thu = sum(d.tong_thanh_toan for d in ds_hoa_don)
    
    # Tạo file Excel trong memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet dữ liệu chính
        df.to_excel(writer, sheet_name='Danh sách hóa đơn', index=False, startrow=4)
        worksheet = writer.sheets['Danh sách hóa đơn']
        
        # Thêm tiêu đề và thông tin
        now = datetime.now()
        worksheet['A1'] = f'DANH SÁCH HÓA ĐƠN CHI TIẾT'
        worksheet['A2'] = f'Khách sạn PTIT - Báo cáo hóa đơn'
        worksheet['A3'] = f'Ngày xuất báo cáo: {now.strftime("%d/%m/%Y %H:%M:%S")}'
        worksheet['A4'] = f'Tổng số hóa đơn: {len(ds_hoa_don)}'
        
        # Styling tiêu đề
        title_font = Font(size=16, bold=True, color="2F7D5A")
        subtitle_font = Font(size=12, italic=True, color="666666")
        header_font = Font(bold=True, color="FFFFFF")
        
        worksheet['A1'].font = title_font
        worksheet['A2'].font = subtitle_font
        worksheet['A3'].font = Font(size=10, color="888888")
        worksheet['A4'].font = Font(size=10, color="666666")
        
        # Merge cells cho tiêu đề
        worksheet.merge_cells('A1:I1')
        worksheet.merge_cells('A2:I2')
        worksheet.merge_cells('A3:I3')
        worksheet.merge_cells('A4:I4')
        
        # Căn chỉnh tiêu đề
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A2'].alignment = Alignment(horizontal='center')
        worksheet['A3'].alignment = Alignment(horizontal='right')
        worksheet['A4'].alignment = Alignment(horizontal='right')
        
        # Styling header của bảng
        header_fill = PatternFill(start_color="2F7D5A", end_color="2F7D5A", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num in range(1, 10):  # A đến I
            cell = worksheet.cell(row=5, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Styling dữ liệu
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        currency_alignment = Alignment(horizontal='right')
        center_alignment = Alignment(horizontal='center')
        left_alignment = Alignment(horizontal='left')
        
        for row_num in range(6, len(df) + 6):  # Dữ liệu từ row 6
            # Cột Mã HĐ (A) - căn giữa
            worksheet.cell(row=row_num, column=1).alignment = center_alignment
            worksheet.cell(row=row_num, column=1).border = data_border
            
            # Cột Khách hàng (B) - căn trái
            worksheet.cell(row=row_num, column=2).alignment = left_alignment
            worksheet.cell(row=row_num, column=2).border = data_border
            
            # Cột Phòng (C) - căn giữa
            worksheet.cell(row=row_num, column=3).alignment = center_alignment
            worksheet.cell(row=row_num, column=3).border = data_border
            
            # Cột Ngày hoàn tất (D) - căn giữa
            worksheet.cell(row=row_num, column=4).alignment = center_alignment
            worksheet.cell(row=row_num, column=4).border = data_border
            
            # Cột Trạng thái (E) - căn giữa
            worksheet.cell(row=row_num, column=5).alignment = center_alignment
            worksheet.cell(row=row_num, column=5).border = data_border
            
            # Các cột tiền tệ (F-I) - căn phải, format số
            for col_num in range(6, 10):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = currency_alignment
                cell.border = data_border
                if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.number_format = '#,##0'
        
        # Căn chỉnh tự động độ rộng cột
        for col_num, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Điều chỉnh độ rộng tối ưu
            if col_num == 1:  # Mã HĐ
                adjusted_width = max(max_length + 2, 8)
            elif col_num == 2:  # Khách hàng
                adjusted_width = max(max_length + 2, 25)
            elif col_num == 3:  # Phòng
                adjusted_width = max(max_length + 2, 10)
            elif col_num == 4:  # Ngày
                adjusted_width = max(max_length + 2, 18)
            elif col_num == 5:  # Trạng thái
                adjusted_width = max(max_length + 2, 15)
            else:  # Các cột tiền
                adjusted_width = max(max_length + 4, 18)
            
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 35)
        
        # Tạo sheet biểu đồ
        chart_sheet = writer.book.create_sheet('Biểu đồ thống kê')
        
        # Biểu đồ tròn cho trạng thái hóa đơn
        pie_chart = PieChart()
        pie_chart.title = "Tỷ lệ trạng thái hóa đơn"
        pie_chart.height = 10
        pie_chart.width = 15
        
        # Dữ liệu cho biểu đồ tròn
        pie_data = [
            ['Trạng thái', 'Số lượng'],
            ['Đã thanh toán', tong_da_thanh_toan],
            ['Hủy (mất cọc)', tong_huy_mat_coc]
        ]
        
        # Ghi dữ liệu cho biểu đồ
        for row_num, row_data in enumerate(pie_data, 1):
            for col_num, value in enumerate(row_data, 1):
                chart_sheet.cell(row=row_num, column=col_num, value=value)
        
        pie_cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=3)
        pie_vals = Reference(chart_sheet, min_col=2, min_row=2, max_row=3)
        pie_chart.add_data(pie_vals, titles_from_data=True)
        pie_chart.set_categories(pie_cats)
        pie_chart.style = 10
        
        chart_sheet.add_chart(pie_chart, "A1")
        
        # Biểu đồ cột cho phân loại doanh thu
        bar_chart = BarChart()
        bar_chart.title = "Phân loại doanh thu"
        bar_chart.y_axis.title = 'Doanh thu (VNĐ)'
        bar_chart.x_axis.title = 'Loại thu nhập'
        bar_chart.height = 10
        bar_chart.width = 15
        
        # Dữ liệu cho biểu đồ cột
        bar_data = [
            ['Loại thu nhập', 'Số tiền'],
            ['Tiền phòng', tong_tien_phong],
            ['Tiền dịch vụ', tong_tien_dv],
            ['Phí phạt', tong_tien_phat]
        ]
        
        # Ghi dữ liệu cho biểu đồ cột
        for row_num, row_data in enumerate(bar_data, 1):
            for col_num, value in enumerate(row_data, 1):
                chart_sheet.cell(row=row_num, column=col_num+3, value=value)  # Bắt đầu từ cột E
        
        bar_cats = Reference(chart_sheet, min_col=5, min_row=2, max_row=4)
        bar_vals = Reference(chart_sheet, min_col=6, min_row=2, max_row=4)
        bar_chart.add_data(bar_vals, titles_from_data=True)
        bar_chart.set_categories(bar_cats)
        bar_chart.style = 11
        
        chart_sheet.add_chart(bar_chart, "H1")
        
        # Tạo sheet tóm tắt
        summary_sheet = writer.book.create_sheet('Tóm tắt')
        
        # Thêm thông tin tóm tắt
        summary_data = [
            ['BÁO CÁO TÓM TẮT HÓA ĐƠN', ''],
            [f'Ngày xuất báo cáo:', now.strftime('%d/%m/%Y %H:%M:%S')],
            [f'Tổng số hóa đơn:', len(ds_hoa_don)],
            ['', ''],
            ['THỐNG KÊ TRẠNG THÁI', ''],
            ['Đã thanh toán:', f'{tong_da_thanh_toan} hóa đơn'],
            ['Hủy (mất cọc):', f'{tong_huy_mat_coc} hóa đơn'],
            ['', ''],
            ['THỐNG KÊ DOANH THU', ''],
            ['Tiền phòng:', f"{tong_tien_phong:,} VNĐ"],
            ['Tiền dịch vụ:', f"{tong_tien_dv:,} VNĐ"],
            ['Phí phạt:', f"{tong_tien_phat:,} VNĐ"],
            ['TỔNG DOANH THU:', f"{tong_doanh_thu:,} VNĐ"],
        ]
        
        # Ghi dữ liệu tóm tắt
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
                if row_num in [1, 5, 9]:  # Header sections
                    cell.font = Font(size=12, bold=True, color="2F7D5A")
                elif row_num == 12:  # Tổng doanh thu
                    cell.font = Font(size=12, bold=True, color="FF6B35")
        
        # Căn chỉnh cột trong sheet tóm tắt
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 30
    
    output.seek(0)
    
    # Tạo response với file Excel
    filename = f"Danh sách hóa đơn_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/xuat-excel-khach-hang')
@login_required
@permission_required('customers.export')
def xuat_excel_khach_hang():
    from io import BytesIO
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    # Lấy tất cả dữ liệu khách hàng
    ds_khach = DatPhong.query.order_by(DatPhong.ngay_nhan.asc()).all()
    
    # Tạo dữ liệu cho Excel
    data = []
    for dp in ds_khach:
        # Lấy danh sách dịch vụ sử dụng
        dich_vu_su_dung = SuDungDichVu.query.filter_by(datphong_id=dp.id).all()
        dich_vu_str = '; '.join([f"{sd.dichvu.ten} (SL: {sd.so_luong}, Thời gian: {sd.thoi_gian.strftime('%d/%m/%Y %H:%M') if sd.thoi_gian else ''})" for sd in dich_vu_su_dung]) if dich_vu_su_dung else 'Không sử dụng dịch vụ'
        
        trang_thai = 'Đã nhận' if dp.trang_thai == 'nhan' else ('Đã thanh toán' if dp.trang_thai == 'da_thanh_toan' else ('Hủy' if dp.trang_thai == 'huy' else 'Chưa nhận'))
        
        data.append({
            'STT': len(data) + 1,
            'Mã đặt phòng': dp.id,
            'Họ tên': dp.khachhang.ho_ten if dp.khachhang else '',
            'CMND': dp.khachhang.cmnd if dp.khachhang else '',
            'SĐT': dp.khachhang.sdt if dp.khachhang else '',
            'Email': dp.khachhang.email if dp.khachhang else '',
            'Địa chỉ': dp.khachhang.dia_chi if dp.khachhang else '',
            'Phòng': dp.phong.ten,
            'Loại phòng': dp.phong.loai.ten if dp.phong.loai else '',
            'Hình thức thuê': 'Theo ngày' if dp.hinh_thuc_thue == 'ngay' else 'Theo giờ',
            'Ngày đặt nhận': dp.ngay_nhan.strftime('%d/%m/%Y %H:%M') if dp.ngay_nhan else '',
            'Ngày đặt trả': dp.ngay_tra.strftime('%d/%m/%Y %H:%M') if dp.ngay_tra else '',
            'Thực tế nhận': dp.thuc_te_nhan.strftime('%d/%m/%Y %H:%M') if dp.thuc_te_nhan else '',
            'Thực tế trả': dp.thuc_te_tra.strftime('%d/%m/%Y %H:%M') if dp.thuc_te_tra else '',
            'Số đêm': dp.so_dem,
            'Trạng thái': trang_thai,
            'Nhân viên check-in': dp.nhanvien.ten if dp.nhanvien else '',
            'Tiền phòng (VNĐ)': dp.tien_phong,
            'Tiền dịch vụ (VNĐ)': dp.tien_dv,
            'Tiền phạt (VNĐ)': dp.tien_phat,
            'Tiền cọc (VNĐ)': dp.tien_coc,
            'Tổng thanh toán (VNĐ)': dp.tong_thanh_toan,
            'Phương thức thanh toán': dp.phuong_thuc_thanh_toan or '',
            'Phương thức cọc': dp.phuong_thuc_coc or '',
            'Cọc đã thanh toán': 'Có' if dp.coc_da_thanh_toan else 'Không',
            'Dịch vụ sử dụng': dich_vu_str,
            'Thời gian tạo': dp.created_at.strftime('%d/%m/%Y %H:%M') if dp.created_at else ''
        })
    
    # Tạo DataFrame
    df = pd.DataFrame(data)
    
    # Tạo file Excel trong memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet dữ liệu chính
        df.to_excel(writer, sheet_name='Danh sách khách hàng', index=False, startrow=3)
        worksheet = writer.sheets['Danh sách khách hàng']
        
        # Thêm tiêu đề
        now = datetime.now()
        worksheet['A1'] = f'DANH SÁCH KHÁCH HÀNG'
        worksheet['A2'] = f'Khách sạn PTIT - Báo cáo khách hàng'
        worksheet['A3'] = f'Ngày xuất báo cáo: {now.strftime("%d/%m/%Y %H:%M:%S")}'
        
        # Styling tiêu đề
        title_font = Font(size=16, bold=True, color="2F7D5A")
        subtitle_font = Font(size=12, italic=True, color="666666")
        
        worksheet['A1'].font = title_font
        worksheet['A2'].font = subtitle_font
        worksheet['A3'].font = Font(size=10, color="888888")
        
        # Merge cells
        worksheet.merge_cells('A1:Y1')
        worksheet.merge_cells('A2:Y2')
        worksheet.merge_cells('A3:Y3')
        
        # Căn chỉnh
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A2'].alignment = Alignment(horizontal='center')
        worksheet['A3'].alignment = Alignment(horizontal='right')
        
        # Header styling
        header_fill = PatternFill(start_color="2F7D5A", end_color="2F7D5A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Border cho toàn bộ bảng
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in worksheet.iter_rows(min_row=4, max_row=len(data)+4, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col_num, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = max(max_length + 2, 10)
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    output.seek(0)
    
    # Tạo response với file Excel
    filename = f"Danh sách khách hàng_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/xuat-excel-lich-su-email')
@login_required
@permission_required('email.logs')
def xuat_excel_lich_su_email():
    from io import BytesIO
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    # Lấy dữ liệu lịch sử email với filter
    status_filter = request.args.get('status', '')
    query = EmailLog.query
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    # Lấy tất cả (không phân trang)
    logs = query.order_by(EmailLog.sent_at.desc()).all()
    
    # Tạo dữ liệu cho Excel
    data = []
    for log in logs:
        data.append({
            'STT': len(data) + 1,
            'ID': log.id,
            'Email người nhận': log.recipient_email,
            'Tên người nhận': log.recipient_name or '',
            'Mẫu email': log.template_key or '',
            'Tiêu đề': log.subject,
            'Trạng thái': 'Thành công' if log.status == 'success' else ('Thất bại' if log.status == 'failed' else 'Đang chờ'),
            'Lỗi': log.error_message or '',
            'Thời gian gửi': log.sent_at.strftime('%d/%m/%Y %H:%M:%S') if log.sent_at else '',
            'Người gửi': log.sender.ten if log.sender else '',
            'Mã đặt phòng': log.datphong_id or '',
            'Tên khách hàng': log.customer.ho_ten if log.customer else '',
            'CMND khách hàng': log.customer.cmnd if log.customer else ''
        })
    
    # Tạo DataFrame
    df = pd.DataFrame(data)
    
    # Tạo file Excel trong memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet dữ liệu chính
        df.to_excel(writer, sheet_name='Lịch sử email', index=False, startrow=3)
        worksheet = writer.sheets['Lịch sử email']
        
        # Thêm tiêu đề
        now = datetime.now()
        worksheet['A1'] = f'LỊCH SỬ GỬI EMAIL'
        worksheet['A2'] = f'Khách sạn PTIT - Báo cáo lịch sử email'
        worksheet['A3'] = f'Ngày xuất báo cáo: {now.strftime("%d/%m/%Y %H:%M:%S")}'
        
        # Styling tiêu đề
        title_font = Font(size=16, bold=True, color="2F7D5A")
        subtitle_font = Font(size=12, italic=True, color="666666")
        
        worksheet['A1'].font = title_font
        worksheet['A2'].font = subtitle_font
        worksheet['A3'].font = Font(size=10, color="888888")
        
        # Merge cells
        worksheet.merge_cells('A1:L1')
        worksheet.merge_cells('A2:L2')
        worksheet.merge_cells('A3:L3')
        
        # Căn chỉnh
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A2'].alignment = Alignment(horizontal='center')
        worksheet['A3'].alignment = Alignment(horizontal='right')
        
        # Header styling
        header_fill = PatternFill(start_color="2F7D5A", end_color="2F7D5A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Border cho toàn bộ bảng
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in worksheet.iter_rows(min_row=4, max_row=len(data)+4, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col_num, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = max(max_length + 2, 10)
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    output.seek(0)
    
    # Tạo response với file Excel
    filename = f"lich_su_email_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/quan-li-dich-vu', methods=['GET', 'POST'])
@login_required
@permission_required('services.manage', 'customers.vouchers')
def quan_li_dich_vu():
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    can_manage_services = current_user.has_permission('services.manage')
    if request.method == 'POST' and not can_manage_services:
        message = 'Ban khong co quyen quan ly dich vu.'
        if wants_json:
            return jsonify({'status': 'error', 'message': message}), 403
        flash(message, 'danger')
        return redirect(url_for('quan_li_dich_vu'))
    if request.method == 'POST':
        try:
            ten = request.form['ten'].strip()
            gia = int(request.form['gia'])
            loai_id = int(request.form['loai_id'])
        except (KeyError, ValueError):
            message = 'Dữ liệu đầu vào không hợp lệ.'
            if wants_json:
                return jsonify({'message': message}), 400
            flash(message, 'danger')
            return redirect(url_for('quan_li_dich_vu'))
        dv = DichVu(ten=ten, gia=gia, loai_id=loai_id)
        db.session.add(dv)
        db.session.commit()
        data = {
            'id': dv.id,
            'ten': dv.ten,
            'gia': dv.gia,
            'loai_id': dv.loai_id,
            'loai_ten': dv.loai.ten if dv.loai else '',
            'message': 'Thêm dịch vụ thành công!'
        }
        if wants_json:
            return jsonify(data), 201
        flash(data['message'], 'success')
        return redirect(url_for('quan_li_dich_vu'))
    discount_percent, expires_days = get_voucher_config()
    return render_template(
        'quan_li_dich_vu.html',
        ds_dv=DichVu.query.order_by(DichVu.loai_id, DichVu.ten).all(),
        ds_loai=DichVuLoai.query.all(),
        dv_edit=None,
        voucher_discount=discount_percent,
        voucher_expires=expires_days,
        can_manage_services=can_manage_services
    )

@app.route('/quan-li-dich-vu/sua/<int:dichvu_id>', methods=['GET', 'POST'])
@app.route('/sua-dich-vu/<int:dichvu_id>', methods=['POST'])
@login_required
@permission_required('services.manage', 'customers.vouchers')
def sua_dich_vu(dichvu_id):
    dv_edit = DichVu.query.get_or_404(dichvu_id)
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    can_manage_services = current_user.has_permission('services.manage')
    if request.method == 'POST' and not can_manage_services:
        message = 'Ban khong co quyen quan ly dich vu.'
        if wants_json:
            return jsonify({'status': 'error', 'message': message}), 403
        flash(message, 'danger')
        return redirect(url_for('quan_li_dich_vu'))
    if request.method == 'POST':
        try:
            ten = request.form['ten'].strip()
            gia = int(request.form['gia'])
            loai_id = int(request.form['loai_id'])
        except (KeyError, ValueError):
            message = 'Dữ liệu đầu vào không hợp lệ.'
            if wants_json:
                return jsonify({'message': message}), 400
            flash(message, 'danger')
            return redirect(url_for('quan_li_dich_vu'))
        dv_edit.ten = ten
        dv_edit.gia = gia
        dv_edit.loai_id = loai_id
        db.session.commit()
        data = {
            'id': dv_edit.id,
            'ten': dv_edit.ten,
            'gia': dv_edit.gia,
            'loai_id': dv_edit.loai_id,
            'loai_ten': dv_edit.loai.ten if dv_edit.loai else '',
            'message': 'Cập nhật dịch vụ thành công!'
        }
        if wants_json:
            return jsonify(data)
        flash(data['message'], 'success')
        return redirect(url_for('quan_li_dich_vu'))
    discount_percent, expires_days = get_voucher_config()
    return render_template(
        'quan_li_dich_vu.html',
        ds_dv=DichVu.query.order_by(DichVu.loai_id, DichVu.ten).all(),
        ds_loai=DichVuLoai.query.all(),
        dv_edit=dv_edit,
        voucher_discount=discount_percent,
        voucher_expires=expires_days,
        can_manage_services=can_manage_services
    )

@app.route('/quan-li-dich-vu/xoa/<int:dichvu_id>', methods=['POST'])
@app.route('/xoa-dich-vu/<int:dichvu_id>', methods=['POST'])
@login_required
@permission_required('services.manage', 'customers.vouchers')
def xoa_dich_vu(dichvu_id):
    dv = DichVu.query.get_or_404(dichvu_id)
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if SuDungDichVu.query.filter_by(dichvu_id=dichvu_id).first():
        message = 'Không thể xóa dịch vụ đang được sử dụng.'
        if wants_json:
            return jsonify({'message': message}), 400
        flash(message, 'danger')
        return redirect(url_for('quan_li_dich_vu'))
    db.session.delete(dv)
    db.session.commit()
    message = 'Đã xóa dịch vụ.'
    if wants_json:
        return jsonify({'message': message, 'id': dichvu_id})
    flash(message, 'success')
    return redirect(url_for('quan_li_dich_vu'))

@app.route('/quan-li-loai-phong', methods=['GET', 'POST'])
@login_required
@permission_required('room_types.manage')
def quan_li_loai_phong():
    if request.method == 'POST':
        ten = request.form.get('ten', '').strip()
        so_nguoi = int(request.form.get('so_nguoi_toi_da', 1))
        gia = int(request.form.get('gia', 0))
        mo_ta = request.form.get('mo_ta', '').strip()
        # Checkbox: nếu được check thì có trong form, không check thì không có
        co_voucher = bool(request.form.get('co_voucher'))
        
        if LoaiPhong.query.filter_by(ten=ten).first():
            flash(f"Tên loại phòng '{ten}' đã tồn tại.", 'danger')
        else:
            loai_moi = LoaiPhong(ten=ten, so_nguoi_toi_da=so_nguoi, gia=gia, mo_ta=mo_ta if mo_ta else None, co_voucher=co_voucher)
            db.session.add(loai_moi)
            db.session.commit()
            flash('Thêm loại phòng mới thành công!', 'success')
        return redirect(url_for('quan_li_loai_phong'))
    return render_template('quan_li_loai_phong.html', ds_loai=LoaiPhong.query.order_by(LoaiPhong.gia).all(), lp_edit=None)

@app.route('/quan-li-loai-phong/sua/<int:loai_id>', methods=['GET', 'POST'])
@login_required
@permission_required('room_types.manage')
def sua_loai_phong(loai_id):
    lp_edit = LoaiPhong.query.get_or_404(loai_id)
    if request.method == 'POST':
        lp_edit.ten = request.form.get('ten', '').strip()
        lp_edit.so_nguoi_toi_da = int(request.form.get('so_nguoi_toi_da', 1))
        lp_edit.gia = int(request.form.get('gia', 0))
        mo_ta = request.form.get('mo_ta', '').strip()
        lp_edit.mo_ta = mo_ta if mo_ta else None
        # Checkbox: nếu được check thì có trong form, không check thì không có
        lp_edit.co_voucher = bool(request.form.get('co_voucher'))
        db.session.commit()
        flash('Cập nhật loại phòng thành công!', 'success')
        return redirect(url_for('quan_li_loai_phong'))
    return render_template('quan_li_loai_phong.html', ds_loai=LoaiPhong.query.order_by(LoaiPhong.gia).all(), lp_edit=lp_edit)

@app.route('/quan-li-loai-phong/xoa/<int:loai_id>', methods=['POST'])
@login_required
@permission_required('room_types.manage')
def xoa_loai_phong(loai_id):
    lp = LoaiPhong.query.get_or_404(loai_id)
    related_rooms = Phong.query.filter_by(loai_id=loai_id).all()

    if related_rooms:
        room_ids = [room.id for room in related_rooms]
        bookings = DatPhong.query.filter(DatPhong.phong_id.in_(room_ids)).all()

        if bookings:
            has_active_booking = any(
                booking.trang_thai in BOOKING_BLOCKING_STATUSES for booking in bookings
            )
            if has_active_booking:
                flash('Không thể xóa loại phòng vì vẫn còn phòng thuộc loại này đang được sử dụng trong các đặt phòng hiện tại.', 'danger')
                return redirect(url_for('quan_li_loai_phong'))

            flash('Không thể xóa loại phòng vì các phòng thuộc loại này đã có lịch sử đặt phòng. Vui lòng lưu trữ thay vì xóa.', 'danger')
            return redirect(url_for('quan_li_loai_phong'))

    deleted_room_count = len(related_rooms)

    for room in related_rooms:
        db.session.delete(room)

    db.session.delete(lp)
    db.session.commit()

    if deleted_room_count:
        flash(f'Đã xóa loại phòng và {deleted_room_count} phòng liên quan.', 'success')
    else:
        flash('Đã xóa loại phòng.', 'success')
    return redirect(url_for('quan_li_loai_phong'))


# ============================================
# QUẢN LÝ PHÒNG - ROOM MANAGEMENT
# ============================================
@app.route('/quan-li-phong')
@login_required
@permission_required('rooms.manage')
def quan_li_phong():
    """
    Hiển thị danh sách tất cả phòng với khả năng thêm/sửa/xóa.
    """
    # Lấy tất cả phòng với thông tin loại phòng và trạng thái
    phongs = Phong.query.options(
        db.joinedload(Phong.loai)
    ).order_by(Phong.ten).all()
    
    # Lấy danh sách loại phòng để hiển thị trong form
    loai_phongs = LoaiPhong.query.order_by(LoaiPhong.ten).all()
    
    # Kiểm tra booking đang hoạt động cho mỗi phòng
    room_booking_status = {}
    for phong in phongs:
        active_booking = DatPhong.query.filter(
            DatPhong.phong_id == phong.id,
            DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES)
        ).first()
        room_booking_status[phong.id] = active_booking
    
    rooms_payload = []
    for phong in phongs:
        loai = getattr(phong, 'loai', None)
        rooms_payload.append({
            'id': phong.id,
            'ten': phong.ten,
            'loai_id': phong.loai_id,
            'loai_ten': loai.ten if loai else None,
            'gia': int(loai.gia) if loai and loai.gia is not None else None,
            'so_nguoi_toi_da': loai.so_nguoi_toi_da if loai else None,
            'trang_thai': phong.trang_thai,
        })

    loai_phongs_payload = [
        {
            'id': loai.id,
            'ten': loai.ten,
            'gia': int(loai.gia) if loai.gia is not None else None,
            'so_nguoi_toi_da': loai.so_nguoi_toi_da,
        }
        for loai in loai_phongs
    ]

    return render_template(
        'quan_li_phong.html',
        phongs=phongs,
        loai_phongs=loai_phongs,
        rooms_payload=rooms_payload,
        loai_phongs_payload=loai_phongs_payload,
        room_booking_status=room_booking_status,
        phong_edit=None
    )


@app.route('/quan-li-phong/them', methods=['POST'])
@login_required
@permission_required('rooms.manage')
def them_phong():
    """Thêm phòng mới."""
    try:
        ten_phong = request.form.get('so_phong', '').strip()  # Form vẫn gửi 'so_phong'
        loai_id = request.form.get('loai_id', type=int)
        trang_thai = request.form.get('trang_thai', 'trong')
        
        # Validation
        if not ten_phong:
            flash('Số phòng không được để trống.', 'danger')
            return redirect(url_for('quan_li_phong'))
        
        if not loai_id:
            flash('Vui lòng chọn loại phòng.', 'danger')
            return redirect(url_for('quan_li_phong'))
        
        # Kiểm tra loại phòng tồn tại
        loai = LoaiPhong.query.get(loai_id)
        if not loai:
            flash('Loại phòng không tồn tại.', 'danger')
            return redirect(url_for('quan_li_phong'))
        
        # Kiểm tra số phòng đã tồn tại
        existing = Phong.query.filter_by(ten=ten_phong).first()
        if existing:
            flash(f'Số phòng "{ten_phong}" đã tồn tại.', 'danger')
            return redirect(url_for('quan_li_phong'))
        
        # Tạo phòng mới
        phong_moi = Phong(
            ten=ten_phong,
            loai_id=loai_id,
            trang_thai=trang_thai
        )
        
        db.session.add(phong_moi)
        db.session.commit()
        
        flash(f'Đã thêm phòng "{ten_phong}" - {loai.ten} thành công!', 'success')
        return redirect(url_for('quan_li_phong'))
        
    except Exception as e:
        app.logger.error(f'Lỗi thêm phòng: {e}')
        db.session.rollback()
        flash('Có lỗi xảy ra khi thêm phòng. Vui lòng thử lại.', 'danger')
        return redirect(url_for('quan_li_phong'))


@app.route('/quan-li-phong/sua/<int:phong_id>', methods=['GET', 'POST'])
@login_required
@permission_required('rooms.manage')
def sua_phong(phong_id):
    """Sửa thông tin phòng."""
    phong = Phong.query.get_or_404(phong_id)
    
    if request.method == 'POST':
        try:
            ten_phong = request.form.get('so_phong', '').strip()  # Form vẫn gửi 'so_phong'
            loai_id = request.form.get('loai_id', type=int)
            trang_thai = request.form.get('trang_thai', 'trong')
            
            # Validation
            if not ten_phong:
                flash('Số phòng không được để trống.', 'danger')
                return redirect(url_for('sua_phong', phong_id=phong_id))
            
            if not loai_id:
                flash('Vui lòng chọn loại phòng.', 'danger')
                return redirect(url_for('sua_phong', phong_id=phong_id))
            
            # Kiểm tra loại phòng tồn tại
            loai = LoaiPhong.query.get(loai_id)
            if not loai:
                flash('Loại phòng không tồn tại.', 'danger')
                return redirect(url_for('sua_phong', phong_id=phong_id))
            
            # Kiểm tra số phòng trùng (ngoại trừ phòng hiện tại)
            existing = Phong.query.filter(
                Phong.ten == ten_phong,
                Phong.id != phong_id
            ).first()
            
            if existing:
                flash(f'Số phòng "{ten_phong}" đã tồn tại.', 'danger')
                return redirect(url_for('sua_phong', phong_id=phong_id))
            
            # Cập nhật thông tin
            phong.ten = ten_phong
            phong.loai_id = loai_id
            phong.trang_thai = trang_thai
            
            db.session.commit()
            
            flash(f'Đã cập nhật phòng "{ten_phong}" thành công!', 'success')
            return redirect(url_for('quan_li_phong'))
            
        except Exception as e:
            app.logger.error(f'Lỗi sửa phòng: {e}')
            db.session.rollback()
            flash('Có lỗi xảy ra khi sửa phòng. Vui lòng thử lại.', 'danger')
            return redirect(url_for('sua_phong', phong_id=phong_id))
    
    # GET request - hiển thị form sửa
    phongs = Phong.query.options(
        db.joinedload(Phong.loai)
    ).order_by(Phong.ten).all()
    loai_phongs = LoaiPhong.query.order_by(LoaiPhong.ten).all()
    
    room_booking_status = {}
    for p in phongs:
        active_booking = DatPhong.query.filter(
            DatPhong.phong_id == p.id,
            DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES)
        ).first()
        room_booking_status[p.id] = active_booking

    rooms_payload = []
    for p in phongs:
        loai = getattr(p, 'loai', None)
        rooms_payload.append({
            'id': p.id,
            'ten': p.ten,
            'loai_id': p.loai_id,
            'loai_ten': loai.ten if loai else None,
            'gia': int(loai.gia) if loai and loai.gia is not None else None,
            'so_nguoi_toi_da': loai.so_nguoi_toi_da if loai else None,
            'trang_thai': p.trang_thai,
        })

    loai_phongs_payload = [
        {
            'id': loai.id,
            'ten': loai.ten,
            'gia': int(loai.gia) if loai.gia is not None else None,
            'so_nguoi_toi_da': loai.so_nguoi_toi_da,
        }
        for loai in loai_phongs
    ]
    
    return render_template(
        'quan_li_phong.html',
        phongs=phongs,
        loai_phongs=loai_phongs,
        rooms_payload=rooms_payload,
        loai_phongs_payload=loai_phongs_payload,
        room_booking_status=room_booking_status,
        phong_edit=phong
    )


@app.route('/quan-li-phong/xoa/<int:phong_id>', methods=['POST'])
@login_required
@permission_required('rooms.manage')
def xoa_phong(phong_id):
    """
    Xóa phòng.
    
    Điều kiện xóa:
    - Không có booking đang hoạt động (dat, nhan, cho_xac_nhan, waiting)
    - Nếu có lịch sử booking cũ (da_huy, da_hoan_thanh) thì cảnh báo nhưng vẫn cho xóa
    """
    try:
        phong = Phong.query.get_or_404(phong_id)
        
        # Kiểm tra booking đang hoạt động
        active_bookings = DatPhong.query.filter(
            DatPhong.phong_id == phong_id,
            DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES)
        ).all()
        
        if active_bookings:
            # Có booking đang hoạt động - không cho xóa
            booking_info = []
            for b in active_bookings[:3]:  # Hiển thị tối đa 3 booking
                booking_info.append(
                    f"{b.khachhang.ho_ten} ({b.trang_thai}, "
                    f"{b.ngay_nhan.strftime('%d/%m/%Y')} - {b.ngay_tra.strftime('%d/%m/%Y')})"
                )
            
            flash(
                f'Không thể xóa phòng "{phong.ten}" vì đang có {len(active_bookings)} booking hoạt động: '
                f'{", ".join(booking_info)}{"..." if len(active_bookings) > 3 else ""}',
                'danger'
            )
            return redirect(url_for('quan_li_phong'))
        
        # Kiểm tra lịch sử booking cũ (chỉ cảnh báo, không chặn)
        historical_bookings = DatPhong.query.filter(
            DatPhong.phong_id == phong_id
        ).count()
        
        # Xóa phòng
        ten_phong = phong.ten
        loai_ten = phong.loai.ten
        
        db.session.delete(phong)
        db.session.commit()
        
        if historical_bookings > 0:
            flash(
                f'Đã xóa phòng "{ten_phong}" ({loai_ten}). '
                f'Lưu ý: Phòng này có {historical_bookings} booking trong lịch sử.',
                'success'
            )
        else:
            flash(f'Đã xóa phòng "{ten_phong}" ({loai_ten}) thành công!', 'success')
        
        return redirect(url_for('quan_li_phong'))
        
    except Exception as e:
        app.logger.error(f'Lỗi xóa phòng: {e}')
        db.session.rollback()
        flash('Có lỗi xảy ra khi xóa phòng. Vui lòng thử lại.', 'danger')
        return redirect(url_for('quan_li_phong'))


@app.route('/xuat-bao-cao/<int:nam>')
@login_required
@permission_required('analytics.revenue')
def xuat_bao_cao_doanh_thu(nam):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, Reference, Series
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
    
    query = db.session.query(
        extract('month', DatPhong.thuc_te_tra).label('thang'),
        func.sum(DatPhong.tien_phong).label('tien_phong'),
        func.sum(DatPhong.tien_dv).label('tien_dv'),
        func.sum(DatPhong.tien_phat).label('tien_phat'),
        func.sum(DatPhong.tong_thanh_toan).label('tong_thanh_toan')
    ).filter(
        DatPhong.trang_thai == 'da_thanh_toan',
        extract('year', DatPhong.thuc_te_tra) == nam
    ).group_by('thang').order_by('thang')
    
    engine = db.session.get_bind()
    if engine is None:
        engine = db.engine
    df = pd.read_sql(query.statement, engine)
    
    if df.empty:
        flash(f'Không có dữ liệu doanh thu cho năm {nam} để xuất báo cáo.', 'warning')
        return redirect(url_for('thong_ke_doanh_thu', nam=nam))
    
    # Điền đầy đủ 12 tháng
    all_months = pd.DataFrame({'thang': range(1, 13)})
    df = pd.merge(all_months, df, on='thang', how='left').fillna(0)
    df[['tien_phong', 'tien_dv', 'tien_phat', 'tong_thanh_toan']] = df[['tien_phong', 'tien_dv', 'tien_phat', 'tong_thanh_toan']].astype(int)
    
    # Tính tổng
    tong_cong = df.sum(numeric_only=True)
    tong_cong['thang'] = 'TỔNG CỘNG'
    
    # Thêm hàng tổng vào cuối
    df_total = pd.DataFrame([tong_cong])
    df = pd.concat([df, df_total], ignore_index=True)
    
    # Đổi tên cột
    df.rename(columns={
        'thang': 'Tháng', 
        'tien_phong': 'Tiền Phòng (VNĐ)', 
        'tien_dv': 'Tiền Dịch Vụ (VNĐ)', 
        'tien_phat': 'Phí Phạt (VNĐ)', 
        'tong_thanh_toan': 'Tổng Doanh Thu (VNĐ)'
    }, inplace=True)
    
    # Tạo file Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet dữ liệu chính
        df.to_excel(writer, sheet_name='Doanh Thu', index=False, startrow=4)
        worksheet = writer.sheets['Doanh Thu']
        
        # Thêm tiêu đề và thông tin
        worksheet['A1'] = f'BÁO CÁO DOANH THU NĂM {nam}'
        worksheet['A2'] = f'Khách sạn PTIT - Thống kê chi tiết theo tháng'
        worksheet['A3'] = f'Ngày xuất báo cáo: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        
        # Styling tiêu đề
        title_font = Font(size=16, bold=True, color="2F7D5A")
        subtitle_font = Font(size=12, italic=True, color="666666")
        header_font = Font(bold=True, color="FFFFFF")
        
        worksheet['A1'].font = title_font
        worksheet['A2'].font = subtitle_font
        worksheet['A3'].font = Font(size=10, color="888888")
        
        # Merge cells cho tiêu đề
        worksheet.merge_cells('A1:E1')
        worksheet.merge_cells('A2:E2')
        worksheet.merge_cells('A3:E3')
        
        # Căn chỉnh tiêu đề
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A2'].alignment = Alignment(horizontal='center')
        worksheet['A3'].alignment = Alignment(horizontal='right')
        
        # Styling header của bảng
        header_fill = PatternFill(start_color="2F7D5A", end_color="2F7D5A", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num in range(1, 6):  # A đến E
            cell = worksheet.cell(row=5, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Styling dữ liệu
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        currency_alignment = Alignment(horizontal='right')
        center_alignment = Alignment(horizontal='center')
        
        for row_num in range(6, len(df) + 6):  # Dữ liệu từ row 6
            # Cột Tháng (A) - căn giữa
            worksheet.cell(row=row_num, column=1).alignment = center_alignment
            worksheet.cell(row=row_num, column=1).border = data_border
            
            # Các cột tiền tệ (B-E) - căn phải
            for col_num in range(2, 6):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = currency_alignment
                cell.border = data_border
                # Format số với dấu phẩy
                if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.number_format = '#,##0'
        
        # Căn chỉnh tự động độ rộng cột
        for col_num, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Điều chỉnh độ rộng tối ưu
            if col_num == 1:  # Cột Tháng
                adjusted_width = max(max_length + 2, 8)
            else:  # Các cột tiền
                adjusted_width = max(max_length + 4, 20)  # Cần rộng hơn cho số tiền
            
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
        
        # Thêm biểu đồ
        chart_sheet = writer.book.create_sheet('Biểu đồ')
        
        # Tạo biểu đồ cột cho doanh thu theo tháng
        chart = BarChart()
        chart.title = f"Doanh thu theo tháng - Năm {nam}"
        chart.y_axis.title = 'Doanh thu (VNĐ)'
        chart.x_axis.title = 'Tháng'
        
        # Dữ liệu cho biểu đồ (loại bỏ hàng tổng cộng)
        chart_data = Reference(worksheet, min_col=5, min_row=5, max_row=len(df)+3, max_col=5)  # Cột Tổng Doanh Thu
        chart_categories = Reference(worksheet, min_col=1, min_row=6, max_row=len(df)+3, max_col=1)  # Cột Tháng
        
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(chart_categories)
        
        # Styling biểu đồ
        chart.style = 10
        chart.height = 15
        chart.width = 25
        
        # Thêm biểu đồ vào sheet
        chart_sheet.add_chart(chart, "A1")
        
        # Thêm sheet tóm tắt
        summary_sheet = writer.book.create_sheet('Tóm tắt')
        
        # Thêm thông tin tóm tắt
        summary_data = [
            ['BÁO CÁO TÓM TẮT DOANH THU', ''],
            [f'Năm:', nam],
            [f'Ngày xuất báo cáo:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ['', ''],
            ['THỐNG KÊ CHI TIẾT', ''],
            ['Tổng doanh thu năm:', f"{tong_cong['tong_thanh_toan']:,} VNĐ"],
            ['Tiền phòng:', f"{tong_cong['tien_phong']:,} VNĐ"],
            ['Tiền dịch vụ:', f"{tong_cong['tien_dv']:,} VNĐ"],
            ['Phí phạt:', f"{tong_cong['tien_phat']:,} VNĐ"],
            ['', ''],
            ['THÁNG CÓ DOANH THU CAO NHẤT', ''],
        ]
        
        # Tìm tháng có doanh thu cao nhất
        max_month = df.iloc[:-1]['Tổng Doanh Thu (VNĐ)'].idxmax() + 1  # +1 vì index bắt đầu từ 0
        max_revenue = df.iloc[max_month-1]['Tổng Doanh Thu (VNĐ)']
        summary_data.extend([
            ['Tháng:', max_month],
            ['Doanh thu:', f"{max_revenue:,} VNĐ"],
        ])
        
        # Ghi dữ liệu tóm tắt
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
                if row_num <= 3:  # Tiêu đề
                    cell.font = Font(size=14, bold=True, color="2F7D5A")
                elif row_num in [5, 10]:  # Header sections
                    cell.font = Font(size=12, bold=True, color="1F4E3D")
        
        # Căn chỉnh cột trong sheet tóm tắt
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 25
    
    output.seek(0)
    return send_file(
        output, 
        as_attachment=True, 
        download_name=f'Báo cáo doanh thu_{nam}.xlsx', 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Chat và api
@app.route('/qr-chat/<int:dat_id>')
@login_required
@permission_required('communications.chat')
def qr_chat(dat_id):
    dp = DatPhong.query.get_or_404(dat_id)
    if not dp.chat_token:
        flash('Phòng này chưa có mã chat. Có lỗi xảy ra.', 'danger')
        return redirect(url_for('nhan_phong'))

    base_url = resolve_public_base_url()
    path = url_for('chat_khach', token=dp.chat_token, _external=False)
    chat_url = urljoin(f"{base_url.rstrip('/')}/", path.lstrip("/"))
    params = {'data': chat_url, 'size': '250x250'}
    qr_code_url = f"https://api.qrserver.com/v1/create-qr-code/?{urlencode(params)}"
    
    return render_template('qr_code_chat.html', dp=dp, qr_code_url=qr_code_url, chat_url=chat_url)

@app.route('/chat/<token>')
def chat_khach(token):
    dp = DatPhong.query.filter_by(chat_token=token, trang_thai='nhan').first_or_404()
    return render_template('chat_khach.html', dp=dp)

@app.route('/tin-nhan')
@login_required
@permission_required('communications.chat')
def tin_nhan():
    conversations = db.session.query(
        DatPhong, 
        func.count(TinNhan.id)
    ).outerjoin(
        TinNhan, 
        (DatPhong.id == TinNhan.datphong_id) & 
        (TinNhan.trang_thai == 'chua_doc') & 
        (TinNhan.nguoi_gui == 'khach')
    ).filter(
        DatPhong.trang_thai == 'nhan',
        DatPhong.id.in_(db.session.query(TinNhan.datphong_id).distinct())
    ).group_by(DatPhong.id).order_by(DatPhong.phong_id).all()
    
    discount_percent, _ = get_voucher_config()
    return render_template('tin_nhan.html', conversations=conversations, voucher_discount=discount_percent)

@app.route('/xoa-hoi-thoai/<int:datphong_id>', methods=['POST'])
@login_required
@permission_required('communications.chat')
@permission_required('chat.delete')
def xoa_hoi_thoai(datphong_id):
    TinNhan.query.filter_by(datphong_id=datphong_id).delete()
    db.session.commit()
    flash('Đã xóa cuộc hội thoại thành công.', 'success')
    return redirect(url_for('tin_nhan'))

@app.route('/api/public/tin-nhan/gui', methods=['POST'])
def api_public_send_message():
    data = request.json
    token = data.get('token')
    noi_dung = data.get('noi_dung')

    dp = DatPhong.query.filter_by(chat_token=token, trang_thai='nhan').first()
    if not dp or not noi_dung:
        return jsonify({'status': 'error', 'message': 'Invalid data'}), 400

    msg = persist_message(dp.id, 'khach', noi_dung)
    payload = serialize_message(msg)

    socketio.emit('new_message_from_guest', {
        'datphong_id': dp.id,
        'phong': dp.phong.ten,
        **payload
    })
    return jsonify({'status': 'success'})

@app.route('/api/tin-nhan/gui', methods=['POST'])
@login_required
@permission_required('communications.chat')
def api_send_message():
    data = request.json
    datphong_id, noi_dung = data.get('datphong_id'), data.get('noi_dung')

    if not datphong_id or not noi_dung:
        return jsonify({'status': 'error', 'message': 'Thiếu thông tin'}), 400

    msg = persist_message(datphong_id, 'nhanvien', noi_dung, user_id=current_user.id)
    payload = serialize_message(msg)
    
    dp = DatPhong.query.get(datphong_id)
    if dp and dp.chat_token:
        socketio.emit('new_message_from_staff', {
            **payload,
            'nguoi_gui': 'staff'
        }, to=dp.chat_token)

    return jsonify({'status': 'success'})


@app.route('/api/public/tin-nhan/gui-file', methods=['POST'])
def api_public_send_file():
    token = request.form.get('token')
    file = request.files.get('file')
    if not token or not file:
        return jsonify({'status': 'error', 'message': 'Thiếu dữ liệu'}), 400

    dp = DatPhong.query.filter_by(chat_token=token, trang_thai='nhan').first()
    if not dp:
        return jsonify({'status': 'error', 'message': 'Phiên chat không hợp lệ'}), 400

    try:
        payload = create_file_message(file, 'guest')
    except ValueError as exc:
        return jsonify({'status': 'error', 'message': str(exc)}), 400

    msg = persist_message(dp.id, 'khach', payload)
    data = serialize_message(msg)
    socketio.emit('new_message_from_guest', {
        'datphong_id': dp.id,
        'phong': dp.phong.ten,
        **data
    })
    return jsonify({'status': 'success', 'message': data})


@app.route('/api/tin-nhan/gui-file', methods=['POST'])
@login_required
@permission_required('communications.chat')
def api_send_file():
    datphong_id = request.form.get('datphong_id', type=int)
    file = request.files.get('file')
    if not datphong_id or not file:
        return jsonify({'status': 'error', 'message': 'Thiếu dữ liệu'}), 400

    dp = DatPhong.query.get(datphong_id)
    if not dp:
        return jsonify({'status': 'error', 'message': 'Đặt phòng không tồn tại'}), 404

    try:
        payload = create_file_message(file, 'staff')
    except ValueError as exc:
        return jsonify({'status': 'error', 'message': str(exc)}), 400

    msg = persist_message(dp.id, 'nhanvien', payload, user_id=current_user.id)
    data = serialize_message(msg)

    if dp.chat_token:
        socketio.emit('new_message_from_staff', {
            **data,
            'nguoi_gui': 'staff'
        }, to=dp.chat_token)

    return jsonify({'status': 'success', 'message': data})

@app.route('/api/public/tin-nhan/<token>')
def api_public_get_messages(token):
    dp = DatPhong.query.filter_by(chat_token=token, trang_thai='nhan').first()
    if not dp: return jsonify({'error': 'Invalid session'}), 404
        
    messages = TinNhan.query.filter_by(datphong_id=dp.id).order_by(TinNhan.thoi_gian.asc()).all()
    return jsonify([serialize_message(m) for m in messages])

@app.route('/api/tin-nhan/<int:datphong_id>')
@login_required
@permission_required('communications.chat')
def api_get_messages(datphong_id):
    messages = TinNhan.query.filter_by(datphong_id=datphong_id).order_by(TinNhan.thoi_gian.asc()).all()
    TinNhan.query.filter_by(datphong_id=datphong_id, nguoi_gui='khach', trang_thai='chua_doc').update({'trang_thai': 'da_doc'})
    db.session.commit()

    return jsonify([serialize_message(m) for m in messages])

@app.route('/api/dat-phong-online/pending-count')
@login_required
@permission_required('bookings.manage_online')
def api_pending_online_count():
    """API tra ve so dat phong online dang cho xac nhan."""
    count = DatPhong.query.join(
        TinNhan,
        db.and_(
            TinNhan.datphong_id == DatPhong.id,
            TinNhan.nguoi_gui == 'khach',
            TinNhan.noi_dung == ONLINE_DEPOSIT_REQUEST_MESSAGE
        )
    ).filter(
        DatPhong.trang_thai == 'cho_xac_nhan'
    ).distinct().count()
    return jsonify({'count': count})

@app.route('/api/tin-nhan/dem-chua-doc')
@login_required
@permission_required('communications.chat')
def api_dem_tin_nhan_chua_doc():
    """API để đếm số tin nhắn chưa đọc từ khách"""
    unread_count = TinNhan.query.join(DatPhong).filter(
        TinNhan.trang_thai == 'chua_doc',
        TinNhan.nguoi_gui == 'khach',
        DatPhong.trang_thai == 'nhan'
    ).count()
    return jsonify({'count': unread_count})


@app.route('/api/public/dich-vu/menu/<token>')
def api_public_menu(token):
    dp = get_active_booking_by_token(token)
    if not dp:
        return jsonify({'error': 'Phiên chat không hợp lệ'}), 404

    categories = []
    for loai in DichVuLoai.query.order_by(DichVuLoai.ten.asc()).all():
        categories.append({
            'id': loai.id,
            'ten': loai.ten,
            'items': [
                {
                    'id': dv.id,
                    'ten': dv.ten,
                    'gia': dv.gia,
                    'gia_text': vnd(dv.gia)
                }
                for dv in DichVu.query.filter_by(loai_id=loai.id).order_by(DichVu.ten.asc()).all()
            ]
        })
    return jsonify({'phong': dp.phong.ten, 'categories': categories})


@app.route('/api/public/dich-vu/history/<token>')
def api_public_service_history(token):
    dp = get_active_booking_by_token(token)
    if not dp:
        return jsonify({'error': 'Phiên chat không hợp lệ'}), 404

    records = SuDungDichVu.query.filter_by(datphong_id=dp.id).order_by(SuDungDichVu.thoi_gian.desc()).all()
    return jsonify([
        {
            'id': r.id,
            'ten': r.dichvu.ten,
            'so_luong': r.so_luong,
            'gia': r.dichvu.gia,
            'tong': r.dichvu.gia * r.so_luong,
            'trang_thai': r.trang_thai,
            'thoi_gian': r.thoi_gian.strftime('%H:%M %d/%m/%Y')
        }
        for r in records
    ])


@app.route('/api/dat-phong/<int:datphong_id>/dich-vu')
@login_required
def api_datphong_services(datphong_id):
    dp = DatPhong.query.get_or_404(datphong_id)
    records = SuDungDichVu.query.filter_by(datphong_id=dp.id).order_by(SuDungDichVu.thoi_gian.desc()).all()
    return jsonify([
        {
            'id': r.id,
            'ten': r.dichvu.ten,
            'so_luong': r.so_luong,
            'gia': r.dichvu.gia,
            'tong': r.dichvu.gia * r.so_luong,
            'trang_thai': r.trang_thai,
            'thoi_gian': r.thoi_gian.strftime('%H:%M %d/%m/%Y')
        }
        for r in records
    ])


@app.route('/api/public/dich-vu/dat', methods=['POST'])
def api_public_order_service():
    data = request.json or {}
    token = data.get('token')
    items = data.get('items', [])
    note = data.get('note', '')

    dp = get_active_booking_by_token(token)
    if not dp:
        return jsonify({'status': 'error', 'message': 'Phiên chat không hợp lệ'}), 404

    if not items:
        return jsonify({'status': 'error', 'message': 'Vui lòng chọn ít nhất một sản phẩm'}), 400

    created = []
    total = 0
    now = datetime.now()
    created_ids = []
    try:
        for item in items:
            dichvu_id = int(item.get('id'))
            so_luong = max(1, int(item.get('so_luong', 1)))
            dichvu = DichVu.query.get(dichvu_id)
            if not dichvu:
                continue
            record = SuDungDichVu(
                datphong_id=dp.id,
                dichvu_id=dichvu.id,
                so_luong=so_luong,
                thoi_gian=now,
                trang_thai='chua_thanh_toan'
            )
            db.session.add(record)
            db.session.flush()  # Get ID before commit
            created.append({'id': record.id, 'ten': dichvu.ten, 'so_luong': so_luong, 'gia': dichvu.gia})
            created_ids.append(record.id)
            total += dichvu.gia * so_luong
        if not created:
            raise ValueError('Không có sản phẩm hợp lệ')
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(exc)}), 400

    # Gửi tin nhắn tự động tới nhân viên
    items_text = ', '.join(f"{c['ten']} x{c['so_luong']}" for c in created)
    message_text = f"Khách đã đặt dịch vụ: {items_text}. Tổng tạm tính {vnd(total)}."
    if note:
        message_text += f" Ghi chú: {note}."
    msg = persist_message(dp.id, 'khach', message_text)
    payload = serialize_message(msg)
    socketio.emit('new_message_from_guest', {
        'datphong_id': dp.id,
        'phong': dp.phong.ten,
        **payload
    })

    description = quote(f"DV {dp.id} {dp.khachhang.ho_ten}")
    qr_code_url = (f"https://img.vietqr.io/image/{BANK_ID}-{BANK_ACCOUNT_NO}-compact2.png"
                   f"?amount={int(total)}&addInfo={description}")

    return jsonify({
        'status': 'success',
        'total': total,
        'total_text': vnd(total),
        'items': created,
        'qr_code_url': qr_code_url
    })


@app.route('/api/public/dich-vu/<int:service_id>/yeu-cau-xac-nhan', methods=['POST'])
def api_public_request_confirmation(service_id):
    """API cho khách hàng gửi yêu cầu xác nhận thanh toán sau khi đã chuyển khoản"""
    try:
        service = SuDungDichVu.query.get_or_404(service_id)
        
        # Kiểm tra service thuộc về booking hợp lệ
        dp = service.datphong
        if not dp or not dp.chat_token:
            return jsonify({
                'status': 'error',
                'message': 'Phiên không hợp lệ'
            }), 404
        
        if service.trang_thai != 'chua_thanh_toan':
            return jsonify({
                'status': 'error',
                'message': 'Đơn hàng này đã được xử lý'
            }), 400
        
        # Chuyển sang trạng thái CHỜ XÁC NHẬN
        service.trang_thai = 'cho_xac_nhan'
        db.session.commit()
        
        # Gửi tin nhắn thông báo cho nhân viên
        msg_text = f"🔔 Khách yêu cầu xác nhận thanh toán: {service.dichvu.ten} x{service.so_luong} = {vnd(service.dichvu.gia * service.so_luong)}"
        msg = persist_message(dp.id, 'he_thong', msg_text)
        payload = serialize_message(msg)
        socketio.emit('new_message_from_guest', {
            'datphong_id': dp.id,
            'phong': dp.phong.ten,
            **payload
        })
        
        # Gửi thông báo realtime cho khách hàng
        if dp.chat_token:
            socketio.emit('order_status_updated', {
                'service_id': service.id,
                'ten': service.dichvu.ten,
                'trang_thai': 'cho_xac_nhan',
                'message': f'Yêu cầu xác nhận "{service.dichvu.ten}" đã được gửi. Vui lòng chờ nhân viên kiểm tra.'
            }, to=dp.chat_token)
        
        return jsonify({
            'status': 'success',
            'message': 'Đã gửi yêu cầu xác nhận',
            'service': {
                'id': service.id,
                'trang_thai': service.trang_thai
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/public/dich-vu/yeu-cau-xac-nhan-nhieu', methods=['POST'])
def api_public_request_confirmation_multiple():
    """API cho khách hàng gửi yêu cầu xác nhận thanh toán cho NHIỀU dịch vụ cùng lúc"""
    try:
        data = request.json or {}
        service_ids = data.get('service_ids', [])
        
        if not service_ids:
            return jsonify({
                'status': 'error',
                'message': 'Không có dịch vụ nào được chọn'
            }), 400
        
        # Lấy tất cả services
        services = SuDungDichVu.query.filter(SuDungDichVu.id.in_(service_ids)).all()
        
        if not services:
            return jsonify({
                'status': 'error',
                'message': 'Không tìm thấy dịch vụ'
            }), 404
        
        # Kiểm tra tất cả services thuộc cùng 1 booking
        dp = services[0].datphong
        if not all(s.datphong_id == dp.id for s in services):
            return jsonify({
                'status': 'error',
                'message': 'Các dịch vụ không thuộc cùng một đơn'
            }), 400
        
        # Kiểm tra token
        if not dp or not dp.chat_token:
            return jsonify({
                'status': 'error',
                'message': 'Phiên không hợp lệ'
            }), 404
        
        # Lọc các service chưa thanh toán
        services_to_confirm = [s for s in services if s.trang_thai == 'chua_thanh_toan']
        
        if not services_to_confirm:
            return jsonify({
                'status': 'error',
                'message': 'Tất cả dịch vụ đã được xử lý'
            }), 400
        
        # Chuyển tất cả sang trạng thái CHỜ XÁC NHẬN
        total_amount = 0
        items_text_parts = []
        for service in services_to_confirm:
            service.trang_thai = 'cho_xac_nhan'
            subtotal = service.dichvu.gia * service.so_luong
            total_amount += subtotal
            items_text_parts.append(f"{service.dichvu.ten} x{service.so_luong}")
        
        db.session.commit()
        
        # Gửi tin nhắn thông báo cho nhân viên
        items_text = ', '.join(items_text_parts)
        msg_text = f"🔔 Khách yêu cầu xác nhận thanh toán: {items_text} = {vnd(total_amount)}"
        msg = persist_message(dp.id, 'he_thong', msg_text)
        payload = serialize_message(msg)
        socketio.emit('new_message_from_guest', {
            'datphong_id': dp.id,
            'phong': dp.phong.ten,
            **payload
        })
        
        # Gửi thông báo realtime cho khách hàng
        if dp.chat_token:
            socketio.emit('order_status_updated', {
                'service_ids': [s.id for s in services_to_confirm],
                'message': f'Yêu cầu xác nhận đơn hàng đã được gửi. Vui lòng chờ nhân viên kiểm tra.'
            }, to=dp.chat_token)
        
        return jsonify({
            'status': 'success',
            'message': 'Đã gửi yêu cầu xác nhận',
            'services': [{
                'id': s.id,
                'trang_thai': s.trang_thai
            } for s in services_to_confirm],
            'total': total_amount
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/phong-theo-loai/<int:loai_id>')
@login_required
def api_phong_theo_loai(loai_id):
    phongs = Phong.query.filter_by(loai_id=loai_id).all()
    return jsonify([{'id': p.id, 'ten': p.ten, 'trang_thai': p.trang_thai} for p in phongs])



def compute_available_rooms(loai_id, ngay_nhan, ngay_tra):
    """Return list of room availability dictionaries for the given range."""
    # Lấy thời gian timeout từ cấu hình
    timeout_setting = HeThongCauHinh.query.filter_by(key='payment_timeout_minutes').first()
    timeout_minutes = int(timeout_setting.value) if timeout_setting and timeout_setting.value else 5
    timeout_seconds = timeout_minutes * 60
    
    phongs = Phong.query.filter_by(loai_id=loai_id).all()
    result = []
    for p in phongs:
        # First, get all overlapping bookings
        overlaps = DatPhong.query.filter(
            DatPhong.phong_id == p.id,
            DatPhong.trang_thai.in_(BOOKING_BLOCKING_STATUSES),
            ~db.or_(DatPhong.ngay_tra <= ngay_nhan, DatPhong.ngay_nhan >= ngay_tra)
        ).all()
        
        # Filter out expired cho_xac_nhan bookings
        valid_overlaps = []
        for overlap in overlaps:
            if overlap.trang_thai == 'cho_xac_nhan':
                # Check if booking has active payment session or is within 5 minutes
                is_valid = False
                if overlap.payment_token:
                    session = PaymentSession.query.filter_by(token=overlap.payment_token).first()
                    if session and not payment_session_expired(session.created_at):
                        is_valid = True
                else:
                    # No payment session yet, check if booking is within 5 minutes
                    if not payment_session_expired(overlap.created_at):
                        is_valid = True
                
                if is_valid:
                    valid_overlaps.append(overlap)
                else:
                    # Mark expired cho_xac_nhan as cancelled
                    overlap.trang_thai = 'huy'
                    db.session.commit()
            else:
                valid_overlaps.append(overlap)
        
        overlap = valid_overlaps[0] if valid_overlaps else None
        
        if overlap:
            if overlap.trang_thai == 'cho_xac_nhan':
                # Phòng đang chờ thanh toán
                is_available = False
                # Tính thời gian còn lại
                remaining_time = None
                if overlap.payment_token:
                    session = PaymentSession.query.filter_by(token=overlap.payment_token).first()
                    if session and not payment_session_expired(session.created_at):
                        remaining_time = max(0, timeout_seconds - (datetime.now() - session.created_at).total_seconds())
                        print(f"DEBUG: Payment session found, remaining_time: {remaining_time}")
                else:
                    # No payment session yet, check if booking is within timeout
                    if not payment_session_expired(overlap.created_at):
                        remaining_time = max(0, timeout_seconds - (datetime.now() - overlap.created_at).total_seconds())
                
                if remaining_time and remaining_time > 0:
                    minutes = int(remaining_time // 60)
                    seconds = int(remaining_time % 60)
                    reason = f"Phòng đang chờ thanh toán ({minutes}:{seconds:02d} còn lại) từ {fmt_dt(overlap.ngay_nhan)} đến {fmt_dt(overlap.ngay_tra)}"
                else:
                    reason = f"Phòng đang chờ thanh toán từ {fmt_dt(overlap.ngay_nhan)} đến {fmt_dt(overlap.ngay_tra)}"
                status = 'cho_thanh_toan'
            else:
                # Phòng đã được đặt hoặc đang ở
                is_available = False
                reason = f"Phòng đã được giữ từ {fmt_dt(overlap.ngay_nhan)} đến {fmt_dt(overlap.ngay_tra)}"
                status = 'da_dat'
        else:
            is_available = True
            reason = ''
            status = 'trong'
            
        result.append({
            'id': p.id,
            'ten': p.ten,
            'trang_thai': status,  # Thay đổi: trả về trạng thái computed thay vì p.trang_thai
            'available': is_available,
            'reason': reason
        })
    return result

@app.route('/api/phong-trong-theo-ngay', methods=['POST'])
@login_required
def api_phong_trong_theo_ngay():
    """
    API kiểm tra phòng trống trong khoảng thời gian cụ thể
    Input: {loai_id: int, ngay_nhan: ISO datetime, ngay_tra: ISO datetime}
    Output: Danh sách phòng với trạng thái available
    """
    try:
        data = request.json
        loai_id = int(data.get('loai_id'))
        ngay_nhan_str = data.get('ngay_nhan')
        ngay_tra_str = data.get('ngay_tra')
        
        if not ngay_nhan_str or not ngay_tra_str:
            return jsonify({'error': 'Thiếu thông tin ngày nhận hoặc ngày trả'}), 400
        
        # Parse datetime
        ngay_nhan = datetime.fromisoformat(ngay_nhan_str.replace('Z', '+00:00'))
        ngay_tra = datetime.fromisoformat(ngay_tra_str.replace('Z', '+00:00'))
        
        # Kiểm tra logic ngày
        if ngay_tra <= ngay_nhan:
            return jsonify({'error': 'Ngày trả phải sau ngày nhận'}), 400
        
        result = compute_available_rooms(loai_id, ngay_nhan, ngay_tra)
        return jsonify(result)
    
    except ValueError as e:
        return jsonify({'error': f'Định dạng ngày không hợp lệ: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Lỗi server: {str(e)}'}), 500

@app.route('/api/public/phong-trong', methods=['POST'])
def api_public_phong_trong():
    """Public API for checking available rooms without authentication."""
    try:
        data = request.get_json(silent=True) or {}
        loai_id = data.get('loai_id')
        ngay_nhan_str = data.get('ngay_nhan')
        ngay_tra_str = data.get('ngay_tra')

        if loai_id is None:
            loai_id = request.form.get('loai_id')
        if ngay_nhan_str is None:
            ngay_nhan_str = request.form.get('ngay_nhan')
        if ngay_tra_str is None:
            ngay_tra_str = request.form.get('ngay_tra')

        loai_id = int(loai_id)

        if not ngay_nhan_str or not ngay_tra_str:
            return jsonify({'error': 'Thiếu thông tin ngày nhận hoặc ngày trả'}), 400

        ngay_nhan = datetime.fromisoformat(ngay_nhan_str.replace('Z', '+00:00'))
        ngay_tra = datetime.fromisoformat(ngay_tra_str.replace('Z', '+00:00'))

        if ngay_tra <= ngay_nhan:
            return jsonify({'error': 'Ngày trả phải sau ngày nhận'}), 400

        rooms = compute_available_rooms(loai_id, ngay_nhan, ngay_tra)
        return jsonify(rooms)
    except ValueError as exc:
        return jsonify({'error': f'Định dạng dữ liệu không hợp lệ: {exc}'}), 400
    except Exception as exc:
        app.logger.exception('Lỗi API phong-trong công khai: %s', exc)
        return jsonify({'error': 'Lỗi máy chủ'}), 500

@app.route('/api/public/validate-voucher', methods=['POST'])
def api_public_validate_voucher():
    """API public để validate mã voucher cho đặt phòng online (không cần đăng nhập)."""
    try:
        data = request.get_json(silent=True) or {}
        code = data.get('code', '').strip().upper()

        if not code:
            return jsonify({
                'valid': False,
                'message': 'Vui lòng nhập mã voucher'
            }), 200

        voucher = Voucher.query.filter_by(code=code, is_used=False).first()
        if not voucher:
            return jsonify({
                'valid': False,
                'message': 'Mã voucher không tồn tại hoặc đã được sử dụng'
            }), 200

        if voucher.expires_at and voucher.expires_at < datetime.now():
            return jsonify({
                'valid': False,
                'message': 'Mã voucher đã hết hạn'
            }), 200

        # Tính discount amount (sẽ được tính lại ở frontend dựa trên giá phòng)
        return jsonify({
            'valid': True,
            'voucher': {
                'id': voucher.id,
                'code': voucher.code,
                'discount_percent': voucher.discount_percent,
                'expires_at': voucher.expires_at.isoformat() if voucher.expires_at else None
            },
            'message': f'Mã voucher hợp lệ! Giảm {voucher.discount_percent}%'
        }), 200

    except Exception as exc:
        app.logger.exception('Lỗi validate voucher: %s', exc)
        return jsonify({
            'valid': False,
            'message': 'Lỗi hệ thống. Vui lòng thử lại sau.'
        }), 500

@app.route('/api/dich-vu/<int:service_id>/xac-nhan-thanh-toan', methods=['POST'])
@login_required
@permission_required('payments.process', 'services.orders')
def api_confirm_service_payment(service_id):
    """API để nhân viên XÁC NHẬN đã nhận tiền từ khách (bước cuối)"""
    try:
        service = SuDungDichVu.query.get_or_404(service_id)
        
        if service.trang_thai == 'da_thanh_toan':
            return jsonify({
                'status': 'error',
                'message': 'Dịch vụ này đã được xác nhận thanh toán trước đó'
            }), 400
        
        # Cập nhật trạng thái thành ĐÃ THANH TOÁN (xác nhận cuối cùng)
        service.trang_thai = 'da_thanh_toan'
        db.session.commit()
        
        # Gửi thông báo qua Socket.IO cho khách hàng
        dp = service.datphong
        if dp and dp.chat_token:
            socketio.emit('payment_confirmed', {
                'service_id': service.id,
                'ten': service.dichvu.ten,
                'message': f'Đã xác nhận thanh toán cho dịch vụ "{service.dichvu.ten}"'
            }, to=dp.chat_token)
        
        return jsonify({
            'status': 'success',
            'message': 'Đã xác nhận nhận tiền thành công',
            'service': {
                'id': service.id,
                'ten': service.dichvu.ten,
                'trang_thai': service.trang_thai
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/dich-vu/<int:service_id>/danh-dau-da-thanh-toan', methods=['POST'])
@login_required
@permission_required('payments.process', 'services.orders')
def api_mark_service_paid(service_id):
    """API KHÔNG CÒN SỬ DỤNG - Khách tự gửi yêu cầu xác nhận"""
    return jsonify({
        'status': 'error',
        'message': 'API này không còn được sử dụng. Khách hàng tự gửi yêu cầu xác nhận.'
    }), 410  # 410 Gone


@app.route('/api/dich-vu/<int:service_id>/huy-yeu-cau', methods=['POST'])
@login_required
@permission_required('payments.process', 'services.orders')
def api_cancel_payment_request(service_id):
    """API để nhân viên HỦY YÊU CẦU xác nhận thanh toán
       Trạng thái: cho_xac_nhan -> chua_thanh_toan
    """
    try:
        service = SuDungDichVu.query.get_or_404(service_id)
        
        if service.trang_thai != 'cho_xac_nhan':
            return jsonify({
                'status': 'error',
                'message': 'Chỉ có thể hủy yêu cầu đang chờ xác nhận'
            }), 400
        
        # Chuyển về trạng thái CHƯA THANH TOÁN
        service.trang_thai = 'chua_thanh_toan'
        db.session.commit()
        
        # Gửi tin nhắn thông báo
        dp = service.datphong
        if dp:
            msg_text = f"❌ Nhân viên đã hủy yêu cầu xác nhận thanh toán: {service.dichvu.ten}. Vui lòng thanh toán lại hoặc liên hệ nhân viên."
            msg = persist_message(dp.id, 'he_thong', msg_text)
            payload = serialize_message(msg)
            socketio.emit('new_message_from_guest', {
                'datphong_id': dp.id,
                'phong': dp.phong.ten,
                **payload
            })
            
            # Gửi thông báo realtime cho khách hàng
            if dp.chat_token:
                socketio.emit('order_status_updated', {
                    'service_id': service.id,
                    'ten': service.dichvu.ten,
                    'trang_thai': 'chua_thanh_toan',
                    'message': f'Yêu cầu xác nhận "{service.dichvu.ten}" đã bị hủy. Vui lòng thanh toán lại.'
                }, to=dp.chat_token)
        
        return jsonify({
            'status': 'success',
            'message': 'Đã hủy yêu cầu xác nhận',
            'service': {
                'id': service.id,
                'trang_thai': service.trang_thai
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/dich-vu/huy-yeu-cau-nhieu', methods=['POST'])
@login_required
@permission_required('payments.process', 'services.orders')
def api_cancel_payment_request_multiple():
    """API để nhân viên HỦY NHIỀU YÊU CẦU xác nhận thanh toán cùng lúc
       Trạng thái: cho_xac_nhan -> chua_thanh_toan
    """
    try:
        data = request.json or {}
        service_ids = data.get('service_ids', [])
        
        if not service_ids:
            return jsonify({
                'status': 'error',
                'message': 'Không có dịch vụ nào được chọn'
            }), 400
        
        # Lấy tất cả services
        services = SuDungDichVu.query.filter(SuDungDichVu.id.in_(service_ids)).all()
        
        if not services:
            return jsonify({
                'status': 'error',
                'message': 'Không tìm thấy dịch vụ'
            }), 404
        
        # Kiểm tra tất cả services đang chờ xác nhận
        services_to_cancel = [s for s in services if s.trang_thai == 'cho_xac_nhan']
        
        if not services_to_cancel:
            return jsonify({
                'status': 'error',
                'message': 'Không có dịch vụ nào đang chờ xác nhận'
            }), 400
        
        # Chuyển tất cả về trạng thái CHƯA THANH TOÁN
        items_text_parts = []
        dp = services_to_cancel[0].datphong
        
        for service in services_to_cancel:
            service.trang_thai = 'chua_thanh_toan'
            items_text_parts.append(f"{service.dichvu.ten} x{service.so_luong}")
        
        db.session.commit()
        
        # Gửi tin nhắn thông báo
        if dp:
            items_text = ', '.join(items_text_parts)
            msg_text = f"❌ Nhân viên đã hủy yêu cầu xác nhận thanh toán: {items_text}. Vui lòng thanh toán lại hoặc liên hệ nhân viên."
            msg = persist_message(dp.id, 'he_thong', msg_text)
            payload = serialize_message(msg)
            socketio.emit('new_message_from_guest', {
                'datphong_id': dp.id,
                'phong': dp.phong.ten,
                **payload
            })
            
            # Gửi thông báo realtime cho khách hàng
            if dp.chat_token:
                socketio.emit('order_status_updated', {
                    'service_ids': [s.id for s in services_to_cancel],
                    'trang_thai': 'chua_thanh_toan',
                    'message': f'Yêu cầu xác nhận đơn hàng đã bị hủy. Vui lòng thanh toán lại.'
                }, to=dp.chat_token)
        
        return jsonify({
            'status': 'success',
            'message': 'Đã hủy yêu cầu xác nhận',
            'services': [{
                'id': s.id,
                'trang_thai': s.trang_thai
            } for s in services_to_cancel]
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/public/dich-vu/thong-tin', methods=['POST'])
def api_public_get_service_info():
    """API lấy thông tin các dịch vụ để thanh toán lại"""
    try:
        data = request.json or {}
        service_ids = data.get('service_ids', [])
        
        if not service_ids:
            return jsonify({
                'status': 'error',
                'message': 'Không có dịch vụ nào'
            }), 400
        
        # Lấy tất cả services
        services = SuDungDichVu.query.filter(SuDungDichVu.id.in_(service_ids)).all()
        
        if not services:
            return jsonify({
                'status': 'error',
                'message': 'Không tìm thấy dịch vụ'
            }), 404
        
        # Tính tổng và tạo danh sách items
        items = []
        total = 0
        dp = services[0].datphong
        
        for service in services:
            items.append({
                'id': service.id,
                'ten': service.dichvu.ten,
                'gia': service.dichvu.gia,
                'so_luong': service.so_luong
            })
            total += service.dichvu.gia * service.so_luong
        
        # Tạo QR code
        description = quote(f"DV {dp.id} {dp.khachhang.ho_ten}")
        qr_code_url = (f"https://img.vietqr.io/image/{BANK_ID}-{BANK_ACCOUNT_NO}-compact2.png"
                       f"?amount={int(total)}&addInfo={description}")
        
        return jsonify({
            'status': 'success',
            'items': items,
            'total': total,
            'qr_code_url': qr_code_url
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/public/dich-vu/huy-don', methods=['POST'])
def api_public_cancel_order():
    """API cho khách hàng HỦY đơn hàng chưa thanh toán"""
    try:
        data = request.json or {}
        service_ids = data.get('service_ids', [])
        
        if not service_ids:
            return jsonify({
                'status': 'error',
                'message': 'Không có dịch vụ nào được chọn'
            }), 400
        
        # Lấy tất cả services
        services = SuDungDichVu.query.filter(SuDungDichVu.id.in_(service_ids)).all()
        
        if not services:
            return jsonify({
                'status': 'error',
                'message': 'Không tìm thấy dịch vụ'
            }), 404
        
        # Kiểm tra tất cả services thuộc cùng 1 booking
        dp = services[0].datphong
        if not all(s.datphong_id == dp.id for s in services):
            return jsonify({
                'status': 'error',
                'message': 'Các dịch vụ không thuộc cùng một đơn'
            }), 400
        
        # CHỈ cho phép hủy dịch vụ chưa thanh toán hoặc chờ xác nhận
        services_to_cancel = [s for s in services if s.trang_thai in ['chua_thanh_toan', 'cho_xac_nhan']]
        
        if not services_to_cancel:
            return jsonify({
                'status': 'error',
                'message': 'Không có dịch vụ nào có thể hủy (đã thanh toán)'
            }), 400
        
        # XÓA các dịch vụ khỏi database
        items_text_parts = []
        for service in services_to_cancel:
            items_text_parts.append(f"{service.dichvu.ten} x{service.so_luong}")
            db.session.delete(service)
        
        db.session.commit()
        
        # Gửi tin nhắn thông báo cho nhân viên
        if dp:
            items_text = ', '.join(items_text_parts)
            msg_text = f"🚫 Khách đã hủy đơn hàng: {items_text}"
            msg = persist_message(dp.id, 'he_thong', msg_text)
            payload = serialize_message(msg)
            socketio.emit('new_message_from_guest', {
                'datphong_id': dp.id,
                'phong': dp.phong.ten,
                **payload
            })
            
            # Thông báo cho staff reload orders
            socketio.emit('order_cancelled', {
                'datphong_id': dp.id,
                'service_ids': [s.id for s in services_to_cancel]
            }, to='staff-dashboard')
        
        return jsonify({
            'status': 'success',
            'message': 'Đã hủy đơn hàng',
            'cancelled_count': len(services_to_cancel)
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/public/dich-vu/chua-thanh-toan/<token>')
def api_public_get_unpaid_orders(token):
    """API lấy danh sách đơn hàng chưa thanh toán của khách"""
    try:
        dp = get_active_booking_by_token(token)
        if not dp:
            return jsonify({
                'status': 'error',
                'message': 'Phiên không hợp lệ'
            }), 404
        
        # Lấy tất cả dịch vụ chưa thanh toán hoặc chờ xác nhận
        orders = SuDungDichVu.query.filter(
            SuDungDichVu.datphong_id == dp.id,
            SuDungDichVu.trang_thai.in_(['chua_thanh_toan', 'cho_xac_nhan'])
        ).order_by(SuDungDichVu.thoi_gian.desc()).all()
        
        result = []
        for order in orders:
            result.append({
                'id': order.id,
                'ten': order.dichvu.ten,
                'gia': order.dichvu.gia,
                'so_luong': order.so_luong,
                'tong': order.dichvu.gia * order.so_luong,
                'trang_thai': order.trang_thai,
                'thoi_gian': order.thoi_gian.strftime('%H:%M %d/%m/%Y') if order.thoi_gian else ''
            })
        
        return jsonify({
            'status': 'success',
            'orders': result,
            'count': len(result)
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/validate-voucher', methods=['POST'])
@app.route('/api/validate-voucher', methods=['POST'])
@login_required
def api_validate_voucher():
    """
    API kiểm tra tính hợp lệ của voucher
    Input: {code: string}
    Output: {valid: bool, discount_percent: int, message: string}
    """
    try:
        data = request.json
        code = data.get('code', '').strip().upper()
        
        if not code:
            return jsonify({
                'valid': False,
                'discount_percent': 0,
                'message': 'Vui lòng nhập mã voucher!'
            }), 400
        
        # Tìm voucher trong database
        voucher = Voucher.query.filter_by(code=code, is_used=False).first()
        
        if not voucher:
            return jsonify({
                'valid': False,
                'discount_percent': 0,
                'message': 'Mã voucher không hợp lệ hoặc đã được sử dụng!'
            }), 200
        
        # Kiểm tra hạn sử dụng
        if voucher.expires_at < datetime.now():
            return jsonify({
                'valid': False,
                'discount_percent': 0,
                'message': 'Mã voucher đã hết hạn!'
            }), 200
        
        # Voucher hợp lệ
        return jsonify({
            'valid': True,
            'discount_percent': voucher.discount_percent,
            'message': f'Áp dụng thành công! Giảm {voucher.discount_percent}%',
            'expires_at': voucher.expires_at.strftime('%d/%m/%Y')
        }), 200
    
    except Exception as e:
        return jsonify({
            'valid': False,
            'discount_percent': 0,
            'message': f'Lỗi: {str(e)}'
        }), 500

@app.route('/api/dat-theo-phong/<int:phong_id>')
@login_required
@permission_required('payments.process', 'services.orders')
def api_dat_theo_phong(phong_id):
    dp = DatPhong.query.filter_by(phong_id=phong_id, trang_thai='nhan').order_by(DatPhong.id.desc()).first()
    if not dp:
        return jsonify({'error': 'Không có đặt phòng hợp lệ'}), 404
    return jsonify({
        'id': dp.id,
        'booking': build_service_booking_payload(dp)
    })

@app.route('/api/dichvu-theo-loai/<int:loai_id>')
@login_required
@permission_required('payments.process', 'services.orders')
def api_dichvu_theo_loai(loai_id):
    dvs = DichVu.query.all() if loai_id == 0 else DichVu.query.filter_by(loai_id=loai_id).all()
    return jsonify([{'id': d.id, 'ten': d.ten, 'gia': d.gia} for d in dvs])

@app.route('/them-dich-vu', methods=['POST'])
@login_required
@permission_required('payments.process', 'services.orders')
def them_dich_vu():
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    dat_raw = (request.form.get('dat_id') or '').strip()
    try:
        dat_id = int(dat_raw)
    except (TypeError, ValueError):
        dat_id = None

    if not dat_id:
        message = 'Vui lòng chọn phòng đang sử dụng trước khi thêm dịch vụ.'
        if wants_json:
            return jsonify({'message': message}), 400
        flash(message, 'danger')
        return redirect(url_for('dich_vu_thanh_toan'))

    dp = DatPhong.query.get_or_404(dat_id)
    if dp.trang_thai != 'nhan':
        message = 'Không thể thêm dịch vụ cho phòng chưa nhận hoặc đã trả.'
        if wants_json:
            return jsonify({'message': message}), 400
        flash(message, 'danger')
        return redirect(url_for('dich_vu_thanh_toan'))

    try:
        dv_id = int(request.form['dv_id'])
    except (KeyError, TypeError, ValueError):
        message = 'Vui lòng chọn dịch vụ hợp lệ.'
        if wants_json:
            return jsonify({'message': message}), 400
        flash(message, 'danger')
        return redirect(url_for('dich_vu_thanh_toan', dat_id=dat_id))

    try:
        so_luong = int(request.form.get('so_luong', '1'))
    except (TypeError, ValueError):
        so_luong = 1

    sd = SuDungDichVu(datphong_id=dat_id, dichvu_id=dv_id, so_luong=max(1, so_luong))
    db.session.add(sd)
    db.session.commit()

    booking_payload = build_service_booking_payload(dp)
    message = 'Đã thêm dịch vụ.'
    if wants_json:
        return jsonify({'message': message, 'booking': booking_payload})
    flash(message, 'success')
    return redirect(url_for('dich_vu_thanh_toan', dat_id=dat_id))

@app.route('/export-luong/<int:nhanvien_id>')
@login_required
def export_luong_nhan_vien(nhanvien_id):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    nv = NguoiDung.query.get_or_404(nhanvien_id)
    
    # Tính lương cho tháng hiện tại
    now = datetime.now()
    start_month = datetime(now.year, now.month, 1)
    next_month = (start_month + timedelta(days=32)).replace(day=1)
    
    # Lấy dữ liệu lương cơ bản
    salary_record = LuongNhanVien.query.filter_by(nguoidung_id=nhanvien_id).first()
    base_salary = salary_record.luong_co_ban if salary_record else 0
    allowance = salary_record.phu_cap if salary_record else 0
    salary_mode = get_salary_mode()
    
    # Tính số ngày công
    work_days = db.session.query(func.count()).filter(
        Attendance.user_id == nhanvien_id,
        Attendance.status == 'approved',
        Attendance.checkin_time >= start_month,
        Attendance.checkin_time < next_month
    ).scalar() or 0
    work_days = int(work_days or 0)

    min_days = get_min_work_days()
    actual_allowance = allowance if work_days >= min_days else 0
    base_effective = compute_effective_base_salary(base_salary, work_days, salary_mode)
    daily_rate = compute_daily_rate(base_salary) if salary_mode == SALARY_MODE_DAILY else 0
    
    # Tính doanh thu tháng
    month_revenue = db.session.query(func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0)).filter(
        DatPhong.nhanvien_id == nhanvien_id,
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).scalar() or 0
    
    # Tính thưởng doanh thu
    tiers = LuongThuongCauHinh.query.order_by(LuongThuongCauHinh.moc_duoi.asc()).all()
    bonus, rate = tinh_thuong_doanh_thu(month_revenue, tiers)
    
    # Tính thưởng top
    # First get all revenues per employee, then find the max
    revenue_subquery = db.session.query(
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('revenue')
    ).filter(
        DatPhong.nhanvien_id.isnot(None),
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).group_by(DatPhong.nhanvien_id).subquery()
    
    top_revenue = db.session.query(func.max(revenue_subquery.c.revenue)).scalar() or 0
    
    top_bonus = 0
    if top_revenue and month_revenue == top_revenue:
        top_bonus = get_top_bonus()
    
    total_salary = base_effective + actual_allowance + bonus + top_bonus

    # Tạo DataFrame
    base_note = f'Lương tháng {now.month}/{now.year}'
    if salary_mode == SALARY_MODE_DAILY and work_days > 0:
        base_note += f' | {work_days} ngày x {daily_rate:,.0f} VNĐ (chia {SALARY_DAILY_DIVISOR})'
    data = {
        'Mục': ['Lương cơ bản', 'Phụ cấp', 'Thưởng doanh thu', 'Thưởng top', 'Tổng lương'],
        'Số tiền (VNĐ)': [base_effective, actual_allowance, bonus, top_bonus, total_salary],
        'Ghi chú': [
            base_note,
            f'Ngày công: {work_days}/{min_days}',
            f'Doanh thu: {vnd(month_revenue)} ({rate*1:.1f}%)',
            'Top doanh thu' if top_bonus > 0 else '',
            ''
        ]
    }
    
    df = pd.DataFrame(data)
    
    # Tạo file Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet chính
        df.to_excel(writer, sheet_name='Bảng lương', index=False, startrow=4)
        worksheet = writer.sheets['Bảng lương']
        
        # Thêm tiêu đề và thông tin
        worksheet['A1'] = f'BẢNG LƯƠNG NHÂN VIÊN'
        worksheet['A2'] = f'Tên: {nv.ten}'
        worksheet['A3'] = f'Tháng: {now.month}/{now.year}'
        worksheet['A4'] = f'Ngày xuất báo cáo: {now.strftime("%d/%m/%Y %H:%M:%S")}'
        
        # Styling tiêu đề
        title_font = Font(size=16, bold=True, color="2F7D5A")
        subtitle_font = Font(size=12, italic=True, color="666666")
        header_font = Font(bold=True, color="FFFFFF")
        
        worksheet['A1'].font = title_font
        worksheet['A2'].font = subtitle_font
        worksheet['A3'].font = subtitle_font
        worksheet['A4'].font = Font(size=10, color="888888")
        
        # Merge cells cho tiêu đề
        worksheet.merge_cells('A1:C1')
        worksheet.merge_cells('A2:C2')
        worksheet.merge_cells('A3:C3')
        worksheet.merge_cells('A4:C4')
        
        # Căn chỉnh tiêu đề
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A2'].alignment = Alignment(horizontal='left')
        worksheet['A3'].alignment = Alignment(horizontal='left')
        worksheet['A4'].alignment = Alignment(horizontal='right')
        
        # Styling header của bảng
        header_fill = PatternFill(start_color="2F7D5A", end_color="2F7D5A", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num in range(1, 4):  # A đến C
            cell = worksheet.cell(row=5, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Styling dữ liệu
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        currency_alignment = Alignment(horizontal='right')
        center_alignment = Alignment(horizontal='center')
        left_alignment = Alignment(horizontal='left')
        
        for row_num in range(6, len(df) + 6):  # Dữ liệu từ row 6
            # Cột Mục (A) - căn trái
            worksheet.cell(row=row_num, column=1).alignment = left_alignment
            worksheet.cell(row=row_num, column=1).border = data_border
            
            # Cột Số tiền (B) - căn phải, format số
            cell_money = worksheet.cell(row=row_num, column=2)
            cell_money.alignment = currency_alignment
            cell_money.border = data_border
            if cell_money.value and isinstance(cell_money.value, (int, float)) and cell_money.value > 0:
                cell_money.number_format = '#,##0'
            
            # Cột Ghi chú (C) - căn trái
            worksheet.cell(row=row_num, column=3).alignment = left_alignment
            worksheet.cell(row=row_num, column=3).border = data_border
        
        # Căn chỉnh tự động độ rộng cột
        for col_num, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Điều chỉnh độ rộng tối ưu
            if col_num == 1:  # Cột Mục
                adjusted_width = max(max_length + 2, 15)
            elif col_num == 2:  # Cột Số tiền
                adjusted_width = max(max_length + 4, 20)
            else:  # Cột Ghi chú
                adjusted_width = max(max_length + 2, 25)
            
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 40)
    
    output.seek(0)

    filename = f'Lương {nv.ten}_{now.month}_{now.year}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export-luong-all')
@login_required
def export_luong_all():
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    staffs = NguoiDung.query.order_by(NguoiDung.ten.asc()).all()
    
    # Tính lương cho tháng hiện tại
    now = datetime.now()
    start_month = datetime(now.year, now.month, 1)
    next_month = (start_month + timedelta(days=32)).replace(day=1)
    
    # Lấy dữ liệu lương
    salary_records = {item.nguoidung_id: item for item in LuongNhanVien.query.all()}
    tiers = LuongThuongCauHinh.query.order_by(LuongThuongCauHinh.moc_duoi.asc()).all()
    salary_mode = get_salary_mode()

    # Tính doanh thu cho tất cả nhân viên
    revenues = {}
    revenue_rows = db.session.query(
        DatPhong.nhanvien_id,
        func.coalesce(func.sum(DatPhong.tong_thanh_toan), 0).label('doanh_thu')
    ).filter(
        DatPhong.nhanvien_id.isnot(None),
        DatPhong.trang_thai == 'da_thanh_toan',
        DatPhong.thuc_te_tra >= start_month,
        DatPhong.thuc_te_tra < next_month
    ).group_by(DatPhong.nhanvien_id).all()
    
    for row in revenue_rows:
        revenues[row.nhanvien_id] = row.doanh_thu
    
    # Tính top revenue
    top_revenue = max(revenues.values()) if revenues else 0
    top_bonus = get_top_bonus()
    
    # Tính work days
    work_days_map = {}
    min_days = get_min_work_days()
    for staff in staffs:
        work_days = db.session.query(func.count()).filter(
            Attendance.user_id == staff.id,
            Attendance.status == 'approved',
            Attendance.checkin_time >= start_month,
            Attendance.checkin_time < next_month
        ).scalar() or 0
        work_days_map[staff.id] = int(work_days or 0)
    
    # Tạo data cho tất cả nhân viên
    data = []
    for staff in staffs:
        record = salary_records.get(staff.id)
        base_monthly = record.luong_co_ban if record else 0
        allowance_base = record.phu_cap if record else 0
        work_days = int(work_days_map.get(staff.id, 0) or 0)
        allowance = allowance_base if work_days >= min_days else 0

        revenue = revenues.get(staff.id, 0)
        bonus, rate = tinh_thuong_doanh_thu(revenue, tiers)
        staff_top_bonus = top_bonus if revenue == top_revenue and top_revenue > 0 else 0
        base_effective = compute_effective_base_salary(base_monthly, work_days, salary_mode)
        total = base_effective + allowance + bonus + staff_top_bonus

        data.append({
            'Tên nhân viên': staff.ten,
            'Lương cơ bản (VNĐ)': base_effective,
            'Phụ cấp (VNĐ)': allowance,
            'Thưởng doanh thu (VNĐ)': bonus,
            'Thưởng top (VNĐ)': staff_top_bonus,
            'Tổng lương (VNĐ)': total,
            'Doanh thu (VNĐ)': revenue,
            'Ngày công': f'{work_days}/{min_days}'
        })
    
    df = pd.DataFrame(data)
    
    # Tính tổng
    tong_cong = df[['Lương cơ bản (VNĐ)', 'Phụ cấp (VNĐ)', 'Thưởng doanh thu (VNĐ)', 'Thưởng top (VNĐ)', 'Tổng lương (VNĐ)', 'Doanh thu (VNĐ)']].sum()
    tong_cong['Tên nhân viên'] = 'TỔNG CỘNG'
    tong_cong['Ngày công'] = ''
    
    # Thêm hàng tổng vào cuối
    df_total = pd.DataFrame([tong_cong])
    df = pd.concat([df, df_total], ignore_index=True)
    
    # Tạo file Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet chính
        df.to_excel(writer, sheet_name='Bảng lương tổng hợp', index=False, startrow=4)
        worksheet = writer.sheets['Bảng lương tổng hợp']
        
        # Thêm tiêu đề và thông tin
        worksheet['A1'] = f'BẢNG LƯƠNG TỔNG HỢP NHÂN VIÊN'
        worksheet['A2'] = f'Khách sạn PTIT - Tháng {now.month}/{now.year}'
        worksheet['A3'] = f'Ngày xuất báo cáo: {now.strftime("%d/%m/%Y %H:%M:%S")}'
        
        # Styling tiêu đề
        title_font = Font(size=16, bold=True, color="2F7D5A")
        subtitle_font = Font(size=12, italic=True, color="666666")
        header_font = Font(bold=True, color="FFFFFF")
        
        worksheet['A1'].font = title_font
        worksheet['A2'].font = subtitle_font
        worksheet['A3'].font = Font(size=10, color="888888")
        
        # Merge cells cho tiêu đề
        worksheet.merge_cells('A1:H1')
        worksheet.merge_cells('A2:H2')
        worksheet.merge_cells('A3:H3')
        
        # Căn chỉnh tiêu đề
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A2'].alignment = Alignment(horizontal='center')
        worksheet['A3'].alignment = Alignment(horizontal='right')
        
        # Styling header của bảng
        header_fill = PatternFill(start_color="2F7D5A", end_color="2F7D5A", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num in range(1, 9):  # A đến H
            cell = worksheet.cell(row=5, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Styling dữ liệu
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        currency_alignment = Alignment(horizontal='right')
        center_alignment = Alignment(horizontal='center')
        left_alignment = Alignment(horizontal='left')
        
        for row_num in range(6, len(df) + 6):  # Dữ liệu từ row 6
            # Cột Tên nhân viên (A) - căn trái
            worksheet.cell(row=row_num, column=1).alignment = left_alignment
            worksheet.cell(row=row_num, column=1).border = data_border
            
            # Các cột tiền tệ (B-G) - căn phải, format số
            for col_num in range(2, 8):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = currency_alignment
                cell.border = data_border
                if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.number_format = '#,##0'
            
            # Cột Ngày công (H) - căn giữa
            worksheet.cell(row=row_num, column=8).alignment = center_alignment
            worksheet.cell(row=row_num, column=8).border = data_border
        
        # Căn chỉnh tự động độ rộng cột
        for col_num, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Điều chỉnh độ rộng tối ưu
            if col_num == 1:  # Cột Tên nhân viên
                adjusted_width = max(max_length + 2, 20)
            elif col_num in [2, 3, 4, 5, 6, 7]:  # Các cột tiền
                adjusted_width = max(max_length + 4, 18)
            else:  # Cột Ngày công
                adjusted_width = max(max_length + 2, 12)
            
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
        
        # Thêm sheet tóm tắt
        summary_sheet = writer.book.create_sheet('Tóm tắt')
        
        # Thêm thông tin tóm tắt
        summary_data = [
            ['BÁO CÁO TÓM TẮT LƯƠNG', ''],
            [f'Tháng:', f'{now.month}/{now.year}'],
            [f'Ngày xuất báo cáo:', now.strftime('%d/%m/%Y %H:%M:%S')],
            [f'Tổng số nhân viên:', len(staffs)],
            ['', ''],
            ['THỐNG KÊ TỔNG QUAN', ''],
            ['Tổng lương cơ bản:', f"{tong_cong['Lương cơ bản (VNĐ)']:,} VNĐ"],
            ['Tổng phụ cấp:', f"{tong_cong['Phụ cấp (VNĐ)']:,} VNĐ"],
            ['Tổng thưởng doanh thu:', f"{tong_cong['Thưởng doanh thu (VNĐ)']:,} VNĐ"],
            ['Tổng thưởng top:', f"{tong_cong['Thưởng top (VNĐ)']:,} VNĐ"],
            ['TỔNG LƯƠNG TOÀN BỘ:', f"{tong_cong['Tổng lương (VNĐ)']:,} VNĐ"],
            ['', ''],
            ['THỐNG KÊ DOANH THU', ''],
            ['Tổng doanh thu toàn bộ:', f"{tong_cong['Doanh thu (VNĐ)']:,} VNĐ"],
        ]
        
        # Ghi dữ liệu tóm tắt
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
                if row_num in [1, 6, 12]:  # Header sections
                    cell.font = Font(size=12, bold=True, color="2F7D5A")
                elif row_num == 11:  # Tổng lương
                    cell.font = Font(size=12, bold=True, color="FF6B35")
        
        # Căn chỉnh cột trong sheet tóm tắt
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 25
    
    output.seek(0)

    filename = f'Lương tổng hợp tháng {now.month} năm {now.year}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========================= SOCKET.IO EVENT HANDLERS =========================
@socketio.on('connect')
def handle_connect():
    print('Client connected')

@socketio.on('join_chat_room')
def handle_join_room(data):
    token = data.get('token')
    if token:
        join_room(token)
        print(f'Client joined room: {token}')

# ========================= ERROR HANDLERS =========================
@app.errorhandler(404)
def handle_not_found(error):
    return render_template('404.html'), 404

# ========================= STATIC FILES OPTIMIZATION =========================
@app.after_request
def add_cache_headers(response):
    if request.path.startswith('/static/'):
        # Disable cache for static files in development
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# ========================= APP RUN =========================
if __name__ == "__main__":
    try:
        socketio.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True, allow_unsafe_werkzeug=True)
    finally:
        scheduler.shutdown()

